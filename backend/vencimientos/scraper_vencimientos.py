"""
Script de scraping para extraer tablas de vencimientos de Estudio del Amo, AGIP y ARBA
"""

import requests
from bs4 import BeautifulSoup
import pandas as pd
from datetime import datetime
from urllib.parse import urlparse
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import time
import re

# URLs a scrapear
URLS = [
    "https://estudiodelamo.com/vencimientos-retenciones-sicore/",
    "https://estudiodelamo.com/vencimientos-autonomos/",
    "https://estudiodelamo.com/vencimientos-iva/",
    "https://estudiodelamo.com/vencimientos-convenio-multilateral-ingresos-brutos/",
    "https://estudiodelamo.com/vencimientos-cargas-sociales-relacion-de-dependencia/",
    "https://estudiodelamo.com/vencimientos-cargas-sociales-servicio-domestico/",
    "https://estudiodelamo.com/vencimientos-impuesto-ganancias-personas-fisicas-bienes-personales-anticipos/",
    "https://estudiodelamo.com/vencimientos-impuesto-a-las-ganancias-personas-juridicas-sociedades-anticipos/",
    # AGIP CABA - Vencimientos Ciudad de Buenos Aires
    "https://www.agip.gob.ar/vencimientos",  # Incluye Regímenes de Información y Recaudación + Contribuyentes Locales CABA
    # ARBA - Vencimientos Provincia de Buenos Aires
    "https://web.arba.gov.ar/vencimientos-contribuyentes-locales",  # IIBB Contribuyentes Locales Provincia BA
]

# Headers para simular un navegador
HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
}


def obtener_nombre_archivo(url):
    """Extrae un nombre de archivo descriptivo de la URL"""
    parsed = urlparse(url)
    
    # Para AGIP CABA, usar nombre específico
    if 'agip.gob.ar' in url:
        return 'agip-caba'
    
    # Para ARBA Provincia BA, usar nombre específico
    if 'arba.gov.ar' in url:
        return 'arba-provincia-ba'
    
    path = parsed.path.strip('/')
    nombre = path.replace('vencimientos-', '').replace('/', '_')
    return nombre if nombre else 'vencimientos'


def es_tabla_valida(df):
    """
    Verifica si una tabla tiene contenido válido
    (no está vacía y tiene al menos 2 filas y 2 columnas)
    """
    if df is None or df.empty:
        return False
    if len(df) < 1:
        return False
    if len(df.columns) < 1:
        return False
    return True


def son_tablas_iguales(df1, df2):
    """
    Compara dos DataFrames para determinar si son iguales
    (mismo contenido, ignorando el orden de las filas)
    """
    try:
        # Normalizar: resetear índices y ordenar columnas
        df1_norm = df1.copy().reset_index(drop=True)
        df2_norm = df2.copy().reset_index(drop=True)
        
        # Comparar forma
        if df1_norm.shape != df2_norm.shape:
            return False
        
        # Comparar columnas
        if list(df1_norm.columns) != list(df2_norm.columns):
            return False
        
        # Comparar contenido (ignorando tipos de datos en comparación de strings)
        df1_str = df1_norm.astype(str)
        df2_str = df2_norm.astype(str)
        
        return df1_str.equals(df2_str)
    except Exception:
        return False


def extraer_agip_datos_estructurados(driver, soup):
    """
    Extrae datos de AGIP desde divs estructurados (no tablas HTML)
    """
    dataframes = []
    
    try:
        # Obtener mes y año actual del calendario
        mes_actual = None
        anio_actual = None
        try:
            mes_elem = soup.find('h2', string=lambda x: x and any(m in str(x) for m in ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']))
            if mes_elem:
                texto_mes = mes_elem.get_text(strip=True)
                # Buscar año en el siguiente elemento
                anio_elem = mes_elem.find_next('span', class_=lambda x: x and 'text-sm' in str(x))
                if anio_elem:
                    anio_actual = anio_elem.get_text(strip=True)
                mes_actual = texto_mes
        except:
            pass
        
        if not mes_actual:
            mes_actual = 'Diciembre 2025'  # Fallback
        else:
            if anio_actual:
                mes_actual = f"{mes_actual} {anio_actual}"
        
        # Buscar todas las secciones (botones)
        secciones = soup.find_all('button', class_=lambda x: x and 'flex justify-center w-full' in str(x))
        
        datos_regimenes = []
        datos_contribuyentes = []
        
        for seccion in secciones:
            texto_boton = seccion.get_text(strip=True)
            
            # Verificar si está expandida (chevron con rotate-180)
            chevron = seccion.find('svg', class_=lambda x: x and 'rotate-180' in str(x))
            if not chevron:
                continue
            
            # Buscar el div contenedor siguiente (hermano siguiente con border-t)
            contenedor = seccion.find_next_sibling('div')
            if not contenedor or 'border-t' not in str(contenedor.get('class', [])):
                continue
            
            # Extraer items dentro del contenedor
            items = contenedor.find_all('div', class_=lambda x: x and 'p-4' in str(x))
            
            for item in items:
                # Extraer descripción (primer div con texto)
                descripcion_elem = item.find('div')
                if descripcion_elem:
                    parrafo = descripcion_elem.find('p')
                    if parrafo:
                        descripcion = parrafo.get_text(strip=True)
                    else:
                        descripcion = descripcion_elem.get_text(strip=True)
                else:
                    continue
                
                # Extraer fecha (div con text-sm text-gray-600)
                fecha_elem = item.find('div', class_=lambda x: x and 'text-sm' in str(x) and 'text-gray-600' in str(x))
                fecha = fecha_elem.get_text(strip=True) if fecha_elem else ''
                
                if descripcion and fecha:
                    if 'Regímenes de Información' in texto_boton or 'Regimenes de Informacion' in texto_boton:
                        # Para Regímenes, crear una fila por cada fecha
                        datos_regimenes.append({
                            'Mes de devengamiento': mes_actual,
                            'Periodo': descripcion,
                            'Fechas de vencimiento según terminación de CUIT': fecha
                        })
                    elif 'Contribuyentes Locales' in texto_boton:
                        # Para Contribuyentes, extraer terminación y crear columnas separadas
                        # El formato es: "Anticipo Nº 11 - Terminación 0 y 1"
                        match_terminacion = re.search(r'Terminación (\d+) y (\d+)', descripcion)
                        if match_terminacion:
                            term1, term2 = match_terminacion.groups()
                            # Determinar qué columna usar según la terminación
                            columna = None
                            if term1 in ['0', '1']:
                                columna = 'Fechas de vencimiento según terminación de CUIT'
                            elif term1 in ['2', '3']:
                                columna = 'Fechas de vencimiento según terminación de CUIT.1'
                            elif term1 in ['4', '5']:
                                columna = 'Fechas de vencimiento según terminación de CUIT.2'
                            elif term1 in ['6', '7']:
                                columna = 'Fechas de vencimiento según terminación de CUIT.3'
                            elif term1 in ['8', '9']:
                                columna = 'Fechas de vencimiento según terminación de CUIT.4'
                            
                            if columna:
                                # Buscar si ya existe una fila para este mes y anticipo
                                anticipo_base = re.sub(r' - Terminación \d+ y \d+.*', '', descripcion)
                                fila_existente = None
                                for dato in datos_contribuyentes:
                                    if dato.get('Mes de devengamiento') == mes_actual and dato.get('Anticipo', '').startswith(anticipo_base.split(' - ')[0] if ' - ' in anticipo_base else anticipo_base):
                                        fila_existente = dato
                                        break
                                
                                if fila_existente:
                                    fila_existente[columna] = fecha
                                else:
                                    nuevo_dato = {
                                        'Mes de devengamiento': mes_actual,
                                        'Anticipo': anticipo_base
                                    }
                                    nuevo_dato[columna] = fecha
                                    datos_contribuyentes.append(nuevo_dato)
        
        # Crear DataFrames
        if datos_regimenes:
            df_regimenes = pd.DataFrame(datos_regimenes)
            # Si hay múltiples fechas, crear una fila por fecha con todas las columnas
            dataframes.append(df_regimenes)
            print(f"    [OK] Regímenes: {len(datos_regimenes)} filas")
        
        if datos_contribuyentes:
            df_contribuyentes = pd.DataFrame(datos_contribuyentes)
            # Asegurar que todas las columnas de terminación existan
            columnas_terminacion = [
                'Fechas de vencimiento según terminación de CUIT',
                'Fechas de vencimiento según terminación de CUIT.1',
                'Fechas de vencimiento según terminación de CUIT.2',
                'Fechas de vencimiento según terminación de CUIT.3',
                'Fechas de vencimiento según terminación de CUIT.4'
            ]
            for col in columnas_terminacion:
                if col not in df_contribuyentes.columns:
                    df_contribuyentes[col] = None
            # Reordenar columnas
            columnas_orden = ['Mes de devengamiento', 'Anticipo'] + columnas_terminacion
            df_contribuyentes = df_contribuyentes[[c for c in columnas_orden if c in df_contribuyentes.columns]]
            dataframes.append(df_contribuyentes)
            print(f"    [OK] Contribuyentes: {len(datos_contribuyentes)} filas")
        
        return dataframes
        
    except Exception as e:
        print(f"  [ERROR] Error extrayendo datos estructurados: {e}")
        import traceback
        traceback.print_exc()
        return []


def extraer_tablas_selenium(url):
    """
    Extrae tablas de sitios dinámicos usando Selenium
    Maneja React SPA y desplegables/accordions que ocultan las tablas
    """
    try:
        print(f"Procesando con Selenium: {url}")
        
        # Configurar Chrome en modo headless
        chrome_options = Options()
        chrome_options.add_argument('--headless')
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')
        chrome_options.add_argument('--disable-gpu')
        chrome_options.add_argument('--window-size=1920,1080')
        chrome_options.add_argument('user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36')
        
        # Inicializar driver
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=chrome_options)
        wait = WebDriverWait(driver, 20)  # Espera hasta 20 segundos
        
        try:
            driver.get(url)
            
            # Para React SPA, esperar a que la página esté completamente cargada
            print("  ⏳ Esperando carga inicial de React SPA...")
            wait.until(lambda d: d.execute_script("return document.readyState") == "complete")
            time.sleep(5)  # Esperar adicional para React
            
            # Para AGIP (React SPA), expandir todos los desplegables
            if 'agip.gob.ar' in url:
                print("  🔍 Buscando y expandiendo desplegables de AGIP (React SPA)...")
                
                # Esperar a que los elementos interactivos se rendericen
                time.sleep(3)
                
                # Estrategia 1: Buscar botones de acordeón/desplegable con múltiples selectores
                selectores_desplegables = [
                    "button[aria-expanded='false']",
                    "button[aria-expanded='true']",  # También intentar los expandidos por si acaso
                    ".accordion-button",
                    ".collapse-toggle",
                    "[data-toggle='collapse']",
                    "[data-bs-toggle='collapse']",
                    ".MuiAccordionSummary-root",
                    "div[role='button'][aria-expanded='false']",
                    "[class*='accordion']",
                    "[class*='collapse']",
                    "[class*='expand']",
                ]
                
                desplegables_expandidos = 0
                elementos_procesados = set()
                
                # Intentar múltiples veces para asegurar que se expandan todos
                for intento in range(3):
                    print(f"    Intento {intento + 1} de expansión...")
                    for selector in selectores_desplegables:
                        try:
                            elementos = driver.find_elements(By.CSS_SELECTOR, selector)
                            for elemento in elementos:
                                try:
                                    # Obtener ID único del elemento para evitar duplicados
                                    elemento_id = elemento.get_attribute('id') or elemento.get_attribute('outerHTML')[:100]
                                    if elemento_id in elementos_procesados:
                                        continue
                                    elementos_procesados.add(elemento_id)
                                    
                                    # Scroll al elemento
                                    driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", elemento)
                                    time.sleep(1)
                                    
                                    # Verificar si es visible y clickeable
                                    if elemento.is_displayed() and elemento.is_enabled():
                                        # Intentar clic con JavaScript (más confiable para React)
                                        driver.execute_script("arguments[0].click();", elemento)
                                        desplegables_expandidos += 1
                                        print(f"      ✓ Elemento expandido (intento {intento + 1})")
                                        time.sleep(2)  # Esperar a que React renderice
                                        
                                        # Esperar a que aparezcan tablas después de expandir
                                        try:
                                            wait.until(EC.presence_of_element_located((By.TAG_NAME, "table")))
                                        except:
                                            pass
                                except Exception as e:
                                    pass
                        except Exception as e:
                            continue
                    
                    # Esperar entre intentos
                    time.sleep(2)
                
                # Estrategia 2: Buscar todos los botones y hacer clic en los relevantes
                print("  🔍 Buscando botones con texto relacionado...")
                try:
                    botones = driver.find_elements(By.TAG_NAME, "button")
                    for boton in botones:
                        try:
                            texto = boton.text.lower()
                            if any(palabra in texto for palabra in ['vencimiento', 'regimen', 'contribuyente', 'información', 'recaudación', 'ver más', 'expandir', 'mostrar', 'desplegar']):
                                if boton.is_displayed():
                                    driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", boton)
                                    time.sleep(1)
                                    driver.execute_script("arguments[0].click();", boton)
                                    desplegables_expandidos += 1
                                    print(f"    ✓ Botón '{boton.text[:40]}...' clickeado")
                                    time.sleep(2)
                        except:
                            pass
                except Exception as e:
                    print(f"    ⚠️  Error buscando botones: {e}")
                
                # Estrategia 3: Buscar divs clickeables
                print("  🔍 Buscando divs clickeables...")
                try:
                    divs_clickeables = driver.find_elements(By.CSS_SELECTOR, "div[role='button'], div[onclick], div[class*='click'], div[class*='toggle']")
                    for div in divs_clickeables[:10]:  # Limitar a 10 para no hacer demasiados clics
                        try:
                            if div.is_displayed():
                                driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", div)
                                time.sleep(1)
                                driver.execute_script("arguments[0].click();", div)
                                desplegables_expandidos += 1
                                print(f"    ✓ Div clickeable expandido")
                                time.sleep(2)
                        except:
                            pass
                except:
                    pass
                
                print(f"  ✓ Total elementos expandidos: {desplegables_expandidos}")
                
                # Esperar a que React renderice todo el contenido
                print("  ⏳ Esperando renderizado completo de React...")
                time.sleep(5)
                
                # Hacer scroll completo de la página para activar lazy loading
                print("  📜 Haciendo scroll completo de la página...")
                driver.execute_script("window.scrollTo(0, 0);")
                time.sleep(1)
                driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                time.sleep(2)
                driver.execute_script("window.scrollTo(0, 0);")
                time.sleep(2)
            
            # Para ARBA, también intentar expandir desplegables
            elif 'arba.gov.ar' in url:
                print("  🔍 Buscando desplegables de ARBA...")
                try:
                    botones = driver.find_elements(By.CSS_SELECTOR, "button[aria-expanded='false'], .accordion-button, [data-toggle='collapse']")
                    for boton in botones:
                        try:
                            driver.execute_script("arguments[0].scrollIntoView(true);", boton)
                            time.sleep(1)
                            driver.execute_script("arguments[0].click();", boton)
                            time.sleep(2)
                        except:
                            pass
                except:
                    pass
                time.sleep(5)
            
            # Esperar a que las tablas estén presentes
            print("  ⏳ Esperando tablas...")
            try:
                wait.until(EC.presence_of_element_located((By.TAG_NAME, "table")))
            except:
                print("  ⚠️  No se encontraron tablas con wait, continuando...")
            
            time.sleep(3)  # Espera final antes de extraer
            
            # Obtener el HTML renderizado después de expandir
            html_content = driver.page_source
            soup = BeautifulSoup(html_content, 'html.parser')
            
            # Para AGIP, usar extracción de datos estructurados
            if 'agip.gob.ar' in url:
                print("  🔍 Extrayendo datos estructurados de AGIP...")
                dataframes = extraer_agip_datos_estructurados(driver, soup)
                if dataframes:
                    print(f"  ✓ {len(dataframes)} tablas extraídas de AGIP")
                    for i, df in enumerate(dataframes, 1):
                        print(f"    Tabla {i}: {df.shape} - {df.columns.tolist()}")
                    return dataframes
                else:
                    print("  ⚠️  No se pudieron extraer datos estructurados de AGIP")
            
            # Para otros sitios, buscar tablas HTML tradicionales
            tablas = soup.find_all('table')
            print(f"  ✓ {len(tablas)} tablas encontradas en el HTML")
            
            if not tablas:
                # Si no hay tablas, buscar en el DOM con Selenium directamente
                print("  🔍 Buscando tablas directamente en el DOM...")
                try:
                    tablas_selenium = driver.find_elements(By.TAG_NAME, "table")
                    print(f"  ✓ {len(tablas_selenium)} tablas encontradas con Selenium")
                    if tablas_selenium:
                        # Forzar renderizado haciendo clic en cada tabla
                        for tabla_sel in tablas_selenium:
                            try:
                                driver.execute_script("arguments[0].scrollIntoView(true);", tabla_sel)
                                time.sleep(1)
                            except:
                                pass
                        time.sleep(2)
                        html_content = driver.page_source
                        soup = BeautifulSoup(html_content, 'html.parser')
                        tablas = soup.find_all('table')
                except Exception as e:
                    print(f"  ⚠️  Error buscando tablas: {e}")
            
            if not tablas:
                print("  ⚠️  No se encontraron tablas. El contenido puede estar en otro formato.")
                return []
            
            dataframes = []
            for i, tabla in enumerate(tablas):
                try:
                    dfs_leidos = pd.read_html(str(tabla))
                    if dfs_leidos:
                        df = dfs_leidos[0]
                        if es_tabla_valida(df):
                            dataframes.append(df)
                            print(f"  ✓ Tabla {i+1} extraída: {df.shape}")
                except Exception as e:
                    print(f"  ⚠️  Error en tabla {i+1}: {e}")
            
            return dataframes
            
        finally:
            driver.quit()
            
    except Exception as e:
        print(f"  ❌ Error con Selenium: {e}")
        import traceback
        traceback.print_exc()
        return []


def extraer_tablas(url):
    """
    Extrae todas las tablas de una URL específica
    Retorna una lista de DataFrames únicos (sin duplicados)
    """
    try:
        # Para AGIP y ARBA, usar Selenium (contenido dinámico)
        if 'agip.gob.ar' in url or 'arba.gov.ar' in url:
            return extraer_tablas_selenium(url)
        
        # Para otros sitios, usar requests normal
        print(f"Procesando: {url}")
        response = requests.get(url, headers=HEADERS, timeout=30)
        response.raise_for_status()
        
        soup = BeautifulSoup(response.content, 'html.parser')
        tablas = soup.find_all('table')
        
        if not tablas:
            print(f"  ⚠️  No se encontraron tablas en {url}")
            return []
        
        dataframes = []
        tablas_procesadas = 0
        
        for i, tabla in enumerate(tablas):
            try:
                # Verificar si esta tabla está anidada dentro de otra tabla ya procesada
                es_anidada = False
                for tabla_anterior in tablas[:i]:
                    if tabla in tabla_anterior.descendants:
                        es_anidada = True
                        break
                
                if es_anidada:
                    print(f"  ⊘ Tabla {i+1} omitida (anidada dentro de otra tabla)")
                    continue
                
                # Intentar leer la tabla con pandas
                dfs_leidos = pd.read_html(str(tabla))
                if not dfs_leidos:
                    continue
                
                df = dfs_leidos[0]
                
                # Validar que la tabla tenga contenido útil
                if not es_tabla_valida(df):
                    print(f"  ⊘ Tabla {i+1} omitida (vacía o sin contenido válido)")
                    continue
                
                # Verificar duplicados comparando con tablas ya extraídas
                es_duplicado = False
                for df_existente in dataframes:
                    if son_tablas_iguales(df, df_existente):
                        es_duplicado = True
                        print(f"  ⊘ Tabla {i+1} omitida (duplicado de tabla anterior)")
                        break
                
                if not es_duplicado:
                    dataframes.append(df)
                    tablas_procesadas += 1
                    print(f"  ✓ Tabla {tablas_procesadas} extraída: {len(df)} filas, {len(df.columns)} columnas")
                
            except Exception as e:
                print(f"  ✗ Error al procesar tabla {i+1}: {str(e)}")
                continue
        
        print(f"  📊 Total: {len(tablas)} tablas encontradas, {len(dataframes)} únicas extraídas")
        return dataframes
    
    except requests.exceptions.RequestException as e:
        print(f"  ✗ Error al obtener la página {url}: {str(e)}")
        return []
    except Exception as e:
        print(f"  ✗ Error inesperado en {url}: {str(e)}")
        return []


def formatear_excel(archivo):
    """
    Aplica formato profesional a un archivo Excel
    """
    try:
        wb = load_workbook(archivo)
        
        # Estilos
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        alignment_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Formatear encabezados (primera fila)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = alignment_center
                cell.border = border_style
            
            # Ajustar ancho de columnas automáticamente
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # Ajustar ancho (mínimo 10, máximo 50)
                adjusted_width = min(max(max_length + 2, 10), 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Formatear celdas de datos
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = border_style
                    # Primera columna alineada a la izquierda, resto centrado
                    if cell.column == 1:
                        cell.alignment = alignment_left
                    else:
                        cell.alignment = alignment_center
            
            # Congelar primera fila (encabezados)
            ws.freeze_panes = 'A2'
            
            # Ajustar altura de la fila de encabezado
            ws.row_dimensions[1].height = 25
        
        wb.save(archivo)
    except Exception as e:
        print(f"  ⚠️  Advertencia: No se pudo aplicar formato completo: {str(e)}")


def guardar_resultados(dataframes, nombre_base):
    """
    Guarda los DataFrames en un archivo Excel formateado
    """
    if not dataframes:
        print(f"  ⚠️  No hay datos para guardar: {nombre_base}")
        return None
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    archivo = f"{nombre_base}_{timestamp}.xlsx"
    
    with pd.ExcelWriter(archivo, engine='openpyxl') as writer:
        for i, df in enumerate(dataframes):
            nombre_hoja = f"Tabla_{i+1}" if len(dataframes) > 1 else "Datos"
            df.to_excel(writer, sheet_name=nombre_hoja, index=False)
    
    # Aplicar formato
    formatear_excel(archivo)
    print(f"  ✓ Guardado en Excel formateado: {archivo}")
    return archivo


def main():
    """Función principal"""
    print("=" * 60)
    print("SCRAPER DE VENCIMIENTOS - ESTUDIO DEL AMO")
    print("=" * 60)
    print(f"Total de URLs a procesar: {len(URLS)}\n")
    
    resultados_totales = {}
    
    for url in URLS:
        nombre = obtener_nombre_archivo(url)
        dataframes = extraer_tablas(url)
        
        if dataframes:
            resultados_totales[nombre] = {
                'url': url,
                'dataframes': dataframes
            }
            # Guardar individualmente con formato
            guardar_resultados(dataframes, nombre)
        print()
    
    # Guardar resumen consolidado con formato
    if resultados_totales:
        print("=" * 60)
        print("RESUMEN DE EXTRACCIÓN")
        print("=" * 60)
        for nombre, datos in resultados_totales.items():
            print(f"{nombre}: {len(datos['dataframes'])} tabla(s) extraída(s)")
        
        # Crear un archivo Excel consolidado con todas las tablas
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        archivo_consolidado = f"vencimientos_consolidado_{timestamp}.xlsx"
        with pd.ExcelWriter(archivo_consolidado, engine='openpyxl') as writer:
            for nombre, datos in resultados_totales.items():
                for i, df in enumerate(datos['dataframes']):
                    nombre_hoja = f"{nombre}_T{i+1}"[:31]  # Excel limita a 31 caracteres
                    df.to_excel(writer, sheet_name=nombre_hoja, index=False)
        
        # Aplicar formato al archivo consolidado
        formatear_excel(archivo_consolidado)
        print(f"\n✓ Archivo consolidado formateado creado: {archivo_consolidado}")
    
    print("\n" + "=" * 60)
    print("PROCESO COMPLETADO")
    print("=" * 60)


if __name__ == "__main__":
    main()

