#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import camelot
import pandas as pd
import re
import os
import sys

def safe_print(texto):
    """Imprimir texto de forma segura en Windows"""
    try:
        mensaje = str(texto)
        # Limpiar caracteres problemáticos
        mensaje = mensaje.encode('utf-8', errors='replace').decode('utf-8', errors='replace')
        # Reemplazar emojis comunes
        mensaje = mensaje.replace('\U0001f504', '')
        mensaje = mensaje.replace('\U0001f4ca', '')
        mensaje = mensaje.replace('\U0001f4c4', '')
        mensaje = mensaje.replace('\u2705', 'OK')
        mensaje = mensaje.replace('\u274c', 'ERROR')
        mensaje = mensaje.replace('\u26a0', 'WARNING')
        mensaje = mensaje.replace('\U0001f4b0', '')
        # Intentar imprimir con codificación segura
        try:
            print(mensaje)
        except UnicodeEncodeError:
            # Si aún falla, usar ASCII completamente
            print(mensaje.encode('ascii', errors='replace').decode('ascii'))
    except Exception:
        # Si todo falla, imprimir solo texto ASCII
        try:
            print(str(texto).encode('ascii', errors='replace').decode('ascii'))
        except Exception:
            print("Error al imprimir mensaje")

class ExtractorBancoDelSolV1:
    """Extractor de tablas para Banco del Sol usando Camelot - Solo extrae tablas"""
    
    def __init__(self):
        pass
    
    def extraer_datos(self, pdf_path, excel_path=None):
        """Extraer datos del PDF de Banco del Sol usando Camelot para extraer tablas"""
        try:
            safe_print(f"Extrayendo tablas del PDF: {pdf_path}")
            
            # Extraer todas las tablas del PDF usando Camelot
            # Usar método 'lattice' para tablas con bordes definidos
            tables = camelot.read_pdf(pdf_path, pages='all', flavor='lattice')
            
            if not tables:
                safe_print("No se encontraron tablas con bordes definidos, intentando con método 'stream'...")
                # Si no encuentra tablas con lattice, probar con stream
                tables = camelot.read_pdf(pdf_path, pages='all', flavor='stream')
            
            if not tables:
                safe_print("No se encontraron tablas en el PDF")
                return pd.DataFrame()
            
            safe_print(f"Se encontraron {len(tables)} tablas")
            
            # Procesar todas las tablas encontradas
            todas_las_tablas = []
            for i, table in enumerate(tables):
                safe_print(f"Procesando tabla {i+1}/{len(tables)}")
                df_tabla = table.df
                
                # Verificar si es una tabla real (no texto descriptivo)
                if self._es_tabla_estructurada(df_tabla):
                    # Limpiar y procesar la tabla
                    df_limpia = self._procesar_tabla_camelot(df_tabla, i+1)
                    if not df_limpia.empty:
                        todas_las_tablas.append(df_limpia)
                        safe_print(f"  OK Tabla {i+1}: {len(df_limpia)} filas procesadas")
                else:
                    safe_print(f"  WARNING Tabla {i+1} descartada (texto descriptivo): {df_tabla.shape}")
            
            if not todas_las_tablas:
                safe_print("No se pudieron procesar las tablas")
                return pd.DataFrame()
            
            # Combinar todas las tablas en una sola
            df_final = pd.concat(todas_las_tablas, ignore_index=True)
            
            # Limpiar datos finales
            df_final = self._limpiar_datos_finales(df_final)
            
            # Guardar Excel con múltiples hojas
            if excel_path:
                self._guardar_excel_multiples_hojas(df_final, excel_path)
            
            safe_print(f"Total de registros extraídos: {len(df_final)}")
            return df_final
            
        except Exception as e:
            safe_print(f"Error extrayendo datos con Camelot: {e}")
            return pd.DataFrame()
    
    def _es_tabla_estructurada(self, df):
        """Verificar si es una tabla estructurada (no texto descriptivo)"""
        try:
            if df.empty:
                return False
            
            # Verificar dimensiones mínimas
            if df.shape[0] < 2 or df.shape[1] < 2:
                return False
            
            # Obtener todo el texto de la tabla
            texto_completo = ' '.join([str(valor) for valor in df.values.flatten() if pd.notna(valor)]).lower()
            
            # Indicadores de texto descriptivo/legal que debemos descartar
            indicadores_texto_descriptivo = [
                'garantía', 'depósito', 'acuerdo', 'impuesto', 'cheque', 'físico',
                'consultar', 'navegando', 'correspondiente', 'consulta',
                'entidad', 'debe', 'informar', 'regulador', 'usuario', 'estándar',
                'moneda', 'extranjera', 'garantizados', 'hasta',
                'conjunta', 'exclusiones', 'interés', 'alto', 'adquiridos',
                'endoso', 'partes', 'relacionadas', 'ley', 'decreto', 'comunicación',
                'legales', 'intercambio', 'información', 'giro', 'descubierto',
                'solicitud', 'echeqs', 'fondos', 'comunes',
                'inversión', 'unidad', 'financiera', 'aviso', 'importante',
                'disponibles', 'horas', 'día', 'hábil', 'siguiente',
                'características', 'confirmar', 'período', 'días', 'calendario',
                'entidad', 'destruir', 'originales', 'reproducciones',
                'diferido', 'electrónicamente', 'detalles', 'pendientes', 'banca',
                'línea', 'navegando', 'correspondiente', 'consulta'
            ]
            
            # Contar coincidencias con texto descriptivo
            coincidencias_texto = sum(1 for indicador in indicadores_texto_descriptivo if indicador in texto_completo)
            
            # Si tiene muchas coincidencias con texto descriptivo, descartar
            if coincidencias_texto >= 8:
                return False
            
            # Verificar que tenga datos estructurados (números, fechas, montos)
            tiene_datos_estructurados = False
            
            for col in df.columns:
                col_data = df[col].dropna()
                if len(col_data) == 0:
                    continue
                
                # Verificar si hay números, fechas o montos en la columna
                for valor in col_data.head(10):  # Revisar primeras 10 filas
                    valor_str = str(valor).strip()
                    
                    # Verificar números con decimales
                    if re.match(r'^\d+[.,]\d+$', valor_str):
                        tiene_datos_estructurados = True
                        break
                    # Verificar fechas
                    if re.match(r'^\d{1,2}/\d{1,2}(?:/\d{2,4})?$', valor_str):
                        tiene_datos_estructurados = True
                        break
                    # Verificar montos con $
                    if re.match(r'^\$?\s*\d+[.,]\d+$', valor_str):
                        tiene_datos_estructurados = True
                        break
                    # Verificar números enteros largos
                    if re.match(r'^\d+$', valor_str) and len(valor_str) > 2:
                        tiene_datos_estructurados = True
                        break
                
                if tiene_datos_estructurados:
                    break
            
            return tiene_datos_estructurados
            
        except Exception as e:
            safe_print(f"Error verificando si es tabla estructurada: {e}")
            return False
    
    def _procesar_tabla_camelot(self, df_tabla, numero_tabla):
        """Procesar tabla extraída por Camelot - Solo tablas"""
        try:
            if df_tabla.empty:
                return pd.DataFrame()
            
            safe_print(f"  Tabla {numero_tabla} - Dimensiones: {df_tabla.shape}")
            
            # Solo limpiar valores básicos
            df_limpia = df_tabla.copy()
            
            # Remover filas completamente vacías
            df_limpia = df_limpia.dropna(how='all')
            
            # Remover columnas completamente vacías
            df_limpia = df_limpia.dropna(axis=1, how='all')
            
            # Agregar identificador de tabla
            df_limpia['Numero_Tabla'] = numero_tabla
            
            return df_limpia
        
        except Exception as e:
            safe_print(f"Error procesando tabla {numero_tabla}: {e}")
            return pd.DataFrame()
    
    def _limpiar_datos_finales(self, df):
        """Limpiar datos finales - Solo limpieza básica"""
        try:
            if df.empty:
                return df
            
            # Solo remover filas completamente vacías
            df = df.dropna(how='all')
            
            return df
            
        except Exception as e:
            safe_print(f"Error limpiando datos finales: {e}")
            return df
    
    def _guardar_excel_multiples_hojas(self, df, excel_path):
        """Guardar Excel con múltiples hojas: tablas, movimientos consolidados y totales por concepto"""
        try:
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                # 1. Hoja original con todas las tablas
                df.to_excel(writer, sheet_name='Tablas Extraidas', index=False)
                worksheet_original = writer.sheets['Tablas Extraidas']
                self._aplicar_formato_basico(worksheet_original, df)
                
                # 2. Hoja de Movimientos Consolidados
                df_consolidado = self._crear_movimientos_consolidados(df)
                if not df_consolidado.empty:
                    df_consolidado.to_excel(writer, sheet_name='Movimientos Consolidados', index=False)
                    worksheet_consolidado = writer.sheets['Movimientos Consolidados']
                    self._aplicar_formato_consolidado(worksheet_consolidado, df_consolidado)
                
                # 3. Hoja de Totales por Concepto
                df_totales = self._crear_totales_por_concepto(df_consolidado)
                if not df_totales.empty:
                    df_totales.to_excel(writer, sheet_name='Totales por Concepto', index=False)
                    worksheet_totales = writer.sheets['Totales por Concepto']
                    self._aplicar_formato_totales(worksheet_totales, df_totales)
            
            safe_print(f"Excel guardado con múltiples hojas: {excel_path}")
            
        except Exception as e:
            safe_print(f"Error guardando Excel: {e}")
    
    def _aplicar_formato_basico(self, worksheet, df):
        """Aplicar formato básico a la hoja de Excel"""
        try:
            from openpyxl.styles import Font, Alignment, Border, Side
            
            # Configurar anchos de columna automáticamente
            for col in range(1, len(df.columns) + 1):
                worksheet.column_dimensions[chr(64 + col)].width = 15
            
            # Aplicar formato a encabezados
            header_font = Font(bold=True, size=11)
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Aplicar formato a datos
            data_font = Font(size=10)
            data_alignment = Alignment(horizontal='left', vertical='center')
            
            # Bordes
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in range(1, len(df) + 2):  # +2 para incluir header
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.border = thin_border
                    cell.font = data_font
                    cell.alignment = data_alignment
                    
        except Exception as e:
            safe_print(f"Error aplicando formato: {e}")
    
    def _crear_movimientos_consolidados(self, df):
        """Crear hoja de movimientos consolidados"""
        try:
            if df.empty:
                return pd.DataFrame()
            
            # Buscar columnas que contengan datos de movimientos bancarios
            df_consolidado = pd.DataFrame()
            
            # Mapear columnas comunes de movimientos bancarios
            mapeo_columnas = {
                'fecha': 'Fecha',
                'comprobante': 'Origen', 
                'movimiento': 'Descripcion',
                'debito': 'Importe',
                'credito': 'Importe',
                'saldo': 'Saldo',
                'débito': 'Importe',
                'crédito': 'Importe',
                'saldo en cuenta': 'Saldo'
            }
            
            # Buscar y mapear columnas
            for col_original in df.columns:
                if col_original in ['Numero_Tabla']:
                    continue
                
                col_original_lower = str(col_original).lower().strip()
                for patron, col_estandar in mapeo_columnas.items():
                    if patron in col_original_lower:
                        df_consolidado[col_estandar] = df[col_original]
                        break
            
            # Si no se encontraron mapeos, usar estructura original
            if df_consolidado.empty:
                # Intentar usar las primeras columnas como datos
                columnas_importantes = []
                for col in df.columns:
                    if col not in ['Numero_Tabla'] and len(columnas_importantes) < 6:
                        columnas_importantes.append(col)
                
                if columnas_importantes:
                    df_consolidado = df[columnas_importantes].copy()
                    # Renombrar columnas genéricamente
                    nombres_estandar = ['Fecha', 'Origen', 'Descripcion', 'Importe', 'Saldo', 'Movimiento']
                    for i, col in enumerate(df_consolidado.columns):
                        if i < len(nombres_estandar):
                            df_consolidado = df_consolidado.rename(columns={col: nombres_estandar[i]})
            
            # Limpiar datos
            df_consolidado = df_consolidado.dropna(how='all')
            
            # Agregar columna de clasificación si no existe
            if 'Clasificacion' not in df_consolidado.columns:
                df_consolidado['Clasificacion'] = self._clasificar_movimientos(df_consolidado)
            
            safe_print(f"Movimientos consolidados: {len(df_consolidado)} registros")
            return df_consolidado
            
        except Exception as e:
            safe_print(f"Error creando movimientos consolidados: {e}")
            return pd.DataFrame()
    
    def _clasificar_movimientos(self, df):
        """Clasificar movimientos por tipo"""
        try:
            clasificaciones = []
            
            for _, row in df.iterrows():
                descripcion = str(row.get('Descripcion', '')).lower()
                concepto = str(row.get('Concepto', '')).lower()
                
                # Clasificar por palabras clave
                if any(palabra in descripcion for palabra in ['impuesto', 'ley', '25.413']):
                    clasificaciones.append('Impuestos')
                elif any(palabra in descripcion for palabra in ['sircreb', 'regimen', 'recaudacion']):
                    clasificaciones.append('SIRCREB')
                elif any(palabra in descripcion for palabra in ['transferencia', 'transfer']):
                    clasificaciones.append('Transferencias')
                elif any(palabra in descripcion for palabra in ['cheque', 'cheques']):
                    clasificaciones.append('Cheques')
                elif any(palabra in descripcion for palabra in ['tarjeta', 'tarjeta de credito']):
                    clasificaciones.append('Tarjetas')
                elif any(palabra in descripcion for palabra in ['cobro', 'cobros']):
                    clasificaciones.append('Cobros')
                elif any(palabra in descripcion for palabra in ['pago', 'pagos']):
                    clasificaciones.append('Pagos')
                elif any(palabra in descripcion for palabra in ['comision', 'comisiones']):
                    clasificaciones.append('Comisiones')
                else:
                    clasificaciones.append('Otros')
            
            return clasificaciones
            
        except Exception as e:
            safe_print(f"Error clasificando movimientos: {e}")
            return ['Otros'] * len(df)
    
    def _crear_totales_por_concepto(self, df_consolidado):
        """Crear hoja de totales por concepto - Balance de cuenta usando datos consolidados"""
        try:
            if df_consolidado.empty:
                return pd.DataFrame()
            
            # Buscar saldo inicial en la columna 'Saldo' (primera fila con saldo válido)
            saldo_inicial = 0
            total_debitos = 0
            total_creditos = 0
            
            # Buscar saldo inicial en todas las columnas
            for _, row in df_consolidado.iterrows():
                # Buscar en la columna 'Movimiento' primero (donde está el saldo inicial real)
                movimiento_valor = row.get('Movimiento', '')
                if movimiento_valor and str(movimiento_valor).strip() != '':
                    # Verificar si contiene "saldo inicial" o es un monto muy grande (probable saldo inicial)
                    valor_str = str(movimiento_valor).lower()
                    if 'saldo inicial' in valor_str:
                        monto = self._extraer_monto_de_texto(str(movimiento_valor))
                        if monto is not None and monto > 0:
                            saldo_inicial = monto
                            safe_print(f"Saldo inicial encontrado en Movimiento: ${saldo_inicial:,.2f}")
                            break
                
                # Si no se encontró, buscar en la columna 'Saldo'
                if saldo_inicial == 0:
                    saldo_valor = row.get('Saldo', '')
                    if saldo_valor and str(saldo_valor).strip() != '':
                        monto = self._extraer_monto_de_texto(str(saldo_valor))
                        if monto is not None and monto > 0:
                            saldo_inicial = monto
                            safe_print(f"Saldo inicial encontrado en Saldo: ${saldo_inicial:,.2f}")
                            break
            
            # Sumar débitos y créditos de las columnas 'Importe' y 'Movimiento'
            for _, row in df_consolidado.iterrows():
                # Buscar débitos en la columna 'Importe'
                importe_valor = row.get('Importe', '')
                if importe_valor and str(importe_valor).strip() != '':
                    monto = self._extraer_monto_de_texto(str(importe_valor))
                    if monto is not None and monto > 0:
                        # Verificar si es débito o crédito basado en el contexto
                        descripcion = str(row.get('Descripcion', '')).lower()
                        if any(palabra in descripcion for palabra in ['debito', 'débito', 'pago', 'transferencia', 'impuesto']):
                            total_debitos += monto
                        elif any(palabra in descripcion for palabra in ['credito', 'crédito', 'deposito', 'cobro']):
                            total_creditos += monto
                        else:
                            # Por defecto asumir débito
                            total_debitos += monto
                
                # Buscar créditos en la columna 'Movimiento'
                movimiento_valor = row.get('Movimiento', '')
                if movimiento_valor and str(movimiento_valor).strip() != '':
                    monto = self._extraer_monto_de_texto(str(movimiento_valor))
                    if monto is not None and monto > 0:
                        # Verificar si es débito o crédito basado en el contexto
                        descripcion = str(row.get('Descripcion', '')).lower()
                        if any(palabra in descripcion for palabra in ['debito', 'débito', 'pago', 'transferencia', 'impuesto']):
                            total_debitos += monto
                        elif any(palabra in descripcion for palabra in ['credito', 'crédito', 'deposito', 'cobro']):
                            total_creditos += monto
                        else:
                            # Por defecto asumir débito
                            total_debitos += monto
            
            # Calcular saldo final
            saldo_final = saldo_inicial - total_debitos + total_creditos
            
            # Crear DataFrame con balance de cuenta
            datos_totales = [
                {
                    'Concepto': 'Saldo Inicial',
                    'Monto': round(saldo_inicial, 2)
                },
                {
                    'Concepto': 'Total Débitos',
                    'Monto': round(total_debitos, 2)
                },
                {
                    'Concepto': 'Total Créditos',
                    'Monto': round(total_creditos, 2)
                },
                {
                    'Concepto': 'Saldo Final',
                    'Monto': round(saldo_final, 2)
                }
            ]
            
            df_totales = pd.DataFrame(datos_totales)
            
            safe_print(f"Balance: Saldo Inicial: ${saldo_inicial:,.2f}, Débitos: ${total_debitos:,.2f}, Créditos: ${total_creditos:,.2f}, Saldo Final: ${saldo_final:,.2f}")
            return df_totales
            
        except Exception as e:
            safe_print(f"Error creando totales por concepto: {e}")
            return pd.DataFrame()
    
    def _extraer_monto_de_texto(self, texto):
        """Extraer monto numérico de un texto"""
        try:
            if not texto or texto == 'nan':
                return None
            
            # Limpiar texto
            texto_limpio = str(texto).strip()
            
            # Solo buscar montos que tengan el símbolo $ o que estén claramente identificados como montos
            if '$' not in texto_limpio and not any(palabra in texto_limpio.lower() for palabra in ['debito', 'débito', 'credito', 'crédito', 'saldo']):
                return None
            
            # Buscar patrones específicos de montos bancarios
            # Formato argentino: $ 1.234.567,89
            patron_monto_argentino = r'\$\s*(\d{1,3}(?:\.\d{3})*,\d{2})'
            match = re.search(patron_monto_argentino, texto_limpio)
            if match:
                numero_str = match.group(1).replace('.', '').replace(',', '.')
                return float(numero_str)
            
            # Formato simple: $ 1234567,89
            patron_monto_simple = r'\$\s*(\d+,\d{2})'
            match = re.search(patron_monto_simple, texto_limpio)
            if match:
                numero_str = match.group(1).replace(',', '.')
                return float(numero_str)
            
            # Solo números con decimales en formato argentino (sin $)
            patron_decimal = r'(\d{1,3}(?:\.\d{3})*,\d{2})'
            match = re.search(patron_decimal, texto_limpio)
            if match:
                numero_str = match.group(1).replace('.', '').replace(',', '.')
                # Solo aceptar si el número es razonable (menos de 100 millones)
                if float(numero_str) < 100000000:
                    return float(numero_str)
            
            return None
            
        except Exception as e:
            return None
    
    def _aplicar_formato_consolidado(self, worksheet, df):
        """Aplicar formato a la hoja de movimientos consolidados"""
        try:
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            
            # Configurar anchos de columna
            column_widths = {
                'A': 12,  # Fecha
                'B': 15,  # Origen
                'C': 40,  # Descripcion
                'D': 15,  # Importe
                'E': 15,  # Saldo
                'F': 15,  # Movimiento
                'G': 20,  # Clasificacion
            }
            
            for col, width in column_widths.items():
                if col in worksheet.column_dimensions:
                    worksheet.column_dimensions[col].width = width
            
            # Aplicar formato a encabezados
            header_font = Font(bold=True, size=11)
            header_fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Aplicar formato a datos
            data_font = Font(size=10)
            data_alignment = Alignment(horizontal='left', vertical='center')
            money_alignment = Alignment(horizontal='right', vertical='center')
            
            # Bordes
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in range(1, len(df) + 2):
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.border = thin_border
                    cell.font = data_font
                    
                    # Alineación específica por columna
                    if col in [4, 5]:  # Columnas de montos
                        cell.alignment = money_alignment
                    else:
                        cell.alignment = data_alignment
                        
        except Exception as e:
            safe_print(f"Error aplicando formato consolidado: {e}")
    
    def _aplicar_formato_totales(self, worksheet, df):
        """Aplicar formato a la hoja de totales por concepto"""
        try:
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            
            # Configurar anchos de columna simples
            column_widths = {
                'A': 20,  # Concepto
                'B': 20,  # Monto
            }
            
            for col, width in column_widths.items():
                if col in worksheet.column_dimensions:
                    worksheet.column_dimensions[col].width = width
            
            # Aplicar formato a encabezados
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Aplicar formato a datos
            data_font = Font(size=11)
            data_alignment = Alignment(horizontal='left', vertical='center')
            money_alignment = Alignment(horizontal='right', vertical='center')
            total_font = Font(bold=True, size=12)
            
            # Bordes
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in range(1, len(df) + 2):
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.border = thin_border
                    
                    # Formato especial para fila de saldo final
                    if row == len(df) + 1:  # Última fila (Saldo Final)
                        cell.font = total_font
                        cell.fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
                    
                    # Alineación específica
                    if col == 2:  # Columna de monto
                        cell.alignment = money_alignment
                    else:
                        cell.alignment = data_alignment
                        
        except Exception as e:
            safe_print(f"Error aplicando formato totales: {e}")

# Función principal para compatibilidad
def extraer_datos_banco_del_sol_v1(pdf_path, excel_path=None):
    """Función principal para extraer datos de Banco del Sol V1"""
    extractor = ExtractorBancoDelSolV1()
    return extractor.extraer_datos(pdf_path, excel_path)








