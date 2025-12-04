#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import os
import tempfile
import importlib.util
from pathlib import Path
import sys
import requests
import threading
import uuid
import pandas as pd
from datetime import datetime
import glob
import re

# Logging para diagnóstico
import logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    stream=sys.stdout
)
logger = logging.getLogger(__name__)

# Información de inicio
logger.info("=" * 60)
logger.info("SERVIDOR DE EXTRACTORES DE BANCOS - INICIANDO")
logger.info("=" * 60)
logger.info(f"Python Version: {sys.version}")
logger.info(f"Working Directory: {os.getcwd()}")
logger.info(f"Variable PORT: {os.environ.get('PORT', 'NO DEFINIDA')}")
logger.info(f"Variable EXTRACTOR_PORT: {os.environ.get('EXTRACTOR_PORT', 'NO DEFINIDA')}")

# Crear aplicación Flask
app = Flask(__name__)

# Configurar CORS - permitir todos los orígenes en producción
# Flask-CORS manejará los headers automáticamente, no agregar manualmente
CORS(app, resources={r"/*": {
    "origins": "*",
    "methods": ["GET", "POST", "PUT", "DELETE", "OPTIONS"],
    "allow_headers": ["Content-Type", "Authorization", "ngrok-skip-browser-warning", "User-Agent"],
    "expose_headers": ["Content-Type"],
    "supports_credentials": True
}})

logger.info("Flask app creada correctamente")
logger.info("CORS configurado para permitir todos los orígenes")

# Middleware para logging de todas las peticiones
@app.before_request
def log_request_info():
    logger.info(f"Request recibido: {request.method} {request.path}")
    logger.info(f"Origin: {request.headers.get('Origin', 'No Origin')}")
    if request.method == 'POST':
        logger.info(f"Form data keys: {list(request.form.keys())}")
        logger.info(f"Files: {list(request.files.keys())}")

@app.after_request
def log_response_info(response):
    logger.info(f"Response enviado: {response.status_code} para {request.path}")
    # Flask-CORS ya maneja los headers CORS, no agregar manualmente para evitar duplicados
    return response

@app.route('/', methods=['GET'])
def root():
    """Endpoint raíz"""
    logger.info("Request recibido en endpoint raíz")
    try:
        response = jsonify({
            'message': 'Servidor de Extractores de Bancos',
            'status': 'running',
            'endpoints': {
                'health': '/health',
                'extractors': '/extractors',
                'extract': '/extract',
                'google_client_id': '/api/google/client-id',
                'google_oauth_token': '/api/google/oauth/token',
                'google_oauth_refresh': '/api/google/oauth/refresh'
            }
        })
        logger.info("Response enviado desde endpoint raíz")
        return response, 200
    except Exception as e:
        logger.error(f"Error en endpoint raíz: {str(e)}", exc_info=True)
        return jsonify({'error': str(e)}), 500

# Directorio de extractores
EXTRACTORES_DIR = Path(__file__).parent / 'extractores'
TEMP_DIR = Path(tempfile.gettempdir()) / 'extractores_temp'
# Crear directorio temporal si no existe (con permisos completos)
try:
    TEMP_DIR.mkdir(parents=True, exist_ok=True)
    logger.info(f"Directorio temporal creado/verificado: {TEMP_DIR}")
except Exception as e:
    logger.error(f"Error al crear directorio temporal: {e}")
    # Fallback a directorio en la carpeta del proyecto
    TEMP_DIR = Path(__file__).parent / 'temp'
    TEMP_DIR.mkdir(parents=True, exist_ok=True)
    logger.info(f"Usando directorio temporal alternativo: {TEMP_DIR}")

logger.info(f"Directorio de extractores: {EXTRACTORES_DIR}")
logger.info(f"Directorio temporal: {TEMP_DIR}")
logger.info(f"Directorio de extractores existe: {EXTRACTORES_DIR.exists()}")

# Mapeo de bancos a sus extractores
logger.info("Cargando configuración de extractores...")
BANCO_EXTRACTORS = {
    'banco_galicia': {
        'script': 'extractor_banco_galicia.py',
        'function': 'extraer_datos_banco_galicia'
    },
    'banco_galicia_mas': {
        'script': 'extractor_banco_galicia_mas.py',
        'function': 'extraer_datos_banco_galicia_mas'
    },
    'mercado_pago': {
        'script': 'extractor_mercado_pago_directo.py',
        'function': 'extraer_datos_mercado_pago_directo'
    },
    'banco_comafi': {
        'script': 'extractor_banco_comafi.py',
        'function': 'extraer_datos_banco_comafi'
    },
    'banco_jpmorgan': {
        'script': 'extractor_banco_jpmorgan.py',
        'function': 'extraer_datos_banco_jpmorgan'
    },
    'banco_bind': {
        'script': 'extractor_banco_bind.py',
        'function': 'extraer_datos_banco_bind'
    },
    'banco_supervielle': {
        'script': 'extractor_banco_supervielle.py',
        'function': 'extraer_datos_banco_supervielle'
    },
    'banco_cabal': {
        'script': 'extractor_banco_cabal.py',
        'function': 'extraer_datos_banco_cabal'
    },
    'banco_credicoop': {
        'script': 'extractor_banco_credicoop_v3.py',
        'function': 'extraer_datos_banco_credicoop'
    },
    'banco_cmf': {
        'script': 'extractor_banco_cmf.py',
        'function': 'extraer_datos_banco_cmf'
    },
    'banco_santander': {
        'script': 'extractor_santander_simple.py',
        'function': 'extraer_datos_santander_v3'
    },
    'banco_del_sol': {
        'script': 'extractor_banco_del_sol_v1.py',
        'function': 'extraer_datos_banco_del_sol_v1'
    },
    'banco_ciudad': {
        'script': 'extractor_banco_ciudad.py',
        'function': 'extraer_datos_banco_ciudad'
    },
    'banco_bbva': {
        'script': 'extractor_bbva_mejorado.py',
        'function': 'extraer_datos_bbva'
    },
    'banco_icbc': {
        'script': 'extractor_banco_icbc.py',
        'function': 'extraer_datos_banco_icbc'
    },
    'banco_macro': {
        'script': 'extractor_banco_macro.py',
        'function': 'extraer_datos_banco_macro'
    },
    'banco_nacion': {
        'script': 'nacion.py',
        'function': 'extraer_datos_banco_nacion'
    },
    'colppy': {
        'script': 'Colppy.py',
        'function': 'extraer_datos_colppy'
    },
}

logger.info(f"Extractores configurados: {len(BANCO_EXTRACTORS)}")
logger.info("Aplicación Flask inicializada correctamente")

# ==================== SISTEMA DE VENCIMIENTOS ====================
# Usar ruta absoluta para evitar problemas con el directorio de trabajo
VENCIMIENTOS_DIR = Path(__file__).parent.resolve() / 'vencimientos'
VENCIMIENTOS_DIR.mkdir(parents=True, exist_ok=True)
logger.info(f"Directorio de vencimientos: {VENCIMIENTOS_DIR}")
logger.info(f"Directorio de vencimientos existe: {VENCIMIENTOS_DIR.exists()}")

# Sistema de jobs asíncronos para vencimientos
vencimientos_jobs = {}
vencimientos_jobs_lock = threading.Lock()

def get_vencimientos_job(job_id):
    """Obtiene el estado de un job de vencimientos"""
    with vencimientos_jobs_lock:
        return vencimientos_jobs.get(job_id)

def update_vencimientos_job(job_id, status, progress=0, message='', error=None):
    """Actualiza el estado de un job de vencimientos"""
    with vencimientos_jobs_lock:
        if job_id in vencimientos_jobs:
            vencimientos_jobs[job_id].update({
                'status': status,
                'progress': progress,
                'message': message,
                'error': error,
                'updated_at': datetime.now().isoformat()
            })

def ejecutar_scraper_vencimientos(job_id):
    """Ejecuta el scraper de vencimientos en segundo plano"""
    try:
        update_vencimientos_job(job_id, 'processing', 10, 'Iniciando scraper...')
        
        # Usar rutas absolutas para evitar problemas con el directorio de trabajo
        scraper_path = VENCIMIENTOS_DIR.resolve() / 'scraper_vencimientos.py'
        
        logger.info(f"Ejecutando scraper desde: {scraper_path}")
        logger.info(f"Directorio de trabajo actual: {os.getcwd()}")
        
        if not scraper_path.exists():
            raise FileNotFoundError(f"Scraper no encontrado: {scraper_path}")
        
        # Guardar el directorio de trabajo original
        original_cwd = os.getcwd()
        logger.info(f"Directorio original guardado: {original_cwd}")
        
        try:
            # Cambiar al directorio de vencimientos para que el scraper guarde los archivos ahí
            vencimientos_dir_abs = str(VENCIMIENTOS_DIR.resolve())
            os.chdir(vencimientos_dir_abs)
            logger.info(f"Cambiado al directorio: {vencimientos_dir_abs}")
            
            # Agregar el directorio al path de Python
            vencimientos_dir_str = str(VENCIMIENTOS_DIR.resolve())
            if vencimientos_dir_str not in sys.path:
                sys.path.insert(0, vencimientos_dir_str)
            
            # Importar y ejecutar el scraper usando ruta absoluta
            spec = importlib.util.spec_from_file_location(
                'scraper_vencimientos', 
                str(scraper_path.resolve())
            )
            if spec is None or spec.loader is None:
                raise ImportError(f"No se pudo cargar el módulo scraper desde {scraper_path}")
            
            module = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(module)
            
            update_vencimientos_job(job_id, 'processing', 30, 'Ejecutando scraper...')
            
            # Ejecutar la función main del scraper
            if hasattr(module, 'main'):
                module.main()
            else:
                raise AttributeError("El scraper no tiene una función 'main()'")
            
            update_vencimientos_job(job_id, 'processing', 90, 'Finalizando...')
            
            # Buscar el archivo consolidado más reciente usando ruta absoluta
            archivos_consolidados = list(VENCIMIENTOS_DIR.resolve().glob('vencimientos_consolidado_*.xlsx'))
            if archivos_consolidados:
                archivo_mas_reciente = max(archivos_consolidados, key=lambda p: p.stat().st_mtime)
                update_vencimientos_job(
                    job_id, 
                    'completed', 
                    100, 
                    f'Scraper completado. Archivo: {archivo_mas_reciente.name}',
                    None
                )
            else:
                update_vencimientos_job(job_id, 'completed', 100, 'Scraper completado', None)
                
        finally:
            # Restaurar el directorio de trabajo original
            try:
                os.chdir(original_cwd)
                logger.info(f"Directorio restaurado a: {original_cwd}")
            except Exception as e:
                logger.warning(f"Error al restaurar directorio de trabajo: {e}")
            
    except Exception as e:
        logger.error(f"Error ejecutando scraper de vencimientos: {str(e)}", exc_info=True)
        update_vencimientos_job(job_id, 'error', 0, f'Error: {str(e)}', str(e))

def encontrar_archivo_consolidado_mas_reciente():
    """Encuentra el archivo consolidado más reciente"""
    # Usar ruta absoluta para evitar problemas con el directorio de trabajo
    archivos = list(VENCIMIENTOS_DIR.resolve().glob('vencimientos_consolidado_*.xlsx'))
    if not archivos:
        return None
    return max(archivos, key=lambda p: p.stat().st_mtime)

def extraer_ultimo_digito_cuil(cuil):
    """Extrae el último dígito de un CUIL (formato XX-XXXXXXXX-X o XXXXXXXXXXX)"""
    # Remover guiones y espacios
    cuil_limpio = str(cuil).replace('-', '').replace(' ', '').strip()
    if not cuil_limpio:
        return None
    # Retornar el último carácter
    return cuil_limpio[-1] if len(cuil_limpio) > 0 else None

def encontrar_columna_fecha_por_digito(columnas, fila, ultimo_digito):
    """Encuentra la columna que contiene la fecha correspondiente al último dígito del CUIL"""
    if ultimo_digito is None:
        return None
    
    # Buscar columnas de fechas por terminación
    columnas_fechas = [col for col in columnas if 'vencimiento según terminación' in str(col).lower() or 'terminación' in str(col).lower()]
    
    # Mapeo de dígitos a rangos comunes
    # 0-1-2-3 -> primera columna
    # 4-5-6 -> segunda columna  
    # 7-8-9 -> tercera columna
    digito_int = int(ultimo_digito) if ultimo_digito.isdigit() else None
    
    if digito_int is not None:
        if digito_int in [0, 1, 2, 3]:
            # Buscar primera columna de fechas (sin .1 o .2)
            for col in columnas_fechas:
                if '.1' not in str(col) and '.2' not in str(col):
                    # Verificar que esta columna contenga el rango correcto
                    valor = fila.get(col, '')
                    if pd.notna(valor):
                        valor_str = str(valor).strip()
                        # Verificar si contiene el dígito en el rango
                        if any(d in valor_str for d in ['0', '1', '2', '3']) or '0-1-2-3' in valor_str or '0, 1, 2' in valor_str:
                            return col
            # Si no se encuentra específica, usar la primera
            if columnas_fechas:
                primera = [c for c in columnas_fechas if '.1' not in str(c) and '.2' not in str(c)]
                if primera:
                    return primera[0]
        
        elif digito_int in [4, 5, 6]:
            # Buscar segunda columna de fechas (.1)
            for col in columnas_fechas:
                if '.1' in str(col):
                    valor = fila.get(col, '')
                    if pd.notna(valor):
                        valor_str = str(valor).strip()
                        if any(d in valor_str for d in ['4', '5', '6']) or '4-5-6' in valor_str or '4, 5, 6' in valor_str:
                            return col
            # Si no se encuentra específica, usar la que tiene .1
            segunda = [c for c in columnas_fechas if '.1' in str(c)]
            if segunda:
                return segunda[0]
        
        elif digito_int in [7, 8, 9]:
            # Buscar tercera columna de fechas (.2)
            for col in columnas_fechas:
                if '.2' in str(col):
                    valor = fila.get(col, '')
                    if pd.notna(valor):
                        valor_str = str(valor).strip()
                        if any(d in valor_str for d in ['7', '8', '9']) or '7-8-9' in valor_str or '7, 8, 9' in valor_str:
                            return col
            # Si no se encuentra específica, usar la que tiene .2
            tercera = [c for c in columnas_fechas if '.2' in str(c)]
            if tercera:
                return tercera[0]
    
    # Fallback: buscar en todas las columnas de fechas
    for col in columnas_fechas:
        valor = fila.get(col, '')
        if pd.notna(valor):
            valor_str = str(valor).strip()
            if ultimo_digito in valor_str:
                return col
    
    return None

def buscar_por_ultimo_digito(df, ultimo_digito):
    """Busca filas en un DataFrame donde alguna columna contenga el último dígito"""
    if ultimo_digito is None:
        return pd.DataFrame()
    
    filas_coincidentes = []
    filas_procesadas = set()  # Para evitar duplicados
    
    # Buscar en columnas de fechas por terminación primero (más específico)
    columnas_fechas_terminacion = [col for col in df.columns if 'vencimiento según terminación' in str(col).lower() or 'terminación' in str(col).lower()]
    
    # También buscar en columnas que puedan contener rangos de dígitos
    columnas_digitos = []
    for col in df.columns:
        col_str = str(col).lower()
        if any(term in col_str for term in ['dígito', 'digito', 'verificador', 'dv', 'd.v.', 'terminación', 'terminacion']):
            columnas_digitos.append(col)
    
    # Buscar en todas las columnas
    todas_columnas = list(df.columns)
    
    for idx, row in df.iterrows():
        if idx in filas_procesadas:
            continue
            
        coincidencia_encontrada = False
        
        # Primero buscar en columnas de fechas por terminación
        for col in columnas_fechas_terminacion:
            valor = row[col]
            if pd.notna(valor):
                valor_str = str(valor).strip()
                # Buscar si contiene el dígito o el rango que incluye el dígito
                if ultimo_digito in valor_str:
                    # Verificar rangos comunes
                    digito_int = int(ultimo_digito) if ultimo_digito.isdigit() else None
                    if digito_int is not None:
                        # Verificar si el dígito está en el rango de esta columna
                        if digito_int in [0, 1, 2, 3] and ('.1' not in str(col) and '.2' not in str(col)):
                            filas_coincidentes.append(idx)
                            filas_procesadas.add(idx)
                            coincidencia_encontrada = True
                            break
                        elif digito_int in [4, 5, 6] and '.1' in str(col):
                            filas_coincidentes.append(idx)
                            filas_procesadas.add(idx)
                            coincidencia_encontrada = True
                            break
                        elif digito_int in [7, 8, 9] and '.2' in str(col):
                            filas_coincidentes.append(idx)
                            filas_procesadas.add(idx)
                            coincidencia_encontrada = True
                            break
                        # Si contiene el dígito directamente
                        elif ultimo_digito in valor_str:
                            filas_coincidentes.append(idx)
                            filas_procesadas.add(idx)
                            coincidencia_encontrada = True
                            break
        
        # Si no se encontró en columnas de fechas, buscar en otras columnas
        if not coincidencia_encontrada:
            for col in todas_columnas:
                if col in columnas_fechas_terminacion:
                    continue  # Ya las procesamos
                    
                valor = row[col]
                if pd.notna(valor):
                    valor_str = str(valor).strip()
                    # Buscar si el último dígito del valor coincide
                    if valor_str and len(valor_str) > 0:
                        # Extraer solo dígitos del final
                        digitos_finales = ''.join(filter(str.isdigit, valor_str[-3:]))
                        if digitos_finales and digitos_finales[-1] == ultimo_digito:
                            filas_coincidentes.append(idx)
                            filas_procesadas.add(idx)
                            break
    
    if filas_coincidentes:
        return df.loc[filas_coincidentes]
    return pd.DataFrame()

@app.route('/vencimientos/listar', methods=['GET'])
def vencimientos_listar():
    """Lista los vencimientos disponibles desde el archivo consolidado más reciente (compartido para todos los usuarios)"""
    try:
        logger.info(f"Buscando archivo consolidado en: {VENCIMIENTOS_DIR.resolve()}")
        archivo = encontrar_archivo_consolidado_mas_reciente()
        
        if not archivo:
            logger.warning(f"No se encontró archivo consolidado en {VENCIMIENTOS_DIR.resolve()}")
            return jsonify({
                'success': False,
                'message': 'No se encontró ningún archivo consolidado de vencimientos. Ejecuta "Refrescar Vencimientos" para generar los datos.',
                'data': {}
            }), 404
        
        logger.info(f"Archivo consolidado encontrado: {archivo}")
        
        # Leer el archivo Excel
        excel_file = pd.ExcelFile(archivo)
        fecha_modificacion = datetime.fromtimestamp(archivo.stat().st_mtime)
        resultado = {
            'archivo': archivo.name,
            'fecha_actualizacion': fecha_modificacion.isoformat(),
            'fecha_actualizacion_formateada': fecha_modificacion.strftime('%d/%m/%Y %H:%M:%S'),
            'compartido': True,  # Indica que los datos son compartidos entre todos los usuarios
            'hojas': {}
        }
        
        # Leer cada hoja
        for sheet_name in excel_file.sheet_names:
            df = pd.read_excel(excel_file, sheet_name=sheet_name)
            # Convertir a formato JSON (solo primeras 100 filas para no sobrecargar)
            df_limited = df.head(100)
            # Reemplazar NaN con None para que se serialice correctamente como null en JSON
            df_limited = df_limited.where(pd.notna(df_limited), None)
            # Convertir a diccionario
            datos_dict = df_limited.to_dict('records')
            # Limpiar cualquier NaN restante (por si acaso)
            datos_limpios = []
            for fila in datos_dict:
                fila_limpia = {k: (None if (isinstance(v, float) and pd.isna(v)) else v) for k, v in fila.items()}
                datos_limpios.append(fila_limpia)
            
            resultado['hojas'][sheet_name] = {
                'total_filas': len(df),
                'columnas': list(df.columns),
                'datos': datos_limpios
            }
        
        return jsonify({
            'success': True,
            'data': resultado
        }), 200
        
    except Exception as e:
        logger.error(f"Error listando vencimientos: {str(e)}", exc_info=True)
        return jsonify({
            'success': False,
            'message': f'Error al listar vencimientos: {str(e)}'
        }), 500

@app.route('/vencimientos/refrescar', methods=['POST'])
def vencimientos_refrescar():
    """Ejecuta el scraper de vencimientos de forma asíncrona. Los datos generados serán compartidos para todos los usuarios."""
    try:
        # Verificar que el directorio de vencimientos existe
        if not VENCIMIENTOS_DIR.exists():
            logger.error(f"Directorio de vencimientos no existe: {VENCIMIENTOS_DIR}")
            return jsonify({
                'success': False,
                'message': f'Directorio de vencimientos no encontrado: {VENCIMIENTOS_DIR}'
            }), 500
        
        # Verificar que el scraper existe
        scraper_path = VENCIMIENTOS_DIR / 'scraper_vencimientos.py'
        if not scraper_path.exists():
            logger.error(f"Scraper no encontrado: {scraper_path}")
            return jsonify({
                'success': False,
                'message': f'Scraper no encontrado: {scraper_path}'
            }), 404
        
        # Verificar si ya hay un job en proceso
        with vencimientos_jobs_lock:
            jobs_activos = [j for j in vencimientos_jobs.values() if j.get('status') in ['pending', 'processing']]
            if jobs_activos:
                job_activo = jobs_activos[0]
                logger.info(f"Ya hay un job activo: {job_activo['id']}")
                return jsonify({
                    'success': False,
                    'message': 'Ya hay una actualización en proceso. Espera a que finalice antes de iniciar otra.',
                    'job_id': job_activo['id']
                }), 409  # Conflict
        
        # Crear un nuevo job
        job_id = str(uuid.uuid4())
        
        with vencimientos_jobs_lock:
            vencimientos_jobs[job_id] = {
                'id': job_id,
                'status': 'pending',
                'progress': 0,
                'message': 'Job creado, esperando inicio...',
                'error': None,
                'created_at': datetime.now().isoformat(),
                'updated_at': datetime.now().isoformat(),
                'compartido': True  # Indica que los datos serán compartidos
            }
        
        logger.info(f"Iniciando scraper de vencimientos con job_id: {job_id}")
        logger.info(f"Los datos se guardarán en: {VENCIMIENTOS_DIR.resolve()} (compartido para todos los usuarios)")
        
        # Iniciar el scraper en un thread separado
        thread = threading.Thread(target=ejecutar_scraper_vencimientos, args=(job_id,))
        thread.daemon = True
        thread.start()
        
        return jsonify({
            'success': True,
            'job_id': job_id,
            'message': 'Scraper iniciado en segundo plano. Los datos generados serán compartidos para todos los usuarios.',
            'compartido': True
        }), 200
        
    except Exception as e:
        logger.error(f"Error iniciando scraper de vencimientos: {str(e)}", exc_info=True)
        return jsonify({
            'success': False,
            'message': f'Error al iniciar scraper: {str(e)}'
        }), 500

@app.route('/vencimientos/status/<job_id>', methods=['GET'])
def vencimientos_status(job_id):
    """Consulta el estado de un job de refresco de vencimientos"""
    try:
        logger.info(f"Consultando estado de job: {job_id}")
        logger.info(f"Jobs disponibles: {list(vencimientos_jobs.keys())}")
        
        job = get_vencimientos_job(job_id)
        
        if not job:
            logger.warning(f"Job {job_id} no encontrado. Jobs disponibles: {list(vencimientos_jobs.keys())}")
            return jsonify({
                'success': False,
                'message': f'Job no encontrado: {job_id}',
                'available_jobs': list(vencimientos_jobs.keys())
            }), 404
        
        logger.info(f"Job encontrado: {job_id}, estado: {job.get('status')}")
        return jsonify({
            'success': True,
            'job': job
        }), 200
        
    except Exception as e:
        logger.error(f"Error consultando estado de job: {str(e)}", exc_info=True)
        return jsonify({
            'success': False,
            'message': f'Error al consultar estado: {str(e)}'
        }), 500

@app.route('/vencimientos/filtrar-por-cuils', methods=['POST'])
def vencimientos_filtrar_por_cuils():
    """Filtra vencimientos por CUILs desde un archivo Excel"""
    try:
        # Validar que se recibió el archivo
        if 'archivo' not in request.files:
            return jsonify({
                'success': False,
                'message': 'No se recibió ningún archivo Excel'
            }), 400
        
        archivo_cuils = request.files['archivo']
        
        if archivo_cuils.filename == '':
            return jsonify({
                'success': False,
                'message': 'El archivo está vacío'
            }), 400
        
        # Validar extensión
        ext = Path(archivo_cuils.filename).suffix.lower()
        if ext not in ['.xlsx', '.xls']:
            return jsonify({
                'success': False,
                'message': 'El archivo debe ser Excel (.xlsx o .xls)'
            }), 400
        
        # Guardar archivo temporalmente
        temp_file = TEMP_DIR / f'cuils_{int(datetime.now().timestamp())}{ext}'
        archivo_cuils.save(str(temp_file))
        
        try:
            # Leer archivo de CUILs
            logger.info(f"Leyendo archivo de CUILs: {temp_file}")
            df_cuils = pd.read_excel(temp_file)
            
            logger.info(f"Columnas encontradas en el archivo: {list(df_cuils.columns)}")
            logger.info(f"Total de filas: {len(df_cuils)}")
            
            # Buscar columna con CUILs (puede llamarse 'CUIL', 'cuil', 'CUIT', 'cuit', o la primera columna)
            columna_cuil = None
            posibles_nombres = ['CUIL', 'cuil', 'CUIT', 'cuit', 'CUIL/CUIT', 'CUIT/CUIL', 'CUIT/CUIL', 'CUIL_CUIT', 'CUIT_CUIL']
            
            # Buscar por nombre exacto o parcial
            for col in df_cuils.columns:
                col_str = str(col).upper().strip()
                # Buscar coincidencia exacta o parcial
                for nombre_posible in posibles_nombres:
                    if nombre_posible.upper() in col_str or col_str in nombre_posible.upper():
                        columna_cuil = col
                        logger.info(f"Columna CUIL encontrada por nombre: {col}")
                        break
                if columna_cuil:
                    break
            
            # Si no se encuentra, usar la primera columna que tenga datos
            if columna_cuil is None:
                logger.info("No se encontró columna con nombre CUIL/CUIT, usando la primera columna con datos")
                for col in df_cuils.columns:
                    valores_no_vacios = df_cuils[col].dropna()
                    if len(valores_no_vacios) > 0:
                        columna_cuil = col
                        logger.info(f"Usando primera columna con datos: {col}")
                        break
            
            if columna_cuil is None:
                return jsonify({
                    'success': False,
                    'message': 'No se encontró ninguna columna con datos en el archivo',
                    'columnas_disponibles': list(df_cuils.columns)
                }), 400
            
            logger.info(f"Columna seleccionada para CUILs: {columna_cuil}")
            
            # Extraer CUILs - limpiar y convertir a string
            cuils_raw = df_cuils[columna_cuil].dropna()
            logger.info(f"Valores encontrados en la columna (primeros 5): {cuils_raw.head().tolist()}")
            
            # Convertir a string y limpiar
            cuils = []
            for valor in cuils_raw:
                if pd.notna(valor):
                    valor_str = str(valor).strip()
                    # Remover espacios y caracteres especiales, pero mantener números y guiones
                    if valor_str and len(valor_str) > 0:
                        cuils.append(valor_str)
            
            # Eliminar duplicados
            cuils = list(set(cuils))
            
            logger.info(f"Total de CUILs únicos encontrados: {len(cuils)}")
            if len(cuils) > 0:
                logger.info(f"Ejemplos de CUILs: {cuils[:5]}")
            
            if len(cuils) == 0:
                return jsonify({
                    'success': False,
                    'message': f'No se encontraron CUILs válidos en la columna "{columna_cuil}". Verifica que el archivo contenga CUILs en formato válido.',
                    'columna_usada': columna_cuil,
                    'columnas_disponibles': list(df_cuils.columns),
                    'total_filas': len(df_cuils)
                }), 400
            
            # Obtener archivo consolidado más reciente
            archivo_consolidado = encontrar_archivo_consolidado_mas_reciente()
            
            if not archivo_consolidado:
                return jsonify({
                    'success': False,
                    'message': 'No se encontró ningún archivo consolidado de vencimientos. Ejecuta primero "Refrescar Vencimientos".'
                }), 404
            
            # Leer archivo consolidado
            excel_file = pd.ExcelFile(archivo_consolidado)
            
            # Procesar cada CUIL
            resultados = []
            
            for cuil in cuils:
                ultimo_digito = extraer_ultimo_digito_cuil(cuil)
                
                if ultimo_digito is None:
                    resultados.append({
                        'CUIL': cuil,
                        'estado': 'Sin vencimientos',
                        'motivo': 'CUIL inválido',
                        'datos': []
                    })
                    continue
                
                # Buscar en todas las hojas
                vencimientos_encontrados = []
                
                for sheet_name in excel_file.sheet_names:
                    df = pd.read_excel(excel_file, sheet_name=sheet_name)
                    
                    # Identificar la columna de fecha correspondiente al dígito ANTES de procesar filas
                    # Buscar columnas de fechas por terminación
                    columnas_fechas = [col for col in df.columns if 'vencimiento según terminación' in str(col).lower() or 'terminación' in str(col).lower()]
                    
                    # Determinar qué columna usar según el dígito
                    columna_fecha_digito = None
                    digito_int = int(ultimo_digito) if ultimo_digito.isdigit() else None
                    
                    if digito_int is not None and columnas_fechas:
                        if digito_int in [0, 1, 2, 3]:
                            # Primera columna (sin .1 o .2)
                            primera = [c for c in columnas_fechas if '.1' not in str(c) and '.2' not in str(c)]
                            if primera:
                                columna_fecha_digito = primera[0]
                        elif digito_int in [4, 5, 6]:
                            # Segunda columna (.1)
                            segunda = [c for c in columnas_fechas if '.1' in str(c)]
                            if segunda:
                                columna_fecha_digito = segunda[0]
                        elif digito_int in [7, 8, 9]:
                            # Tercera columna (.2)
                            tercera = [c for c in columnas_fechas if '.2' in str(c)]
                            if tercera:
                                columna_fecha_digito = tercera[0]
                    
                    logger.info(f"Hoja {sheet_name}: columna_fecha_digito={columna_fecha_digito} para dígito {ultimo_digito}")
                    
                    if columna_fecha_digito:
                        # Procesar TODAS las filas y extraer la fecha de la columna correspondiente
                        for idx, fila in df.iterrows():
                            # Obtener la fecha de la columna correspondiente al dígito
                            fecha_vencimiento = fila.get(columna_fecha_digito)
                            
                            # Solo procesar si hay una fecha válida
                            if pd.notna(fecha_vencimiento) and str(fecha_vencimiento).strip():
                                fecha_limpia = None if (isinstance(fecha_vencimiento, float) and pd.isna(fecha_vencimiento)) else fecha_vencimiento
                                
                                if fecha_limpia and str(fecha_limpia).strip():
                                    # Convertir fila a diccionario y limpiar NaN
                                    fila_dict = fila.to_dict()
                                    fila_limpia = {}
                                    
                                    # Agregar columnas de información general
                                    columnas_info = ['Mes de devengamiento', 'Periodo devengado', 'MES DE DEVENGAMIENTO', 
                                                    'PERIODO DEVENGADO', 'Mes', 'Período', 'Periodo', 'Devengamiento']
                                    
                                    for k, v in fila_dict.items():
                                        # Limpiar NaN
                                        valor_limpio = None if (isinstance(v, float) and pd.isna(v)) else v
                                        
                                        col_str = str(k).lower()
                                        
                                        # Incluir columnas de información general
                                        if any(info_col.lower() in col_str for info_col in columnas_info):
                                            fila_limpia[k] = valor_limpio
                                        # Incluir la fecha de vencimiento
                                        elif k == columna_fecha_digito:
                                            fila_limpia['Fecha_Vencimiento'] = fecha_limpia
                                        # Excluir otras columnas de fechas por terminación
                                        elif 'vencimiento según terminación' in col_str or 'terminación' in col_str:
                                            continue
                                        else:
                                            # Incluir otras columnas que no sean de fechas por terminación
                                            if valor_limpio is not None:
                                                fila_limpia[k] = valor_limpio
                                    
                                    # Agregar el vencimiento encontrado
                                    vencimientos_encontrados.append({
                                        'tipo_vencimiento': sheet_name,
                                        **fila_limpia
                                    })
                    
                    logger.info(f"Total vencimientos encontrados en {sheet_name}: {len([v for v in vencimientos_encontrados if v.get('tipo_vencimiento') == sheet_name])}")
                
                if vencimientos_encontrados:
                    resultados.append({
                        'CUIL': cuil,
                        'ultimo_digito': ultimo_digito,
                        'estado': 'Con vencimientos',
                        'total': len(vencimientos_encontrados),
                        'datos': vencimientos_encontrados
                    })
                else:
                    resultados.append({
                        'CUIL': cuil,
                        'ultimo_digito': ultimo_digito,
                        'estado': 'Sin vencimientos',
                        'datos': []
                    })
            
            # Crear DataFrame con resultados organizados: CUIL -> Tipo -> Fechas
            filas_resultado = []
            for resultado in resultados:
                cuil = resultado['CUIL']
                ultimo_digito = resultado.get('ultimo_digito', '')
                estado = resultado['estado']
                
                if resultado['datos']:
                    # Fila de encabezado del CUIL
                    filas_resultado.append({
                        'CUIL': cuil,
                        'Ultimo_Digito': ultimo_digito,
                        'Estado': estado,
                        'Tipo_Vencimiento': f"TOTAL: {len(resultado['datos'])} vencimiento(s)",
                        'Es_Encabezado_CUIL': True
                    })
                    
                    # Agrupar vencimientos por tipo
                    vencimientos_por_tipo = {}
                    for vencimiento in resultado['datos']:
                        tipo = vencimiento.get('tipo_vencimiento', 'Sin tipo')
                        if tipo not in vencimientos_por_tipo:
                            vencimientos_por_tipo[tipo] = []
                        vencimientos_por_tipo[tipo].append(vencimiento)
                    
                    # Crear filas organizadas: Tipo -> Fechas
                    for tipo_vencimiento, vencimientos in vencimientos_por_tipo.items():
                        # Fila de encabezado del tipo de vencimiento
                        filas_resultado.append({
                            'CUIL': '',
                            'Ultimo_Digito': '',
                            'Estado': '',
                            'Tipo_Vencimiento': f"  → {tipo_vencimiento} ({len(vencimientos)} vencimiento(s))",
                            'Es_Encabezado_CUIL': False,
                            'Es_Encabezado_Tipo': True
                        })
                        
                        # Filas con los datos de cada vencimiento (fechas)
                        for vencimiento in vencimientos:
                            # Extraer todas las columnas excepto tipo_vencimiento
                            fila_vencimiento = {
                                'CUIL': '',
                                'Ultimo_Digito': '',
                                'Estado': '',
                                'Tipo_Vencimiento': '',  # Vacío porque ya está en el encabezado
                                'Es_Encabezado_CUIL': False,
                                'Es_Encabezado_Tipo': False
                            }
                            
                            # Agregar todas las columnas del vencimiento
                            for k, v in vencimiento.items():
                                if k != 'tipo_vencimiento':
                                    fila_vencimiento[k] = v
                            
                            filas_resultado.append(fila_vencimiento)
                    
                    # Fila separadora después de cada CUIL
                    filas_resultado.append({
                        'CUIL': '',
                        'Ultimo_Digito': '',
                        'Estado': '',
                        'Tipo_Vencimiento': '',
                        'Es_Encabezado_CUIL': False,
                        'Es_Encabezado_Tipo': False
                    })
                else:
                    # Si no hay vencimientos, crear una fila indicando esto
                    fila = {
                        'CUIL': cuil,
                        'Ultimo_Digito': ultimo_digito,
                        'Estado': estado,
                        'Motivo': resultado.get('motivo', 'No se encontraron coincidencias'),
                        'Tipo_Vencimiento': '',
                        'Es_Encabezado_CUIL': False,
                        'Es_Encabezado_Tipo': False
                    }
                    filas_resultado.append(fila)
                    # Fila separadora
                    filas_resultado.append({
                        'CUIL': '',
                        'Ultimo_Digito': '',
                        'Estado': '',
                        'Tipo_Vencimiento': '',
                        'Es_Encabezado_CUIL': False,
                        'Es_Encabezado_Tipo': False
                    })
            
            df_resultado = pd.DataFrame(filas_resultado)
            
            # Guardar resultado en Excel con formato
            resultado_filename = f'vencimientos_filtrados_{int(datetime.now().timestamp())}.xlsx'
            resultado_path = TEMP_DIR / resultado_filename
            
            with pd.ExcelWriter(resultado_path, engine='openpyxl') as writer:
                df_resultado.to_excel(writer, sheet_name='Vencimientos_Filtrados', index=False)
            
            # Aplicar formato al Excel (colores para encabezados)
            try:
                from openpyxl import load_workbook
                from openpyxl.styles import Font, PatternFill, Alignment
                
                wb = load_workbook(resultado_path)
                ws = wb['Vencimientos_Filtrados']
                
                # Estilos
                header_cuil_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_tipo_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=11)
                tipo_font = Font(bold=True, color="000000", size=10)
                
                # Aplicar formato a las filas
                for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
                    # Verificar si es encabezado de CUIL
                    es_encabezado_cuil = False
                    es_encabezado_tipo = False
                    
                    # Buscar columna Es_Encabezado_CUIL y Es_Encabezado_Tipo
                    for col_idx, cell in enumerate(row, start=1):
                        header = ws.cell(row=1, column=col_idx).value
                        if header == 'Es_Encabezado_CUIL' and cell.value == True:
                            es_encabezado_cuil = True
                        if header == 'Es_Encabezado_Tipo' and cell.value == True:
                            es_encabezado_tipo = True
                    
                    if es_encabezado_cuil:
                        # Formato para encabezado de CUIL
                        for cell in row:
                            if cell.value:  # Solo aplicar si tiene valor
                                cell.fill = header_cuil_fill
                                cell.font = header_font
                                cell.alignment = Alignment(horizontal='left', vertical='center')
                    elif es_encabezado_tipo:
                        # Formato para encabezado de tipo
                        for cell in row:
                            if cell.value:  # Solo aplicar si tiene valor
                                cell.fill = header_tipo_fill
                                cell.font = tipo_font
                                cell.alignment = Alignment(horizontal='left', vertical='center')
                
                # Ocultar columnas de control
                for col_idx, header in enumerate(ws[1], start=1):
                    if header.value in ['Es_Encabezado_CUIL', 'Es_Encabezado_Tipo']:
                        ws.column_dimensions[header.column_letter].hidden = True
                
                wb.save(resultado_path)
            except Exception as e:
                logger.warning(f"No se pudo aplicar formato al Excel: {e}")
            
            base_url = request.host_url.rstrip('/')
            
            return jsonify({
                'success': True,
                'message': f'Filtrado completado. {len(cuils)} CUILs procesados.',
                'filename': resultado_filename,
                'total_cuils': len(cuils),
                'cuils_con_vencimientos': sum(1 for r in resultados if r['estado'] == 'Con vencimientos'),
                'cuils_sin_vencimientos': sum(1 for r in resultados if r['estado'] == 'Sin vencimientos'),
                'downloadUrl': f'{base_url}/download/{resultado_filename}'
            }), 200
            
        finally:
            # Limpiar archivo temporal
            if temp_file.exists():
                try:
                    temp_file.unlink()
                except Exception as e:
                    logger.warning(f"Error al eliminar archivo temporal: {e}")
        
    except Exception as e:
        logger.error(f"Error filtrando vencimientos por CUILs: {str(e)}", exc_info=True)
        return jsonify({
            'success': False,
            'message': f'Error al filtrar vencimientos: {str(e)}'
        }), 500

# ==================== FIN SISTEMA DE VENCIMIENTOS ====================

def load_extractor_module(script_name):
    """Carga dinámicamente un módulo extractor"""
    script_path = EXTRACTORES_DIR / script_name
    
    if not script_path.exists():
        raise FileNotFoundError(f"El archivo extractor no existe: {script_path}")
    
    logger.info(f"Cargando módulo desde: {script_path}")
    
    try:
        spec = importlib.util.spec_from_file_location(script_name.replace('.py', ''), script_path)
        if spec is None:
            raise ImportError(f"No se pudo crear el spec para {script_name}")
        
        module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(module)
        logger.info(f"Módulo {script_name} cargado exitosamente")
        return module
    except Exception as e:
        logger.error(f"Error cargando módulo {script_name}: {str(e)}", exc_info=True)
        raise

@app.route('/health', methods=['GET'])
def health():
    """Endpoint de salud"""
    try:
        logger.info("Health check recibido")
        port = os.environ.get('PORT', os.environ.get('EXTRACTOR_PORT', 'NO DEFINIDA'))
        response = jsonify({
            'status': 'ok',
            'message': 'Servidor funcionando correctamente',
            'extractors_count': len(BANCO_EXTRACTORS),
            'port': port,
            'host': request.host,
            'remote_addr': request.remote_addr
        })
        logger.info("Health check respondido exitosamente")
        return response, 200
    except Exception as e:
        logger.error(f"Error en health check: {str(e)}", exc_info=True)
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500

@app.route('/extractors', methods=['GET'])
def list_extractors():
    """Lista todos los extractores disponibles"""
    return jsonify({
        'extractors': list(BANCO_EXTRACTORS.keys()),
        'count': len(BANCO_EXTRACTORS)
    })

@app.route('/extract', methods=['POST'])
def extract():
    """Endpoint principal para extraer datos"""
    try:
        # Validar que se recibió el archivo y el banco
        if 'pdf' not in request.files:
            return jsonify({'success': False, 'message': 'No se recibió ningún archivo PDF'}), 400
        
        if 'banco' not in request.form:
            return jsonify({'success': False, 'message': 'No se especificó el banco'}), 400
        
        pdf_file = request.files['pdf']
        banco_id = request.form['banco']
        
        # Validar que el banco existe
        if banco_id not in BANCO_EXTRACTORS:
            return jsonify({'success': False, 'message': f'Banco no soportado: {banco_id}'}), 400
        
        # Guardar el PDF temporalmente
        pdf_filename = f"{banco_id}_{pdf_file.filename}"
        # Limpiar nombre de archivo (remover caracteres problemáticos)
        pdf_filename = pdf_filename.replace(' ', '_').replace('/', '_').replace('\\', '_')
        pdf_path = TEMP_DIR / pdf_filename
        
        # Asegurar que el directorio existe
        TEMP_DIR.mkdir(parents=True, exist_ok=True)
        
        # Guardar archivo
        pdf_file.save(str(pdf_path))
        logger.info(f"PDF guardado temporalmente en: {pdf_path}")
        
        # Generar nombre del archivo Excel de salida
        excel_filename = pdf_filename.replace('.pdf', '_extraido.xlsx')
        excel_path = TEMP_DIR / excel_filename
        
        try:
            # Cargar el módulo extractor
            extractor_info = BANCO_EXTRACTORS[banco_id]
            logger.info(f"Cargando extractor: {extractor_info['script']}")
            
            try:
                module = load_extractor_module(extractor_info['script'])
            except Exception as load_error:
                logger.error(f"Error cargando módulo {extractor_info['script']}: {str(load_error)}", exc_info=True)
                return jsonify({
                    'success': False,
                    'message': f'Error al cargar el extractor: {str(load_error)}'
                }), 500
            
            try:
                extractor_function = getattr(module, extractor_info['function'])
            except AttributeError as attr_error:
                logger.error(f"Función {extractor_info['function']} no encontrada en {extractor_info['script']}: {str(attr_error)}")
                return jsonify({
                    'success': False,
                    'message': f'Función {extractor_info["function"]} no encontrada en el extractor'
                }), 500
            
            # Ejecutar la extracción
            logger.info(f"Extrayendo datos de {banco_id}...")
            try:
                df = extractor_function(str(pdf_path), str(excel_path))
            except Exception as extract_error:
                logger.error(f"Error durante la extracción: {str(extract_error)}", exc_info=True)
                return jsonify({
                    'success': False,
                    'message': f'Error al extraer datos: {str(extract_error)}'
                }), 500
            
            # Verificar que se generó el archivo Excel
            if not excel_path.exists():
                logger.warning(f"El archivo Excel no se generó en: {excel_path}")
                return jsonify({
                    'success': False,
                    'message': 'No se pudo generar el archivo Excel'
                }), 500
            
            # Obtener información del resultado
            rows = len(df) if df is not None and hasattr(df, '__len__') else 0
            logger.info(f"Extracción completada: {rows} filas extraídas")

            base_url = request.host_url.rstrip('/')

            return jsonify({
                'success': True,
                'message': 'Extracción completada exitosamente',
                'filename': excel_filename,
                'rows': rows,
                'downloadUrl': f'{base_url}/download/{excel_filename}'
            })
            
        except Exception as e:
            logger.error(f"Error durante la extracción: {str(e)}", exc_info=True)
            return jsonify({
                'success': False,
                'message': f'Error al procesar el PDF: {str(e)}'
            }), 500
        
        finally:
            # Limpiar el PDF temporal
            if pdf_path.exists():
                try:
                    pdf_path.unlink()
                except Exception as e:
                    print(f"Error al eliminar PDF temporal: {e}")
    
    except Exception as e:
        print(f"Error general: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error del servidor: {str(e)}'
        }), 500

@app.route('/pdf-to-ocr', methods=['POST'])
def pdf_to_ocr():
    """Endpoint para convertir PDF escaneado a PDF con OCR"""
    try:
        # Validar que se recibió el archivo
        if 'pdf' not in request.files:
            return jsonify({'success': False, 'message': 'No se recibió ningún archivo PDF'}), 400
        
        pdf_file = request.files['pdf']
        
        # Guardar el PDF temporalmente
        pdf_filename = f"ocr_input_{pdf_file.filename}"
        # Limpiar nombre de archivo (remover caracteres problemáticos)
        pdf_filename = pdf_filename.replace(' ', '_').replace('/', '_').replace('\\', '_')
        pdf_path = TEMP_DIR / pdf_filename
        
        # Asegurar que el directorio existe
        TEMP_DIR.mkdir(parents=True, exist_ok=True)
        
        # Guardar archivo
        pdf_file.save(str(pdf_path))
        logger.info(f"PDF guardado temporalmente en: {pdf_path}")
        
        # Generar nombre del archivo PDF de salida
        output_filename = pdf_filename.replace('ocr_input_', 'ocr_output_').replace('.pdf', '_OCR.pdf')
        output_path = TEMP_DIR / output_filename
        
        try:
            # Cargar el módulo extractor de OCR
            import sys
            sys.path.insert(0, str(EXTRACTORES_DIR))
            from extractor_pdf_ocr import extraer_texto_pdf_ocr
            
            # Ejecutar la conversión OCR
            print(f"Convirtiendo PDF a OCR: {pdf_filename}...")
            success = extraer_texto_pdf_ocr(str(pdf_path), str(output_path))
            
            # Verificar que se generó el archivo
            if not output_path.exists():
                return jsonify({
                    'success': False,
                    'message': 'No se pudo generar el archivo PDF con OCR'
                }), 500
            
            base_url = request.host_url.rstrip('/')

            return jsonify({
                'success': True,
                'message': 'Conversión OCR completada exitosamente',
                'filename': output_filename,
                'downloadUrl': f'{base_url}/download-pdf/{output_filename}'
            })
            
        except Exception as e:
            print(f"Error durante la conversión OCR: {str(e)}")
            return jsonify({
                'success': False,
                'message': f'Error al procesar el PDF: {str(e)}'
            }), 500
        
        finally:
            # Limpiar el PDF temporal de entrada
            if pdf_path.exists():
                try:
                    pdf_path.unlink()
                except Exception as e:
                    print(f"Error al eliminar PDF temporal: {e}")
    
    except Exception as e:
        print(f"Error general: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error del servidor: {str(e)}'
        }), 500

@app.route('/download/<filename>', methods=['GET'])
def download(filename):
    """Endpoint para descargar archivos Excel generados"""
    try:
        file_path = TEMP_DIR / filename
        
        if not file_path.exists():
            return jsonify({'success': False, 'message': 'Archivo no encontrado'}), 404
        
        return send_file(
            str(file_path),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/download-pdf/<filename>', methods=['GET'])
def download_pdf(filename):
    """Endpoint para descargar archivos PDF generados"""
    try:
        file_path = TEMP_DIR / filename
        
        if not file_path.exists():
            return jsonify({'success': False, 'message': 'Archivo no encontrado'}), 404
        
        return send_file(
            str(file_path),
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/cleanup', methods=['POST'])
def cleanup():
    """Limpia archivos temporales antiguos"""
    try:
        import time
        cleaned = 0
        now = time.time()
        
        for file in TEMP_DIR.glob('*'):
            if file.is_file():
                # Eliminar archivos más antiguos de 1 hora
                if now - file.stat().st_mtime > 3600:
                    file.unlink()
                    cleaned += 1
        
        return jsonify({
            'success': True,
            'message': f'Se limpiaron {cleaned} archivos'
        })
    
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/google/oauth/token', methods=['POST'])
def exchange_google_token():
    """Intercambia código de autorización por token de acceso"""
    try:
        logger.info("Request recibido en /api/google/oauth/token")
        data = request.json
        
        if not data:
            logger.error("No se recibió JSON en el request")
            return jsonify({'error': 'Se requiere JSON en el body'}), 400
        
        code = data.get('code')
        redirect_uri = data.get('redirect_uri')
        
        if not code or not redirect_uri:
            logger.error(f"Faltan parámetros: code={bool(code)}, redirect_uri={bool(redirect_uri)}")
            return jsonify({'error': 'code y redirect_uri son requeridos'}), 400
        
        # Obtener credenciales desde variables de entorno del backend
        client_id = os.getenv('GOOGLE_CLIENT_ID')
        client_secret = os.getenv('GOOGLE_CLIENT_SECRET')
        
        if not client_id or not client_secret:
            logger.error("Credenciales de Google no configuradas en variables de entorno")
            return jsonify({
                'error': 'Credenciales de Google no configuradas',
                'message': 'Configura GOOGLE_CLIENT_ID y GOOGLE_CLIENT_SECRET en las variables de entorno del backend'
            }), 500
        
        logger.info(f"Intercambiando código por token (redirect_uri: {redirect_uri})")
        
        # Intercambiar código por token
        token_response = requests.post('https://oauth2.googleapis.com/token', data={
            'code': code,
            'client_id': client_id,
            'client_secret': client_secret,
            'redirect_uri': redirect_uri,
            'grant_type': 'authorization_code',
        })
        
        if token_response.status_code != 200:
            try:
                error_data = token_response.json()
                logger.error(f"Error de Google OAuth: {error_data}")
                # Incluir información adicional para debugging
                error_data['debug_info'] = {
                    'client_id_configured': bool(client_id),
                    'client_secret_configured': bool(client_secret),
                    'redirect_uri_used': redirect_uri,
                }
                return jsonify(error_data), token_response.status_code
            except:
                # Si no se puede parsear JSON, devolver el texto
                logger.error(f"Error de Google OAuth (no JSON): {token_response.text}")
                return jsonify({
                    'error': 'oauth_error',
                    'error_description': token_response.text[:200],
                    'status_code': token_response.status_code
                }), token_response.status_code
        
        token_data = token_response.json()
        logger.info("Token obtenido exitosamente")
        
        return jsonify(token_data), 200
        
    except Exception as e:
        logger.error(f"Error en exchange_google_token: {str(e)}", exc_info=True)
        return jsonify({'error': str(e)}), 500

@app.route('/api/google/oauth/refresh', methods=['POST'])
def refresh_google_token():
    """Refresca un token de acceso usando refresh token"""
    try:
        logger.info("Request recibido en /api/google/oauth/refresh")
        data = request.json
        
        if not data:
            logger.error("No se recibió JSON en el request")
            return jsonify({'error': 'Se requiere JSON en el body'}), 400
        
        refresh_token = data.get('refresh_token')
        
        if not refresh_token:
            logger.error("No se recibió refresh_token")
            return jsonify({'error': 'refresh_token es requerido'}), 400
        
        client_id = os.getenv('GOOGLE_CLIENT_ID')
        client_secret = os.getenv('GOOGLE_CLIENT_SECRET')
        
        if not client_id or not client_secret:
            logger.error("Credenciales de Google no configuradas en variables de entorno")
            return jsonify({
                'error': 'Credenciales de Google no configuradas',
                'message': 'Configura GOOGLE_CLIENT_ID y GOOGLE_CLIENT_SECRET en las variables de entorno del backend'
            }), 500
        
        logger.info("Refrescando token de acceso")
        
        token_response = requests.post('https://oauth2.googleapis.com/token', data={
            'client_id': client_id,
            'client_secret': client_secret,
            'refresh_token': refresh_token,
            'grant_type': 'refresh_token',
        })
        
        if token_response.status_code != 200:
            error_data = token_response.json()
            logger.error(f"Error al refrescar token: {error_data}")
            return jsonify(error_data), token_response.status_code
        
        token_data = token_response.json()
        logger.info("Token refrescado exitosamente")
        
        return jsonify(token_data), 200
        
    except Exception as e:
        logger.error(f"Error en refresh_google_token: {str(e)}", exc_info=True)
        return jsonify({'error': str(e)}), 500

@app.route('/api/google/client-id', methods=['GET'])
def get_google_client_id():
    """Devuelve el Client ID de Google (sin el secret)"""
    try:
        logger.info("Request recibido en /api/google/client-id")
        client_id = os.getenv('GOOGLE_CLIENT_ID')
        
        if not client_id:
            logger.error("GOOGLE_CLIENT_ID no configurado en variables de entorno")
            return jsonify({
                'error': 'Client ID no configurado',
                'message': 'Configura GOOGLE_CLIENT_ID en las variables de entorno del backend'
            }), 500
        
        logger.info("Client ID devuelto exitosamente")
        # Solo devolvemos el Client ID, nunca el secret
        return jsonify({'client_id': client_id}), 200
        
    except Exception as e:
        logger.error(f"Error en get_google_client_id: {str(e)}", exc_info=True)
        return jsonify({'error': str(e)}), 500

@app.route('/consilador/comparar', methods=['POST'])
def consilador_comparar():
    """Endpoint para comparar dos archivos Excel usando el consilador"""
    try:
        # Validar que se recibieron los dos archivos Excel
        if 'archivo1' not in request.files or 'archivo2' not in request.files:
            return jsonify({
                'success': False, 
                'message': 'Se requieren dos archivos Excel (archivo1 y archivo2)'
            }), 400
        
        archivo1 = request.files['archivo1']
        archivo2 = request.files['archivo2']
        
        # Validar que los archivos tienen nombres
        if archivo1.filename == '' or archivo2.filename == '':
            return jsonify({
                'success': False, 
                'message': 'Ambos archivos deben tener un nombre válido'
            }), 400
        
        # Validar extensiones
        allowed_extensions = {'.xlsx', '.xls'}
        ext1 = Path(archivo1.filename).suffix.lower()
        ext2 = Path(archivo2.filename).suffix.lower()
        
        if ext1 not in allowed_extensions or ext2 not in allowed_extensions:
            return jsonify({
                'success': False, 
                'message': 'Ambos archivos deben ser Excel (.xlsx o .xls)'
            }), 400
        
        # Crear directorio temporal para la comparación
        import time
        comparacion_dir = TEMP_DIR / f'consilador_{int(time.time())}'
        comparacion_dir.mkdir(parents=True, exist_ok=True)
        
        # Guardar los archivos temporalmente
        archivo1_path = comparacion_dir / f'archivo1{ext1}'
        archivo2_path = comparacion_dir / f'archivo2{ext2}'
        
        archivo1.save(str(archivo1_path))
        archivo2.save(str(archivo2_path))
        
        logger.info(f"Archivos guardados en: {comparacion_dir}")
        
        try:
            # Importar el módulo de comparación
            consilador_dir = Path(__file__).parent / 'Consilador'
            comparar_automatico_path = consilador_dir / 'comparar_automatico.py'
            
            if not comparar_automatico_path.exists():
                return jsonify({
                    'success': False,
                    'message': 'Módulo de comparación no encontrado'
                }), 500
            
            # Agregar el directorio al path
            sys.path.insert(0, str(consilador_dir))
            
            # Cambiar al directorio temporal para que el script encuentre los archivos
            original_cwd = os.getcwd()
            os.chdir(str(comparacion_dir))
            
            try:
                # Importar y ejecutar la función de comparación
                spec = importlib.util.spec_from_file_location(
                    'comparar_automatico', 
                    str(comparar_automatico_path)
                )
                module = importlib.util.module_from_spec(spec)
                spec.loader.exec_module(module)
                
                # Ejecutar la comparación
                logger.info("Ejecutando comparación de archivos...")
                resultado = module.comparar_archivos_actuales()
                
                if not resultado:
                    return jsonify({
                        'success': False,
                        'message': 'Error al comparar los archivos'
                    }), 500
                
                # Buscar el archivo de resultado
                resultado_path = comparacion_dir / 'resultado_comparacion.xlsx'
                
                if not resultado_path.exists():
                    return jsonify({
                        'success': False,
                        'message': 'No se generó el archivo de resultado'
                    }), 500
                
                # Generar nombre único para el archivo de resultado
                resultado_filename = f'resultado_comparacion_{int(time.time())}.xlsx'
                resultado_final = TEMP_DIR / resultado_filename
                
                # Mover el resultado al directorio temporal principal
                import shutil
                shutil.move(str(resultado_path), str(resultado_final))
                
                base_url = request.host_url.rstrip('/')
                
                return jsonify({
                    'success': True,
                    'message': 'Comparación completada exitosamente',
                    'filename': resultado_filename,
                    'downloadUrl': f'{base_url}/download/{resultado_filename}'
                })
                
            finally:
                # Restaurar directorio original
                os.chdir(original_cwd)
                # Limpiar archivos temporales del directorio de comparación
                try:
                    import shutil
                    if comparacion_dir.exists():
                        shutil.rmtree(comparacion_dir, ignore_errors=True)
                        logger.info(f"Directorio temporal limpiado: {comparacion_dir}")
                except Exception as e:
                    logger.warning(f"Error al limpiar archivos temporales: {e}")
                
        except Exception as e:
            logger.error(f"Error durante la comparación: {str(e)}", exc_info=True)
            return jsonify({
                'success': False,
                'message': f'Error al procesar los archivos: {str(e)}'
            }), 500
        
    except Exception as e:
        logger.error(f"Error general en consilador_comparar: {str(e)}", exc_info=True)
        return jsonify({
            'success': False,
            'message': f'Error del servidor: {str(e)}'
        }), 500

if __name__ == '__main__':
    # NOTA: Este bloque SOLO se ejecuta cuando se corre directamente con python server.py
    # Railway/Gunicorn NO ejecuta este bloque, importa la app directamente
    
    host = os.environ.get('EXTRACTOR_HOST', '0.0.0.0')
    # Railway usa la variable PORT, si no existe usa EXTRACTOR_PORT o 5000
    port = int(os.environ.get('PORT', os.environ.get('EXTRACTOR_PORT', '5000')))
    debug = os.environ.get('EXTRACTOR_DEBUG', 'false').lower() == 'true'

    logger.info("=" * 50)
    logger.info("MODO DESARROLLO - Servidor Flask Directo")
    logger.info("=" * 50)
    logger.info(f"Extractores disponibles: {len(BANCO_EXTRACTORS)}")
    logger.info(f"Directorio temporal: {TEMP_DIR}")
    logger.info(f"Escuchando en http://{host}:{port}")
    logger.info(f"Debug mode: {debug}")
    logger.info("=" * 50)
    logger.warning("ADVERTENCIA: En producción usa Gunicorn, no este modo")
    
    # Deshabilitar carga automática de .env para evitar errores de codificación
    # Flask 3.0 intenta cargar .env automáticamente, pero si el archivo está corrupto falla
    # Las variables de entorno se pueden configurar manualmente o en el sistema
    try:
        # Intentar deshabilitar la carga de .env
        import flask.cli
        original_load_dotenv = flask.cli.load_dotenv
        flask.cli.load_dotenv = lambda *args, **kwargs: None
    except Exception as e:
        logger.warning(f"No se pudo deshabilitar load_dotenv: {e}")
    
    try:
        app.run(host=host, port=port, debug=debug)
    finally:
        # Restaurar función original si fue modificada
        try:
            flask.cli.load_dotenv = original_load_dotenv
        except:
            pass

