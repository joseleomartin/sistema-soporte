import tkinter as tk
from tkinter import ttk, messagebox
import pandas as pd
import json
import os
import logging
from PIL import Image, ImageTk
try:
    # Gráficos embebidos
    from matplotlib.figure import Figure
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
except Exception:
    Figure = None

# Configuración de estilos PROFESIONAL PREMIUM - Diseño empresarial elegante
STYLE_CONFIG = {
    'bg_color': '#f5f7fa',  # Fondo principal más refinado y suave
    'bg_gradient_start': '#5a67d8',  # Gradiente más vibrante
    'bg_gradient_end': '#7c3aed',  # Púrpura elegante
    'bg_dark': '#1a202c',  # Azul oscuro más profundo y elegante
    'bg_darker': '#0d1117',  # Casi negro para máximo contraste
    'primary_color': '#6366f1',  # Indigo premium
    'primary_hover': '#4f46e5',  # Hover más intenso
    'primary_light': '#818cf8',  # Versión clara
    'primary_dark': '#4338ca',  # Versión oscura para profundidad
    'secondary_color': '#06b6d4',  # Cyan vibrante
    'accent_color': '#ec4899',  # Rosa vibrante para acentos
    'accent_light': '#f472b6',  # Rosa claro
    'success_color': '#10b981',  # Verde éxito
    'success_light': '#34d399',  # Verde claro
    'success_dark': '#059669',  # Verde oscuro
    'warning_color': '#f59e0b',  # Naranja advertencia
    'warning_light': '#fbbf24',  # Naranja claro
    'warning_dark': '#d97706',  # Naranja oscuro
    'danger_color': '#ef4444',  # Rojo peligro
    'danger_light': '#f87171',  # Rojo claro
    'danger_dark': '#dc2626',  # Rojo oscuro
    'text_color': '#1a202c',  # Texto principal más oscuro
    'text_light': '#4a5568',  # Texto secundario
    'text_muted': '#718096',  # Texto atenuado
    'text_white': '#ffffff',  # Texto blanco
    'card_bg': '#ffffff',  # Fondo de cards blanco puro
    'card_shadow': '#e2e8f0',  # Sombra suave
    'card_shadow_dark': '#cbd5e1',  # Sombra más pronunciada
    'card_shadow_light': '#f1f5f9',  # Sombra muy sutil
    'border_color': '#e2e8f0',  # Borde estándar
    'border_light': '#f1f5f9',  # Borde muy sutil
    'border_dark': '#cbd5e1',  # Borde más visible
    'header_bg': '#1a202c',  # Header oscuro elegante
    'header_gradient': 'linear-gradient(135deg, #5a67d8 0%, #7c3aed 100%)',
    'header_text': '#ffffff',  # Texto del header
    'header_subtitle': '#cbd5e1',  # Subtítulo del header
    'light_bg': '#f7fafc',  # Fondo claro muy suave
    'highlight': '#fef3c7',  # Resaltado amarillo
    'sidebar_bg': '#f9fafb',  # Fondo sidebar
    'sidebar_active': '#6366f1',  # Item activo sidebar
    'metric_bg_1': '#eef2ff',  # Azul claro muy suave
    'metric_bg_2': '#ecfdf5',  # Verde claro muy suave
    'metric_bg_3': '#fffbeb',  # Amarillo claro muy suave
    'metric_bg_4': '#fdf2f8',  # Rosa claro muy suave
    'hover_overlay': 'rgba(99, 102, 241, 0.1)',  # Overlay hover
}

DATA_FILE = 'data.json'

class Producto:
    def __init__(self, nombre, peso_unidad, precio_venta=None, cantidad_fabricar=None, cantidad_por_hora=None, valor_hora_ajustado_promedio=None, material=None, materiales=None, precios_materiales=None, iibb_porcentaje=None, moneda_precio='ARS'):
        self.nombre = nombre
        self.peso_unidad = float(peso_unidad)
        self.precio_venta = float(precio_venta) if precio_venta not in (None, '') else None
        self.moneda_precio = moneda_precio  # Nueva propiedad para guardar la moneda del precio
        self.cantidad_fabricar = int(float(cantidad_fabricar)) if cantidad_fabricar else 0
        self.cantidad_por_hora = float(cantidad_por_hora) if cantidad_por_hora else 0
        self.material = material or ''  # Compatibilidad con versiones previas (no usado si hay "materiales")
        self.materiales = materiales or []  # [{'nombre': str, 'kg_por_unidad': float}]
        self.precios_materiales = precios_materiales or {}
        self.iibb_porcentaje = float(iibb_porcentaje) if iibb_porcentaje not in (None, '') else 0.0  # IIBB porcentaje
        self.productos_por_kilo = 1 / self.peso_unidad if self.peso_unidad else 0
        self.costo_unitario_mp = self._calcular_costo_materiales()
        self.costo_total_mp = self.costo_unitario_mp * self.cantidad_fabricar if self.cantidad_fabricar else 0
        self.horas_necesarias = self.cantidad_fabricar / self.cantidad_por_hora if self.cantidad_fabricar and self.cantidad_por_hora else 0
        self.valor_hora_ajustado_promedio = float(valor_hora_ajustado_promedio) if valor_hora_ajustado_promedio else 0
        self.incidencia_mano_obra = self.horas_necesarias * self.valor_hora_ajustado_promedio if self.horas_necesarias and self.valor_hora_ajustado_promedio else 0
        # Costo unitario de mano de obra (Incidencia total / Cantidad a fabricar)
        self.costo_unitario_mano_obra = self.incidencia_mano_obra / self.cantidad_fabricar if self.cantidad_fabricar else 0
        # Costo base (MP + MO) sin IIBB
        self.costo_base_unitario = self.costo_unitario_mp + self.costo_unitario_mano_obra
        # IIBB unitario calculado como porcentaje del PRECIO DE VENTA
        self.iibb_unitario = (self.precio_venta * self.iibb_porcentaje / 100) if self.precio_venta and self.iibb_porcentaje > 0 else 0.0
        # Total IIBB (IIBB unitario * cantidad a fabricar)
        self.total_iibb = self.iibb_unitario * self.cantidad_fabricar if self.cantidad_fabricar else 0
        # Rentabilidad neta por unidad (precio de venta - costos - IIBB)
        self.rentabilidad_neta = None
        if self.precio_venta is not None:
            self.rentabilidad_neta = self.precio_venta - self.costo_base_unitario - self.iibb_unitario
        self.rentabilidad_neta_total = self.rentabilidad_neta * self.cantidad_fabricar if self.rentabilidad_neta is not None and self.cantidad_fabricar else None

    def _calcular_costo_materiales(self, valor_dolar_costos=None):
        total = 0.0
        try:
            for m in (self.materiales or []):
                nombre_mat = m.get('nombre')
                kg_por_unidad = float(m.get('kg_por_unidad', 0))
                
                # Verificar si tiene precio manual (desde pop-up de costos)
                if 'precio_manual' in m:
                    # Usar precio manual ingresado (ya está en pesos)
                    costo_pesos = float(m['precio_manual'])
                    total += kg_por_unidad * costo_pesos
                else:
                    # Usar precio del stock (comportamiento anterior)
                    precios = self.precios_materiales.get(nombre_mat, {})
                    costo_usd = float(precios.get('costo_kilo_usd', 0))
                    
                    # Usar valor del dólar del formulario de costos si está disponible, sino del stock
                    if valor_dolar_costos and valor_dolar_costos > 0:
                        valor_dolar = valor_dolar_costos
                    else:
                        valor_dolar = float(precios.get('valor_dolar', 0))
                    
                    total += kg_por_unidad * costo_usd * valor_dolar
        except Exception:
            pass
        return total

    def actualizar_precios_materiales(self, precios_materiales, valor_dolar_costos=None):
        self.precios_materiales = precios_materiales or {}
        self.costo_unitario_mp = self._calcular_costo_materiales(valor_dolar_costos)
        self.costo_total_mp = self.costo_unitario_mp * self.cantidad_fabricar if self.cantidad_fabricar else 0
        # Recalcular costo base e IIBB
        self.costo_base_unitario = self.costo_unitario_mp + self.costo_unitario_mano_obra
        self.iibb_unitario = (self.precio_venta * self.iibb_porcentaje / 100) if self.precio_venta and self.iibb_porcentaje > 0 else 0.0
        self.total_iibb = self.iibb_unitario * self.cantidad_fabricar if self.cantidad_fabricar else 0
        # Recalcular rentabilidad neta
        if self.precio_venta is not None:
            self.rentabilidad_neta = self.precio_venta - self.costo_base_unitario - self.iibb_unitario
            self.rentabilidad_neta_total = self.rentabilidad_neta * self.cantidad_fabricar if self.cantidad_fabricar else None

    def actualizar_precio_venta(self, nuevo_precio_venta):
        """Actualizar el precio de venta y recalcular IIBB y rentabilidad"""
        self.precio_venta = float(nuevo_precio_venta) if nuevo_precio_venta not in (None, '') else None
        # Recalcular IIBB basado en el nuevo precio de venta
        self.iibb_unitario = (self.precio_venta * self.iibb_porcentaje / 100) if self.precio_venta and self.iibb_porcentaje > 0 else 0.0
        self.total_iibb = self.iibb_unitario * self.cantidad_fabricar if self.cantidad_fabricar else 0
        # Recalcular rentabilidad neta
        if self.precio_venta is not None:
            self.rentabilidad_neta = self.precio_venta - self.costo_base_unitario - self.iibb_unitario
            self.rentabilidad_neta_total = self.rentabilidad_neta * self.cantidad_fabricar if self.cantidad_fabricar else None
        else:
            self.rentabilidad_neta = None
            self.rentabilidad_neta_total = None

    def actualizar_valor_hora_promedio(self, nuevo_valor_hora_promedio):
        """Actualizar el valor hora promedio y recalcular todos los costos relacionados"""
        self.valor_hora_ajustado_promedio = float(nuevo_valor_hora_promedio)
        self.incidencia_mano_obra = self.horas_necesarias * self.valor_hora_ajustado_promedio if self.horas_necesarias and self.valor_hora_ajustado_promedio else 0
        self.costo_unitario_mano_obra = self.incidencia_mano_obra / self.cantidad_fabricar if self.cantidad_fabricar else 0
        # Recalcular costo base y IIBB basado en precio de venta
        self.costo_base_unitario = self.costo_unitario_mp + self.costo_unitario_mano_obra
        self.iibb_unitario = (self.precio_venta * self.iibb_porcentaje / 100) if self.precio_venta and self.iibb_porcentaje > 0 else 0.0
        self.total_iibb = self.iibb_unitario * self.cantidad_fabricar if self.cantidad_fabricar else 0
        # Recalcular rentabilidad neta (incluye IIBB)
        if self.precio_venta is not None:
            self.rentabilidad_neta = self.precio_venta - self.costo_base_unitario - self.iibb_unitario
        else:
            self.rentabilidad_neta = None
        self.rentabilidad_neta_total = self.rentabilidad_neta * self.cantidad_fabricar if self.rentabilidad_neta is not None and self.cantidad_fabricar else None

    def to_dict(self):
        # Extraer componentes del nombre (formato: "Familia - Medida - Característica")
        nombre_parts = self.nombre.split(' - ')
        if len(nombre_parts) >= 3:
            familia = nombre_parts[0]
            medida = nombre_parts[1]
            caracteristica = ' - '.join(nombre_parts[2:])
        else:
            # Fallback para nombres antiguos
            familia = self.nombre
            medida = ''
            caracteristica = ''
        
        return {
            'Familia': familia,
            'Medida': medida,
            'Característica': caracteristica,
            'Peso por unidad (kg)': round(self.peso_unidad, 5),
            'Cantidad a fabricar': self.cantidad_fabricar,
            'Cantidad producida por hora': round(self.cantidad_por_hora, 5),
            'Productos por kilo': round(self.productos_por_kilo, 5),
            'Costo unitario MP': round(self.costo_unitario_mp, 5),
            'Costo total MP': round(self.costo_total_mp, 5),
            'Horas necesarias': round(self.horas_necesarias, 5),
            'Valor hora ajustado promedio': round(self.valor_hora_ajustado_promedio, 5),
            'Incidencia mano de obra': round(self.incidencia_mano_obra, 5),
            'Costo unitario mano de obra': round(self.costo_unitario_mano_obra, 5),
            'Precio de venta': round(self.precio_venta, 5) if self.precio_venta else '',
            'Moneda precio': getattr(self, 'moneda_precio', 'ARS'),
            'Precio con moneda': f"${round(self.precio_venta, 2)} ({getattr(self, 'moneda_precio', 'ARS')})" if self.precio_venta else '',
            'Rentabilidad neta': round(self.rentabilidad_neta, 5) if self.rentabilidad_neta is not None else '',
            'Rentabilidad neta total': round(self.rentabilidad_neta_total, 5) if self.rentabilidad_neta_total is not None else ''
        }

    def to_save_dict(self):
        return {
            'nombre': self.nombre,
            'material': self.material,
            'peso_unidad': self.peso_unidad,
            'precio_venta': self.precio_venta,
            'cantidad_fabricar': self.cantidad_fabricar,
            'cantidad_por_hora': self.cantidad_por_hora,
            'materiales': self.materiales,
            'iibb_porcentaje': self.iibb_porcentaje,
            'moneda_precio': getattr(self, 'moneda_precio', 'ARS')
        }

    @staticmethod
    def from_save_dict(data, valor_hora_promedio, precios_materiales=None):
        return Producto(
            data.get('nombre',''),
            data.get('peso_unidad',0),
            data.get('precio_venta',None),
            data.get('cantidad_fabricar',0),
            data.get('cantidad_por_hora',0),
            valor_hora_promedio,
            data.get('material',''),
            data.get('materiales', []),
            precios_materiales,
            data.get('iibb_porcentaje', 0),
            data.get('moneda_precio', 'ARS')
        )

class Empleado:
    def __init__(self, nombre, valor_hora, dias_trabajados, horas_dia, ausencias, vacaciones, feriados, lic_enfermedad, otras_licencias, horas_descanso, carga_social, horas_extras, feriados_trabajados):
        self.nombre = nombre
        self.valor_hora = float(valor_hora)
        self.dias_trabajados = int(dias_trabajados)
        self.horas_dia = float(horas_dia)
        self.ausencias = int(ausencias)
        self.vacaciones = int(vacaciones)
        self.feriados = int(feriados)
        self.lic_enfermedad = int(lic_enfermedad)
        self.otras_licencias = int(otras_licencias)
        self.horas_descanso = float(horas_descanso)
        self.carga_social = float(carga_social)
        self.horas_extras = float(horas_extras)
        self.feriados_trabajados = int(feriados_trabajados)
        # Cálculos base de horas según matriz de incidencia
        # Días netos de trabajo = días posibles de trabajo - días no productivos
        # Días posibles de trabajo = 365 - 104 (fines de semana) = 261
        dias_posibles_trabajo = 261  # Días del año menos fines de semana
        dias_no_productivos = self.feriados + self.vacaciones + self.lic_enfermedad + self.otras_licencias + self.ausencias
        self.dias_efectivos = dias_posibles_trabajo - dias_no_productivos
        # Horas productivas = días efectivos × (horas_dia - horas_descanso)
        # El descanso no cuenta como parte de las horas productivas
        self.horas_productivas = self.dias_efectivos * (self.horas_dia - self.horas_descanso)
        # Vacaciones, feriados, licencias y ausencias se calculan con el día completo
        self.horas_vacaciones = self.vacaciones * self.horas_dia
        self.horas_feriados = self.feriados * self.horas_dia
        self.horas_lic_enfermedad = self.lic_enfermedad * self.horas_dia
        self.horas_otras_licencias = self.otras_licencias * self.horas_dia
        self.horas_ausencias = self.ausencias * self.horas_dia
        # Horas de descanso totales (para registro, no para cálculo de porcentaje)
        self.horas_descanso_total = self.horas_descanso * self.dias_efectivos
        # Total horas pagas = días efectivos × horas por día
        self.total_horas_pagas = self.dias_efectivos * self.horas_dia
        self.horas_no_productivas = self.horas_vacaciones + self.horas_feriados + self.horas_lic_enfermedad + self.horas_otras_licencias + self.horas_ausencias
        self.horas_totales = self.horas_productivas + self.horas_no_productivas + self.horas_extras + (self.feriados_trabajados * self.horas_dia)
        # Derivados
        self._calcular_derivados()

    def _calcular_derivados(self):
        # MATRIZ DE MANO DE OBRA - Metodología según matriz de Excel
        # Base de cálculo: Total horas pagas (no horas productivas)
        base_horas_pagas = self.total_horas_pagas if self.total_horas_pagas else 1
        
        # COLUMNA 1: Porcentajes base (sobre horas pagas)
        self.jornal_productivas = 100.00  # Base siempre 100%
        self.jornal_feriados = (self.horas_feriados / base_horas_pagas) * 100
        self.jornal_vacaciones = (self.horas_vacaciones / base_horas_pagas) * 100
        self.jornal_lic_enfermedad = (self.horas_lic_enfermedad / base_horas_pagas) * 100
        self.jornal_otras_licencias = (self.horas_otras_licencias / base_horas_pagas) * 100
        self.jornal_ausencias = (self.horas_ausencias / base_horas_pagas) * 100
        self.jornal_descanso = (self.horas_descanso_total / base_horas_pagas) * 100
        
        # Horas extras y feriados trabajados (con recargos)
        self.jornal_horas_extras = (self.horas_extras / base_horas_pagas) * 100 * 1.5  # 50% adicional
        self.jornal_feriados_trabajados = ((self.feriados_trabajados * self.horas_dia) / base_horas_pagas) * 100 * 2.0  # 100% adicional
        
        # Subtotal Columna 1 (solo componentes base, sin carga social)
        self.subtotal_col1 = (
            self.jornal_productivas + self.jornal_feriados + self.jornal_vacaciones +
            self.jornal_lic_enfermedad + self.jornal_otras_licencias + self.jornal_ausencias + 
            self.jornal_descanso + self.jornal_horas_extras + self.jornal_feriados_trabajados
        )
        
        # SAC sobre subtotal Columna 1
        self.sac_col1 = self.subtotal_col1 / 12
        
        # Total Columna 1 (incluye SAC)
        self.total_col1 = self.subtotal_col1 + self.sac_col1
        
        # COLUMNA 2: Carga social (43% aplicada a TODOS los componentes)
        c = self.carga_social / 100
        
        # Carga social aplicada a TODOS los componentes de la Columna 1
        self.ccss_productivas = self.jornal_productivas * c
        self.ccss_feriados = self.jornal_feriados * c
        self.ccss_vacaciones = self.jornal_vacaciones * c
        self.ccss_lic_enfermedad = self.jornal_lic_enfermedad * c
        self.ccss_otras_licencias = self.jornal_otras_licencias * c
        self.ccss_ausencias = self.jornal_ausencias * c
        self.ccss_descanso = self.jornal_descanso * c
        self.ccss_horas_extras = self.jornal_horas_extras * c
        self.ccss_feriados_trabajados = self.jornal_feriados_trabajados * c
        
        self.subtotal_col2 = (
            self.ccss_productivas + self.ccss_feriados + self.ccss_vacaciones +
            self.ccss_lic_enfermedad + self.ccss_otras_licencias + self.ccss_ausencias + 
            self.ccss_descanso + self.ccss_horas_extras + self.ccss_feriados_trabajados
        )
        self.sac_col2 = self.subtotal_col2 / 12
        self.total_col2 = self.subtotal_col2 + self.sac_col2
        
        # COLUMNA 3: Total final (Columna 1 + Columna 2)
        self.indice_ajustado = self.total_col1 + self.total_col2
        
        # Valor hora ajustado
        self.valor_hora_ajustado = round(self.valor_hora * (self.indice_ajustado / 100), 5)
        
        # Horas trabajadas (productivas + extras + feriados trabajados)
        self.horas_trabajadas = self.horas_productivas + self.horas_extras + (self.feriados_trabajados * self.horas_dia)
        self.horas_descansadas = self.horas_no_productivas
        
        # Costo total ajustado
        self.costo_total_ajustado = self.horas_productivas * self.valor_hora_ajustado
        
        # Valor hora efectivo (igual al ajustado en este caso)
        self.valor_hora_efectivo = self.valor_hora_ajustado
        
        # Mantener compatibilidad con nombres anteriores
        self.subtotal = self.subtotal_col1
        self.sac = self.sac_col1
        
        # Agregar atributos para compatibilidad con to_dict()
        self.total_productivas = self.jornal_productivas + self.ccss_productivas
        self.total_vacaciones = self.jornal_vacaciones + self.ccss_vacaciones
        self.total_feriados = self.jornal_feriados + self.ccss_feriados
        self.total_lic_enfermedad = self.jornal_lic_enfermedad + self.ccss_lic_enfermedad
        self.total_otras_licencias = self.jornal_otras_licencias + self.ccss_otras_licencias
        self.total_ausencias = self.jornal_ausencias + self.ccss_ausencias
        self.total_descanso = self.jornal_descanso + self.ccss_descanso
        self.total_horas_extras = self.jornal_horas_extras + self.ccss_horas_extras
        self.total_feriados_trabajados = self.jornal_feriados_trabajados + self.ccss_feriados_trabajados

    def to_save_dict(self):
        return {
            'nombre': self.nombre,
            'valor_hora': self.valor_hora,
            'dias_trabajados': self.dias_trabajados,
            'horas_dia': self.horas_dia,
            'ausencias': self.ausencias,
            'vacaciones': self.vacaciones,
            'feriados': self.feriados,
            'lic_enfermedad': self.lic_enfermedad,
            'otras_licencias': self.otras_licencias,
            'horas_descanso': self.horas_descanso,
            'carga_social': self.carga_social,
            'horas_extras': self.horas_extras,
            'feriados_trabajados': self.feriados_trabajados
        }

    @staticmethod
    def from_save_dict(data):
        return Empleado(
            data.get('nombre',''),
            data.get('valor_hora',0),
            data.get('dias_trabajados',0),
            data.get('horas_dia',0),
            data.get('ausencias',0),
            data.get('vacaciones',0),
            data.get('feriados',0),
            data.get('lic_enfermedad',0),
            data.get('otras_licencias',0),
            data.get('horas_descanso',0),
            data.get('carga_social',0),
            data.get('horas_extras',0),
            data.get('feriados_trabajados',0)
        )

    def to_dict(self):
        return {
            'Nombre': self.nombre,
            'Valor hora': self.valor_hora,
            'Días': self.dias_trabajados,
            'Horas/día': self.horas_dia,
            'Ausencias': self.ausencias,
            'Vacaciones': self.vacaciones,
            'Feriados': self.feriados,
            'Lic. Enfermedad': self.lic_enfermedad,
            'Otras licencias': self.otras_licencias,
            'Horas descanso/día': self.horas_descanso,
            'Carga social (%)': self.carga_social,
            'Horas extras': self.horas_extras,
            'Feriados trabajados': self.feriados_trabajados,
            'Horas trabajadas': round(self.horas_trabajadas,5),
            'Horas descansadas': round(self.horas_descansadas,5),
            'Jornal Productivas (%)': round(self.jornal_productivas,5),
            'Jornal Vacaciones (%)': round(self.jornal_vacaciones,5),
            'Jornal Feriados (%)': round(self.jornal_feriados,5),
            'Jornal Lic. Enfermedad (%)': round(self.jornal_lic_enfermedad,5),
            'Jornal Otras licencias (%)': round(self.jornal_otras_licencias,5),
            'Jornal Ausencias (%)': round(self.jornal_ausencias,5),
            'Jornal Descanso (%)': round(self.jornal_descanso,5),
            'Jornal Horas extras (%)': round(self.jornal_horas_extras,5),
            'Jornal Feriados trabajados (%)': round(self.jornal_feriados_trabajados,5),
            'CCSS Productivas (%)': round(self.ccss_productivas,5),
            'CCSS Vacaciones (%)': round(self.ccss_vacaciones,5),
            'CCSS Feriados (%)': round(self.ccss_feriados,5),
            'CCSS Lic. Enfermedad (%)': round(self.ccss_lic_enfermedad,5),
            'CCSS Otras licencias (%)': round(self.ccss_otras_licencias,5),
            'CCSS Ausencias (%)': round(self.ccss_ausencias,5),
            'CCSS Descanso (%)': round(self.ccss_descanso,5),
            'CCSS Horas extras (%)': round(self.ccss_horas_extras,5),
            'CCSS Feriados trabajados (%)': round(self.ccss_feriados_trabajados,5),
            'Total Productivas (%)': round(self.total_productivas,5),
            'Total Vacaciones (%)': round(self.total_vacaciones,5),
            'Total Feriados (%)': round(self.total_feriados,5),
            'Total Lic. Enfermedad (%)': round(self.total_lic_enfermedad,5),
            'Total Otras licencias (%)': round(self.total_otras_licencias,5),
            'Total Ausencias (%)': round(self.total_ausencias,5),
            'Total Descanso (%)': round(self.total_descanso,5),
            'Total Horas extras (%)': round(self.total_horas_extras,5),
            'Total Feriados trabajados (%)': round(self.total_feriados_trabajados,5),
            'Sub-Total (%)': round(self.subtotal,5),
            'SAC (%)': round(self.sac,5),
            'Índice ajustado (%)': round(self.indice_ajustado,5),
            'Costo total ajustado': round(self.costo_total_ajustado,5),
            'Valor hora ajustado': round(self.valor_hora_ajustado,5),
            'Valor hora efectivo': round(self.valor_hora_efectivo,5)
        }

class App:
    def __init__(self, root):
        self.root = root
        self.root.title('FABINSA CONTROL')
        self.root.geometry('1600x900')
        self.root.configure(bg=STYLE_CONFIG['bg_color'])
        self.root.state('zoomed')  # Maximizar ventana
        
        # Configurar icono de la ventana
        try:
            icon_path = 'logo_fabinsa.png'
            if os.path.exists(icon_path):
                icon_image = Image.open(icon_path)
                icon_photo = ImageTk.PhotoImage(icon_image)
                self.root.iconphoto(True, icon_photo)
        except Exception as e:
            pass  # Si falla, mantener icono por defecto
        
        # Configurar el estilo de ttk
        self.configurar_estilos()
        
        self.productos = []
        self.empleados = []
        # Stock
        self.stock_mp = []  # {'nombre': str, 'kg': float, 'costo_kg_usd': float, 'valor_dolar': float}
        self.stock_prod = []  # {'nombre': str, 'cantidad': int, 'peso_unidad': float} - Productos fabricados
        self.productos_reventa = []  # [{'nombre': str, 'cantidad': int, 'costo_unitario': float}] - Productos de reventa
        # Materiales en formulario de producto actual
        self.materiales_producto_form = []  # [{'nombre': str, 'kg_por_unidad': float}]
        # Métricas de pedidos completados
        self.metricas = []  # [{'fecha': str, 'producto': str, 'cantidad': int, ...}]
        # Ventas
        self.ventas = []  # [{'fecha': str, 'producto': str, 'cantidad': int, 'precio_unit': float, 'descuento_pct': float, 'ingreso_bruto': float, 'ingreso_neto': float}]
        self.venta_display_map = {}
        # Planilla de costos (independiente, no afecta stock)
        self.items_costos = []  # Lista de objetos Producto para simulación
        self.materiales_costos_form = []  # Materiales temporales para el formulario de costos
        
        # Variables para rastrear elementos en edición (mantener visible mientras se edita)
        self.editando_producto_index = None
        self.editando_empleado_index = None
        self.editando_mp_index = None
        self.editando_pr_index = None
        self.editando_costos_index = None
        
        # Logs de compras
        self.logs_compra_mp = []  # Logs de compras de materia prima
        self.logs_compra_prod = []  # Logs de compras de productos
        
        # Registrar validación para limitar decimales
        self.vcmd_decimal = (root.register(self.validar_decimal), '%P')
        
        self.crear_widgets()
        # Cargar estado desde disco
        self.load_state()

    def validar_decimal(self, value):
        """Valida que el valor tenga máximo 5 decimales"""
        if value == '' or value == '-':
            return True
        try:
            # Permitir números con punto decimal
            if '.' in value:
                partes = value.split('.')
                if len(partes) == 2 and len(partes[1]) <= 5:
                    float(value)
                    return True
                elif len(partes) == 2 and len(partes[1]) > 5:
                    return False
                else:
                    return False
            else:
                float(value)
                return True
        except ValueError:
            return False

    def crear_frame_con_scroll(self, parent):
        """Crear un frame con scroll vertical para cualquier pestaña"""
        # Canvas principal
        canvas = tk.Canvas(parent, bg=STYLE_CONFIG['bg_color'], highlightthickness=0)
        scrollbar = ttk.Scrollbar(parent, orient='vertical', command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg=STYLE_CONFIG['bg_color'])
        
        # Configurar el canvas para evitar scroll horizontal
        def configure_scroll_region(event):
            canvas.configure(scrollregion=canvas.bbox('all'))
            # Asegurar que el frame no sea más ancho que el canvas
            canvas_width = canvas.winfo_width()
            if canvas_width > 1:  # Evitar división por cero
                canvas.itemconfig(canvas.find_all()[0], width=canvas_width)
        
        scrollable_frame.bind('<Configure>', configure_scroll_region)
        canvas.bind('<Configure>', configure_scroll_region)
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor='nw')
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Pack layout
        canvas.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        # Habilitar scroll con rueda del mouse
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        
        scrollable_frame.bind('<Enter>', lambda e: canvas.bind_all("<MouseWheel>", _on_mousewheel))
        scrollable_frame.bind('<Leave>', lambda e: canvas.unbind_all("<MouseWheel>"))
        
        return scrollable_frame
    
    def crear_card_moderno(self, parent, **kwargs):
        """Crear un card moderno con sombra simulada (doble frame)"""
        # Frame exterior (sombra)
        shadow_frame = tk.Frame(parent, bg=STYLE_CONFIG['card_shadow'])
        
        # Frame interior (card real)
        card_frame = tk.Frame(shadow_frame, bg=STYLE_CONFIG['card_bg'])
        card_frame.pack(padx=3, pady=3, fill='both', expand=True)
        
        return shadow_frame, card_frame
    
    def crear_metric_card(self, parent, titulo, valor, icono='📊', bg_color=None):
        """Crear un widget de métrica estilo dashboard moderno"""
        bg = bg_color or STYLE_CONFIG['metric_bg_1']
        
        # Frame de la métrica con color de fondo
        metric_frame = tk.Frame(parent, bg=bg, relief='solid', borderwidth=1, 
                               highlightbackground=STYLE_CONFIG['border_color'],
                               highlightthickness=1)
        metric_frame.pack(side='left', padx=8, pady=8, ipadx=20, ipady=15)
        
        # Icono
        icon_label = tk.Label(metric_frame, text=icono, font=('Segoe UI', 28), 
                             bg=bg, fg=STYLE_CONFIG['primary_color'])
        icon_label.pack(side='left', padx=(0, 15))
        
        # Container de texto
        text_container = tk.Frame(metric_frame, bg=bg)
        text_container.pack(side='left')
        
        # Título
        title_label = tk.Label(text_container, text=titulo, 
                              font=('Segoe UI', 9, 'bold'),
                              fg=STYLE_CONFIG['text_light'], bg=bg)
        title_label.pack(anchor='w')
        
        # Valor
        value_label = tk.Label(text_container, text=valor,
                              font=('Segoe UI', 18, 'bold'),
                              fg=STYLE_CONFIG['text_color'], bg=bg)
        value_label.pack(anchor='w')
        
        return metric_frame, value_label
    
    def crear_boton_navegacion(self, texto, icono, index):
        """Crear un botón de navegación premium para el sidebar"""
        # Container del botón con indicador mejorado
        btn_container = tk.Frame(self.nav_buttons_frame, bg=STYLE_CONFIG['bg_darker'])
        btn_container.pack(fill='x', padx=10, pady=4)
        
        # Indicador lateral (barra izquierda) más ancha
        indicator = tk.Frame(btn_container, bg=STYLE_CONFIG['bg_darker'], width=5)
        indicator.pack(side='left', fill='y')
        
        # Botón principal con mejor diseño
        btn = tk.Button(
            btn_container,
            text=f'  {icono}  {texto}',
            font=('Segoe UI', 12, 'bold'),
            fg=STYLE_CONFIG['text_white'],
            bg=STYLE_CONFIG['bg_darker'],
            activebackground=STYLE_CONFIG['primary_color'],
            activeforeground=STYLE_CONFIG['text_white'],
            relief='flat',
            bd=0,
            padx=20,
            pady=18,
            anchor='w',
            cursor='hand2',
            command=lambda: self.cambiar_tab(index)
        )
        btn.pack(side='left', fill='both', expand=True)
        
        # Efecto hover mejorado
        def on_enter(e):
            if self.current_tab_index != index:
                btn.config(bg=STYLE_CONFIG['primary_hover'])
        
        def on_leave(e):
            if self.current_tab_index != index:
                btn.config(bg=STYLE_CONFIG['bg_darker'])
        
        btn.bind('<Enter>', on_enter)
        btn.bind('<Leave>', on_leave)
        
        # Guardar referencia al botón y al indicador
        self.nav_buttons[index] = {'btn': btn, 'indicator': indicator, 'container': btn_container}
        return btn
    
    def cambiar_tab(self, index):
        """Cambiar a una pestaña específica y actualizar el estilo del sidebar"""
        self.tabs.select(index)
        self.current_tab_index = index
        
        # Actualizar estilos de los botones e indicadores
        for idx, components in self.nav_buttons.items():
            btn = components['btn']
            indicator = components['indicator']
            
            if idx == index:
                # Botón activo con estilo premium
                btn.config(bg=STYLE_CONFIG['primary_color'], fg=STYLE_CONFIG['text_white'])
                indicator.config(bg=STYLE_CONFIG['accent_color'], width=5)
            else:
                # Botón inactivo
                btn.config(bg=STYLE_CONFIG['bg_darker'], fg=STYLE_CONFIG['text_white'])
                indicator.config(bg=STYLE_CONFIG['bg_darker'], width=5)
    
    def inicializar_navegacion(self):
        """Inicializar los botones de navegación después de crear todas las pestañas"""
        tabs_info = [
            ('Producción', '🏭', 0),
            ('Empleados', '👥', 1),
            ('Stock', '📦', 2),
            ('Métricas', '📊', 3),
            ('Ventas', '💵', 4),
            ('Compra', '🛒', 5),
            ('Costos', '💰', 6)
        ]
        
        for texto, icono, index in tabs_info:
            self.crear_boton_navegacion(texto, icono, index)
        
        # Seleccionar la primera pestaña por defecto
        self.cambiar_tab(0)

    def configurar_estilos(self):
        """Configurar estilos ULTRA PREMIUM de nivel empresarial moderno"""
        style = ttk.Style()
        style.theme_use('clam')
        
        # Estilo para las pestañas PRINCIPALES (Notebook) - OCULTAS (controladas por sidebar)
        style.configure('TNotebook', 
                       background=STYLE_CONFIG['bg_color'], 
                       borderwidth=0, 
                       relief='flat',
                       tabmargins=[0, 0, 0, 0])
        
        # Ocultar las pestañas del notebook principal
        style.layout('TNotebook', [])
        style.layout('TNotebook.Tab', [])
        
        # Estilo para SUB-NOTEBOOKS (como el de Stock) - VISIBLES CON LAYOUT COMPLETO
        style.configure('SubNotebook.TNotebook', 
                       background=STYLE_CONFIG['bg_color'], 
                       borderwidth=0, 
                       relief='flat',
                       tabmargins=[8, 8, 8, 0])
        
        # IMPORTANTE: Restaurar el layout para SubNotebook para que las pestañas sean visibles
        style.layout('SubNotebook.TNotebook.Tab', [
            ('Notebook.tab', {
                'sticky': 'nswe',
                'children': [
                    ('Notebook.padding', {
                        'side': 'top',
                        'sticky': 'nswe',
                        'children': [
                            ('Notebook.label', {'side': 'top', 'sticky': ''})
                        ]
                    })
                ]
            })
        ])
        
        style.configure('SubNotebook.TNotebook.Tab', 
                       background=STYLE_CONFIG['light_bg'],
                       foreground=STYLE_CONFIG['text_muted'],
                       padding=[24, 14],
                       font=('Segoe UI', 11, 'bold'),
                       borderwidth=0,
                       relief='flat')
        style.map('SubNotebook.TNotebook.Tab',
                 background=[('selected', STYLE_CONFIG['primary_color']),
                           ('active', STYLE_CONFIG['primary_light']),
                           ('!selected', STYLE_CONFIG['light_bg'])],
                 foreground=[('selected', 'white'),
                           ('active', 'white'),
                           ('!selected', STYLE_CONFIG['text_muted'])],
                 padding=[('selected', [24, 14]),
                         ('!selected', [24, 14])],
                 font=[('selected', ('Segoe UI', 11, 'bold')),
                      ('!selected', ('Segoe UI', 11, 'bold'))])
        
        # Botones personalizados - DISEÑO PROFESIONAL PREMIUM
        style.configure('Custom.TButton',
                       background=STYLE_CONFIG['primary_color'],
                       foreground=STYLE_CONFIG['text_white'],
                       font=('Segoe UI', 11, 'bold'),
                       padding=[24, 14],
                       relief='flat',
                       borderwidth=0,
                       focuscolor='none')
        style.map('Custom.TButton',
                 background=[('active', STYLE_CONFIG['primary_hover']),
                           ('pressed', STYLE_CONFIG['primary_dark']),
                           ('disabled', STYLE_CONFIG['text_muted'])],
                 foreground=[('disabled', STYLE_CONFIG['text_muted'])])
        
        # Estilo para los LabelFrame - Cards elegantes con mejor diseño
        style.configure('TLabelframe', 
                       background=STYLE_CONFIG['card_bg'],
                       borderwidth=1,
                       relief='solid',
                       bordercolor=STYLE_CONFIG['border_color'])
        style.configure('TLabelframe.Label',
                       background=STYLE_CONFIG['card_bg'],
                       foreground=STYLE_CONFIG['primary_color'],
                       font=('Segoe UI', 14, 'bold'),
                       padding=[10, 6])
        
        # Estilo para botones de acción (verde) - PREMIUM
        style.configure('Action.TButton',
                       background=STYLE_CONFIG['success_color'],
                       foreground=STYLE_CONFIG['text_white'],
                       font=('Segoe UI', 11, 'bold'),
                       padding=[22, 13],
                       relief='flat',
                       borderwidth=0)
        style.map('Action.TButton',
                 background=[('active', STYLE_CONFIG['success_light']),
                          ('pressed', STYLE_CONFIG['success_dark'])])
        
        # Estilo para botones de peligro (rojo) - PREMIUM
        style.configure('Danger.TButton',
                       background=STYLE_CONFIG['danger_color'],
                       foreground=STYLE_CONFIG['text_white'],
                       font=('Segoe UI', 11, 'bold'),
                       padding=[22, 13],
                       relief='flat',
                       borderwidth=0)
        style.map('Danger.TButton',
                 background=[('active', STYLE_CONFIG['danger_light']),
                          ('pressed', STYLE_CONFIG['danger_dark'])])
        
        # Estilo para botones de advertencia (naranja) - PREMIUM
        style.configure('Warning.TButton',
                       background=STYLE_CONFIG['warning_color'],
                       foreground=STYLE_CONFIG['text_white'],
                       font=('Segoe UI', 11, 'bold'),
                       padding=[22, 13],
                       relief='flat',
                       borderwidth=0)
        style.map('Warning.TButton',
                 background=[('active', STYLE_CONFIG['warning_light']),
                          ('pressed', STYLE_CONFIG['warning_dark'])])
        
        # Estilo para botón del header (exportación) - DESTACADO PREMIUM
        style.configure('Header.TButton',
                       background=STYLE_CONFIG['success_color'],
                       foreground=STYLE_CONFIG['text_white'],
                       font=('Segoe UI', 12, 'bold'),
                       padding=[28, 16],
                       relief='flat',
                       borderwidth=0)
        style.map('Header.TButton',
                 background=[('active', STYLE_CONFIG['success_light']),
                          ('pressed', STYLE_CONFIG['success_dark'])])
        
        # Estilo para los Treeview (tablas) - DISEÑO PROFESIONAL
        style.configure('Custom.Treeview',
                       background=STYLE_CONFIG['card_bg'],
                       foreground=STYLE_CONFIG['text_color'],
                       fieldbackground=STYLE_CONFIG['card_bg'],
                       font=('Segoe UI', 10),
                       rowheight=40,  # Más espacio vertical
                       borderwidth=1,
                       relief='solid',
                       bordercolor=STYLE_CONFIG['border_color'])
        style.configure('Custom.Treeview.Heading',
                       background=STYLE_CONFIG['bg_dark'],
                       foreground=STYLE_CONFIG['header_text'],
                       font=('Segoe UI', 11, 'bold'),
                       relief='flat',
                       borderwidth=0,
                       padding=[14, 12])  # Más padding en headers
        style.map('Custom.Treeview.Heading',
                 background=[('active', STYLE_CONFIG['primary_color'])],
                 foreground=[('active', STYLE_CONFIG['text_white'])])
        style.map('Custom.Treeview',
                 background=[('selected', STYLE_CONFIG['primary_color'])],
                 foreground=[('selected', STYLE_CONFIG['text_white'])],
                 fieldbackground=[('selected', STYLE_CONFIG['primary_color'])])
        
        # Estilo para Entry (campos de texto) - MEJORADO
        style.configure('TEntry',
                       fieldbackground=STYLE_CONFIG['light_bg'],
                       foreground=STYLE_CONFIG['text_color'],
                       borderwidth=1,
                       relief='solid',
                       bordercolor=STYLE_CONFIG['border_color'],
                       font=('Segoe UI', 11))
        
        # Estilo para Combobox - MEJORADO
        style.configure('TCombobox',
                       fieldbackground=STYLE_CONFIG['light_bg'],
                       background=STYLE_CONFIG['card_bg'],
                       foreground=STYLE_CONFIG['text_color'],
                       arrowcolor=STYLE_CONFIG['primary_color'],
                       borderwidth=1,
                       font=('Segoe UI', 11))
        style.map('TCombobox',
                 fieldbackground=[('readonly', STYLE_CONFIG['light_bg'])],
                 background=[('readonly', STYLE_CONFIG['card_bg'])],
                 arrowcolor=[('active', STYLE_CONFIG['primary_hover'])])

    def crear_widgets(self):
        # ========== HEADER PROFESIONAL PREMIUM ==========
        # Frame principal del header con altura optimizada
        header_main = tk.Frame(self.root, bg=STYLE_CONFIG['bg_darker'], height=110)
        header_main.pack(fill='x', padx=0, pady=0)
        header_main.pack_propagate(False)
        
        # Frame principal con gradiente elegante
        gradient_frame = tk.Frame(header_main, bg=STYLE_CONFIG['header_bg'], height=110)
        gradient_frame.pack(fill='both', expand=True)
        
        # Logo en la parte superior izquierda con mejor tamaño
        try:
            logo_path = 'logo_fabinsa.png'
            if os.path.exists(logo_path):
                logo_image = Image.open(logo_path)
                # Logo más grande y profesional
                logo_height = 75
                aspect_ratio = logo_image.width / logo_image.height
                logo_width = int(logo_height * aspect_ratio)
                logo_image = logo_image.resize((logo_width, logo_height), Image.Resampling.LANCZOS)
                logo_photo = ImageTk.PhotoImage(logo_image)
                
                logo_label = tk.Label(gradient_frame, image=logo_photo, bg=STYLE_CONFIG['header_bg'])
                logo_label.image = logo_photo  # Mantener referencia
                logo_label.pack(side='left', padx=35, pady=17)
        except Exception as e:
            print(f"Error cargando logo: {e}")
        
        # Container del título con diseño centrado mejorado
        title_container = tk.Frame(gradient_frame, bg=STYLE_CONFIG['header_bg'])
        title_container.place(relx=0.5, rely=0.5, anchor='center')
        
        # Título principal con tipografía premium
        title_label = tk.Label(title_container, 
                              text='📊 FABINSA CONTROL',
                              font=('Segoe UI', 28, 'bold'),
                              fg=STYLE_CONFIG['text_white'],
                              bg=STYLE_CONFIG['header_bg'])
        title_label.pack(side='top')
        
        # Subtítulo más elegante y profesional
        subtitle_label = tk.Label(title_container,
                                 text='Sistema Profesional de Gestión y Análisis Empresarial',
                                 font=('Segoe UI', 12),
                                 fg=STYLE_CONFIG['header_subtitle'],
                                 bg=STYLE_CONFIG['header_bg'])
        subtitle_label.pack(side='top', pady=(6, 0))
        
        # Botón de exportación con diseño destacado
        btn_exportar_general = ttk.Button(gradient_frame, 
                                         text='📊 Exportar a Excel', 
                                         command=self.exportar_excel, 
                                         style='Header.TButton')
        btn_exportar_general.pack(side='right', padx=35, pady=27)
        
        # Separador premium con gradiente de color
        separator = tk.Frame(self.root, bg=STYLE_CONFIG['primary_color'], height=4)
        separator.pack(fill='x')
        
        # Línea sutil adicional para profundidad
        separator2 = tk.Frame(self.root, bg=STYLE_CONFIG['border_light'], height=1)
        separator2.pack(fill='x')
        
        # ========== LAYOUT PRINCIPAL CON SIDEBAR ==========
        main_container = tk.Frame(self.root, bg=STYLE_CONFIG['bg_color'])
        main_container.pack(fill='both', expand=True)
        
        # ===== SIDEBAR IZQUIERDA (Navegación Premium) =====
        self.sidebar = tk.Frame(main_container, bg=STYLE_CONFIG['bg_darker'], width=260)
        self.sidebar.pack(side='left', fill='y')
        self.sidebar.pack_propagate(False)
        
        # Borde derecho del sidebar con color premium
        tk.Frame(self.sidebar, bg=STYLE_CONFIG['primary_color'], width=3).pack(side='right', fill='y')
        
        # Container interior del sidebar
        sidebar_inner = tk.Frame(self.sidebar, bg=STYLE_CONFIG['bg_darker'])
        sidebar_inner.pack(fill='both', expand=True)
        
        # Título del sidebar con mejor diseño
        sidebar_title = tk.Label(sidebar_inner,
                                text='📋 NAVEGACIÓN',
                                font=('Segoe UI', 14, 'bold'),
                                fg=STYLE_CONFIG['text_white'],
                                bg=STYLE_CONFIG['bg_darker'],
                                pady=25)
        sidebar_title.pack(fill='x')
        
        # Separador superior elegante
        tk.Frame(sidebar_inner, bg=STYLE_CONFIG['primary_color'], height=4).pack(fill='x')
        
        # Container para los botones de navegación con mejor espaciado
        self.nav_buttons_frame = tk.Frame(sidebar_inner, bg=STYLE_CONFIG['bg_darker'])
        self.nav_buttons_frame.pack(fill='both', expand=True, pady=20)
        
        # Footer del sidebar mejorado
        footer_frame = tk.Frame(sidebar_inner, bg=STYLE_CONFIG['bg_darker'])
        footer_frame.pack(side='bottom', fill='x', pady=20)
        
        # Separador antes del footer más sutil
        tk.Frame(footer_frame, bg=STYLE_CONFIG['border_color'], height=1).pack(fill='x', pady=(0, 12))
        
        # Información del sistema con mejor tipografía
        version_label = tk.Label(footer_frame,
                                text='v1.0 Professional',
                                font=('Segoe UI', 9),
                                fg=STYLE_CONFIG['text_muted'],
                                bg=STYLE_CONFIG['bg_darker'])
        version_label.pack(pady=(0, 4))
        
        copyright_label = tk.Label(footer_frame,
                                   text='© 2025 FABINSA',
                                   font=('Segoe UI', 9),
                                   fg=STYLE_CONFIG['text_muted'],
                                   bg=STYLE_CONFIG['bg_darker'])
        copyright_label.pack()
        
        # Frame principal con padding premium mejorado
        main_frame = tk.Frame(main_container, bg=STYLE_CONFIG['bg_color'])
        main_frame.pack(side='left', fill='both', expand=True, padx=20, pady=18)
        
        # Pestañas principales (ahora ocultas, controladas por sidebar)
        self.tabs = ttk.Notebook(main_frame)
        self.tabs.pack(expand=1, fill='both')
        
        # Diccionario para tracking de botones de navegación
        self.nav_buttons = {}
        self.current_tab_index = 0
        
        # Pestaña productos con diseño mejorado y scroll
        frame_prod_parent = tk.Frame(self.tabs, bg=STYLE_CONFIG['bg_color'])
        self.tabs.add(frame_prod_parent, text='📦 Produccion')
        frame_prod = self.crear_frame_con_scroll(frame_prod_parent)
        
        # ========== LAYOUT SUPERIOR: FORMULARIO + MÉTRICAS ==========
        top_section = tk.Frame(frame_prod, bg=STYLE_CONFIG['bg_color'])
        top_section.grid(row=0, column=0, sticky='ew', padx=5, pady=5)
        top_section.grid_columnconfigure(0, weight=1)
        top_section.grid_columnconfigure(1, weight=0)
        
        # ========== CARD PREMIUM PARA FORMULARIO DE PRODUCTOS ==========
        card_shadow_prod = tk.Frame(top_section, bg=STYLE_CONFIG['card_shadow_dark'])
        card_shadow_prod.grid(row=0, column=0, sticky='nsew', padx=(0, 15), pady=0)
        
        input_frame_prod = ttk.LabelFrame(card_shadow_prod, text='📝 Datos del Producto', padding=25)
        input_frame_prod.grid(row=0, column=0, sticky='nsew', padx=4, pady=4)
        
        # Configurar labels con estilo premium mejorado
        labels_prod = [
            '📦 Familia',
            '📏 Medida',
            '🔍 Característica',
            '💵 Precio de venta (opcional)',
            '🏛️ IIBB % (opcional)',
            '🔢 Cantidad a fabricar',
            '⏱️ Cantidad producida por hora',
            '🧱 Material (MP)',
            '⚖️ Kg por unidad (material)'
        ]
        
        for i, label_text in enumerate(labels_prod):
            label = tk.Label(input_frame_prod, 
                           text=label_text,
                           font=('Segoe UI', 11, 'bold'),
                           fg=STYLE_CONFIG['text_color'],
                           bg=STYLE_CONFIG['card_bg'])
            label.grid(row=i, column=0, sticky='w', pady=4)
        self.familia_var = tk.StringVar()
        self.medida_var = tk.StringVar()
        self.caracteristica_var = tk.StringVar()
        self.peso_var = tk.StringVar()  # calculado por materiales si se deja vacío
        self.costo_var = tk.StringVar()  # deprecado, ya no visible
        self.dolar_var = tk.StringVar()  # deprecado, ya no visible
        self.venta_var = tk.StringVar()
        self.iibb_porcentaje_var = tk.StringVar()
        self.cant_fab_var = tk.StringVar()
        self.cant_hora_var = tk.StringVar()
        
        # Entries con estilo PROFESIONAL PREMIUM
        # Familia (sin validación numérica)
        entry_familia = tk.Entry(input_frame_prod, 
                       textvariable=self.familia_var, 
                       width=22,
                       font=('Segoe UI', 12),
                       relief='flat',
                       borderwidth=0,
                       bg=STYLE_CONFIG['light_bg'],
                       fg=STYLE_CONFIG['text_color'],
                       highlightthickness=1,
                       highlightcolor=STYLE_CONFIG['primary_color'],
                       highlightbackground=STYLE_CONFIG['border_color'],
                       insertbackground=STYLE_CONFIG['primary_color'])
        entry_familia.grid(row=0, column=1, pady=3, padx=(18, 0), ipady=8)
        
        # Medida (sin validación numérica)
        entry_medida = tk.Entry(input_frame_prod, 
                       textvariable=self.medida_var, 
                       width=22,
                       font=('Segoe UI', 12),
                       relief='flat',
                       borderwidth=0,
                       bg=STYLE_CONFIG['light_bg'],
                       fg=STYLE_CONFIG['text_color'],
                       highlightthickness=1,
                       highlightcolor=STYLE_CONFIG['primary_color'],
                       highlightbackground=STYLE_CONFIG['border_color'],
                       insertbackground=STYLE_CONFIG['primary_color'])
        entry_medida.grid(row=1, column=1, pady=3, padx=(18, 0), ipady=8)
        
        # Característica (sin validación numérica)
        entry_caracteristica = tk.Entry(input_frame_prod, 
                       textvariable=self.caracteristica_var, 
                       width=22,
                       font=('Segoe UI', 12),
                       relief='flat',
                       borderwidth=0,
                       bg=STYLE_CONFIG['light_bg'],
                       fg=STYLE_CONFIG['text_color'],
                       highlightthickness=1,
                       highlightcolor=STYLE_CONFIG['primary_color'],
                       highlightbackground=STYLE_CONFIG['border_color'],
                       insertbackground=STYLE_CONFIG['primary_color'])
        entry_caracteristica.grid(row=2, column=1, pady=3, padx=(18, 0), ipady=8)
        
        # Precio de venta (con validación numérica)
        entry_venta = tk.Entry(input_frame_prod, 
                       textvariable=self.venta_var, 
                       width=22,
                       font=('Segoe UI', 12),
                       relief='flat',
                       borderwidth=0,
                       bg=STYLE_CONFIG['light_bg'],
                       fg=STYLE_CONFIG['text_color'],
                       highlightthickness=1,
                       highlightcolor=STYLE_CONFIG['primary_color'],
                       highlightbackground=STYLE_CONFIG['border_color'],
                       insertbackground=STYLE_CONFIG['primary_color'],
                       validate='key',
                       validatecommand=self.vcmd_decimal)
        entry_venta.grid(row=3, column=1, pady=3, padx=(18, 0), ipady=8)
        
        # IIBB porcentaje (con validación numérica)
        entry_iibb_porcentaje = tk.Entry(input_frame_prod, 
                       textvariable=self.iibb_porcentaje_var, 
                       width=22,
                       font=('Segoe UI', 12),
                       relief='flat',
                       borderwidth=0,
                       bg=STYLE_CONFIG['light_bg'],
                       fg=STYLE_CONFIG['text_color'],
                       highlightthickness=1,
                       highlightcolor=STYLE_CONFIG['primary_color'],
                       highlightbackground=STYLE_CONFIG['border_color'],
                       insertbackground=STYLE_CONFIG['primary_color'],
                       validate='key',
                       validatecommand=self.vcmd_decimal)
        entry_iibb_porcentaje.grid(row=4, column=1, pady=3, padx=(18, 0), ipady=8)
        
        # Cantidad a fabricar (con validación numérica)
        entry_cant_fab = tk.Entry(input_frame_prod, 
                       textvariable=self.cant_fab_var, 
                       width=22,
                       font=('Segoe UI', 12),
                       relief='flat',
                       borderwidth=0,
                       bg=STYLE_CONFIG['light_bg'],
                       fg=STYLE_CONFIG['text_color'],
                       highlightthickness=1,
                       highlightcolor=STYLE_CONFIG['primary_color'],
                       highlightbackground=STYLE_CONFIG['border_color'],
                       insertbackground=STYLE_CONFIG['primary_color'],
                       validate='key',
                       validatecommand=self.vcmd_decimal)
        entry_cant_fab.grid(row=5, column=1, pady=3, padx=(18, 0), ipady=8)
        
        # Cantidad por hora (con validación numérica)
        entry_cant_hora = tk.Entry(input_frame_prod, 
                       textvariable=self.cant_hora_var, 
                       width=22,
                       font=('Segoe UI', 12),
                       relief='flat',
                       borderwidth=0,
                       bg=STYLE_CONFIG['light_bg'],
                       fg=STYLE_CONFIG['text_color'],
                       highlightthickness=1,
                       highlightcolor=STYLE_CONFIG['primary_color'],
                       highlightbackground=STYLE_CONFIG['border_color'],
                       insertbackground=STYLE_CONFIG['primary_color'],
                       validate='key',
                       validatecommand=self.vcmd_decimal)
        entry_cant_hora.grid(row=6, column=1, pady=3, padx=(18, 0), ipady=8)
        
        # Selector de material con estilo mejorado
        self.sel_mp_var = tk.StringVar()
        self.combo_mp = ttk.Combobox(input_frame_prod, textvariable=self.sel_mp_var, state='readonly', width=20, font=('Segoe UI', 12))
        self.combo_mp.grid(row=7, column=1, pady=3, padx=(18, 0), ipady=6)

        # Kg por unidad (material)
        self.kg_por_unidad_var = tk.StringVar()
        kg_material_entry = tk.Entry(input_frame_prod, 
                       textvariable=self.kg_por_unidad_var, 
                       width=22,
                       font=('Segoe UI', 12),
                       relief='flat',
                       borderwidth=0,
                       bg=STYLE_CONFIG['light_bg'],
                       fg=STYLE_CONFIG['text_color'],
                       highlightthickness=1,
                       highlightcolor=STYLE_CONFIG['primary_color'],
                       highlightbackground=STYLE_CONFIG['border_color'],
                       insertbackground=STYLE_CONFIG['primary_color'],
                       validate='key',
                       validatecommand=self.vcmd_decimal)
        kg_material_entry.grid(row=8, column=1, pady=3, padx=(18, 0), ipady=8)

        btns_comp_frame = tk.Frame(input_frame_prod, bg=STYLE_CONFIG['bg_color'])
        btns_comp_frame.grid(row=9, column=0, columnspan=2, sticky='w', pady=(10, 5))
        btn_add_mat = ttk.Button(btns_comp_frame, text='➕ Agregar material', command=self.agregar_material_a_producto_form, style='Action.TButton')
        btn_add_mat.pack(side='left', padx=5, pady=5)
        btn_del_mat = ttk.Button(btns_comp_frame, text='🗑️ Quitar material', command=self.eliminar_material_de_producto_form, style='Danger.TButton')
        btn_del_mat.pack(side='left', padx=5, pady=5)

        # Tabla de materiales del producto
        comp_tree_frame = tk.Frame(input_frame_prod, bg=STYLE_CONFIG['bg_color'])
        comp_tree_frame.grid(row=10, column=0, columnspan=2, sticky='nsew', pady=(5, 0))
        xscroll_comp = tk.Scrollbar(comp_tree_frame, orient='horizontal')
        yscroll_comp = tk.Scrollbar(comp_tree_frame, orient='vertical')
        self.tree_materiales_prod = ttk.Treeview(comp_tree_frame,
                                                 columns=('Material', 'Kg por unidad'),
                                                 show='headings',
                                                 style='Custom.Treeview',
                                                 height=3,
                                                 xscrollcommand=xscroll_comp.set,
                                                 yscrollcommand=yscroll_comp.set)
        for col in self.tree_materiales_prod['columns']:
            self.tree_materiales_prod.heading(col, text=col)
            self.tree_materiales_prod.column(col, width=130, anchor='center', minwidth=100)
        self.tree_materiales_prod.grid(row=0, column=0, sticky='nsew')
        yscroll_comp.config(command=self.tree_materiales_prod.yview)
        xscroll_comp.config(command=self.tree_materiales_prod.xview)
        yscroll_comp.grid(row=0, column=1, sticky='ns')
        xscroll_comp.grid(row=1, column=0, sticky='ew')

        # Botón principal con mejor diseño
        btn_frame_add = tk.Frame(input_frame_prod, bg=STYLE_CONFIG['card_bg'])
        btn_frame_add.grid(row=11, column=0, columnspan=2, pady=15)
        
        btn_agregar_prod = ttk.Button(btn_frame_add, 
                                    text='➕ Agregar Producto', 
                                    command=self.agregar_producto,
                                    style='Custom.TButton')
        btn_agregar_prod.pack()
        
        # Inicializar opciones de MP en combobox
        self.refresh_mp_combobox()
        
        # ========== DASHBOARD DE MÉTRICAS PREMIUM (AL LADO DEL FORMULARIO) ==========
        # Card contenedor del dashboard con sombra elegante
        metrics_shadow = tk.Frame(top_section, bg=STYLE_CONFIG['card_shadow_dark'])
        metrics_shadow.grid(row=0, column=1, sticky='nsew', padx=(15, 0), pady=0)
        
        metrics_container = ttk.LabelFrame(metrics_shadow, text='📊 Dashboard de Productos', padding=20)
        metrics_container.grid(row=0, column=0, sticky='nsew', padx=4, pady=4)
        
        # Container de cards de métricas en grid 2x2
        cards_frame = tk.Frame(metrics_container, bg=STYLE_CONFIG['card_bg'])
        cards_frame.pack(fill='both', expand=True)
        
        # Configurar grid 2x2
        cards_frame.grid_columnconfigure(0, weight=1)
        cards_frame.grid_columnconfigure(1, weight=1)
        cards_frame.grid_rowconfigure(0, weight=1)
        cards_frame.grid_rowconfigure(1, weight=1)
        
        # Crear cards de métricas con colores en grid 2x2
        self.resumen_prod_labels = {}
        metricas_info = [
            ('💰 Rentabilidad neta total', '$ 0.00', '💵', STYLE_CONFIG['metric_bg_2'], 0, 0),
            ('📈 Rentabilidad promedio', '$ 0.00', '📊', STYLE_CONFIG['metric_bg_1'], 0, 1),
            ('🔢 Cantidad total', '0 un.', '📦', STYLE_CONFIG['metric_bg_3'], 1, 0),
            ('⚖️ Material total (kg)', '0.00 kg', '⚖️', STYLE_CONFIG['metric_bg_4'], 1, 1)
        ]
        
        for titulo, valor_inicial, icono, bg, row, col in metricas_info:
            # Crear card individual con sombra
            card_bg = tk.Frame(cards_frame, bg=STYLE_CONFIG['card_shadow'])
            card_bg.grid(row=row, column=col, padx=6, pady=6, sticky='nsew')
            
            card = tk.Frame(card_bg, bg=bg, relief='flat', borderwidth=0)
            card.pack(padx=2, pady=2, fill='both', expand=True)
            
            # Container interno con padding aumentado
            inner = tk.Frame(card, bg=bg)
            inner.pack(fill='both', expand=True, padx=18, pady=16)
            
            # Icono arriba más grande
            icon_label = tk.Label(inner, text=icono, font=('Segoe UI', 36), 
                                 bg=bg, fg=STYLE_CONFIG['primary_color'])
            icon_label.pack(pady=(0, 10))
            
            # Título con mejor tipografía
            title_label = tk.Label(inner, text=titulo.split(' ', 1)[1],  # Quitar emoji del texto
                                  font=('Segoe UI', 10, 'bold'),
                                  fg=STYLE_CONFIG['text_light'], bg=bg)
            title_label.pack()
            
            # Valor con tipografía más grande y destacada
            value_label = tk.Label(inner, text=valor_inicial,
                                  font=('Segoe UI', 20, 'bold'),
                                  fg=STYLE_CONFIG['text_color'], bg=bg)
            value_label.pack(pady=(8, 0))
            
            self.resumen_prod_labels[titulo] = value_label
        
        # ========== SEPARADOR VISUAL ==========
        tk.Frame(frame_prod, bg=STYLE_CONFIG['border_color'], height=2).grid(row=1, column=0, sticky='ew', padx=20, pady=(10, 20))
        
        # ========== SECCIÓN DE TABLA DE PRODUCTOS ==========
        # Título de la tabla
        table_title_frame = tk.Frame(frame_prod, bg=STYLE_CONFIG['bg_color'])
        table_title_frame.grid(row=2, column=0, sticky='ew', padx=20, pady=(0, 10))
        
        table_title = tk.Label(table_title_frame,
                              text='📋 Lista de Productos en Producción',
                              font=('Segoe UI', 16, 'bold'),
                              fg=STYLE_CONFIG['text_color'],
                              bg=STYLE_CONFIG['bg_color'])
        table_title.pack(side='left')
        
        # Contador de productos con mejor estilo
        self.prod_count_label = tk.Label(table_title_frame,
                                         text='(0 productos)',
                                         font=('Segoe UI', 12),
                                         fg=STYLE_CONFIG['text_light'],
                                         bg=STYLE_CONFIG['bg_color'])
        self.prod_count_label.pack(side='left', padx=12)
        
        # Tabla de productos con estilo mejorado y scrollbars
        tree_frame_prod = tk.Frame(frame_prod, bg=STYLE_CONFIG['bg_color'])
        tree_frame_prod.grid(row=3, column=0, padx=20, pady=(0, 15), sticky='nsew')
        
        # Configurar el grid para que la tabla se expanda
        frame_prod.grid_rowconfigure(3, weight=1)
        frame_prod.grid_columnconfigure(0, weight=1)
        
        xscroll_prod = tk.Scrollbar(tree_frame_prod, orient='horizontal')
        yscroll_prod = tk.Scrollbar(tree_frame_prod, orient='vertical')
        
        self.tree = ttk.Treeview(tree_frame_prod, 
                               columns=('Familia', 'Medida', 'Característica', 'Peso por unidad (kg)', 'Cantidad a fabricar', 'Cantidad producida por hora',
                                       'Productos por kilo', 'Costo unitario MP', 'Costo total MP', 'Horas necesarias', 'Valor hora ajustado promedio', 'Incidencia mano de obra',
                                       'Costo unitario mano de obra', 'Precio de venta', 'Rentabilidad neta', 'Rentabilidad neta total'), 
                               show='headings',
                               style='Custom.Treeview',
                               height=8,
                               xscrollcommand=xscroll_prod.set, 
                               yscrollcommand=yscroll_prod.set)
        
        for col in self.tree['columns']:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=120, anchor='center', minwidth=100)
        
        # Configurar scrollbars
        xscroll_prod.config(command=self.tree.xview)
        yscroll_prod.config(command=self.tree.yview)
        
        # Disposición con grid para que las barras se vean siempre
        tree_frame_prod.grid_rowconfigure(0, weight=1)
        tree_frame_prod.grid_columnconfigure(0, weight=1)
        self.tree.grid(row=0, column=0, sticky='nsew')
        yscroll_prod.grid(row=0, column=1, sticky='ns')
        xscroll_prod.grid(row=1, column=0, sticky='ew')
        
        # ========== BOTONES DE ACCIÓN MODERNOS ==========
        btn_frame_prod = tk.Frame(frame_prod, bg=STYLE_CONFIG['bg_color'])
        btn_frame_prod.grid(row=4, column=0, pady=15)
        
        btn_editar_prod = ttk.Button(btn_frame_prod, text='✏️ Editar Producto', command=self.editar_producto, style='Action.TButton')
        btn_editar_prod.pack(side='left', padx=8)
        
        btn_eliminar_prod = ttk.Button(btn_frame_prod, text='🗑️ Eliminar Producto', command=self.eliminar_producto, style='Danger.TButton')
        btn_eliminar_prod.pack(side='left', padx=8)
        
        # Separador entre botones
        tk.Frame(btn_frame_prod, bg=STYLE_CONFIG['border_color'], width=2).pack(side='left', fill='y', padx=15, pady=5)
        
        # Botón para marcar producto como completado (mover a stock de productos)
        btn_completado_prod = ttk.Button(btn_frame_prod, text='✅ Marcar como Completado', command=self.marcar_completado, style='Action.TButton')
        btn_completado_prod.pack(side='left', padx=8)
        # Pestaña empleados con scroll
        frame_emp_parent = tk.Frame(self.tabs, bg=STYLE_CONFIG['bg_color'])
        self.tabs.add(frame_emp_parent, text='👥 Empleados')
        frame_emp = self.crear_frame_con_scroll(frame_emp_parent)
        
        # ========== CARD MODERNO PARA FORMULARIO DE EMPLEADOS ==========
        card_shadow_emp = tk.Frame(frame_emp, bg=STYLE_CONFIG['card_shadow_dark'])
        card_shadow_emp.grid(row=0, column=0, sticky='nw', padx=5, pady=5)
        
        input_frame = ttk.LabelFrame(card_shadow_emp, text='📝 Datos del Empleado', padding=5)
        input_frame.grid(row=0, column=0, sticky='nw', padx=3, pady=3)
        
        # Labels con emojis y mejor estilo
        labels_emp = [
            '👤 Nombre',
            '💰 Valor hora',
            '📅 Días trabajados',
            '⏰ Horas por día',
            '❌ Ausencias',
            '🏖️ Vacaciones',
            '🎉 Feriados',
            '🏥 Lic. Enfermedad',
            '📋 Otras licencias',
            '😴 Horas descanso/día',
            '🏛️ Carga social (%)',
            '⏱️ Horas extras',
            '💼 Feriados trabajados'
        ]
        
        self.emp_vars = [tk.StringVar() for _ in labels_emp]
        
        for i, label_text in enumerate(labels_emp):
            label = tk.Label(input_frame, 
                           text=label_text,
                           font=('Segoe UI', 10, 'bold'),
                           fg=STYLE_CONFIG['text_color'],
                           bg=STYLE_CONFIG['card_bg'])
            label.grid(row=i, column=0, sticky='w', pady=2)
            
            # Solo el primer campo (nombre) NO tiene validación numérica
            if i == 0:
                entry = tk.Entry(input_frame, 
                               textvariable=self.emp_vars[i], 
                               width=20,
                               font=('Segoe UI', 10),
                               relief='flat',
                               borderwidth=0,
                               bg=STYLE_CONFIG['light_bg'],
                               fg=STYLE_CONFIG['text_color'],
                               highlightthickness=2,
                               highlightcolor=STYLE_CONFIG['primary_color'],
                               highlightbackground=STYLE_CONFIG['border_color'],
                               insertbackground=STYLE_CONFIG['primary_color'])
            else:
                # Todos los demás campos tienen validación numérica
                entry = tk.Entry(input_frame, 
                               textvariable=self.emp_vars[i], 
                               width=20,
                               font=('Segoe UI', 10),
                               relief='flat',
                               borderwidth=0,
                               bg=STYLE_CONFIG['light_bg'],
                               fg=STYLE_CONFIG['text_color'],
                               highlightthickness=2,
                               highlightcolor=STYLE_CONFIG['primary_color'],
                               highlightbackground=STYLE_CONFIG['border_color'],
                               insertbackground=STYLE_CONFIG['primary_color'],
                               validate='key',
                               validatecommand=self.vcmd_decimal)
            entry.grid(row=i, column=1, pady=2, padx=(10, 0), ipady=2)
        
        # Botón de agregar con estilo
        btn_agregar_emp = ttk.Button(input_frame, 
                                   text='➕ Agregar Empleado', 
                                   command=self.agregar_empleado,
                                   style='Custom.TButton')
        btn_agregar_emp.grid(row=len(labels_emp), column=0, columnspan=2, pady=5)
        
        # ========== DASHBOARD DE MÉTRICAS DE EMPLEADOS MODERNO ==========
        metrics_emp_container = tk.Frame(frame_emp, bg=STYLE_CONFIG['bg_color'])
        metrics_emp_container.grid(row=0, column=1, sticky='nw', padx=5, pady=5)
        
        # Título de la sección
        metrics_emp_title = tk.Label(metrics_emp_container, 
                                    text='👥 Dashboard de Empleados',
                                    font=('Segoe UI', 14, 'bold'),
                                    fg=STYLE_CONFIG['text_color'],
                                    bg=STYLE_CONFIG['bg_color'])
        metrics_emp_title.pack(anchor='w', pady=(0, 5))
        
        # Container de cards de métricas
        cards_emp_frame = tk.Frame(metrics_emp_container, bg=STYLE_CONFIG['bg_color'])
        cards_emp_frame.pack(fill='both', expand=True)
        
        # Crear cards de métricas con colores
        self.resumen_labels = {}
        metricas_emp_info = [
            ('📈 Media índice ajustado (%)', '0', '📊', STYLE_CONFIG['metric_bg_1']),
            ('💰 Costo total ajustado total', '0', '💵', STYLE_CONFIG['metric_bg_2']),
            ('⏰ Valor hora efectivo total', '0', '⏱️', STYLE_CONFIG['metric_bg_3']),
            ('🔢 Horas trabajadas totales', '0', '⏰', STYLE_CONFIG['metric_bg_4'])
        ]
        
        for titulo, valor_inicial, icono, bg in metricas_emp_info:
            # Crear card individual
            card_bg = tk.Frame(cards_emp_frame, bg=STYLE_CONFIG['card_shadow'])
            card_bg.pack(fill='x', pady=2)
            
            card = tk.Frame(card_bg, bg=bg, relief='flat', borderwidth=0)
            card.pack(padx=2, pady=2, fill='both', expand=True, ipadx=10, ipady=6)
            
            # Icono a la izquierda
            icon_label = tk.Label(card, text=icono, font=('Segoe UI', 24), 
                                 bg=bg, fg=STYLE_CONFIG['primary_color'])
            icon_label.pack(side='left', padx=(5, 15))
            
            # Contenedor de texto
            text_frame = tk.Frame(card, bg=bg)
            text_frame.pack(side='left', fill='both', expand=True)
            
            # Título
            title_label = tk.Label(text_frame, text=titulo.split(' ', 1)[1],  # Quitar emoji del texto
                                  font=('Segoe UI', 9, 'bold'),
                                  fg=STYLE_CONFIG['text_light'], bg=bg, anchor='w')
            title_label.pack(anchor='w')
            
            # Valor
            value_label = tk.Label(text_frame, text=valor_inicial,
                                  font=('Segoe UI', 16, 'bold'),
                                  fg=STYLE_CONFIG['text_color'], bg=bg, anchor='w')
            value_label.pack(anchor='w')
            
            self.resumen_labels[titulo] = value_label
        # Tabla de empleados con scrollbars y mejor estilo
        tree_frame = tk.Frame(frame_emp, bg=STYLE_CONFIG['bg_color'])
        tree_frame.grid(row=1, column=0, columnspan=2, padx=5, pady=5, sticky='nsew')
        
        xscroll = tk.Scrollbar(tree_frame, orient='horizontal')
        yscroll = tk.Scrollbar(tree_frame, orient='vertical')

        self.tree_emp = ttk.Treeview(tree_frame, 
                                   columns=('Nombre', 'Valor hora', 'Días', 'Horas/día', 'Ausencias', 'Vacaciones', 'Feriados', 'Lic. Enfermedad', 'Otras licencias', 'Horas descanso/día', 'Carga social (%)', 'Horas extras', 'Feriados trabajados',
                                           'Horas trabajadas', 'Horas descansadas',
                                           'Jornal Productivas (%)', 'Jornal Vacaciones (%)', 'Jornal Feriados (%)', 'Jornal Lic. Enfermedad (%)', 'Jornal Otras licencias (%)', 'Jornal Ausencias (%)', 'Jornal Descanso (%)', 'Jornal Horas extras (%)', 'Jornal Feriados trabajados (%)',
                                           'CCSS Productivas (%)', 'CCSS Vacaciones (%)', 'CCSS Feriados (%)', 'CCSS Lic. Enfermedad (%)', 'CCSS Otras licencias (%)', 'CCSS Ausencias (%)', 'CCSS Descanso (%)', 'CCSS Horas extras (%)', 'CCSS Feriados trabajados (%)',
                                           'Total Productivas (%)', 'Total Vacaciones (%)', 'Total Feriados (%)', 'Total Lic. Enfermedad (%)', 'Total Otras licencias (%)', 'Total Ausencias (%)', 'Total Descanso (%)', 'Total Horas extras (%)', 'Total Feriados trabajados (%)',
                                           'Sub-Total (%)', 'SAC (%)', 'Índice ajustado (%)', 'Costo total ajustado', 'Valor hora ajustado', 'Valor hora efectivo'), 
                                   show='headings', 
                                   height=7,
                                   style='Custom.Treeview',
                                   xscrollcommand=xscroll.set, 
                                   yscrollcommand=yscroll.set)
        
        for col in self.tree_emp['columns']:
            self.tree_emp.heading(col, text=col)
        
        xscroll.config(command=self.tree_emp.xview)
        yscroll.config(command=self.tree_emp.yview)
        xscroll.pack(side='bottom', fill='x')
        yscroll.pack(side='right', fill='y')
        self.tree_emp.pack(side='left', fill='both', expand=True)
        
        frame_emp.grid_rowconfigure(1, weight=1)
        frame_emp.grid_columnconfigure(0, weight=1)
        
        # ========== BOTONES DE ACCIÓN EMPLEADOS MODERNOS ==========
        # Separador visual
        separator_emp = tk.Frame(frame_emp, bg=STYLE_CONFIG['border_color'], height=2)
        separator_emp.grid(row=2, column=0, columnspan=2, sticky='ew', padx=5, pady=(2, 5))
        
        btn_frame_emp = tk.Frame(frame_emp, bg=STYLE_CONFIG['bg_color'])
        btn_frame_emp.grid(row=2, column=0, columnspan=2, pady=5)
        
        btn_editar_emp = ttk.Button(btn_frame_emp, text='✏️ Editar Empleado', command=self.editar_empleado, style='Action.TButton')
        btn_editar_emp.pack(side='left', padx=8)
        
        btn_eliminar_emp = ttk.Button(btn_frame_emp, text='🗑️ Eliminar Empleado', command=self.eliminar_empleado, style='Danger.TButton')
        btn_eliminar_emp.pack(side='left', padx=8)
        

        # Pestaña Stock
        frame_stock = tk.Frame(self.tabs, bg=STYLE_CONFIG['bg_color'])
        self.tabs.add(frame_stock, text='📦 Stock')
        
        # Crear sub-notebook para las diferentes secciones de stock (CON PESTAÑAS VISIBLES)
        self.stock_tabs = ttk.Notebook(frame_stock, style='SubNotebook.TNotebook')
        self.stock_tabs.pack(fill='both', expand=True, padx=10, pady=10)
        
        # ==================== PESTAÑA 1: MATERIA PRIMA ====================
        frame_mp_parent = tk.Frame(self.stock_tabs, bg=STYLE_CONFIG['bg_color'])
        self.stock_tabs.add(frame_mp_parent, text='🧱 Materia Prima')
        frame_mp = self.crear_frame_con_scroll(frame_mp_parent)
        
        # Variables para Materia Prima
        self.mp_familia_var = tk.StringVar()
        self.mp_medida_var = tk.StringVar()
        self.mp_caracteristica_var = tk.StringVar()
        self.mp_kg_var = tk.StringVar()
        self.mp_costo_var = tk.StringVar()
        self.mp_dolar_var = tk.StringVar()
        self.mp_stock_min_var = tk.StringVar()
        
        # Variables para Producto de Reventa
        self.pr_familia_var = tk.StringVar()
        self.pr_medida_var = tk.StringVar()
        self.pr_caracteristica_var = tk.StringVar()
        self.pr_cantidad_var = tk.StringVar()
        self.pr_costo_unitario_var = tk.StringVar()
        self.pr_moneda_var = tk.StringVar(value='ARS')
        self.pr_valor_dolar_var = tk.StringVar()
        self.pr_otros_costos_var = tk.StringVar()
        
        # Formulario Materia Prima
        input_frame_mp = ttk.LabelFrame(frame_mp, text='📥 Agregar Materia Prima', padding=15)
        input_frame_mp.pack(fill='x', padx=15, pady=15)
        
        mp_labels = ['📦 Familia', '📏 Medida', '🔍 Característica', '⚖️ Stock (kg)', '💰 Costo kilo (USD)', '💱 Valor dólar', '⚠️ Stock Mínimo (kg)']
        
        for i, txt in enumerate(mp_labels):
            label = tk.Label(input_frame_mp,
                text=txt,
                font=('Arial', 9, 'bold'),
                fg=STYLE_CONFIG['text_color'],
                bg=STYLE_CONFIG['bg_color'])
            label.grid(row=i, column=0, sticky='w', pady=3)
        
        entry_mp_familia = tk.Entry(input_frame_mp, textvariable=self.mp_familia_var, width=25, font=('Arial', 9), relief='solid', borderwidth=1)
        entry_mp_familia.grid(row=0, column=1, pady=3, padx=(10, 0))
        entry_mp_medida = tk.Entry(input_frame_mp, textvariable=self.mp_medida_var, width=25, font=('Arial', 9), relief='solid', borderwidth=1)
        entry_mp_medida.grid(row=1, column=1, pady=3, padx=(10, 0))
        entry_mp_caracteristica = tk.Entry(input_frame_mp, textvariable=self.mp_caracteristica_var, width=25, font=('Arial', 9), relief='solid', borderwidth=1)
        entry_mp_caracteristica.grid(row=2, column=1, pady=3, padx=(10, 0))
        entry_mp_kg = tk.Entry(input_frame_mp, textvariable=self.mp_kg_var, width=25, font=('Arial', 9), relief='solid', borderwidth=1, validate='key', validatecommand=self.vcmd_decimal)
        entry_mp_kg.grid(row=3, column=1, pady=3, padx=(10, 0))
        entry_mp_costo = tk.Entry(input_frame_mp, textvariable=self.mp_costo_var, width=25, font=('Arial', 9), relief='solid', borderwidth=1, validate='key', validatecommand=self.vcmd_decimal)
        entry_mp_costo.grid(row=4, column=1, pady=3, padx=(10, 0))
        entry_mp_dolar = tk.Entry(input_frame_mp, textvariable=self.mp_dolar_var, width=25, font=('Arial', 9), relief='solid', borderwidth=1, validate='key', validatecommand=self.vcmd_decimal)
        entry_mp_dolar.grid(row=5, column=1, pady=3, padx=(10, 0))
        entry_mp_stock_min = tk.Entry(input_frame_mp, textvariable=self.mp_stock_min_var, width=25, font=('Arial', 9), relief='solid', borderwidth=1, validate='key', validatecommand=self.vcmd_decimal)
        entry_mp_stock_min.grid(row=6, column=1, pady=3, padx=(10, 0))
        
        btn_agregar_mp = ttk.Button(input_frame_mp, text='➕ Agregar MP', command=self.agregar_materia_prima, style='Custom.TButton')
        btn_agregar_mp.grid(row=7, column=0, columnspan=2, pady=15)
        
        # Tabla Materia Prima
        tree_frame_mp = tk.Frame(frame_mp, bg=STYLE_CONFIG['bg_color'])
        tree_frame_mp.pack(fill='both', expand=True, padx=15, pady=(0, 15))
        
        yscroll_mp = tk.Scrollbar(tree_frame_mp, orient='vertical')
        xscroll_mp = tk.Scrollbar(tree_frame_mp, orient='horizontal')
        self.tree_mp = ttk.Treeview(tree_frame_mp,
            columns=('Familia', 'Medida', 'Característica', 'Stock (kg)', 'Costo kilo (USD)', 'Valor dólar', 'Costo Total', 'Stock Mínimo', 'Último Movimiento'),
            show='headings',
            style='Custom.Treeview',
            height=12,
            yscrollcommand=yscroll_mp.set,
            xscrollcommand=xscroll_mp.set)
        for col in self.tree_mp['columns']:
            self.tree_mp.heading(col, text=col)
            self.tree_mp.column(col, width=150, anchor='center', minwidth=100)
        
        # Configurar tags para colorear filas
        self.tree_mp.tag_configure('stock_bajo_minimo', background='#ffcccc')  # Fondo rojo claro
        
        yscroll_mp.config(command=self.tree_mp.yview)
        xscroll_mp.config(command=self.tree_mp.xview)
        self.tree_mp.grid(row=0, column=0, sticky='nsew')
        yscroll_mp.grid(row=0, column=1, sticky='ns')
        xscroll_mp.grid(row=1, column=0, sticky='ew')
        tree_frame_mp.grid_rowconfigure(0, weight=1)
        tree_frame_mp.grid_columnconfigure(0, weight=1)
        
        btn_frame_mp = tk.Frame(frame_mp, bg=STYLE_CONFIG['bg_color'])
        btn_frame_mp.pack(pady=10)
        btn_editar_mp = ttk.Button(btn_frame_mp, text='✏️ Editar', command=self.editar_materia_prima, style='Action.TButton')
        btn_editar_mp.pack(side='left', padx=5)
        btn_eliminar_mp = ttk.Button(btn_frame_mp, text='🗑️ Eliminar', command=self.eliminar_materia_prima, style='Danger.TButton')
        btn_eliminar_mp.pack(side='left', padx=5)
        btn_importar_mp = ttk.Button(btn_frame_mp, text='📥 Importar Materia Prima', command=self.importar_materia_prima_excel, style='Custom.TButton')
        btn_importar_mp.pack(side='left', padx=5)
        
        # ==================== PESTAÑA 2: PRODUCTOS DE REVENTA ====================
        frame_pr_parent = tk.Frame(self.stock_tabs, bg=STYLE_CONFIG['bg_color'])
        self.stock_tabs.add(frame_pr_parent, text='📦 Productos de Reventa')
        frame_pr = self.crear_frame_con_scroll(frame_pr_parent)
        
        # Formulario Producto de Reventa
        input_frame_pr = ttk.LabelFrame(frame_pr, text='📥 Agregar Producto de Reventa', padding=15)
        input_frame_pr.pack(fill='x', padx=15, pady=15)
        
        # Familia
        tk.Label(input_frame_pr, text='📦 Familia',
                font=('Arial', 9, 'bold'),
                fg=STYLE_CONFIG['text_color'],
                bg=STYLE_CONFIG['bg_color']).grid(row=0, column=0, sticky='w', pady=3)
        entry_pr_familia = tk.Entry(input_frame_pr, textvariable=self.pr_familia_var, width=25, font=('Arial', 9), relief='solid', borderwidth=1)
        entry_pr_familia.grid(row=0, column=1, pady=3, padx=(10, 0))
        
        # Medida
        tk.Label(input_frame_pr, text='📏 Medida',
                font=('Arial', 9, 'bold'),
                fg=STYLE_CONFIG['text_color'],
                bg=STYLE_CONFIG['bg_color']).grid(row=1, column=0, sticky='w', pady=3)
        entry_pr_medida = tk.Entry(input_frame_pr, textvariable=self.pr_medida_var, width=25, font=('Arial', 9), relief='solid', borderwidth=1)
        entry_pr_medida.grid(row=1, column=1, pady=3, padx=(10, 0))
        
        # Característica
        tk.Label(input_frame_pr, text='🔍 Característica',
                font=('Arial', 9, 'bold'),
                fg=STYLE_CONFIG['text_color'],
                bg=STYLE_CONFIG['bg_color']).grid(row=2, column=0, sticky='w', pady=3)
        entry_pr_caracteristica = tk.Entry(input_frame_pr, textvariable=self.pr_caracteristica_var, width=25, font=('Arial', 9), relief='solid', borderwidth=1)
        entry_pr_caracteristica.grid(row=2, column=1, pady=3, padx=(10, 0))
        
        # Cantidad
        tk.Label(input_frame_pr, text='📦 Cantidad (unidades)',
                font=('Arial', 9, 'bold'),
                fg=STYLE_CONFIG['text_color'],
                bg=STYLE_CONFIG['bg_color']).grid(row=3, column=0, sticky='w', pady=3)
        entry_pr_cantidad = tk.Entry(input_frame_pr, textvariable=self.pr_cantidad_var, width=25, font=('Arial', 9), relief='solid', borderwidth=1, validate='key', validatecommand=self.vcmd_decimal)
        entry_pr_cantidad.grid(row=3, column=1, pady=3, padx=(10, 0))
        
        # Moneda
        tk.Label(input_frame_pr, text='💱 Moneda',
                font=('Arial', 9, 'bold'),
                fg=STYLE_CONFIG['text_color'],
                bg=STYLE_CONFIG['bg_color']).grid(row=4, column=0, sticky='w', pady=3)
        
        moneda_frame = tk.Frame(input_frame_pr, bg=STYLE_CONFIG['bg_color'])
        moneda_frame.grid(row=4, column=1, sticky='w', pady=3, padx=(10, 0))
        
        radio_ars = tk.Radiobutton(moneda_frame,
                                   text='ARS',
                                   variable=self.pr_moneda_var,
                                   value='ARS',
                                   font=('Arial', 9),
                                   bg=STYLE_CONFIG['bg_color'],
                                   command=self._actualizar_campos_moneda_pr)
        radio_ars.pack(side='left', padx=(0, 10))
        
        radio_usd = tk.Radiobutton(moneda_frame,
                                   text='USD',
                                   variable=self.pr_moneda_var,
                                   value='USD',
                                   font=('Arial', 9),
                                   bg=STYLE_CONFIG['bg_color'],
                                   command=self._actualizar_campos_moneda_pr)
        radio_usd.pack(side='left')
        
        # Costo unitario
        tk.Label(input_frame_pr, text='💰 Costo unitario',
                font=('Arial', 9, 'bold'),
                fg=STYLE_CONFIG['text_color'],
                bg=STYLE_CONFIG['bg_color']).grid(row=5, column=0, sticky='w', pady=3)
        entry_pr_costo = tk.Entry(input_frame_pr, textvariable=self.pr_costo_unitario_var, width=25, font=('Arial', 9), relief='solid', borderwidth=1, validate='key', validatecommand=self.vcmd_decimal)
        entry_pr_costo.grid(row=5, column=1, pady=3, padx=(10, 0))
        
        # Otros costos
        tk.Label(input_frame_pr, text='📝 Otros Costos',
                font=('Arial', 9, 'bold'),
                fg=STYLE_CONFIG['text_color'],
                bg=STYLE_CONFIG['bg_color']).grid(row=6, column=0, sticky='w', pady=3)
        entry_pr_otros_costos = tk.Entry(input_frame_pr, textvariable=self.pr_otros_costos_var, width=25, font=('Arial', 9), relief='solid', borderwidth=1, validate='key', validatecommand=self.vcmd_decimal)
        entry_pr_otros_costos.grid(row=6, column=1, pady=3, padx=(10, 0))
        
        # Valor del dólar
        self.pr_lbl_dolar = tk.Label(input_frame_pr, text='💵 Valor del dólar',
                                     font=('Arial', 9, 'bold'),
                                     fg=STYLE_CONFIG['text_color'],
                                     bg=STYLE_CONFIG['bg_color'])
        self.pr_entry_dolar = tk.Entry(input_frame_pr, textvariable=self.pr_valor_dolar_var, width=25, font=('Arial', 9), relief='solid', borderwidth=1, validate='key', validatecommand=self.vcmd_decimal)
        
        btn_agregar_pr = ttk.Button(input_frame_pr, text='➕ Agregar Producto', command=self.agregar_producto_reventa, style='Custom.TButton')
        btn_agregar_pr.grid(row=8, column=0, columnspan=2, pady=15)
        
        self._actualizar_campos_moneda_pr()
        
        # Tabla Productos de Reventa
        tree_frame_sp = tk.Frame(frame_pr, bg=STYLE_CONFIG['bg_color'])
        tree_frame_sp.pack(fill='both', expand=True, padx=15, pady=(0, 15))
        
        yscroll_sp = tk.Scrollbar(tree_frame_sp, orient='vertical')
        xscroll_sp = tk.Scrollbar(tree_frame_sp, orient='horizontal')
        self.tree_stock_prod = ttk.Treeview(tree_frame_sp,
            columns=('Familia', 'Medida', 'Característica', 'Cantidad (u)', 'Costo Unitario', 'Costo Total Compra'),
            show='headings',
            style='Custom.Treeview',
            height=12,
            yscrollcommand=yscroll_sp.set,
            xscrollcommand=xscroll_sp.set)
        for col in self.tree_stock_prod['columns']:
            self.tree_stock_prod.heading(col, text=col)
            self.tree_stock_prod.column(col, width=180, anchor='center', minwidth=120)
        yscroll_sp.config(command=self.tree_stock_prod.yview)
        xscroll_sp.config(command=self.tree_stock_prod.xview)
        self.tree_stock_prod.grid(row=0, column=0, sticky='nsew')
        yscroll_sp.grid(row=0, column=1, sticky='ns')
        xscroll_sp.grid(row=1, column=0, sticky='ew')
        tree_frame_sp.grid_rowconfigure(0, weight=1)
        tree_frame_sp.grid_columnconfigure(0, weight=1)
        
        btn_frame_sp = tk.Frame(frame_pr, bg=STYLE_CONFIG['bg_color'])
        btn_frame_sp.pack(pady=10)
        btn_editar_pr = ttk.Button(btn_frame_sp, text='✏️ Editar', command=self.editar_producto_reventa, style='Action.TButton')
        btn_editar_pr.pack(side='left', padx=5)
        btn_eliminar_pr = ttk.Button(btn_frame_sp, text='🗑️ Eliminar', command=self.eliminar_producto_reventa, style='Danger.TButton')
        btn_eliminar_pr.pack(side='left', padx=5)
        btn_importar_stock = ttk.Button(btn_frame_sp, text='📥 Importar Stock', command=self.importar_stock_excel, style='Custom.TButton')
        btn_importar_stock.pack(side='left', padx=5)
        
        # ==================== PESTAÑA 3: PRODUCTOS FABRICADOS ====================
        frame_fab_parent = tk.Frame(self.stock_tabs, bg=STYLE_CONFIG['bg_color'])
        self.stock_tabs.add(frame_fab_parent, text='🏭 Productos Fabricados')
        frame_fab = self.crear_frame_con_scroll(frame_fab_parent)
        
        # Tabla Productos Fabricados
        tree_frame_sf = tk.Frame(frame_fab, bg=STYLE_CONFIG['bg_color'])
        tree_frame_sf.pack(fill='both', expand=True, padx=15, pady=15)
        
        yscroll_sf = tk.Scrollbar(tree_frame_sf, orient='vertical')
        xscroll_sf = tk.Scrollbar(tree_frame_sf, orient='horizontal')
        self.tree_stock_fabricados = ttk.Treeview(tree_frame_sf,
            columns=('Familia', 'Medida', 'Característica', 'Cantidad (u)', 'Peso unidad (kg)', 'Total (kg)', 'Costo prod. unitario', 'Costo Total'),
            show='headings',
            style='Custom.Treeview',
            height=15,
            yscrollcommand=yscroll_sf.set,
            xscrollcommand=xscroll_sf.set)
        for col in self.tree_stock_fabricados['columns']:
            self.tree_stock_fabricados.heading(col, text=col)
            self.tree_stock_fabricados.column(col, width=140, anchor='center', minwidth=100)
        yscroll_sf.config(command=self.tree_stock_fabricados.yview)
        xscroll_sf.config(command=self.tree_stock_fabricados.xview)
        self.tree_stock_fabricados.grid(row=0, column=0, sticky='nsew')
        yscroll_sf.grid(row=0, column=1, sticky='ns')
        xscroll_sf.grid(row=1, column=0, sticky='ew')
        tree_frame_sf.grid_rowconfigure(0, weight=1)
        tree_frame_sf.grid_columnconfigure(0, weight=1)
        
        btn_frame_sf = tk.Frame(frame_fab, bg=STYLE_CONFIG['bg_color'])
        btn_frame_sf.pack(pady=10)
        btn_eliminar_sf = ttk.Button(btn_frame_sf, text='🗑️ Eliminar', command=self.eliminar_stock_fabricado, style='Danger.TButton')
        btn_eliminar_sf.pack(side='left', padx=5)

        # ==================== PESTAÑA 4: CONTROL DE INGRESOS Y EGRESOS ====================
        frame_control_parent = tk.Frame(self.stock_tabs, bg=STYLE_CONFIG['bg_color'])
        self.stock_tabs.add(frame_control_parent, text='📊 Control Ingresos/Egresos')
        frame_control = self.crear_frame_con_scroll(frame_control_parent)
        
        # Variables para el control de ingresos/egresos
        self.control_tipo_var = tk.StringVar(value='ingreso')
        self.control_familia_var = tk.StringVar()
        self.control_medida_var = tk.StringVar()
        self.control_caracteristica_var = tk.StringVar()
        self.control_cantidad_var = tk.StringVar()
        self.control_motivo_var = tk.StringVar()
        
        # Sección de métricas con indicadores visuales
        metrics_frame = ttk.LabelFrame(frame_control, text='📈 Métricas de Stock', padding=15)
        metrics_frame.pack(fill='x', padx=15, pady=15)
        
        # Crear grid para las métricas
        metrics_grid = tk.Frame(metrics_frame, bg=STYLE_CONFIG['bg_color'])
        metrics_grid.pack(fill='x')
        
        # Métricas de Materia Prima
        mp_metrics_frame = tk.Frame(metrics_grid, bg=STYLE_CONFIG['bg_color'], relief='solid', borderwidth=1)
        mp_metrics_frame.grid(row=0, column=0, padx=10, pady=10, sticky='ew')
        
        tk.Label(mp_metrics_frame, text='🧱 MATERIA PRIMA', 
                font=('Arial', 12, 'bold'), fg='#2E86AB', bg=STYLE_CONFIG['bg_color']).pack(pady=5)
        
        # Ingresos MP
        ingreso_mp_frame = tk.Frame(mp_metrics_frame, bg=STYLE_CONFIG['bg_color'])
        ingreso_mp_frame.pack(pady=5)
        self.label_ingreso_mp = tk.Label(ingreso_mp_frame, text='Ingresos: 0 kg', 
                                       font=('Arial', 10, 'bold'), fg='#27AE60', bg=STYLE_CONFIG['bg_color'])
        self.label_ingreso_mp.pack(side='left')
        self.arrow_ingreso_mp = tk.Label(ingreso_mp_frame, text='↑', 
                                       font=('Arial', 16, 'bold'), fg='#27AE60', bg=STYLE_CONFIG['bg_color'])
        self.arrow_ingreso_mp.pack(side='left', padx=5)
        
        # Egresos MP
        egreso_mp_frame = tk.Frame(mp_metrics_frame, bg=STYLE_CONFIG['bg_color'])
        egreso_mp_frame.pack(pady=5)
        self.label_egreso_mp = tk.Label(egreso_mp_frame, text='Egresos: 0 kg', 
                                      font=('Arial', 10, 'bold'), fg='#E74C3C', bg=STYLE_CONFIG['bg_color'])
        self.label_egreso_mp.pack(side='left')
        self.arrow_egreso_mp = tk.Label(egreso_mp_frame, text='↓', 
                                      font=('Arial', 16, 'bold'), fg='#E74C3C', bg=STYLE_CONFIG['bg_color'])
        self.arrow_egreso_mp.pack(side='left', padx=5)
        
        # Saldo MP
        saldo_mp_frame = tk.Frame(mp_metrics_frame, bg=STYLE_CONFIG['bg_color'])
        saldo_mp_frame.pack(pady=5)
        self.label_saldo_mp = tk.Label(saldo_mp_frame, text='Saldo: 0 kg', 
                                     font=('Arial', 10, 'bold'), fg='#3498DB', bg=STYLE_CONFIG['bg_color'])
        self.label_saldo_mp.pack(side='left')
        self.arrow_saldo_mp = tk.Label(saldo_mp_frame, text='=', 
                                     font=('Arial', 16, 'bold'), fg='#3498DB', bg=STYLE_CONFIG['bg_color'])
        self.arrow_saldo_mp.pack(side='left', padx=5)
        
        # Métricas de Productos de Reventa
        pr_metrics_frame = tk.Frame(metrics_grid, bg=STYLE_CONFIG['bg_color'], relief='solid', borderwidth=1)
        pr_metrics_frame.grid(row=0, column=1, padx=10, pady=10, sticky='ew')
        
        tk.Label(pr_metrics_frame, text='📦 PRODUCTOS REVENTA', 
                font=('Arial', 12, 'bold'), fg='#8E44AD', bg=STYLE_CONFIG['bg_color']).pack(pady=5)
        
        # Ingresos PR
        ingreso_pr_frame = tk.Frame(pr_metrics_frame, bg=STYLE_CONFIG['bg_color'])
        ingreso_pr_frame.pack(pady=5)
        self.label_ingreso_pr = tk.Label(ingreso_pr_frame, text='Ingresos: 0 u', 
                                       font=('Arial', 10, 'bold'), fg='#27AE60', bg=STYLE_CONFIG['bg_color'])
        self.label_ingreso_pr.pack(side='left')
        self.arrow_ingreso_pr = tk.Label(ingreso_pr_frame, text='↑', 
                                       font=('Arial', 16, 'bold'), fg='#27AE60', bg=STYLE_CONFIG['bg_color'])
        self.arrow_ingreso_pr.pack(side='left', padx=5)
        
        # Egresos PR
        egreso_pr_frame = tk.Frame(pr_metrics_frame, bg=STYLE_CONFIG['bg_color'])
        egreso_pr_frame.pack(pady=5)
        self.label_egreso_pr = tk.Label(egreso_pr_frame, text='Egresos: 0 u', 
                                      font=('Arial', 10, 'bold'), fg='#E74C3C', bg=STYLE_CONFIG['bg_color'])
        self.label_egreso_pr.pack(side='left')
        self.arrow_egreso_pr = tk.Label(egreso_pr_frame, text='↓', 
                                      font=('Arial', 16, 'bold'), fg='#E74C3C', bg=STYLE_CONFIG['bg_color'])
        self.arrow_egreso_pr.pack(side='left', padx=5)
        
        # Saldo PR
        saldo_pr_frame = tk.Frame(pr_metrics_frame, bg=STYLE_CONFIG['bg_color'])
        saldo_pr_frame.pack(pady=5)
        self.label_saldo_pr = tk.Label(saldo_pr_frame, text='Saldo: 0 u', 
                                     font=('Arial', 10, 'bold'), fg='#3498DB', bg=STYLE_CONFIG['bg_color'])
        self.label_saldo_pr.pack(side='left')
        self.arrow_saldo_pr = tk.Label(saldo_pr_frame, text='=', 
                                     font=('Arial', 16, 'bold'), fg='#3498DB', bg=STYLE_CONFIG['bg_color'])
        self.arrow_saldo_pr.pack(side='left', padx=5)
        
        # Métricas de Productos Fabricados
        fab_metrics_frame = tk.Frame(metrics_grid, bg=STYLE_CONFIG['bg_color'], relief='solid', borderwidth=1)
        fab_metrics_frame.grid(row=0, column=2, padx=10, pady=10, sticky='ew')
        
        tk.Label(fab_metrics_frame, text='🏭 PRODUCTOS FABRICADOS', 
                font=('Arial', 12, 'bold'), fg='#E67E22', bg=STYLE_CONFIG['bg_color']).pack(pady=5)
        
        # Ingresos FAB
        ingreso_fab_frame = tk.Frame(fab_metrics_frame, bg=STYLE_CONFIG['bg_color'])
        ingreso_fab_frame.pack(pady=5)
        self.label_ingreso_fab = tk.Label(ingreso_fab_frame, text='Ingresos: 0 u', 
                                        font=('Arial', 10, 'bold'), fg='#27AE60', bg=STYLE_CONFIG['bg_color'])
        self.label_ingreso_fab.pack(side='left')
        self.arrow_ingreso_fab = tk.Label(ingreso_fab_frame, text='↑', 
                                        font=('Arial', 16, 'bold'), fg='#27AE60', bg=STYLE_CONFIG['bg_color'])
        self.arrow_ingreso_fab.pack(side='left', padx=5)
        
        # Egresos FAB
        egreso_fab_frame = tk.Frame(fab_metrics_frame, bg=STYLE_CONFIG['bg_color'])
        egreso_fab_frame.pack(pady=5)
        self.label_egreso_fab = tk.Label(egreso_fab_frame, text='Egresos: 0 u', 
                                       font=('Arial', 10, 'bold'), fg='#E74C3C', bg=STYLE_CONFIG['bg_color'])
        self.label_egreso_fab.pack(side='left')
        self.arrow_egreso_fab = tk.Label(egreso_fab_frame, text='↓', 
                                       font=('Arial', 16, 'bold'), fg='#E74C3C', bg=STYLE_CONFIG['bg_color'])
        self.arrow_egreso_fab.pack(side='left', padx=5)
        
        # Saldo FAB
        saldo_fab_frame = tk.Frame(fab_metrics_frame, bg=STYLE_CONFIG['bg_color'])
        saldo_fab_frame.pack(pady=5)
        self.label_saldo_fab = tk.Label(saldo_fab_frame, text='Saldo: 0 u', 
                                      font=('Arial', 10, 'bold'), fg='#3498DB', bg=STYLE_CONFIG['bg_color'])
        self.label_saldo_fab.pack(side='left')
        self.arrow_saldo_fab = tk.Label(saldo_fab_frame, text='=', 
                                      font=('Arial', 16, 'bold'), fg='#3498DB', bg=STYLE_CONFIG['bg_color'])
        self.arrow_saldo_fab.pack(side='left', padx=5)
        
        # Configurar grid para que las columnas se expandan
        metrics_grid.grid_columnconfigure(0, weight=1)
        metrics_grid.grid_columnconfigure(1, weight=1)
        metrics_grid.grid_columnconfigure(2, weight=1)
        
        # Sección para registrar ingresos/egresos manuales
        registro_frame = ttk.LabelFrame(frame_control, text='📝 Registrar Movimiento Manual', padding=15)
        registro_frame.pack(fill='x', padx=15, pady=(0, 15))
        
        # Tipo de movimiento
        tipo_frame = tk.Frame(registro_frame, bg=STYLE_CONFIG['bg_color'])
        tipo_frame.pack(fill='x', pady=5)
        tk.Label(tipo_frame, text='Tipo:', font=('Arial', 9, 'bold'), 
                fg=STYLE_CONFIG['text_color'], bg=STYLE_CONFIG['bg_color']).pack(side='left')
        
        tipo_ingreso = tk.Radiobutton(tipo_frame, text='Ingreso ↑', variable=self.control_tipo_var, 
                                    value='ingreso', font=('Arial', 9), fg='#27AE60', bg=STYLE_CONFIG['bg_color'])
        tipo_ingreso.pack(side='left', padx=10)
        
        tipo_egreso = tk.Radiobutton(tipo_frame, text='Egreso ↓', variable=self.control_tipo_var, 
                                   value='egreso', font=('Arial', 9), fg='#E74C3C', bg=STYLE_CONFIG['bg_color'])
        tipo_egreso.pack(side='left', padx=10)
        
        # Campos del producto
        campos_frame = tk.Frame(registro_frame, bg=STYLE_CONFIG['bg_color'])
        campos_frame.pack(fill='x', pady=10)
        
        # Familia
        tk.Label(campos_frame, text='📦 Familia:', font=('Arial', 9, 'bold'), 
                fg=STYLE_CONFIG['text_color'], bg=STYLE_CONFIG['bg_color']).grid(row=0, column=0, sticky='w', pady=2)
        entry_control_familia = tk.Entry(campos_frame, textvariable=self.control_familia_var, 
                                       width=20, font=('Arial', 9), relief='solid', borderwidth=1)
        entry_control_familia.grid(row=0, column=1, padx=(10, 0), pady=2)
        
        # Medida
        tk.Label(campos_frame, text='📏 Medida:', font=('Arial', 9, 'bold'), 
                fg=STYLE_CONFIG['text_color'], bg=STYLE_CONFIG['bg_color']).grid(row=0, column=2, sticky='w', pady=2, padx=(20, 0))
        entry_control_medida = tk.Entry(campos_frame, textvariable=self.control_medida_var, 
                                      width=20, font=('Arial', 9), relief='solid', borderwidth=1)
        entry_control_medida.grid(row=0, column=3, padx=(10, 0), pady=2)
        
        # Característica
        tk.Label(campos_frame, text='🔍 Característica:', font=('Arial', 9, 'bold'), 
                fg=STYLE_CONFIG['text_color'], bg=STYLE_CONFIG['bg_color']).grid(row=1, column=0, sticky='w', pady=2)
        entry_control_caracteristica = tk.Entry(campos_frame, textvariable=self.control_caracteristica_var, 
                                              width=20, font=('Arial', 9), relief='solid', borderwidth=1)
        entry_control_caracteristica.grid(row=1, column=1, padx=(10, 0), pady=2)
        
        # Cantidad
        tk.Label(campos_frame, text='📊 Cantidad:', font=('Arial', 9, 'bold'), 
                fg=STYLE_CONFIG['text_color'], bg=STYLE_CONFIG['bg_color']).grid(row=1, column=2, sticky='w', pady=2, padx=(20, 0))
        entry_control_cantidad = tk.Entry(campos_frame, textvariable=self.control_cantidad_var, 
                                        width=20, font=('Arial', 9), relief='solid', borderwidth=1, 
                                        validate='key', validatecommand=self.vcmd_decimal)
        entry_control_cantidad.grid(row=1, column=3, padx=(10, 0), pady=2)
        
        # Motivo
        tk.Label(campos_frame, text='📝 Motivo:', font=('Arial', 9, 'bold'), 
                fg=STYLE_CONFIG['text_color'], bg=STYLE_CONFIG['bg_color']).grid(row=2, column=0, sticky='w', pady=2)
        entry_control_motivo = tk.Entry(campos_frame, textvariable=self.control_motivo_var, 
                                      width=50, font=('Arial', 9), relief='solid', borderwidth=1)
        entry_control_motivo.grid(row=2, column=1, columnspan=3, padx=(10, 0), pady=2, sticky='ew')
        
        # Botón para registrar
        btn_registrar = ttk.Button(registro_frame, text='📝 Registrar Movimiento', 
                                 command=self.registrar_movimiento_manual, style='Custom.TButton')
        btn_registrar.pack(pady=10)
        
        # Tabla de log de movimientos
        log_frame = ttk.LabelFrame(frame_control, text='📋 Log de Movimientos', padding=15)
        log_frame.pack(fill='both', expand=True, padx=15, pady=(0, 15))
        
        # Crear tabla de log
        log_tree_frame = tk.Frame(log_frame, bg=STYLE_CONFIG['bg_color'])
        log_tree_frame.pack(fill='both', expand=True)
        
        yscroll_log = tk.Scrollbar(log_tree_frame, orient='vertical')
        xscroll_log = tk.Scrollbar(log_tree_frame, orient='horizontal')
        self.tree_log_movimientos = ttk.Treeview(log_tree_frame,
            columns=('Fecha', 'Tipo', 'Familia', 'Medida', 'Característica', 'Cantidad', 'Motivo'),
            show='headings',
            style='Custom.Treeview',
            height=15,
            yscrollcommand=yscroll_log.set,
            xscrollcommand=xscroll_log.set)
        
        for col in self.tree_log_movimientos['columns']:
            self.tree_log_movimientos.heading(col, text=col)
            if col == 'Motivo':
                self.tree_log_movimientos.column(col, width=200, anchor='center', minwidth=150)
            else:
                self.tree_log_movimientos.column(col, width=120, anchor='center', minwidth=100)
        
        yscroll_log.config(command=self.tree_log_movimientos.yview)
        xscroll_log.config(command=self.tree_log_movimientos.xview)
        self.tree_log_movimientos.grid(row=0, column=0, sticky='nsew')
        yscroll_log.grid(row=0, column=1, sticky='ns')
        xscroll_log.grid(row=1, column=0, sticky='ew')
        log_tree_frame.grid_rowconfigure(0, weight=1)
        log_tree_frame.grid_columnconfigure(0, weight=1)
        
        # Inicializar listas para el log de movimientos
        if not hasattr(self, 'log_ingresos_mp'):
            self.log_ingresos_mp = []
        if not hasattr(self, 'log_egresos_mp'):
            self.log_egresos_mp = []
        if not hasattr(self, 'log_ingresos_pr'):
            self.log_ingresos_pr = []
        if not hasattr(self, 'log_egresos_pr'):
            self.log_egresos_pr = []
        if not hasattr(self, 'log_ingresos_fab'):
            self.log_ingresos_fab = []
        if not hasattr(self, 'log_egresos_fab'):
            self.log_egresos_fab = []

        # Pestaña Métricas (separada en 3 tablas) con scroll
        frame_metric_parent = tk.Frame(self.tabs, bg=STYLE_CONFIG['bg_color'])
        self.tabs.add(frame_metric_parent, text='📈 Métricas')
        frame_metric = self.crear_frame_con_scroll(frame_metric_parent)

        # Producción / Ingreso stock
        lf_prod = ttk.LabelFrame(frame_metric, text='Producción (Ingreso stock)', padding=10)
        lf_prod.pack(fill='both', expand=True, padx=15, pady=(15, 8))
        prod_container = tk.Frame(lf_prod, bg=STYLE_CONFIG['bg_color'])
        prod_container.pack(fill='both', expand=True)
        prod_frame = tk.Frame(prod_container, bg=STYLE_CONFIG['bg_color'])
        prod_frame.pack(side='left', fill='both', expand=True)
        xs_p = tk.Scrollbar(prod_frame, orient='horizontal')
        ys_p = tk.Scrollbar(prod_frame, orient='vertical')
        self.tree_metricas_prod = ttk.Treeview(
            prod_frame,
            columns=('Fecha','Familia','Medida','Característica','Cantidad (u)','Peso unidad (kg)','Kg consumidos','Costo unitario MP','Costo unitario MO','Costo prod. unitario','Costo total MP','Precio de venta','Rentabilidad neta','Rentabilidad total'),
            show='headings', style='Custom.Treeview', height=8, xscrollcommand=xs_p.set, yscrollcommand=ys_p.set)
        for col in self.tree_metricas_prod['columns']:
            self.tree_metricas_prod.heading(col, text=col)
            width = 130
            if col in ('Fecha','Familia','Medida','Característica'):
                width = 120
            self.tree_metricas_prod.column(col, width=width, anchor='center', minwidth=90)
        self.tree_metricas_prod.grid(row=0, column=0, sticky='nsew')
        ys_p.config(command=self.tree_metricas_prod.yview); xs_p.config(command=self.tree_metricas_prod.xview)
        ys_p.grid(row=0, column=1, sticky='ns'); xs_p.grid(row=1, column=0, sticky='ew')
        # Panel de métricas/gráfico a la derecha
        prod_side = tk.Frame(prod_container, bg=STYLE_CONFIG['bg_color'])
        prod_side.pack(side='left', fill='y', padx=10)
        self.lbl_prod_resumen = tk.Label(prod_side, text='Resumen Producción', font=('Arial', 9, 'bold'), bg=STYLE_CONFIG['bg_color'], fg=STYLE_CONFIG['primary_color'])
        self.lbl_prod_resumen.pack(anchor='w')
        self.lbl_prod_text = tk.Label(prod_side, text='-', bg=STYLE_CONFIG['bg_color'])
        self.lbl_prod_text.pack(anchor='w')
        if Figure is not None:
            self.fig_prod = Figure(figsize=(3,2), dpi=100)
            self.ax_prod = self.fig_prod.add_subplot(111)
            self.ax_prod.set_title('Kg consumidos')
            self.canvas_prod = FigureCanvasTkAgg(self.fig_prod, master=prod_side)
            self.canvas_prod.get_tk_widget().pack()

        # Ingresos de Materia Prima
        lf_mp = ttk.LabelFrame(frame_metric, text='Ingresos de Materia Prima', padding=10)
        lf_mp.pack(fill='both', expand=True, padx=15, pady=8)
        mp_container = tk.Frame(lf_mp, bg=STYLE_CONFIG['bg_color'])
        mp_container.pack(fill='both', expand=True)
        mp_frame = tk.Frame(mp_container, bg=STYLE_CONFIG['bg_color'])
        mp_frame.pack(side='left', fill='both', expand=True)
        xs_mp = tk.Scrollbar(mp_frame, orient='horizontal')
        ys_mp = tk.Scrollbar(mp_frame, orient='vertical')
        self.tree_metricas_mp = ttk.Treeview(mp_frame,
            columns=('Fecha','Material','Kg','Costo unitario','Costo total'),
            show='headings', style='Custom.Treeview', height=6, xscrollcommand=xs_mp.set, yscrollcommand=ys_mp.set)
        for col in self.tree_metricas_mp['columns']:
            self.tree_metricas_mp.heading(col, text=col)
            self.tree_metricas_mp.column(col, width=140 if col=='Material' else 120, anchor='center', minwidth=90)
        self.tree_metricas_mp.grid(row=0, column=0, sticky='nsew')
        ys_mp.config(command=self.tree_metricas_mp.yview); xs_mp.config(command=self.tree_metricas_mp.xview)
        ys_mp.grid(row=0, column=1, sticky='ns'); xs_mp.grid(row=1, column=0, sticky='ew')
        # Lateral con totales y gráfico
        mp_side = tk.Frame(mp_container, bg=STYLE_CONFIG['bg_color'])
        mp_side.pack(side='left', fill='y', padx=10)
        tk.Label(mp_side, text='Resumen MP', font=('Arial', 9, 'bold'), bg=STYLE_CONFIG['bg_color'], fg=STYLE_CONFIG['primary_color']).pack(anchor='w')
        self.lbl_mp_text = tk.Label(mp_side, text='-', bg=STYLE_CONFIG['bg_color'])
        self.lbl_mp_text.pack(anchor='w')
        if Figure is not None:
            self.fig_mp = Figure(figsize=(3,2), dpi=100)
            self.ax_mp = self.fig_mp.add_subplot(111)
            self.ax_mp.set_title('Kg MP ingresados')
            self.canvas_mp = FigureCanvasTkAgg(self.fig_mp, master=mp_side)
            self.canvas_mp.get_tk_widget().pack()

        # Ventas / Revenue
        lf_sales = ttk.LabelFrame(frame_metric, text='Ventas', padding=10)
        lf_sales.pack(fill='both', expand=True, padx=15, pady=(8, 15))
        sales_container = tk.Frame(lf_sales, bg=STYLE_CONFIG['bg_color'])
        sales_container.pack(fill='both', expand=True)
        sales_frame_m = tk.Frame(sales_container, bg=STYLE_CONFIG['bg_color'])
        sales_frame_m.pack(side='left', fill='both', expand=True)
        xs_s = tk.Scrollbar(sales_frame_m, orient='horizontal')
        ys_s = tk.Scrollbar(sales_frame_m, orient='vertical')
        self.tree_metricas_sales = ttk.Treeview(sales_frame_m,
            columns=('Fecha','Familia','Medida','Característica','Cantidad (u)','Precio unitario','Descuento (%)','IIB (%)','Precio final','Ingreso neto','Ganancia total'),
            show='headings', style='Custom.Treeview', height=7, xscrollcommand=xs_s.set, yscrollcommand=ys_s.set)
        for col in self.tree_metricas_sales['columns']:
            self.tree_metricas_sales.heading(col, text=col)
            self.tree_metricas_sales.column(col, width=120, anchor='center', minwidth=90)
        self.tree_metricas_sales.grid(row=0, column=0, sticky='nsew')
        ys_s.config(command=self.tree_metricas_sales.yview); xs_s.config(command=self.tree_metricas_sales.xview)
        ys_s.grid(row=0, column=1, sticky='ns'); xs_s.grid(row=1, column=0, sticky='ew')
        # Lateral con resumen de ventas y gráfica
        sales_side = tk.Frame(sales_container, bg=STYLE_CONFIG['bg_color'])
        sales_side.pack(side='left', fill='y', padx=10)
        tk.Label(sales_side, text='Resumen Ventas', font=('Arial', 9, 'bold'), bg=STYLE_CONFIG['bg_color'], fg=STYLE_CONFIG['primary_color']).pack(anchor='w')
        self.lbl_sales_text = tk.Label(sales_side, text='-', bg=STYLE_CONFIG['bg_color'])
        self.lbl_sales_text.pack(anchor='w')
        if Figure is not None:
            self.fig_sales = Figure(figsize=(3,2), dpi=100)
            self.ax_sales = self.fig_sales.add_subplot(111)
            self.ax_sales.set_title('Ingresos netos')
            self.canvas_sales = FigureCanvasTkAgg(self.fig_sales, master=sales_side)
            self.canvas_sales.get_tk_widget().pack()

        # Pestaña Ventas con scroll
        frame_sales_parent = tk.Frame(self.tabs, bg=STYLE_CONFIG['bg_color'])
        self.tabs.add(frame_sales_parent, text='🧾 Ventas')
        frame_sales = self.crear_frame_con_scroll(frame_sales_parent)

        sales_form = ttk.LabelFrame(frame_sales, text='📝 Registrar Venta', padding=15)
        sales_form.grid(row=0, column=0, sticky='nw', padx=15, pady=15)
        # Ajustes de columnas para que los elementos de la derecha no se monten
        try:
            sales_form.grid_columnconfigure(2, minsize=140)
            sales_form.grid_columnconfigure(3, minsize=200)
        except Exception:
            pass

        # Selector de tipo de producto
        tk.Label(sales_form, text='📋 Tipo', font=('Arial', 9, 'bold'), fg=STYLE_CONFIG['text_color'], bg=STYLE_CONFIG['bg_color']).grid(row=0, column=0, sticky='w', pady=3)
        self.venta_tipo_var = tk.StringVar(value='fabricado')
        
        tipo_frame = tk.Frame(sales_form, bg=STYLE_CONFIG['bg_color'])
        tipo_frame.grid(row=0, column=1, sticky='w', pady=3, padx=(10, 0))
        
        radio_fab = tk.Radiobutton(tipo_frame,
                                   text='🏭 Fabricado',
                                   variable=self.venta_tipo_var,
                                   value='fabricado',
                                   font=('Arial', 9),
                                   bg=STYLE_CONFIG['bg_color'],
                                   command=self._refresh_combo_venta_productos)
        radio_fab.pack(side='left', padx=(0, 10))
        
        radio_rev = tk.Radiobutton(tipo_frame,
                                   text='📦 Reventa',
                                   variable=self.venta_tipo_var,
                                   value='reventa',
                                   font=('Arial', 9),
                                   bg=STYLE_CONFIG['bg_color'],
                                   command=self._refresh_combo_venta_productos)
        radio_rev.pack(side='left')

        # Selector de producto
        tk.Label(sales_form, text='📦 Producto', font=('Arial', 9, 'bold'), fg=STYLE_CONFIG['text_color'], bg=STYLE_CONFIG['bg_color']).grid(row=1, column=0, sticky='w', pady=3)
        self.venta_producto_var = tk.StringVar()
        self.combo_venta_prod = ttk.Combobox(sales_form, textvariable=self.venta_producto_var, state='readonly', width=23)
        self.combo_venta_prod.grid(row=1, column=1, pady=3, padx=(10, 0))
        self.combo_venta_prod.bind('<<ComboboxSelected>>', lambda e: self.venta_on_producto_change())
        # Costo unitario mostrado
        self.venta_costo_label = tk.Label(sales_form, text='Costo un.: -', font=('Arial', 9, 'bold'), fg=STYLE_CONFIG['primary_color'], bg=STYLE_CONFIG['bg_color'])
        self.venta_costo_label.grid(row=1, column=2, sticky='w', padx=(15,0))

        tk.Label(sales_form, text='🔢 Cantidad', font=('Arial', 9, 'bold'), fg=STYLE_CONFIG['text_color'], bg=STYLE_CONFIG['bg_color']).grid(row=2, column=0, sticky='w', pady=3)
        self.venta_cantidad_var = tk.StringVar()
        e_cant = tk.Entry(sales_form, textvariable=self.venta_cantidad_var, width=25, font=('Arial', 9), relief='solid', borderwidth=1, validate='key', validatecommand=self.vcmd_decimal)
        e_cant.grid(row=2, column=1, pady=3, padx=(10, 0))
        e_cant.bind('<KeyRelease>', lambda e: self.venta_recalc_totales())
        self.venta_disp_label = tk.Label(sales_form, text='Disponible: -', font=('Arial', 9, 'bold'), fg=STYLE_CONFIG['primary_color'], bg=STYLE_CONFIG['bg_color'])
        self.venta_disp_label.grid(row=2, column=2, sticky='w', padx=(15,0))

        tk.Label(sales_form, text='💵 Precio unitario', font=('Arial', 9, 'bold'), fg=STYLE_CONFIG['text_color'], bg=STYLE_CONFIG['bg_color']).grid(row=3, column=0, sticky='w', pady=3)
        self.venta_precio_var = tk.StringVar()
        e_precio = tk.Entry(sales_form, textvariable=self.venta_precio_var, width=25, font=('Arial', 9), relief='solid', borderwidth=1, validate='key', validatecommand=self.vcmd_decimal)
        e_precio.grid(row=3, column=1, pady=3, padx=(10, 0))
        e_precio.bind('<KeyRelease>', lambda e: self.venta_recalc_totales())

        tk.Label(sales_form, text='🏷️ Descuento (%)', font=('Arial', 9, 'bold'), fg=STYLE_CONFIG['text_color'], bg=STYLE_CONFIG['bg_color']).grid(row=4, column=0, sticky='w', pady=3)
        self.venta_descuento_var = tk.StringVar()
        e_desc = tk.Entry(sales_form, textvariable=self.venta_descuento_var, width=25, font=('Arial', 9), relief='solid', borderwidth=1, validate='key', validatecommand=self.vcmd_decimal)
        e_desc.grid(row=4, column=1, pady=3, padx=(10, 0))
        e_desc.bind('<KeyRelease>', lambda e: self.venta_recalc_totales())
        # Simulación de descuento
        ttk.Button(sales_form, text='Simular', command=self.venta_simular_descuento, style='Action.TButton').grid(row=4, column=2, sticky='w')
        self.venta_preview_desc_label = tk.Label(sales_form, text='Precio con desc.: -', font=('Arial', 9, 'bold'), fg=STYLE_CONFIG['primary_color'], bg=STYLE_CONFIG['bg_color'])
        self.venta_preview_desc_label.grid(row=4, column=3, sticky='w')

        # IIB (%)
        tk.Label(sales_form, text='🧾 IIB (%)', font=('Arial', 9, 'bold'), fg=STYLE_CONFIG['text_color'], bg=STYLE_CONFIG['bg_color']).grid(row=5, column=0, sticky='w', pady=3)
        self.venta_iib_var = tk.StringVar()
        e_iib = tk.Entry(sales_form, textvariable=self.venta_iib_var, width=25, font=('Arial', 9), relief='solid', borderwidth=1, validate='key', validatecommand=self.vcmd_decimal)
        e_iib.grid(row=5, column=1, pady=3, padx=(10, 0))
        e_iib.bind('<KeyRelease>', lambda e: self.venta_recalc_totales())
        self.venta_preview_final_label = tk.Label(sales_form, text='Precio final: -', font=('Arial', 9, 'bold'), fg=STYLE_CONFIG['primary_color'], bg=STYLE_CONFIG['bg_color'])
        self.venta_preview_final_label.grid(row=5, column=3, sticky='w')

        # Totales
        totals_frame = tk.Frame(sales_form, bg=STYLE_CONFIG['bg_color'])
        totals_frame.grid(row=6, column=0, columnspan=3, sticky='w', pady=(8,0))
        self.venta_total_bruto_lbl = tk.Label(totals_frame, text='Bruto: 0.00', font=('Arial', 9, 'bold'), fg=STYLE_CONFIG['primary_color'], bg=STYLE_CONFIG['bg_color'])
        self.venta_total_bruto_lbl.pack(side='left', padx=(0,15))
        self.venta_total_neto_lbl = tk.Label(totals_frame, text='Neto: 0.00', font=('Arial', 9, 'bold'), fg=STYLE_CONFIG['primary_color'], bg=STYLE_CONFIG['bg_color'])
        self.venta_total_neto_lbl.pack(side='left')
        # Ganancia
        self.venta_ganancia_un_lbl = tk.Label(totals_frame, text='Ganancia un.: 0.00', font=('Arial', 9, 'bold'), fg=STYLE_CONFIG['primary_color'], bg=STYLE_CONFIG['bg_color'])
        self.venta_ganancia_un_lbl.pack(side='left', padx=(15,15))
        self.venta_ganancia_total_lbl = tk.Label(totals_frame, text='Ganancia total: 0.00', font=('Arial', 9, 'bold'), fg=STYLE_CONFIG['primary_color'], bg=STYLE_CONFIG['bg_color'])
        self.venta_ganancia_total_lbl.pack(side='left')

        ttk.Button(sales_form, text='➕ Registrar Venta', command=self.registrar_venta, style='Action.TButton').grid(row=7, column=0, columnspan=3, pady=12)

        # Tabla de ventas
        sales_tree_frame = tk.Frame(frame_sales, bg=STYLE_CONFIG['bg_color'])
        sales_tree_frame.grid(row=1, column=0, padx=15, pady=15, sticky='nsew')
        frame_sales.grid_rowconfigure(1, weight=1)
        frame_sales.grid_columnconfigure(0, weight=1)
        xscroll_sales = tk.Scrollbar(sales_tree_frame, orient='horizontal')
        yscroll_sales = tk.Scrollbar(sales_tree_frame, orient='vertical')
        self.tree_ventas = ttk.Treeview(sales_tree_frame,
            columns=('Fecha', 'Familia', 'Medida', 'Característica', 'Cantidad (u)', 'Precio unitario', 'Descuento (%)', 'IIB (%)', 'Precio final', 'Costo unitario', 'Ingreso bruto', 'Ingreso neto', 'Ganancia un.', 'Ganancia total', 'Stock antes', 'Stock después'),
            show='headings', style='Custom.Treeview', height=12, xscrollcommand=xscroll_sales.set, yscrollcommand=yscroll_sales.set)
        for col in self.tree_ventas['columns']:
            self.tree_ventas.heading(col, text=col)
        # Ajustar anchos para que quepa la tabla completa
        ventas_col_widths = {
            'Fecha': 150,
            'Producto': 120,
            'Cantidad (u)': 90,
            'Precio unitario': 100,
            'Descuento (%)': 90,
            'IIB (%)': 80,
            'Precio final': 110,
            'Costo unitario': 110,
            'Ingreso bruto': 120,
            'Ingreso neto': 120,
            'Ganancia un.': 110,
            'Ganancia total': 130,
            'Stock antes': 100,
            'Stock después': 110,
        }
        for col in self.tree_ventas['columns']:
            width = ventas_col_widths.get(col, 110)
            self.tree_ventas.column(col, width=width, anchor='center', minwidth=80)
        self.tree_ventas.grid(row=0, column=0, sticky='nsew')
        yscroll_sales.config(command=self.tree_ventas.yview)
        xscroll_sales.config(command=self.tree_ventas.xview)
        yscroll_sales.grid(row=0, column=1, sticky='ns')
        xscroll_sales.grid(row=1, column=0, sticky='ew')
        btn_sales = tk.Frame(frame_sales, bg=STYLE_CONFIG['bg_color'])
        btn_sales.grid(row=2, column=0, pady=5)
        ttk.Button(btn_sales, text='🗑️ Eliminar venta', command=self.eliminar_venta, style='Danger.TButton').pack(side='left', padx=5)
        # Inicializar combo de ventas
        self._refresh_combo_venta_productos()

        # Pestaña Compra sin scroll innecesario
        frame_compra_parent = tk.Frame(self.tabs, bg=STYLE_CONFIG['bg_color'])
        self.tabs.add(frame_compra_parent, text='🛒 Compra')
        frame_compra = tk.Frame(frame_compra_parent, bg=STYLE_CONFIG['bg_color'])

        # Crear sub-notebook para las diferentes secciones de compra
        self.compra_tabs = ttk.Notebook(frame_compra, style='SubNotebook.TNotebook')
        self.compra_tabs.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Pack del frame principal
        frame_compra.pack(fill='both', expand=True)

        # ==================== PESTAÑA 1: MATERIA PRIMA ====================
        frame_compra_mp_parent = tk.Frame(self.compra_tabs, bg=STYLE_CONFIG['bg_color'])
        self.compra_tabs.add(frame_compra_mp_parent, text='🧱 Compra Materia Prima')

        frame_compra_mp = tk.Frame(frame_compra_mp_parent, bg=STYLE_CONFIG['bg_color'])
        frame_compra_mp.pack(fill='both', expand=True)

        # ========== LAYOUT MEJORADO: FORMULARIO + LOGS ==========
        # Configurar grid para formulario a la izquierda y logs a la derecha
        frame_compra_mp.grid_columnconfigure(0, weight=1)
        frame_compra_mp.grid_columnconfigure(1, weight=1)
        frame_compra_mp.grid_rowconfigure(0, weight=1)

        # ========== FORMULARIO DE COMPRA (IZQUIERDA) ==========
        compra_mp_form = ttk.LabelFrame(frame_compra_mp, text='📝 Registrar Compra de Materia Prima', padding=20)
        compra_mp_form.grid(row=0, column=0, sticky='nsew', padx=(15, 7), pady=15)
        compra_mp_form.grid_columnconfigure(1, weight=1)

        # Material
        tk.Label(compra_mp_form, text='🧱 Material', font=('Segoe UI', 11, 'bold'), fg=STYLE_CONFIG['text_color'], bg=STYLE_CONFIG['card_bg']).grid(row=0, column=0, sticky='w', pady=8)
        self.compra_mp_material_var = tk.StringVar()
        material_entry = tk.Entry(compra_mp_form, textvariable=self.compra_mp_material_var, width=40, font=('Segoe UI', 11), relief='flat', borderwidth=0, bg=STYLE_CONFIG['light_bg'], fg=STYLE_CONFIG['text_color'], highlightthickness=2, highlightcolor=STYLE_CONFIG['primary_color'], highlightbackground=STYLE_CONFIG['border_color'], insertbackground=STYLE_CONFIG['primary_color'])
        material_entry.grid(row=0, column=1, sticky='ew', padx=(15, 0), pady=8, ipady=6)

        # Familia (opcional)
        tk.Label(compra_mp_form, text='📁 Familia (opcional)', font=('Segoe UI', 11, 'bold'), fg=STYLE_CONFIG['text_color'], bg=STYLE_CONFIG['card_bg']).grid(row=1, column=0, sticky='w', pady=8)
        self.compra_mp_familia_var = tk.StringVar()
        familia_entry = tk.Entry(compra_mp_form, textvariable=self.compra_mp_familia_var, width=40, font=('Segoe UI', 11), relief='flat', borderwidth=0, bg=STYLE_CONFIG['light_bg'], fg=STYLE_CONFIG['text_color'], highlightthickness=2, highlightcolor=STYLE_CONFIG['primary_color'], highlightbackground=STYLE_CONFIG['border_color'], insertbackground=STYLE_CONFIG['primary_color'])
        familia_entry.grid(row=1, column=1, sticky='ew', padx=(15, 0), pady=8, ipady=6)

        # Medida (opcional)
        tk.Label(compra_mp_form, text='📐 Medida (opcional)', font=('Segoe UI', 11, 'bold'), fg=STYLE_CONFIG['text_color'], bg=STYLE_CONFIG['card_bg']).grid(row=2, column=0, sticky='w', pady=8)
        self.compra_mp_medida_var = tk.StringVar()
        medida_entry = tk.Entry(compra_mp_form, textvariable=self.compra_mp_medida_var, width=40, font=('Segoe UI', 11), relief='flat', borderwidth=0, bg=STYLE_CONFIG['light_bg'], fg=STYLE_CONFIG['text_color'], highlightthickness=2, highlightcolor=STYLE_CONFIG['primary_color'], highlightbackground=STYLE_CONFIG['border_color'], insertbackground=STYLE_CONFIG['primary_color'])
        medida_entry.grid(row=2, column=1, sticky='ew', padx=(15, 0), pady=8, ipady=6)

        # Característica (opcional)
        tk.Label(compra_mp_form, text='🔍 Característica (opcional)', font=('Segoe UI', 11, 'bold'), fg=STYLE_CONFIG['text_color'], bg=STYLE_CONFIG['card_bg']).grid(row=3, column=0, sticky='w', pady=8)
        self.compra_mp_caracteristica_var = tk.StringVar()
        caracteristica_entry = tk.Entry(compra_mp_form, textvariable=self.compra_mp_caracteristica_var, width=40, font=('Segoe UI', 11), relief='flat', borderwidth=0, bg=STYLE_CONFIG['light_bg'], fg=STYLE_CONFIG['text_color'], highlightthickness=2, highlightcolor=STYLE_CONFIG['primary_color'], highlightbackground=STYLE_CONFIG['border_color'], insertbackground=STYLE_CONFIG['primary_color'])
        caracteristica_entry.grid(row=3, column=1, sticky='ew', padx=(15, 0), pady=8, ipady=6)

        # Cantidad
        tk.Label(compra_mp_form, text='📏 Cantidad (kg)', font=('Segoe UI', 11, 'bold'), fg=STYLE_CONFIG['text_color'], bg=STYLE_CONFIG['card_bg']).grid(row=4, column=0, sticky='w', pady=8)
        self.compra_mp_cantidad_var = tk.StringVar()
        cantidad_entry = tk.Entry(compra_mp_form, textvariable=self.compra_mp_cantidad_var, width=40, font=('Segoe UI', 11), relief='flat', borderwidth=0, bg=STYLE_CONFIG['light_bg'], fg=STYLE_CONFIG['text_color'], highlightthickness=2, highlightcolor=STYLE_CONFIG['primary_color'], highlightbackground=STYLE_CONFIG['border_color'], insertbackground=STYLE_CONFIG['primary_color'], validate='key', validatecommand=self.vcmd_decimal)
        cantidad_entry.grid(row=4, column=1, sticky='ew', padx=(15, 0), pady=8, ipady=6)

        # Precio por kg
        tk.Label(compra_mp_form, text='💰 Precio por kg ($)', font=('Segoe UI', 11, 'bold'), fg=STYLE_CONFIG['text_color'], bg=STYLE_CONFIG['card_bg']).grid(row=5, column=0, sticky='w', pady=8)
        self.compra_mp_precio_var = tk.StringVar()
        precio_entry = tk.Entry(compra_mp_form, textvariable=self.compra_mp_precio_var, width=40, font=('Segoe UI', 11), relief='flat', borderwidth=0, bg=STYLE_CONFIG['light_bg'], fg=STYLE_CONFIG['text_color'], highlightthickness=2, highlightcolor=STYLE_CONFIG['primary_color'], highlightbackground=STYLE_CONFIG['border_color'], insertbackground=STYLE_CONFIG['primary_color'], validate='key', validatecommand=self.vcmd_decimal)
        precio_entry.grid(row=5, column=1, sticky='ew', padx=(15, 0), pady=8, ipady=6)

        # Proveedor
        tk.Label(compra_mp_form, text='🏢 Proveedor', font=('Segoe UI', 11, 'bold'), fg=STYLE_CONFIG['text_color'], bg=STYLE_CONFIG['card_bg']).grid(row=6, column=0, sticky='w', pady=8)
        self.compra_mp_proveedor_var = tk.StringVar()
        proveedor_entry = tk.Entry(compra_mp_form, textvariable=self.compra_mp_proveedor_var, width=40, font=('Segoe UI', 11), relief='flat', borderwidth=0, bg=STYLE_CONFIG['light_bg'], fg=STYLE_CONFIG['text_color'], highlightthickness=2, highlightcolor=STYLE_CONFIG['primary_color'], highlightbackground=STYLE_CONFIG['border_color'], insertbackground=STYLE_CONFIG['primary_color'])
        proveedor_entry.grid(row=6, column=1, sticky='ew', padx=(15, 0), pady=8, ipady=6)

        # Fecha
        tk.Label(compra_mp_form, text='📅 Fecha', font=('Segoe UI', 11, 'bold'), fg=STYLE_CONFIG['text_color'], bg=STYLE_CONFIG['card_bg']).grid(row=7, column=0, sticky='w', pady=8)
        self.compra_mp_fecha_var = tk.StringVar()
        fecha_entry = tk.Entry(compra_mp_form, textvariable=self.compra_mp_fecha_var, width=40, font=('Segoe UI', 11), relief='flat', borderwidth=0, bg=STYLE_CONFIG['light_bg'], fg=STYLE_CONFIG['text_color'], highlightthickness=2, highlightcolor=STYLE_CONFIG['primary_color'], highlightbackground=STYLE_CONFIG['border_color'], insertbackground=STYLE_CONFIG['primary_color'])
        fecha_entry.grid(row=7, column=1, sticky='ew', padx=(15, 0), pady=8, ipady=6)

        # Moneda
        tk.Label(compra_mp_form, text='💱 Moneda', font=('Segoe UI', 11, 'bold'), fg=STYLE_CONFIG['text_color'], bg=STYLE_CONFIG['card_bg']).grid(row=8, column=0, sticky='w', padx=15, pady=8)
        
        self.compra_mp_moneda_var = tk.StringVar(value='ARS')
        moneda_mp_frame = tk.Frame(compra_mp_form, bg=STYLE_CONFIG['card_bg'])
        moneda_mp_frame.grid(row=8, column=1, sticky='w', padx=(15, 0), pady=8)
        
        radio_ars_mp = tk.Radiobutton(moneda_mp_frame, text='ARS', variable=self.compra_mp_moneda_var, value='ARS', font=('Segoe UI', 10), bg=STYLE_CONFIG['card_bg'], fg=STYLE_CONFIG['text_color'], command=self._actualizar_campos_moneda_compra_mp)
        radio_ars_mp.pack(side='left', padx=(0, 15))
        
        radio_usd_mp = tk.Radiobutton(moneda_mp_frame, text='USD', variable=self.compra_mp_moneda_var, value='USD', font=('Segoe UI', 10), bg=STYLE_CONFIG['card_bg'], fg=STYLE_CONFIG['text_color'], command=self._actualizar_campos_moneda_compra_mp)
        radio_usd_mp.pack(side='left')
        
        # Valor del dólar
        self.compra_mp_lbl_dolar = tk.Label(compra_mp_form, text='💵 Valor del dólar', font=('Segoe UI', 11, 'bold'), fg=STYLE_CONFIG['text_color'], bg=STYLE_CONFIG['card_bg'])
        self.compra_mp_valor_dolar_var = tk.StringVar()
        self.compra_mp_entry_dolar = tk.Entry(compra_mp_form, textvariable=self.compra_mp_valor_dolar_var, width=40, font=('Segoe UI', 11), relief='flat', borderwidth=0, bg=STYLE_CONFIG['light_bg'], fg=STYLE_CONFIG['text_color'], highlightthickness=2, highlightcolor=STYLE_CONFIG['primary_color'], highlightbackground=STYLE_CONFIG['border_color'], insertbackground=STYLE_CONFIG['primary_color'], validate='key', validatecommand=self.vcmd_decimal)

        # Botones para materia prima
        btn_compra_mp_frame = tk.Frame(compra_mp_form, bg=STYLE_CONFIG['card_bg'])
        btn_compra_mp_frame.grid(row=10, column=0, columnspan=2, pady=20)

        ttk.Button(btn_compra_mp_frame, text='➕ Registrar Compra MP', command=self.registrar_compra_materia_prima, style='Action.TButton').pack(side='left', padx=10)
        ttk.Button(btn_compra_mp_frame, text='🔄 Limpiar Formulario', command=self.limpiar_formulario_compra_mp, style='Warning.TButton').pack(side='left', padx=10)
        
        # Actualizar visibilidad inicial
        self._actualizar_campos_moneda_compra_mp()

        # ========== LOGS DE COMPRAS (DERECHA) ==========
        logs_compra_mp_frame = ttk.LabelFrame(frame_compra_mp, text='📋 Logs de Compras de Materia Prima', padding=15)
        logs_compra_mp_frame.grid(row=0, column=1, sticky='nsew', padx=(7, 15), pady=15)
        
        # Tabla de logs con scroll
        logs_tree_frame = tk.Frame(logs_compra_mp_frame, bg=STYLE_CONFIG['card_bg'])
        logs_tree_frame.pack(fill='both', expand=True)
        
        xscroll_logs_mp = tk.Scrollbar(logs_tree_frame, orient='horizontal')
        yscroll_logs_mp = tk.Scrollbar(logs_tree_frame, orient='vertical')
        
        self.tree_logs_compra_mp = ttk.Treeview(logs_tree_frame,
                                               columns=('Fecha', 'Material', 'Cantidad', 'Precio', 'Proveedor', 'Total'),
                                               show='headings',
                                               style='Custom.Treeview',
                                               height=20,
                                               xscrollcommand=xscroll_logs_mp.set,
                                               yscrollcommand=yscroll_logs_mp.set)
        
        # Configurar columnas
        self.tree_logs_compra_mp.heading('Fecha', text='📅 Fecha')
        self.tree_logs_compra_mp.heading('Material', text='🧱 Material')
        self.tree_logs_compra_mp.heading('Cantidad', text='📏 Cantidad (kg)')
        self.tree_logs_compra_mp.heading('Precio', text='💰 Precio/kg')
        self.tree_logs_compra_mp.heading('Proveedor', text='🏢 Proveedor')
        self.tree_logs_compra_mp.heading('Total', text='💵 Total')
        
        # Ancho de columnas
        self.tree_logs_compra_mp.column('Fecha', width=100, anchor='center')
        self.tree_logs_compra_mp.column('Material', width=200, anchor='w')
        self.tree_logs_compra_mp.column('Cantidad', width=100, anchor='center')
        self.tree_logs_compra_mp.column('Precio', width=100, anchor='center')
        self.tree_logs_compra_mp.column('Proveedor', width=150, anchor='w')
        self.tree_logs_compra_mp.column('Total', width=120, anchor='center')
        
        self.tree_logs_compra_mp.grid(row=0, column=0, sticky='nsew')
        yscroll_logs_mp.config(command=self.tree_logs_compra_mp.yview)
        xscroll_logs_mp.config(command=self.tree_logs_compra_mp.xview)
        yscroll_logs_mp.grid(row=0, column=1, sticky='ns')
        xscroll_logs_mp.grid(row=1, column=0, sticky='ew')
        
        logs_tree_frame.grid_rowconfigure(0, weight=1)
        logs_tree_frame.grid_columnconfigure(0, weight=1)

        # ==================== PESTAÑA 2: PRODUCTOS DE REVENTA ====================
        frame_compra_prod_parent = tk.Frame(self.compra_tabs, bg=STYLE_CONFIG['bg_color'])
        self.compra_tabs.add(frame_compra_prod_parent, text='📦 Compra Productos')

        frame_compra_prod = tk.Frame(frame_compra_prod_parent, bg=STYLE_CONFIG['bg_color'])
        frame_compra_prod.pack(fill='both', expand=True)

        # ========== LAYOUT MEJORADO: FORMULARIO + LOGS ==========
        # Configurar grid para formulario a la izquierda y logs a la derecha
        frame_compra_prod.grid_columnconfigure(0, weight=1)
        frame_compra_prod.grid_columnconfigure(1, weight=1)
        frame_compra_prod.grid_rowconfigure(0, weight=1)

        # ========== FORMULARIO DE COMPRA (IZQUIERDA) ==========
        compra_prod_form = ttk.LabelFrame(frame_compra_prod, text='📝 Registrar Compra de Productos', padding=20)
        compra_prod_form.grid(row=0, column=0, sticky='nsew', padx=(15, 7), pady=15)
        compra_prod_form.grid_columnconfigure(1, weight=1)

        # Producto
        tk.Label(compra_prod_form, text='📦 Producto', font=('Segoe UI', 11, 'bold'), fg=STYLE_CONFIG['text_color'], bg=STYLE_CONFIG['card_bg']).grid(row=0, column=0, sticky='w', pady=8)
        self.compra_prod_producto_var = tk.StringVar()
        producto_entry = tk.Entry(compra_prod_form, textvariable=self.compra_prod_producto_var, width=40, font=('Segoe UI', 11), relief='flat', borderwidth=0, bg=STYLE_CONFIG['light_bg'], fg=STYLE_CONFIG['text_color'], highlightthickness=2, highlightcolor=STYLE_CONFIG['primary_color'], highlightbackground=STYLE_CONFIG['border_color'], insertbackground=STYLE_CONFIG['primary_color'])
        producto_entry.grid(row=0, column=1, sticky='ew', padx=(15, 0), pady=8, ipady=6)

        # Familia (opcional)
        tk.Label(compra_prod_form, text='📁 Familia (opcional)', font=('Segoe UI', 11, 'bold'), fg=STYLE_CONFIG['text_color'], bg=STYLE_CONFIG['card_bg']).grid(row=1, column=0, sticky='w', pady=8)
        self.compra_prod_familia_var = tk.StringVar()
        familia_prod_entry = tk.Entry(compra_prod_form, textvariable=self.compra_prod_familia_var, width=40, font=('Segoe UI', 11), relief='flat', borderwidth=0, bg=STYLE_CONFIG['light_bg'], fg=STYLE_CONFIG['text_color'], highlightthickness=2, highlightcolor=STYLE_CONFIG['primary_color'], highlightbackground=STYLE_CONFIG['border_color'], insertbackground=STYLE_CONFIG['primary_color'])
        familia_prod_entry.grid(row=1, column=1, sticky='ew', padx=(15, 0), pady=8, ipady=6)

        # Medida (opcional)
        tk.Label(compra_prod_form, text='📐 Medida (opcional)', font=('Segoe UI', 11, 'bold'), fg=STYLE_CONFIG['text_color'], bg=STYLE_CONFIG['card_bg']).grid(row=2, column=0, sticky='w', pady=8)
        self.compra_prod_medida_var = tk.StringVar()
        medida_prod_entry = tk.Entry(compra_prod_form, textvariable=self.compra_prod_medida_var, width=40, font=('Segoe UI', 11), relief='flat', borderwidth=0, bg=STYLE_CONFIG['light_bg'], fg=STYLE_CONFIG['text_color'], highlightthickness=2, highlightcolor=STYLE_CONFIG['primary_color'], highlightbackground=STYLE_CONFIG['border_color'], insertbackground=STYLE_CONFIG['primary_color'])
        medida_prod_entry.grid(row=2, column=1, sticky='ew', padx=(15, 0), pady=8, ipady=6)

        # Característica (opcional)
        tk.Label(compra_prod_form, text='🔍 Característica (opcional)', font=('Segoe UI', 11, 'bold'), fg=STYLE_CONFIG['text_color'], bg=STYLE_CONFIG['card_bg']).grid(row=3, column=0, sticky='w', pady=8)
        self.compra_prod_caracteristica_var = tk.StringVar()
        caracteristica_prod_entry = tk.Entry(compra_prod_form, textvariable=self.compra_prod_caracteristica_var, width=40, font=('Segoe UI', 11), relief='flat', borderwidth=0, bg=STYLE_CONFIG['light_bg'], fg=STYLE_CONFIG['text_color'], highlightthickness=2, highlightcolor=STYLE_CONFIG['primary_color'], highlightbackground=STYLE_CONFIG['border_color'], insertbackground=STYLE_CONFIG['primary_color'])
        caracteristica_prod_entry.grid(row=3, column=1, sticky='ew', padx=(15, 0), pady=8, ipady=6)

        # Cantidad
        tk.Label(compra_prod_form, text='📏 Cantidad', font=('Segoe UI', 11, 'bold'), fg=STYLE_CONFIG['text_color'], bg=STYLE_CONFIG['card_bg']).grid(row=4, column=0, sticky='w', pady=8)
        self.compra_prod_cantidad_var = tk.StringVar()
        cantidad_prod_entry = tk.Entry(compra_prod_form, textvariable=self.compra_prod_cantidad_var, width=40, font=('Segoe UI', 11), relief='flat', borderwidth=0, bg=STYLE_CONFIG['light_bg'], fg=STYLE_CONFIG['text_color'], highlightthickness=2, highlightcolor=STYLE_CONFIG['primary_color'], highlightbackground=STYLE_CONFIG['border_color'], insertbackground=STYLE_CONFIG['primary_color'], validate='key', validatecommand=self.vcmd_decimal)
        cantidad_prod_entry.grid(row=4, column=1, sticky='ew', padx=(15, 0), pady=8, ipady=6)

        # Precio unitario
        tk.Label(compra_prod_form, text='💰 Precio unitario ($)', font=('Segoe UI', 11, 'bold'), fg=STYLE_CONFIG['text_color'], bg=STYLE_CONFIG['card_bg']).grid(row=5, column=0, sticky='w', pady=8)
        self.compra_prod_precio_var = tk.StringVar()
        precio_prod_entry = tk.Entry(compra_prod_form, textvariable=self.compra_prod_precio_var, width=40, font=('Segoe UI', 11), relief='flat', borderwidth=0, bg=STYLE_CONFIG['light_bg'], fg=STYLE_CONFIG['text_color'], highlightthickness=2, highlightcolor=STYLE_CONFIG['primary_color'], highlightbackground=STYLE_CONFIG['border_color'], insertbackground=STYLE_CONFIG['primary_color'], validate='key', validatecommand=self.vcmd_decimal)
        precio_prod_entry.grid(row=5, column=1, sticky='ew', padx=(15, 0), pady=8, ipady=6)

        # Proveedor
        tk.Label(compra_prod_form, text='🏢 Proveedor', font=('Segoe UI', 11, 'bold'), fg=STYLE_CONFIG['text_color'], bg=STYLE_CONFIG['card_bg']).grid(row=6, column=0, sticky='w', pady=8)
        self.compra_prod_proveedor_var = tk.StringVar()
        proveedor_prod_entry = tk.Entry(compra_prod_form, textvariable=self.compra_prod_proveedor_var, width=40, font=('Segoe UI', 11), relief='flat', borderwidth=0, bg=STYLE_CONFIG['light_bg'], fg=STYLE_CONFIG['text_color'], highlightthickness=2, highlightcolor=STYLE_CONFIG['primary_color'], highlightbackground=STYLE_CONFIG['border_color'], insertbackground=STYLE_CONFIG['primary_color'])
        proveedor_prod_entry.grid(row=6, column=1, sticky='ew', padx=(15, 0), pady=8, ipady=6)

        # Fecha
        tk.Label(compra_prod_form, text='📅 Fecha', font=('Segoe UI', 11, 'bold'), fg=STYLE_CONFIG['text_color'], bg=STYLE_CONFIG['card_bg']).grid(row=7, column=0, sticky='w', pady=8)
        self.compra_prod_fecha_var = tk.StringVar()
        fecha_prod_entry = tk.Entry(compra_prod_form, textvariable=self.compra_prod_fecha_var, width=40, font=('Segoe UI', 11), relief='flat', borderwidth=0, bg=STYLE_CONFIG['light_bg'], fg=STYLE_CONFIG['text_color'], highlightthickness=2, highlightcolor=STYLE_CONFIG['primary_color'], highlightbackground=STYLE_CONFIG['border_color'], insertbackground=STYLE_CONFIG['primary_color'])
        fecha_prod_entry.grid(row=7, column=1, sticky='ew', padx=(15, 0), pady=8, ipady=6)

        # Moneda
        tk.Label(compra_prod_form, text='💱 Moneda', font=('Segoe UI', 11, 'bold'), fg=STYLE_CONFIG['text_color'], bg=STYLE_CONFIG['card_bg']).grid(row=8, column=0, sticky='w', padx=15, pady=8)
        
        self.compra_prod_moneda_var = tk.StringVar(value='ARS')
        moneda_prod_frame = tk.Frame(compra_prod_form, bg=STYLE_CONFIG['card_bg'])
        moneda_prod_frame.grid(row=8, column=1, sticky='w', padx=(15, 0), pady=8)
        
        radio_ars_prod = tk.Radiobutton(moneda_prod_frame, text='ARS', variable=self.compra_prod_moneda_var, value='ARS', font=('Segoe UI', 10), bg=STYLE_CONFIG['card_bg'], fg=STYLE_CONFIG['text_color'], command=self._actualizar_campos_moneda_compra_prod)
        radio_ars_prod.pack(side='left', padx=(0, 15))
        
        radio_usd_prod = tk.Radiobutton(moneda_prod_frame, text='USD', variable=self.compra_prod_moneda_var, value='USD', font=('Segoe UI', 10), bg=STYLE_CONFIG['card_bg'], fg=STYLE_CONFIG['text_color'], command=self._actualizar_campos_moneda_compra_prod)
        radio_usd_prod.pack(side='left')
        
        # Valor del dólar
        self.compra_prod_lbl_dolar = tk.Label(compra_prod_form, text='💵 Valor del dólar', font=('Segoe UI', 11, 'bold'), fg=STYLE_CONFIG['text_color'], bg=STYLE_CONFIG['card_bg'])
        self.compra_prod_valor_dolar_var = tk.StringVar()
        self.compra_prod_entry_dolar = tk.Entry(compra_prod_form, textvariable=self.compra_prod_valor_dolar_var, width=40, font=('Segoe UI', 11), relief='flat', borderwidth=0, bg=STYLE_CONFIG['light_bg'], fg=STYLE_CONFIG['text_color'], highlightthickness=2, highlightcolor=STYLE_CONFIG['primary_color'], highlightbackground=STYLE_CONFIG['border_color'], insertbackground=STYLE_CONFIG['primary_color'], validate='key', validatecommand=self.vcmd_decimal)

        # Botones para productos
        btn_compra_prod_frame = tk.Frame(compra_prod_form, bg=STYLE_CONFIG['card_bg'])
        btn_compra_prod_frame.grid(row=10, column=0, columnspan=2, pady=20)

        ttk.Button(btn_compra_prod_frame, text='➕ Registrar Compra Producto', command=self.registrar_compra_producto, style='Action.TButton').pack(side='left', padx=10)
        ttk.Button(btn_compra_prod_frame, text='🔄 Limpiar Formulario', command=self.limpiar_formulario_compra_prod, style='Warning.TButton').pack(side='left', padx=10)
        
        # Actualizar visibilidad inicial
        self._actualizar_campos_moneda_compra_prod()

        # ========== LOGS DE COMPRAS (DERECHA) ==========
        logs_compra_prod_frame = ttk.LabelFrame(frame_compra_prod, text='📋 Logs de Compras de Productos', padding=15)
        logs_compra_prod_frame.grid(row=0, column=1, sticky='nsew', padx=(7, 15), pady=15)
        
        # Tabla de logs con scroll
        logs_prod_tree_frame = tk.Frame(logs_compra_prod_frame, bg=STYLE_CONFIG['card_bg'])
        logs_prod_tree_frame.pack(fill='both', expand=True)
        
        xscroll_logs_prod = tk.Scrollbar(logs_prod_tree_frame, orient='horizontal')
        yscroll_logs_prod = tk.Scrollbar(logs_prod_tree_frame, orient='vertical')
        
        self.tree_logs_compra_prod = ttk.Treeview(logs_prod_tree_frame,
                                                 columns=('Fecha', 'Producto', 'Cantidad', 'Precio', 'Proveedor', 'Total'),
                                                 show='headings',
                                                 style='Custom.Treeview',
                                                 height=20,
                                                 xscrollcommand=xscroll_logs_prod.set,
                                                 yscrollcommand=yscroll_logs_prod.set)
        
        # Configurar columnas
        self.tree_logs_compra_prod.heading('Fecha', text='📅 Fecha')
        self.tree_logs_compra_prod.heading('Producto', text='📦 Producto')
        self.tree_logs_compra_prod.heading('Cantidad', text='📏 Cantidad')
        self.tree_logs_compra_prod.heading('Precio', text='💰 Precio/Unidad')
        self.tree_logs_compra_prod.heading('Proveedor', text='🏢 Proveedor')
        self.tree_logs_compra_prod.heading('Total', text='💵 Total')
        
        # Ancho de columnas
        self.tree_logs_compra_prod.column('Fecha', width=100, anchor='center')
        self.tree_logs_compra_prod.column('Producto', width=200, anchor='w')
        self.tree_logs_compra_prod.column('Cantidad', width=100, anchor='center')
        self.tree_logs_compra_prod.column('Precio', width=100, anchor='center')
        self.tree_logs_compra_prod.column('Proveedor', width=150, anchor='w')
        self.tree_logs_compra_prod.column('Total', width=120, anchor='center')
        
        self.tree_logs_compra_prod.grid(row=0, column=0, sticky='nsew')
        yscroll_logs_prod.config(command=self.tree_logs_compra_prod.yview)
        xscroll_logs_prod.config(command=self.tree_logs_compra_prod.xview)
        yscroll_logs_prod.grid(row=0, column=1, sticky='ns')
        xscroll_logs_prod.grid(row=1, column=0, sticky='ew')
        
        logs_prod_tree_frame.grid_rowconfigure(0, weight=1)
        logs_prod_tree_frame.grid_columnconfigure(0, weight=1)

        # Inicializar combos de compra
        self._refresh_combo_compra_materiales()
        self._refresh_combo_compra_productos()

        # Pestaña Costos - Planilla de simulación con scroll
        frame_costos_parent = tk.Frame(self.tabs, bg=STYLE_CONFIG['bg_color'])
        self.tabs.add(frame_costos_parent, text='💰 Costos')
        frame_costos = self.crear_frame_con_scroll(frame_costos_parent)
        
        # Configurar columnas (3 columnas: formulario, gráfico, totales)
        frame_costos.grid_columnconfigure(0, weight=2)  # Formulario más ancho
        frame_costos.grid_columnconfigure(1, weight=1)  # Gráfico responsivo
        frame_costos.grid_columnconfigure(2, weight=1)  # Totales responsivo
        
        # Configurar filas para mejor distribución vertical
        frame_costos.grid_rowconfigure(0, weight=1)  # Fila superior (formulario, gráfico, totales)
        frame_costos.grid_rowconfigure(1, weight=0)  # Fila del título de la tabla
        frame_costos.grid_rowconfigure(2, weight=1)  # Fila de la tabla (se expande)
        
        # Formulario de carga de items (izquierda)
        form_costos = ttk.LabelFrame(frame_costos, text='📝 Actualizar Planilla', padding=15)
        form_costos.grid(row=0, column=0, sticky='nsew', padx=(15, 7), pady=15)
        
        # Familia del producto
        tk.Label(form_costos, text='📦 Familia', font=('Arial', 9, 'bold'), fg=STYLE_CONFIG['text_color'], bg=STYLE_CONFIG['bg_color']).grid(row=0, column=0, sticky='w', pady=3)
        self.costos_familia_var = tk.StringVar()
        tk.Entry(form_costos, textvariable=self.costos_familia_var, width=25, font=('Arial', 9), relief='solid', borderwidth=1).grid(row=0, column=1, pady=3, padx=(10, 0))
        
        # Medida del producto
        tk.Label(form_costos, text='📏 Medida', font=('Arial', 9, 'bold'), fg=STYLE_CONFIG['text_color'], bg=STYLE_CONFIG['bg_color']).grid(row=1, column=0, sticky='w', pady=3)
        self.costos_medida_var = tk.StringVar()
        tk.Entry(form_costos, textvariable=self.costos_medida_var, width=25, font=('Arial', 9), relief='solid', borderwidth=1).grid(row=1, column=1, pady=3, padx=(10, 0))
        
        # Característica del producto
        tk.Label(form_costos, text='🔍 Característica', font=('Arial', 9, 'bold'), fg=STYLE_CONFIG['text_color'], bg=STYLE_CONFIG['bg_color']).grid(row=2, column=0, sticky='w', pady=3)
        self.costos_caracteristica_var = tk.StringVar()
        tk.Entry(form_costos, textvariable=self.costos_caracteristica_var, width=25, font=('Arial', 9), relief='solid', borderwidth=1).grid(row=2, column=1, pady=3, padx=(10, 0))
        
        # Precio de venta
        tk.Label(form_costos, text='💵 Precio Venta', font=('Arial', 9, 'bold'), fg=STYLE_CONFIG['text_color'], bg=STYLE_CONFIG['bg_color']).grid(row=0, column=2, sticky='w', pady=3, padx=(15, 0))
        
        # Frame para precio de venta y moneda
        venta_frame = tk.Frame(form_costos, bg=STYLE_CONFIG['bg_color'])
        venta_frame.grid(row=0, column=3, pady=3, padx=(10, 0), sticky='ew')
        
        self.costos_venta_var = tk.StringVar()
        self.costos_venta_moneda_var = tk.StringVar(value='ARS')
        
        # Campo de entrada para precio
        entry_venta = tk.Entry(venta_frame, textvariable=self.costos_venta_var, width=10, font=('Arial', 9), relief='solid', borderwidth=1, validate='key', validatecommand=self.vcmd_decimal)
        entry_venta.pack(side='left', padx=(0, 8))
        
        # Frame para radio buttons con mejor espaciado
        moneda_venta_frame = tk.Frame(venta_frame, bg=STYLE_CONFIG['bg_color'])
        moneda_venta_frame.pack(side='left')
        
        # Radio buttons para moneda con mejor espaciado
        tk.Radiobutton(moneda_venta_frame, text='ARS', variable=self.costos_venta_moneda_var, value='ARS',
                      font=('Arial', 8), bg=STYLE_CONFIG['bg_color'],
                      fg=STYLE_CONFIG['text_color']).pack(side='left', padx=(0, 6))
        
        tk.Radiobutton(moneda_venta_frame, text='USD', variable=self.costos_venta_moneda_var, value='USD',
                      font=('Arial', 8), bg=STYLE_CONFIG['bg_color'],
                      fg=STYLE_CONFIG['text_color']).pack(side='left')
        
        # Cantidad a fabricar
        tk.Label(form_costos, text='🔢 Cant. Fabricar', font=('Arial', 9, 'bold'), fg=STYLE_CONFIG['text_color'], bg=STYLE_CONFIG['bg_color']).grid(row=1, column=2, sticky='w', pady=3, padx=(15, 0))
        self.costos_cant_fab_var = tk.StringVar()
        tk.Entry(form_costos, textvariable=self.costos_cant_fab_var, width=20, font=('Arial', 9), relief='solid', borderwidth=1, validate='key', validatecommand=self.vcmd_decimal).grid(row=1, column=3, pady=3, padx=(10, 0))
        
        # Cantidad por hora
        tk.Label(form_costos, text='⏱️ Cant. por Hora', font=('Arial', 9, 'bold'), fg=STYLE_CONFIG['text_color'], bg=STYLE_CONFIG['bg_color']).grid(row=2, column=2, sticky='w', pady=3, padx=(15, 0))
        self.costos_cant_hora_var = tk.StringVar()
        tk.Entry(form_costos, textvariable=self.costos_cant_hora_var, width=20, font=('Arial', 9), relief='solid', borderwidth=1, validate='key', validatecommand=self.vcmd_decimal).grid(row=2, column=3, pady=3, padx=(10, 0))
        
        # IIBB porcentaje (campo simple)
        tk.Label(form_costos, text='🏛️ IIBB %', font=('Arial', 9, 'bold'), fg=STYLE_CONFIG['text_color'], bg=STYLE_CONFIG['bg_color']).grid(row=3, column=0, sticky='w', pady=3)
        self.costos_iibb_porcentaje_var = tk.StringVar()
        tk.Entry(form_costos, textvariable=self.costos_iibb_porcentaje_var, width=25, font=('Arial', 9), relief='solid', borderwidth=1, validate='key', validatecommand=self.vcmd_decimal).grid(row=3, column=1, pady=3, padx=(10, 0))
        
        # Descuento porcentaje
        tk.Label(form_costos, text='🎁 Descuento %', font=('Arial', 9, 'bold'), fg=STYLE_CONFIG['text_color'], bg=STYLE_CONFIG['bg_color']).grid(row=3, column=2, sticky='w', pady=3, padx=(15, 0))
        self.costos_descuento_var = tk.StringVar()
        tk.Entry(form_costos, textvariable=self.costos_descuento_var, width=20, font=('Arial', 9), relief='solid', borderwidth=1, validate='key', validatecommand=self.vcmd_decimal).grid(row=3, column=3, pady=3, padx=(10, 0))
        
        # Precio del dólar actual
        tk.Label(form_costos, text='💱 Precio Dólar', font=('Arial', 9, 'bold'), fg=STYLE_CONFIG['text_color'], bg=STYLE_CONFIG['bg_color']).grid(row=4, column=0, sticky='w', pady=3)
        self.costos_dolar_var = tk.StringVar()
        tk.Entry(form_costos, textvariable=self.costos_dolar_var, width=25, font=('Arial', 9), relief='solid', borderwidth=1, validate='key', validatecommand=self.vcmd_decimal).grid(row=4, column=1, pady=3, padx=(10, 0))
        
        # Sección de materiales
        materiales_section = tk.Frame(form_costos, bg=STYLE_CONFIG['bg_color'])
        materiales_section.grid(row=5, column=0, columnspan=4, sticky='ew', pady=(10, 5))
        
        tk.Label(materiales_section, text='🧱 Materiales (agregar manualmente)', font=('Arial', 10, 'bold'), fg=STYLE_CONFIG['primary_color'], bg=STYLE_CONFIG['bg_color']).pack(anchor='w')
        
        materiales_inputs = tk.Frame(form_costos, bg=STYLE_CONFIG['bg_color'])
        materiales_inputs.grid(row=6, column=0, columnspan=4, sticky='ew', pady=5)
        
        # Botón para agregar material (reemplaza el combo y campos anteriores)
        ttk.Button(materiales_inputs, text='➕ Agregar Material Manual', command=self.agregar_material_costos, style='Action.TButton').grid(row=0, column=0, padx=(0, 10))
        
        # Botón para eliminar material seleccionado
        ttk.Button(materiales_inputs, text='🗑️ Eliminar Material', command=self.eliminar_material_costos, style='Danger.TButton').grid(row=0, column=1)
        
        # Tabla de materiales del item actual
        mat_table_frame = tk.Frame(form_costos, bg=STYLE_CONFIG['bg_color'])
        mat_table_frame.grid(row=7, column=0, columnspan=4, pady=(5, 10), sticky='ew')
        
        self.tree_materiales_costos = ttk.Treeview(mat_table_frame, columns=('Material', 'Kg/unidad', 'Costo/kg', 'Moneda', 'Costo Total'), show='headings', style='Custom.Treeview', height=4)
        for col in ('Material', 'Kg/unidad', 'Costo/kg', 'Moneda', 'Costo Total'):
            self.tree_materiales_costos.heading(col, text=col)
            self.tree_materiales_costos.column(col, width=120, anchor='center')
        self.tree_materiales_costos.pack(side='left', fill='both', expand=True)
        
        # Bind doble clic para editar material
        self.tree_materiales_costos.bind('<Double-1>', self.editar_material_costos)
        
        scroll_mat_c = ttk.Scrollbar(mat_table_frame, orient='vertical', command=self.tree_materiales_costos.yview)
        scroll_mat_c.pack(side='right', fill='y')
        self.tree_materiales_costos.configure(yscrollcommand=scroll_mat_c.set)
        
        btn_quitar_mat_c = ttk.Button(form_costos, text='➖ Quitar Material', command=self.quitar_material_costos, style='Danger.TButton')
        btn_quitar_mat_c.grid(row=8, column=0, columnspan=4, pady=(0, 10))
        
        # Peso total calculado
        self.lbl_peso_total_costos = tk.Label(form_costos, text='Peso total/unidad: 0.000 kg', font=('Arial', 9, 'bold'), fg=STYLE_CONFIG['primary_color'], bg=STYLE_CONFIG['bg_color'])
        self.lbl_peso_total_costos.grid(row=9, column=0, columnspan=2, sticky='w', pady=(5, 10))
        
        # Botones de acción del formulario
        btn_form_frame = tk.Frame(form_costos, bg=STYLE_CONFIG['bg_color'])
        btn_form_frame.grid(row=10, column=0, columnspan=4, pady=10)
        
        ttk.Button(btn_form_frame, text='🔄 Actualizar Planilla', command=self.agregar_item_costos, style='Action.TButton').pack(side='left', padx=5)
        ttk.Button(btn_form_frame, text='📥 Importar Datos', command=self.importar_datos_costos, style='Custom.TButton').pack(side='left', padx=5)
        ttk.Button(btn_form_frame, text='🔄 Limpiar Formulario', command=self.limpiar_formulario_costos, style='Warning.TButton').pack(side='left', padx=5)
        
        # Gráfico de torta (en el medio)
        grafico_container = tk.Frame(frame_costos, bg=STYLE_CONFIG['card_bg'], relief='solid', borderwidth=2)
        grafico_container.grid(row=0, column=1, sticky='nsew', padx=5, pady=15)
        
        grafico_title = tk.Label(grafico_container,
                                text='📊 Distribución de Costos',
                                font=('Segoe UI', 12, 'bold'),
                                fg=STYLE_CONFIG['primary_color'],
                                bg=STYLE_CONFIG['card_bg'])
        grafico_title.pack(pady=(10, 5))
        
        if Figure:
            try:
                grafico_frame = tk.Frame(grafico_container, bg='white', relief='flat', width=250, height=250)
                grafico_frame.pack(pady=(5, 8), padx=8, fill='both', expand=True)
                grafico_frame.pack_propagate(False)  # Mantener el tamaño mínimo
                
                self.fig_costos = Figure(figsize=(2.8, 2.8), dpi=80, facecolor='white')
                self.ax_costos = self.fig_costos.add_subplot(111)
                
                # Configurar el gráfico inicial
                self.ax_costos.text(0.5, 0.5, 'Agregue items\npara ver el gráfico', 
                                  ha='center', va='center', fontsize=12, 
                                  color=STYLE_CONFIG['text_light'])
                self.ax_costos.set_xlim(0, 1)
                self.ax_costos.set_ylim(0, 1)
                self.ax_costos.axis('off')
                
                self.canvas_costos = FigureCanvasTkAgg(self.fig_costos, master=grafico_frame)
                self.canvas_costos.draw()
                widget = self.canvas_costos.get_tk_widget()
                widget.pack(fill='both', expand=True)
            except Exception as e:
                tk.Label(grafico_container, text=f'Error al cargar gráfico\n{str(e)}', 
                        fg='red', bg=STYLE_CONFIG['card_bg']).pack(pady=20)
        else:
            tk.Label(grafico_container, text='Matplotlib no está instalado\nPara ver gráficos, instale: pip install matplotlib', 
                    fg='red', bg=STYLE_CONFIG['card_bg'], justify='center').pack(pady=20)
        
        # Panel de resumen de totales (arriba a la derecha)
        totales_frame = tk.Frame(frame_costos, bg=STYLE_CONFIG['card_bg'], relief='flat', borderwidth=1)
        totales_frame.grid(row=0, column=2, sticky='new', padx=(5, 15), pady=15)
        
        totales_title = tk.Label(totales_frame,
                                text='📈 Totales de la Planilla',
                                font=('Segoe UI', 14, 'bold'),
                                fg=STYLE_CONFIG['primary_color'],
                                bg=STYLE_CONFIG['card_bg'])
        totales_title.pack(pady=(15, 10))
        
        # Grid de totales
        totales_grid = tk.Frame(totales_frame, bg=STYLE_CONFIG['card_bg'])
        totales_grid.pack(padx=15, pady=(0, 10))
        
        # Totales - una columna vertical
        self.lbl_total_mp_costos = tk.Label(totales_grid, text='Total Materia Prima: $0.00',
                                            font=('Segoe UI', 11, 'bold'),
                                            fg=STYLE_CONFIG['danger_color'],
                                            bg=STYLE_CONFIG['card_bg'])
        self.lbl_total_mp_costos.pack(anchor='w', pady=5)
        
        self.lbl_total_mo_costos = tk.Label(totales_grid, text='Total Mano de Obra: $0.00',
                                            font=('Segoe UI', 11, 'bold'),
                                            fg=STYLE_CONFIG['warning_color'],
                                            bg=STYLE_CONFIG['card_bg'])
        self.lbl_total_mo_costos.pack(anchor='w', pady=5)
        
        self.lbl_total_iibb_costos = tk.Label(totales_grid, text='Total IIBB: $0.00',
                                            font=('Segoe UI', 11, 'bold'),
                                            fg=STYLE_CONFIG['secondary_color'],
                                            bg=STYLE_CONFIG['card_bg'])
        self.lbl_total_iibb_costos.pack(anchor='w', pady=5)
        
        self.lbl_total_prod_costos = tk.Label(totales_grid, text='Total Producción: $0.00',
                                              font=('Segoe UI', 11, 'bold'),
                                              fg=STYLE_CONFIG['text_color'],
                                              bg=STYLE_CONFIG['card_bg'])
        self.lbl_total_prod_costos.pack(anchor='w', pady=5)
        
        # Separador
        tk.Frame(totales_grid, height=2, bg=STYLE_CONFIG['border_color']).pack(fill='x', pady=10)
        
        self.lbl_total_ingreso_costos = tk.Label(totales_grid, text='Total Ingresos: $0.00',
                                                 font=('Segoe UI', 11, 'bold'),
                                                 fg=STYLE_CONFIG['secondary_color'],
                                                 bg=STYLE_CONFIG['card_bg'])
        self.lbl_total_ingreso_costos.pack(anchor='w', pady=5)
        
        self.lbl_total_rentabilidad_costos = tk.Label(totales_grid, text='Rentabilidad Total: $0.00',
                                                      font=('Segoe UI', 12, 'bold'),
                                                      fg=STYLE_CONFIG['success_color'],
                                                      bg=STYLE_CONFIG['card_bg'])
        self.lbl_total_rentabilidad_costos.pack(anchor='w', pady=5)
        
        self.lbl_margen_promedio_costos = tk.Label(totales_grid, text='Margen Promedio: 0.00%',
                                                   font=('Segoe UI', 11, 'bold'),
                                                   fg=STYLE_CONFIG['primary_color'],
                                                   bg=STYLE_CONFIG['card_bg'])
        self.lbl_margen_promedio_costos.pack(anchor='w', pady=5)
        
        # Tabla de items de la planilla (abajo, ocupando las 3 columnas)
        items_label = tk.Label(frame_costos, text='📋 Items en Planilla de Costos', font=('Segoe UI', 12, 'bold'), fg=STYLE_CONFIG['primary_color'], bg=STYLE_CONFIG['bg_color'])
        items_label.grid(row=1, column=0, columnspan=3, sticky='w', padx=15, pady=(10, 5))
        
        costos_table_frame = tk.Frame(frame_costos, bg=STYLE_CONFIG['bg_color'])
        costos_table_frame.grid(row=2, column=0, columnspan=3, padx=15, pady=(0, 10), sticky='ew')
        
        xscroll_costos = tk.Scrollbar(costos_table_frame, orient='horizontal')
        yscroll_costos = tk.Scrollbar(costos_table_frame, orient='vertical')
        
        self.tree_costos = ttk.Treeview(costos_table_frame,
            columns=('Familia', 'Medida', 'Característica', 'Peso/u (kg)', 'Costo MP/u', 'Costo MO/u', 'IIBB %', 'IIBB/u', 'Costo Tot/u', 
                    'Precio/u', 'Margen Bruto', '% Margen Bruto', 'Descuento', 'Margen Neto', '% Margen Neto'),
            show='headings', style='Custom.Treeview', height=10, 
            xscrollcommand=xscroll_costos.set, yscrollcommand=yscroll_costos.set,
            selectmode='extended')
        
        # Configurar encabezados
        for col in self.tree_costos['columns']:
            self.tree_costos.heading(col, text=col)
        
        # Anchos de columnas
        costos_col_widths = {
            'Familia': 120,
            'Medida': 80,
            'Característica': 150,
            'Peso/u (kg)': 90,
            'Costo MP/u': 100,
            'Costo MO/u': 100,
            'IIBB %': 80,
            'IIBB/u': 90,
            'Costo Tot/u': 110,
            'Precio/u': 100,
            'Margen Bruto': 100,
            '% Margen Bruto': 110,
            'Descuento': 90,
            'Margen Neto': 100,
            '% Margen Neto': 110
        }
        
        for col in self.tree_costos['columns']:
            width = costos_col_widths.get(col, 100)
            self.tree_costos.column(col, width=width, anchor='center', minwidth=80)
        
        self.tree_costos.grid(row=0, column=0, sticky='ew')
        yscroll_costos.config(command=self.tree_costos.yview)
        xscroll_costos.config(command=self.tree_costos.xview)
        yscroll_costos.grid(row=0, column=1, sticky='ns')
        xscroll_costos.grid(row=1, column=0, sticky='ew')
        
        # Bind doble clic para editar item o IIBB según la columna
        self.tree_costos.bind('<Double-1>', self.manejar_doble_clic_costos)
        
        costos_table_frame.grid_rowconfigure(0, weight=0)
        costos_table_frame.grid_columnconfigure(0, weight=1)
        
        # Botones de acción (abajo, ocupando las 3 columnas)
        btn_actions_costos = tk.Frame(frame_costos, bg=STYLE_CONFIG['bg_color'])
        btn_actions_costos.grid(row=3, column=0, columnspan=3, pady=5)
        
        ttk.Button(btn_actions_costos, text='✏️ Editar Item Seleccionado', command=self.editar_item_costos, style='Custom.TButton').pack(side='left', padx=5)
        ttk.Button(btn_actions_costos, text='🗑️ Eliminar Items Seleccionados', command=self.eliminar_item_costos, style='Danger.TButton').pack(side='left', padx=5)
        ttk.Button(btn_actions_costos, text='➕ Enviar a Producción', command=self.enviar_a_produccion, style='Action.TButton').pack(side='left', padx=5)
        
        
        # ========== INICIALIZAR NAVEGACIÓN SIDEBAR ==========
        self.inicializar_navegacion()

    def agregar_producto(self):
        familia = self.familia_var.get().strip()
        medida = self.medida_var.get().strip()
        caracteristica = self.caracteristica_var.get().strip()
        peso = self.peso_var.get()
        costo = ''
        dolar = ''
        venta = self.venta_var.get()
        iibb_porcentaje = self.iibb_porcentaje_var.get()
        cant_fab = self.cant_fab_var.get()
        cant_hora = self.cant_hora_var.get()
        
        # Si hay materiales en la tabla y peso vacío, calcular suma
        if (not peso or str(peso).strip() == '') and self.materiales_producto_form:
            try:
                suma_kg = sum(float(m.get('kg_por_unidad', 0)) for m in self.materiales_producto_form)
                peso = str(suma_kg)
                self.peso_var.set(peso)
            except Exception:
                pass
        
        # Si no hay peso y no hay materiales, usar peso por defecto de 1
        if not peso or str(peso).strip() == '':
            peso = '1.0'
            self.peso_var.set(peso)
        
        # Obtener media valor hora efectivo empleados
        valor_hora_ajustado_promedio = 0
        if self.empleados:
            valor_hora_ajustado_promedio = sum(e.valor_hora_efectivo for e in self.empleados) / len(self.empleados)
        
        # Validar que se completen todos los campos requeridos (sin incluir peso ya que se maneja automáticamente)
        if not (familia and medida and caracteristica and cant_fab and cant_hora):
            messagebox.showerror('Error', 'Por favor complete Familia, Medida, Característica, cantidad a fabricar y cantidad por hora.')
            return
        
        # Crear el nombre combinado para el producto
        nombre = f"{familia} - {medida} - {caracteristica}"
        try:
            prod = Producto(nombre, peso, venta, cant_fab, cant_hora, valor_hora_ajustado_promedio, materiales=list(self.materiales_producto_form), precios_materiales=self._mapa_precios_mp(), iibb_porcentaje=iibb_porcentaje)
        except Exception as e:
            messagebox.showerror('Error', f'Error en los datos: {e}')
            return
        
        # Detectar si estamos en modo edición
        if self.editando_producto_index is not None:
            # Sobrescribir el producto en la posición guardada
            self.productos[self.editando_producto_index] = prod
            self.editando_producto_index = None  # Limpiar el índice de edición
        else:
            # Agregar nuevo producto
            self.productos.append(prod)
        
        self.actualizar_tabla_productos()
        self.actualizar_resumen_productos()
        self.familia_var.set('')
        self.medida_var.set('')
        self.caracteristica_var.set('')
        self.peso_var.set('')
        self.costo_var.set('')
        self.dolar_var.set('')
        self.venta_var.set('')
        self.iibb_porcentaje_var.set('')
        self.cant_fab_var.set('')
        self.cant_hora_var.set('')
        # Limpiar materiales del formulario
        self.materiales_producto_form = []
        self.actualizar_tabla_materiales_form()
        self.sel_mp_var.set('')
        self.save_state()
        # Actualizar stock de productos por si las columnas dependen de productos
        self.actualizar_tabla_stock_prod()

    def agregar_empleado(self):
        vals = [v.get() for v in self.emp_vars]
        if not all(vals):
            messagebox.showerror('Error', 'Por favor complete todos los campos de empleado.')
            return
        try:
            emp = Empleado(*vals)
        except Exception as e:
            messagebox.showerror('Error', f'Error en los datos: {e}')
            return
        
        # Detectar si estamos en modo edición
        if self.editando_empleado_index is not None:
            # Sobrescribir el empleado en la posición guardada
            self.empleados[self.editando_empleado_index] = emp
            self.editando_empleado_index = None  # Limpiar el índice de edición
        else:
            # Agregar nuevo empleado
            self.empleados.append(emp)
        
        self.actualizar_tabla_empleados()
        for v in self.emp_vars:
            v.set('')
        self.actualizar_resumen_empleados()
        # Recalcular productos con el nuevo promedio de empleados
        self.recalcular_productos()
        # Actualizar tabla de costos con los nuevos valores de mano de obra
        self.actualizar_tabla_costos()
        self.save_state()
        self.actualizar_tabla_stock_prod()

    def actualizar_resumen_empleados(self):
        if not self.empleados:
            for lbl in self.resumen_labels.values():
                lbl.config(text='0')
            return
        media_indice = sum(e.indice_ajustado for e in self.empleados) / len(self.empleados)
        suma_costo_total = sum(e.costo_total_ajustado for e in self.empleados)
        suma_valor_hora_efectivo = sum(e.valor_hora_efectivo for e in self.empleados)
        suma_horas_trabajadas = sum(e.horas_trabajadas for e in self.empleados)
        self.resumen_labels['📈 Media índice ajustado (%)'].config(text=f'{media_indice:.2f}')
        self.resumen_labels['💰 Costo total ajustado total'].config(text=f'{suma_costo_total:.2f}')
        self.resumen_labels['⏰ Valor hora efectivo total'].config(text=f'{suma_valor_hora_efectivo:.2f}')
        self.resumen_labels['🔢 Horas trabajadas totales'].config(text=f'{suma_horas_trabajadas:.2f}')

    def recalcular_productos(self):
        """Recalcular todos los productos existentes cuando cambian los empleados"""
        if not self.empleados:
            return
        
        # Calcular nuevo promedio de valor hora efectivo
        nuevo_promedio = sum(e.valor_hora_efectivo for e in self.empleados) / len(self.empleados)
        
        # Actualizar cada producto
        for producto in self.productos:
            producto.actualizar_valor_hora_promedio(nuevo_promedio)
        
        # Actualizar cada item de costos
        for item in self.items_costos:
            item.actualizar_valor_hora_promedio(nuevo_promedio)
        
        # Actualizar la tabla de productos
        self.actualizar_tabla_productos()
        self.save_state()

    def _mapa_precios_mp(self):
        # Construir mapa {nombre_mp: {costo_kilo_usd, valor_dolar}}
        mapa = {}
        for s in self.stock_mp:
            nombre = s.get('nombre') if 'nombre' in s else s.get('Nombre')
            if not nombre:
                continue
            costo = s.get('costo_kilo_usd') or s.get('Costo kilo (USD)')
            dolar = s.get('valor_dolar') or s.get('Valor dólar')
            try:
                costo = float(costo) if costo is not None else 0
                dolar = float(dolar) if dolar is not None else 0
            except Exception:
                costo = 0
                dolar = 0
            mapa[nombre] = {'costo_kilo_usd': costo, 'valor_dolar': dolar}
        return mapa

    def actualizar_tabla_productos(self):
        """Actualizar la tabla de productos con los datos recalculados"""
        for item in self.tree.get_children():
            self.tree.delete(item)
        # Orden de columnas de la tabla (sin 'Moneda precio' ni 'Precio con moneda')
        columnas_tabla = [
            'Familia', 'Medida', 'Característica', 'Peso por unidad (kg)',
            'Cantidad a fabricar', 'Cantidad producida por hora', 'Productos por kilo',
            'Costo unitario MP', 'Costo total MP', 'Horas necesarias',
            'Valor hora ajustado promedio', 'Incidencia mano de obra',
            'Costo unitario mano de obra', 'Precio de venta',
            'Rentabilidad neta', 'Rentabilidad neta total'
        ]
        for prod in self.productos:
            prod_dict = prod.to_dict()
            # Construir valores en el orden correcto de las columnas
            values = [prod_dict[col] for col in columnas_tabla]
            self.tree.insert('', 'end', values=values)
        
        # Actualizar contador de productos
        if hasattr(self, 'prod_count_label'):
            count = len(self.productos)
            texto = f'({count} producto{"s" if count != 1 else ""})'
            self.prod_count_label.config(text=texto)

    def actualizar_tabla_empleados(self):
        """Actualizar la tabla de empleados con los datos recalculados"""
        for item in self.tree_emp.get_children():
            self.tree_emp.delete(item)
        for emp in self.empleados:
            values = list(emp.to_dict().values())
            values.append('Editar | Eliminar')  # Agregar columna de acciones
            self.tree_emp.insert('', 'end', values=values)
        # Actualizar vista de stock productos (por peso en kg)
        self.actualizar_tabla_stock_prod()

    def editar_producto(self):
        """Editar el producto seleccionado"""
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning('Advertencia', 'Por favor seleccione un producto para editar.')
            return
        
        item = selection[0]
        index = self.tree.index(item)
        producto = self.productos[index]
        
        # Llenar los campos con los datos del producto seleccionado
        # Extraer componentes del nombre (formato: "Familia - Medida - Característica")
        nombre_parts = producto.nombre.split(' - ')
        if len(nombre_parts) >= 3:
            self.familia_var.set(nombre_parts[0])
            self.medida_var.set(nombre_parts[1])
            self.caracteristica_var.set(' - '.join(nombre_parts[2:]))
        else:
            # Fallback para nombres antiguos
            self.familia_var.set(producto.nombre)
            self.medida_var.set('')
            self.caracteristica_var.set('')
        
        self.peso_var.set(str(producto.peso_unidad))
        self.venta_var.set(str(producto.precio_venta) if producto.precio_venta else '')
        self.iibb_porcentaje_var.set(str(producto.iibb_porcentaje) if producto.iibb_porcentaje else '')
        self.cant_fab_var.set(str(producto.cantidad_fabricar))
        self.cant_hora_var.set(str(producto.cantidad_por_hora))
        # Cargar materiales
        self.materiales_producto_form = list(producto.materiales or [])
        self.actualizar_tabla_materiales_form()
        
        # Guardar el índice del producto en edición (NO eliminarlo)
        self.editando_producto_index = index
        messagebox.showinfo('Modo Edición', f'Producto "{producto.nombre}" cargado para edición.\nModifique los datos y presione "Agregar Producto" para guardar los cambios.')

    def eliminar_producto(self):
        """Eliminar el producto seleccionado"""
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning('Advertencia', 'Por favor seleccione un producto para eliminar.')
            return
        
        if messagebox.askyesno('Confirmar', '¿Está seguro de que desea eliminar este producto?'):
            item = selection[0]
            index = self.tree.index(item)
            self.productos.pop(index)
            self.actualizar_tabla_productos()
            self.actualizar_resumen_productos()
            self.save_state()

    def editar_empleado(self):
        """Editar el empleado seleccionado"""
        selection = self.tree_emp.selection()
        if not selection:
            messagebox.showwarning('Advertencia', 'Por favor seleccione un empleado para editar.')
            return
        
        item = selection[0]
        index = self.tree_emp.index(item)
        empleado = self.empleados[index]
        
        # Llenar los campos con los datos del empleado seleccionado
        values = [empleado.nombre, empleado.valor_hora, empleado.dias_trabajados, empleado.horas_dia, empleado.ausencias, empleado.vacaciones, empleado.feriados, empleado.lic_enfermedad, empleado.otras_licencias, empleado.horas_descanso, empleado.carga_social, empleado.horas_extras, empleado.feriados_trabajados]
        
        for i, var in enumerate(self.emp_vars):
            var.set(str(values[i]))
        
        # Guardar el índice del empleado en edición (NO eliminarlo)
        self.editando_empleado_index = index
        messagebox.showinfo('Modo Edición', f'Empleado "{empleado.nombre}" cargado para edición.\nModifique los datos y presione "Agregar Empleado" para guardar los cambios.')

    def eliminar_empleado(self):
        """Eliminar el empleado seleccionado"""
        selection = self.tree_emp.selection()
        if not selection:
            messagebox.showwarning('Advertencia', 'Por favor seleccione un empleado para eliminar.')
            return
        
        if messagebox.askyesno('Confirmar', '¿Está seguro de que desea eliminar este empleado?'):
            item = selection[0]
            index = self.tree_emp.index(item)
            self.empleados.pop(index)
            self.actualizar_tabla_empleados()
            self.actualizar_resumen_empleados()
            self.recalcular_productos()
            # Actualizar tabla de costos con los nuevos valores de mano de obra
            self.actualizar_tabla_costos()
            self.save_state()

    def actualizar_resumen_productos(self):
        """Actualizar el resumen de productos"""
        if not self.productos:
            self.resumen_prod_labels['💰 Rentabilidad neta total'].config(text='$ 0.00')
            self.resumen_prod_labels['📈 Rentabilidad promedio'].config(text='$ 0.00')
            self.resumen_prod_labels['🔢 Cantidad total'].config(text='0 un.')
            self.resumen_prod_labels['⚖️ Material total (kg)'].config(text='0.00 kg')
            return
        
        # Calcular métricas
        rentabilidad_neta_total = sum(p.rentabilidad_neta_total for p in self.productos if p.rentabilidad_neta_total is not None)
        rentabilidades_validas = [p.rentabilidad_neta for p in self.productos if p.rentabilidad_neta is not None]
        rentabilidad_promedio = sum(rentabilidades_validas) / len(rentabilidades_validas) if rentabilidades_validas else 0
        cantidad_total = sum(p.cantidad_fabricar for p in self.productos)
        material_total_kg = sum(p.peso_unidad * p.cantidad_fabricar for p in self.productos)
        
        # Actualizar labels con formato mejorado
        self.resumen_prod_labels['💰 Rentabilidad neta total'].config(text=f'$ {rentabilidad_neta_total:,.2f}')
        self.resumen_prod_labels['📈 Rentabilidad promedio'].config(text=f'$ {rentabilidad_promedio:,.2f}')
        self.resumen_prod_labels['🔢 Cantidad total'].config(text=f'{int(cantidad_total):,} un.')
        self.resumen_prod_labels['⚖️ Material total (kg)'].config(text=f'{material_total_kg:,.2f} kg')

    def exportar_excel(self):
        if not self.productos and not self.empleados and not self.stock_mp and not self.stock_prod and not self.productos_reventa and not self.metricas and not self.ventas and not self.items_costos:
            messagebox.showerror('Error', 'No hay datos para exportar.')
            return
        
        try:
            with pd.ExcelWriter('calculadora_rentabilidad.xlsx', engine='openpyxl') as writer:
                # Exportar productos si existen
                if self.productos:
                    df_productos = pd.DataFrame([p.to_dict() for p in self.productos])
                    df_productos.to_excel(writer, index=False, sheet_name='Productos')
                
                # Exportar empleados si existen
                if self.empleados:
                    df_empleados = pd.DataFrame([e.to_dict() for e in self.empleados])
                    df_empleados.to_excel(writer, index=False, sheet_name='Empleados')
                # Exportar stock materia prima
                if self.stock_mp:
                    # Normalizar claves para exportar
                    df_mp = pd.DataFrame([
                        {'Nombre': s.get('nombre') or s.get('Nombre'), 'Stock (kg)': s.get('kg') or s.get('Stock (kg)')}
                        for s in self.stock_mp
                    ])
                    df_mp.to_excel(writer, index=False, sheet_name='Stock MP')
                # Exportar stock de productos finalizados
                if self.stock_prod:
                    df_sp = pd.DataFrame([
                        {
                            'Nombre': s['nombre'],
                            'Cantidad (u)': int(s.get('cantidad', 0)),
                            'Peso unidad (kg)': float(s.get('peso_unidad', 0)),
                            'Total (kg)': float(s.get('peso_unidad', 0)) * int(s.get('cantidad', 0))
                        }
                        for s in self.stock_prod
                    ])
                    df_sp.to_excel(writer, index=False, sheet_name='Stock Productos')
                # Exportar productos de reventa
                if self.productos_reventa:
                    df_pr = pd.DataFrame([
                        {
                            'Nombre': pr['nombre'],
                            'Cantidad (u)': int(pr.get('cantidad', 0)),
                            'Costo Unitario': float(pr.get('costo_unitario', 0)),
                            'Otros Costos': float(pr.get('otros_costos', 0)),
                            'Moneda': pr.get('moneda', 'ARS'),
                            'Valor Dólar': float(pr.get('valor_dolar', 0)) if pr.get('moneda') == 'USD' else '',
                            'Costo Total': float(pr.get('costo_total', 0))
                        }
                        for pr in self.productos_reventa
                    ])
                    df_pr.to_excel(writer, index=False, sheet_name='Productos Reventa')
                # Exportar métricas
                if self.metricas:
                    df_met = pd.DataFrame(self.metricas)
                    df_met.to_excel(writer, index=False, sheet_name='Métricas')
                # Exportar ventas
                if self.ventas:
                    df_sales = pd.DataFrame(self.ventas)
                    df_sales.to_excel(writer, index=False, sheet_name='Ventas')
                # Exportar costos
                if self.items_costos:
                    df_costos = pd.DataFrame([item.to_dict() for item in self.items_costos])
                    df_costos.to_excel(writer, index=False, sheet_name='Costos')
            
            messagebox.showinfo('Éxito', 'Archivo calculadora_rentabilidad.xlsx exportado correctamente con todas las hojas.')
        except Exception as e:
            messagebox.showerror('Error', f'No se pudo exportar: {e}')

    def exportar_empleados_excel(self):
        """Método mantenido para compatibilidad, pero ahora exporta ambos"""
        self.exportar_excel()

    # Persistencia
    def save_state(self):
        try:
            data = {
                'empleados': [e.to_save_dict() for e in self.empleados],
                'productos': [p.to_save_dict() for p in self.productos],
                'stock_mp': self.stock_mp,
                'stock_prod': self.stock_prod,
                'productos_reventa': self.productos_reventa,
                'metricas': self.metricas,
                'ventas': self.ventas,
                'items_costos': [item.to_save_dict() for item in self.items_costos],
                'log_ingresos_mp': getattr(self, 'log_ingresos_mp', []),
                'log_egresos_mp': getattr(self, 'log_egresos_mp', []),
                'log_ingresos_pr': getattr(self, 'log_ingresos_pr', []),
                'log_egresos_pr': getattr(self, 'log_egresos_pr', []),
                'log_ingresos_fab': getattr(self, 'log_ingresos_fab', []),
                'log_egresos_fab': getattr(self, 'log_egresos_fab', []),
                'logs_compra_mp': getattr(self, 'logs_compra_mp', []),
                'logs_compra_prod': getattr(self, 'logs_compra_prod', [])
            }
            with open(DATA_FILE, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    def load_state(self):
        if not os.path.exists(DATA_FILE):
            return
        try:
            with open(DATA_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
            # Cargar empleados primero
            self.empleados = [Empleado.from_save_dict(d) for d in data.get('empleados', [])]
            self.actualizar_tabla_empleados()
            self.actualizar_resumen_empleados()
            # Promedio valor hora efectivo actual
            valor_hora_ajustado_promedio = 0
            if self.empleados:
                valor_hora_ajustado_promedio = sum(e.valor_hora_efectivo for e in self.empleados) / len(self.empleados)
            # Mapa de precios de MP al iniciar
            self.stock_mp = data.get('stock_mp', [])
            precios = self._mapa_precios_mp()
            self.productos = [Producto.from_save_dict(d, valor_hora_ajustado_promedio, precios_materiales=precios) for d in data.get('productos', [])]
            self.actualizar_tabla_productos()
            self.actualizar_resumen_productos()
            # Stock
            self.stock_prod = data.get('stock_prod', [])
            self.productos_reventa = data.get('productos_reventa', [])
            # Métricas y Ventas
            self.metricas = data.get('metricas', [])
            self.ventas = data.get('ventas', [])
            # Cargar logs de movimientos de stock
            self.log_ingresos_mp = data.get('log_ingresos_mp', [])
            self.log_egresos_mp = data.get('log_egresos_mp', [])
            self.log_ingresos_pr = data.get('log_ingresos_pr', [])
            self.log_egresos_pr = data.get('log_egresos_pr', [])
            self.log_ingresos_fab = data.get('log_ingresos_fab', [])
            self.log_egresos_fab = data.get('log_egresos_fab', [])
            # Cargar logs de compras
            self.logs_compra_mp = data.get('logs_compra_mp', [])
            self.logs_compra_prod = data.get('logs_compra_prod', [])
            # Cargar items de la planilla de costos
            self.items_costos = [Producto.from_save_dict(d, valor_hora_ajustado_promedio, precios_materiales=precios) for d in data.get('items_costos', [])]
            
            # Migrar productos existentes que no tienen moneda_precio
            for item in self.items_costos:
                if not hasattr(item, 'moneda_precio'):
                    item.moneda_precio = 'ARS'  # Asignar ARS por defecto a productos existentes
            self.actualizar_tabla_materia_prima()
            self.actualizar_tabla_stock_prod()
            self.actualizar_tabla_stock_fabricados()
            self.actualizar_tabla_metricas()
            self.actualizar_tabla_ventas()
            self.actualizar_tabla_costos()
            self.refresh_mp_combobox()
            
            # Cargar logs de compras a las tablas
            self.cargar_logs_compras_a_tablas()
            
            # Actualizar métricas de stock y log de movimientos
            if hasattr(self, 'calcular_metricas_stock'):
                self.calcular_metricas_stock()
            if hasattr(self, 'actualizar_log_movimientos'):
                self.actualizar_log_movimientos()
            
            # Asegurar que ventas tenga los productos del stock disponibles
            try:
                self._refresh_combo_venta_productos()
            except Exception:
                pass
        except Exception:
            pass

    # -------- Stock: Materia prima --------
    def actualizar_tabla_materia_prima(self):
        if not hasattr(self, 'tree_mp'):
            return
        for item in self.tree_mp.get_children():
            self.tree_mp.delete(item)
        for s in self.stock_mp:
            # Extraer componentes del nombre (formato: "Familia - Medida - Característica")
            nombre_completo = s.get('nombre') if 'nombre' in s else s.get('Nombre')
            nombre_parts = nombre_completo.split(' - ')
            if len(nombre_parts) >= 3:
                familia = nombre_parts[0]
                medida = nombre_parts[1]
                caracteristica = ' - '.join(nombre_parts[2:])
            else:
                # Fallback para nombres antiguos
                familia = nombre_completo
                medida = ''
                caracteristica = ''
            
            kg = s.get('kg') if 'kg' in s else s.get('Stock (kg)', 0)
            costo = s.get('costo_kilo_usd') if 'costo_kilo_usd' in s else s.get('Costo kilo (USD)')
            dolar = s.get('valor_dolar') if 'valor_dolar' in s else s.get('Valor dólar')
            stock_minimo = s.get('stock_minimo', 0)
            
            # Calcular costo total: kg × costo_kilo_usd × valor_dolar
            costo_total = ''
            if kg and costo is not None and dolar is not None:
                try:
                    total = float(kg) * float(costo) * float(dolar)
                    costo_total = f'{total:.2f}'
                except Exception:
                    costo_total = '-'
            else:
                costo_total = '-'
            
            # Obtener el último movimiento para este material
            ultimo_movimiento = self.obtener_ultimo_movimiento_material(familia, medida, caracteristica)
            
            # Verificar si el stock está bajo el mínimo
            stock_actual = float(kg) if kg else 0
            stock_min = float(stock_minimo) if stock_minimo else 0
            
            # Formatear stock mínimo
            if stock_min > 0:
                stock_min_display = f"{stock_min:.1f}"
            else:
                stock_min_display = "Sin definir"
            
            # Insertar en la tabla
            item_id = self.tree_mp.insert('', 'end', values=(
                familia,
                medida,
                caracteristica,
                kg, 
                costo if costo is not None else '', 
                dolar if dolar is not None else '',
                costo_total,
                stock_min_display,
                ultimo_movimiento
            ))
            
            # Marcar con color de alerta si el stock está bajo el mínimo
            if stock_min > 0 and stock_actual < stock_min:
                self.tree_mp.set(item_id, 'Stock (kg)', f"⚠️ {kg}")  # Agregar emoji de alerta
                # Aplicar color de fondo rojo a toda la fila
                self.tree_mp.item(item_id, tags=('stock_bajo_minimo',))

    # Funciones para cambio dinámico de formulario en Stock
    def cambiar_modo_stock(self):
        """Cambia entre formulario de Materia Prima y Producto de Reventa"""
        modo = self.modo_stock_var.get()
        
        # Limpiar formulario actual
        for widget in self.form_stock_container.winfo_children():
            widget.destroy()
        
        # Crear nuevo formulario según el modo
        if modo == 'materia_prima':
            self._crear_formulario_materia_prima()
        else:
            self._crear_formulario_producto_reventa()
    
    def _crear_formulario_materia_prima(self):
        """Crea el formulario para Materia Prima"""
        input_frame_mp = ttk.LabelFrame(self.form_stock_container, text='📥 Materia Prima (kg)', padding=15)
        input_frame_mp.pack(fill='both', expand=True)
        
        mp_labels = ['🏷️ Nombre', '⚖️ Stock (kg)', '💰 Costo kilo (USD)', '💱 Valor dólar']
        
        for i, txt in enumerate(mp_labels):
            label = tk.Label(input_frame_mp,
                text=txt,
                font=('Arial', 9, 'bold'),
                fg=STYLE_CONFIG['text_color'],
                bg=STYLE_CONFIG['bg_color'])
            label.grid(row=i, column=0, sticky='w', pady=3)
        
        entry_mp_nombre = tk.Entry(input_frame_mp, textvariable=self.mp_nombre_var, width=25, font=('Arial', 9), relief='solid', borderwidth=1)
        entry_mp_nombre.grid(row=0, column=1, pady=3, padx=(10, 0))
        entry_mp_kg = tk.Entry(input_frame_mp, textvariable=self.mp_kg_var, width=25, font=('Arial', 9), relief='solid', borderwidth=1, validate='key', validatecommand=self.vcmd_decimal)
        entry_mp_kg.grid(row=1, column=1, pady=3, padx=(10, 0))
        entry_mp_costo = tk.Entry(input_frame_mp, textvariable=self.mp_costo_var, width=25, font=('Arial', 9), relief='solid', borderwidth=1, validate='key', validatecommand=self.vcmd_decimal)
        entry_mp_costo.grid(row=2, column=1, pady=3, padx=(10, 0))
        entry_mp_dolar = tk.Entry(input_frame_mp, textvariable=self.mp_dolar_var, width=25, font=('Arial', 9), relief='solid', borderwidth=1, validate='key', validatecommand=self.vcmd_decimal)
        entry_mp_dolar.grid(row=3, column=1, pady=3, padx=(10, 0))
        
        btn_agregar_mp = ttk.Button(input_frame_mp, text='➕ Agregar MP', command=self.agregar_materia_prima, style='Custom.TButton')
        btn_agregar_mp.grid(row=4, column=0, columnspan=2, pady=15)
    
    def _crear_formulario_producto_reventa(self):
        """Crea el formulario para Producto de Reventa"""
        input_frame_pr = ttk.LabelFrame(self.form_stock_container, text='📦 Producto de Reventa', padding=15)
        input_frame_pr.pack(fill='both', expand=True)
        
        # Nombre
        tk.Label(input_frame_pr, text='🏷️ Nombre',
                font=('Arial', 9, 'bold'),
                fg=STYLE_CONFIG['text_color'],
                bg=STYLE_CONFIG['bg_color']).grid(row=0, column=0, sticky='w', pady=3)
        entry_pr_nombre = tk.Entry(input_frame_pr, textvariable=self.pr_nombre_var, width=25, font=('Arial', 9), relief='solid', borderwidth=1)
        entry_pr_nombre.grid(row=0, column=1, pady=3, padx=(10, 0))
        
        # Cantidad
        tk.Label(input_frame_pr, text='📦 Cantidad (unidades)',
                font=('Arial', 9, 'bold'),
                fg=STYLE_CONFIG['text_color'],
                bg=STYLE_CONFIG['bg_color']).grid(row=1, column=0, sticky='w', pady=3)
        entry_pr_cantidad = tk.Entry(input_frame_pr, textvariable=self.pr_cantidad_var, width=25, font=('Arial', 9), relief='solid', borderwidth=1, validate='key', validatecommand=self.vcmd_decimal)
        entry_pr_cantidad.grid(row=1, column=1, pady=3, padx=(10, 0))
        
        # Moneda (Selector)
        tk.Label(input_frame_pr, text='💱 Moneda',
                font=('Arial', 9, 'bold'),
                fg=STYLE_CONFIG['text_color'],
                bg=STYLE_CONFIG['bg_color']).grid(row=2, column=0, sticky='w', pady=3)
        
        moneda_frame = tk.Frame(input_frame_pr, bg=STYLE_CONFIG['bg_color'])
        moneda_frame.grid(row=2, column=1, sticky='w', pady=3, padx=(10, 0))
        
        radio_ars = tk.Radiobutton(moneda_frame,
                                   text='ARS',
                                   variable=self.pr_moneda_var,
                                   value='ARS',
                                   font=('Arial', 9),
                                   bg=STYLE_CONFIG['bg_color'],
                                   command=self._actualizar_campos_moneda_pr)
        radio_ars.pack(side='left', padx=(0, 10))
        
        radio_usd = tk.Radiobutton(moneda_frame,
                                   text='USD',
                                   variable=self.pr_moneda_var,
                                   value='USD',
                                   font=('Arial', 9),
                                   bg=STYLE_CONFIG['bg_color'],
                                   command=self._actualizar_campos_moneda_pr)
        radio_usd.pack(side='left')
        
        # Costo unitario
        tk.Label(input_frame_pr, text='💰 Costo unitario',
                font=('Arial', 9, 'bold'),
                fg=STYLE_CONFIG['text_color'],
                bg=STYLE_CONFIG['bg_color']).grid(row=3, column=0, sticky='w', pady=3)
        entry_pr_costo = tk.Entry(input_frame_pr, textvariable=self.pr_costo_unitario_var, width=25, font=('Arial', 9), relief='solid', borderwidth=1, validate='key', validatecommand=self.vcmd_decimal)
        entry_pr_costo.grid(row=3, column=1, pady=3, padx=(10, 0))
        
        # Otros costos
        tk.Label(input_frame_pr, text='📝 Otros Costos',
                font=('Arial', 9, 'bold'),
                fg=STYLE_CONFIG['text_color'],
                bg=STYLE_CONFIG['bg_color']).grid(row=4, column=0, sticky='w', pady=3)
        entry_pr_otros_costos = tk.Entry(input_frame_pr, textvariable=self.pr_otros_costos_var, width=25, font=('Arial', 9), relief='solid', borderwidth=1, validate='key', validatecommand=self.vcmd_decimal)
        entry_pr_otros_costos.grid(row=4, column=1, pady=3, padx=(10, 0))
        
        # Valor del dólar (condicional)
        self.pr_lbl_dolar = tk.Label(input_frame_pr, text='💵 Valor del dólar',
                                     font=('Arial', 9, 'bold'),
                                     fg=STYLE_CONFIG['text_color'],
                                     bg=STYLE_CONFIG['bg_color'])
        self.pr_entry_dolar = tk.Entry(input_frame_pr, textvariable=self.pr_valor_dolar_var, width=25, font=('Arial', 9), relief='solid', borderwidth=1, validate='key', validatecommand=self.vcmd_decimal)
        
        # Botón agregar
        btn_agregar_pr = ttk.Button(input_frame_pr, text='➕ Agregar Producto', command=self.agregar_producto_reventa, style='Custom.TButton')
        btn_agregar_pr.grid(row=6, column=0, columnspan=2, pady=15)
        
        # Actualizar visibilidad inicial
        self._actualizar_campos_moneda_pr()
    
    def _actualizar_campos_moneda_pr(self):
        """Muestra u oculta el campo de valor del dólar según la moneda seleccionada"""
        if self.pr_moneda_var.get() == 'USD':
            if hasattr(self, 'pr_lbl_dolar'):
                self.pr_lbl_dolar.grid(row=7, column=0, sticky='w', pady=3)
                self.pr_entry_dolar.grid(row=7, column=1, pady=3, padx=(10, 0))
        else:
            self.pr_lbl_dolar.grid_forget()
            self.pr_entry_dolar.grid_forget()
            self.pr_valor_dolar_var.set('')  # Limpiar el valor

    def _actualizar_campos_moneda_compra_prod(self):
        """Muestra u oculta el campo de valor del dólar según la moneda seleccionada en compras de productos"""
        if self.compra_prod_moneda_var.get() == 'USD':
            if hasattr(self, 'compra_prod_lbl_dolar'):
                self.compra_prod_lbl_dolar.grid(row=9, column=0, sticky='w', padx=15, pady=8)
                self.compra_prod_entry_dolar.grid(row=9, column=1, sticky='ew', padx=(15, 0), pady=8, ipady=6)
        else:
            if hasattr(self, 'compra_prod_lbl_dolar'):
                self.compra_prod_lbl_dolar.grid_forget()
                self.compra_prod_entry_dolar.grid_forget()
                self.compra_prod_valor_dolar_var.set('')  # Limpiar el valor

    def _actualizar_campos_moneda_compra_mp(self):
        """Muestra u oculta el campo de valor del dólar según la moneda seleccionada en compras de materia prima"""
        if self.compra_mp_moneda_var.get() == 'USD':
            if hasattr(self, 'compra_mp_lbl_dolar'):
                self.compra_mp_lbl_dolar.grid(row=9, column=0, sticky='w', padx=15, pady=8)
                self.compra_mp_entry_dolar.grid(row=9, column=1, sticky='ew', padx=(15, 0), pady=8, ipady=6)
        else:
            if hasattr(self, 'compra_mp_lbl_dolar'):
                self.compra_mp_lbl_dolar.grid_forget()
                self.compra_mp_entry_dolar.grid_forget()
                self.compra_mp_valor_dolar_var.set('')  # Limpiar el valor

    def agregar_materia_prima(self):
        familia = self.mp_familia_var.get().strip()
        medida = self.mp_medida_var.get().strip()
        caracteristica = self.mp_caracteristica_var.get().strip()
        kg_str = self.mp_kg_var.get().strip()
        costo_str = (self.mp_costo_var.get() or '').strip()
        dolar_str = (self.mp_dolar_var.get() or '').strip()
        stock_min_str = (self.mp_stock_min_var.get() or '').strip()
        
        # Validar que se completen los campos mínimos requeridos
        if not (familia and kg_str):
            messagebox.showerror('Error', 'Complete Familia y stock (kg).')
            return
        
        # Usar valores por defecto para campos opcionales
        if not medida:
            medida = 'Sin Medida'
        if not caracteristica:
            caracteristica = 'Sin Característica'
        
        # Crear el nombre combinado para la materia prima
        nombre = f"{familia} - {medida} - {caracteristica}"
        try:
            kg = float(kg_str)
        except Exception:
            messagebox.showerror('Error', 'El stock (kg) debe ser numérico.')
            return
        # Costo/dólar/stock_min opcionales
        try:
            costo_val = float(costo_str) if costo_str else None
            dolar_val = float(dolar_str) if dolar_str else None
            stock_min_val = float(stock_min_str) if stock_min_str else 0
        except Exception:
            messagebox.showerror('Error', 'Costo, dólar y stock mínimo deben ser numéricos (si se informan).')
            return
        if kg <= 0:
            messagebox.showerror('Error', 'El stock (kg) debe ser mayor a 0.')
            return
        
        # Detectar si estamos en modo edición
        if self.editando_mp_index is not None:
            # Sobrescribir la materia prima en la posición guardada
            nuevo = {'nombre': nombre, 'kg': kg, 'stock_minimo': stock_min_val}
            if costo_val is not None:
                nuevo['costo_kilo_usd'] = costo_val
            if dolar_val is not None:
                nuevo['valor_dolar'] = dolar_val
            self.stock_mp[self.editando_mp_index] = nuevo
            self.editando_mp_index = None  # Limpiar el índice de edición
        else:
            # Unificar por nombre (case-insensitive)
            found = False
            for s in self.stock_mp:
                actual_nombre = (s.get('nombre') or s.get('Nombre') or '').strip()
                if actual_nombre.lower() == nombre.lower():
                    existente = s.get('kg') if 'kg' in s else s.get('Stock (kg)', 0)
                    s['kg'] = float(existente) + kg
                    s['nombre'] = actual_nombre
                    if costo_val is not None:
                        s['costo_kilo_usd'] = costo_val
                    if dolar_val is not None:
                        s['valor_dolar'] = dolar_val
                    s['stock_minimo'] = stock_min_val
                    s.pop('Stock (kg)', None)
                    s.pop('Nombre', None)
                    found = True
                    break
            if not found:
                nuevo = {'nombre': nombre, 'kg': kg, 'stock_minimo': stock_min_val}
                if costo_val is not None:
                    nuevo['costo_kilo_usd'] = costo_val
                if dolar_val is not None:
                    nuevo['valor_dolar'] = dolar_val
                self.stock_mp.append(nuevo)
        # Registrar métrica de ingreso de materia prima
        try:
            import datetime as _dt
            fecha = _dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        except Exception:
            fecha = ''
        # Determinar costo unitario en moneda local (USD * dólar)
        # Buscar el registro final del material para obtener costos actuales
        costo_usd_eff = costo_val if costo_val is not None else 0
        dolar_eff = dolar_val if dolar_val is not None else 0
        for s2 in self.stock_mp:
            actual_nombre2 = (s2.get('nombre') or s2.get('Nombre') or '').strip()
            if actual_nombre2.lower() == nombre.lower():
                if costo_usd_eff == 0:
                    costo_usd_eff = float(s2.get('costo_kilo_usd') or s2.get('Costo kilo (USD)') or 0)
                if dolar_eff == 0:
                    dolar_eff = float(s2.get('valor_dolar') or s2.get('Valor dólar') or 0)
                break
        costo_unit_local = round(costo_usd_eff * dolar_eff, 5)
        costo_total_local = round(costo_unit_local * kg, 5)
        met_mp = {
            'Fecha': fecha,
            'Tipo': 'Ingreso MP',
            'Producto': nombre,
            'Cantidad (u)': '',
            'Peso unidad (kg)': '',
            'Kg consumidos': kg,
            'Costo unitario MP': costo_unit_local,
            'Costo total MP': costo_total_local,
            'Precio de venta': '',
            'Rentabilidad neta': '',
            'Rentabilidad total': ''
        }
        try:
            self.metricas.append(met_mp)
            self.actualizar_tabla_metricas()
        except Exception:
            pass
        self.mp_familia_var.set('')
        self.mp_medida_var.set('')
        self.mp_caracteristica_var.set('')
        self.mp_kg_var.set('')
        self.mp_costo_var.set('')
        self.mp_dolar_var.set('')
        self.mp_stock_min_var.set('')
        self.actualizar_tabla_materia_prima()
        self.refresh_mp_combobox()
        
        # Registrar ingreso automático en el log de movimientos
        if hasattr(self, 'registrar_ingreso_automatico'):
            motivo = "Compra de materia prima" if not found else "Adición de stock materia prima"
            self.registrar_ingreso_automatico(familia, medida, caracteristica, kg, motivo)
        
        self.save_state()
        # Recalcular costos de productos con nuevos precios
        precios = self._mapa_precios_mp()
        for p in self.productos:
            p.actualizar_precios_materiales(precios)
        self.actualizar_tabla_productos()
        # Añadir métrica también en editar y eliminar según corresponda? Aquí no, porque es edición.

    def editar_materia_prima(self):
        if not hasattr(self, 'tree_mp'):
            return
        selection = self.tree_mp.selection()
        if not selection:
            messagebox.showwarning('Advertencia', 'Seleccione un ítem de materia prima para editar.')
            return
        item = selection[0]
        index = self.tree_mp.index(item)
        sel = self.stock_mp[index]
        
        # Cargar datos en el formulario
        # Extraer componentes del nombre (formato: "Familia - Medida - Característica")
        nombre_completo = sel.get('nombre') or sel.get('Nombre') or ''
        nombre_parts = nombre_completo.split(' - ')
        if len(nombre_parts) >= 3:
            self.mp_familia_var.set(nombre_parts[0])
            self.mp_medida_var.set(nombre_parts[1])
            self.mp_caracteristica_var.set(' - '.join(nombre_parts[2:]))
        else:
            # Fallback para nombres antiguos
            self.mp_familia_var.set(nombre_completo)
            self.mp_medida_var.set('')
            self.mp_caracteristica_var.set('')
        
        self.mp_kg_var.set(str(sel.get('kg') if 'kg' in sel else sel.get('Stock (kg)', '')))
        costo = sel.get('costo_kilo_usd') or sel.get('Costo kilo (USD)')
        dolar = sel.get('valor_dolar') or sel.get('Valor dólar')
        if costo:
            self.mp_costo_var.set(str(costo))
        if dolar:
            self.mp_dolar_var.set(str(dolar))
        
        # Guardar el índice de materia prima en edición (NO eliminarlo)
        self.editando_mp_index = index
        messagebox.showinfo('Modo Edición', f'Materia prima cargada para edición.\nModifique los datos y presione "Agregar MP" para guardar los cambios.')

    def eliminar_materia_prima(self):
        if not hasattr(self, 'tree_mp'):
            return
        selection = self.tree_mp.selection()
        if not selection:
            messagebox.showwarning('Advertencia', 'Seleccione un ítem de materia prima para eliminar.')
            return
        if messagebox.askyesno('Confirmar', '¿Eliminar materia prima seleccionada?'):
            item = selection[0]
            index = self.tree_mp.index(item)
            self.stock_mp.pop(index)
            self.actualizar_tabla_materia_prima()
            self.refresh_mp_combobox()
            self.save_state()
            precios = self._mapa_precios_mp()
            for p in self.productos:
                p.actualizar_precios_materiales(precios)
            self.actualizar_tabla_productos()
    
    def exportar_plantilla_materia_prima(self):
        """Exportar plantilla completa con Materia Prima y Productos de Reventa"""
        try:
            from tkinter import filedialog
            filepath = filedialog.asksaveasfilename(
                defaultextension='.xlsx',
                filetypes=[('Excel files', '*.xlsx')],
                initialfile='plantilla_materia_prima.xlsx'
            )
            if not filepath:
                return
            
            # Crear plantilla de Materia Prima con formato mejorado
            df_mp = pd.DataFrame({
                'Nombre': ['Cemento', 'Arena', 'Cal'],
                'Familia': ['Construcción', 'Construcción', 'Construcción'],
                'Medida': ['50kg', '1m3', '25kg'],
                'Característica': ['Portland', 'Fina', 'Hidratada'],
                'Stock (kg)': [1000.0, 5000.0, 500.0],
                'Costo/kg USD': [0.50, 0.30, 0.40],
                'Valor dólar': [1000, 1000, 1000],
                'Stock Mínimo': [100.0, 1000.0, 50.0]
            })
            
            # Crear plantilla de Productos de Reventa
            df_pr = pd.DataFrame({
                'Nombre': ['Producto Ejemplo 1', 'Producto Ejemplo 2'],
                'Familia': ['Categoría A', 'Categoría B'],
                'Medida': ['Unidad', 'Caja'],
                'Característica': ['Estándar', 'Premium'],
                'Stock': [100, 50],
                'Costo en USD': [10.50, 25.00],
                'Valor dólar': [1000, 1000],
                'Costo total': [10500, 25000]
            })
            
            # Escribir ambas hojas en un solo archivo
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df_mp.to_excel(writer, index=False, sheet_name='Materia Prima')
                df_pr.to_excel(writer, index=False, sheet_name='Productos Reventa')
                
                # Agregar hoja de instrucciones
                instrucciones = pd.DataFrame({
                    'INSTRUCCIONES': [
                        'FORMATO DE ARCHIVO PARA IMPORTAR MATERIA PRIMA',
                        '',
                        'COLUMNAS REQUERIDAS:',
                        '• Nombre: Nombre del material',
                        '• Stock (kg): Cantidad en kilogramos',
                        '',
                        'COLUMNAS OPCIONALES:',
                        '• Familia: Categoría del material',
                        '• Medida: Especificación de medida',
                        '• Característica: Descripción adicional',
                        '• Costo/kg USD: Precio por kilogramo en USD',
                        '• Valor dólar: Tipo de cambio',
                        '• Stock Mínimo: Cantidad mínima de stock',
                        '',
                        'FORMATO DE NOMBRE:',
                        'Si tienes Familia, Medida y Característica separadas,',
                        'el sistema creará automáticamente:',
                        '"Familia - Medida - Característica"',
                        '',
                        'EJEMPLO:',
                        'Cemento - 50kg - Portland',
                        '',
                        'IMPORTANTE:',
                        '• Usa números decimales con punto (.) no coma',
                        '• Las columnas pueden estar en cualquier orden',
                        '• Los nombres de columnas pueden variar ligeramente',
                        '• El sistema detectará automáticamente las columnas'
                    ]
                })
                instrucciones.to_excel(writer, index=False, sheet_name='Instrucciones')
            
            messagebox.showinfo(
                '📋 Plantilla Exportada', 
                f'✅ Plantilla creada en:\n{filepath}\n\n'
                f'📁 El archivo contiene:\n'
                f'• "Materia Prima" - Ejemplos de materias primas\n'
                f'• "Productos Reventa" - Ejemplos de productos\n'
                f'• "Instrucciones" - Guía de uso detallada\n\n'
                f'📝 INSTRUCCIONES:\n'
                f'1. Edita la hoja "Materia Prima" con tus datos\n'
                f'2. Usa el botón "📥 Importar Materia Prima"\n'
                f'3. Selecciona tu archivo Excel\n'
                f'4. El sistema detectará automáticamente las columnas'
            )
        except Exception as e:
            messagebox.showerror('Error', f'Error al exportar plantilla: {str(e)}')
    
    def importar_materia_prima_excel(self):
        """Importar materia prima desde archivo Excel con formato mejorado"""
        try:
            from tkinter import filedialog
            filepath = filedialog.askopenfilename(
                title='Seleccionar archivo Excel de Stock',
                filetypes=[('Excel files', '*.xlsx *.xls')]
            )
            if not filepath:
                return
            
            # Leer el archivo Excel - buscar hoja "Materia Prima"
            try:
                # Intentar leer de la hoja "Materia Prima" primero
                try:
                    df = pd.read_excel(filepath, sheet_name='Materia Prima')
                except:
                    # Si no existe, intentar con la primera hoja
                    df = pd.read_excel(filepath, sheet_name=0)
                    messagebox.showwarning('Aviso', 
                        'No se encontró la hoja "Materia Prima". Se está usando la primera hoja del archivo.\n\n'
                        'Para un archivo con múltiples hojas, nombre las hojas como:\n'
                        '- "Materia Prima" para materias primas\n'
                        '- "Productos Reventa" para productos de reventa')
            except Exception as e:
                messagebox.showerror('Error', f'No se pudo leer el archivo Excel: {str(e)}')
                return
            
            # Mostrar las columnas disponibles para debug
            print(f"Columnas disponibles en el Excel: {list(df.columns)}")
            
            # Mapeo de columnas posibles (flexible)
            columnas_posibles = {
                'nombre': ['Nombre', 'nombre', 'Material', 'material', 'Materia Prima', 'Familia'],
                'familia': ['Familia', 'familia', 'Categoría', 'categoria'],
                'medida': ['Medida', 'medida', 'Tamaño', 'tamaño'],
                'caracteristica': ['Característica', 'caracteristica', 'Descripción', 'descripcion', 'Especificación'],
                'stock': ['Stock (kg)', 'stock', 'Stock', 'Cantidad', 'cantidad', 'Kg', 'kg', 'Stock minimo'],
                'costo': ['Costo/kg USD', 'Costo USD', 'Precio USD', 'costo', 'precio', 'Importe'],
                'dolar': ['Valor dólar', 'Dólar', 'USD', 'dolar', 'usd', 'Moneda'],
                'stock_minimo': ['Stock Mínimo', 'stock_minimo', 'Mínimo', 'minimo', 'Stock Min']
            }
            
            # Encontrar las columnas correctas
            columnas_encontradas = {}
            for tipo, opciones in columnas_posibles.items():
                for opcion in opciones:
                    if opcion in df.columns:
                        columnas_encontradas[tipo] = opcion
                        break
            
            # Si no hay columna 'nombre', intentar crear el nombre usando otras columnas
            if 'nombre' not in columnas_encontradas:
                if 'familia' in columnas_encontradas:
                    columnas_encontradas['nombre'] = columnas_encontradas['familia']
                elif 'medida' in columnas_encontradas:
                    columnas_encontradas['nombre'] = columnas_encontradas['medida']
                elif 'caracteristica' in columnas_encontradas:
                    columnas_encontradas['nombre'] = columnas_encontradas['caracteristica']
                else:
                    # Si no hay ninguna columna identificadora, usar la primera columna
                    columnas_encontradas['nombre'] = df.columns[0]
            
            # Validar columnas mínimas requeridas
            if 'stock' not in columnas_encontradas:
                messagebox.showerror('Error', 
                    f'El archivo debe contener al menos una columna de cantidad:\n'
                    f'- Stock, Cantidad, Kg, etc.\n\n'
                    f'Columnas encontradas: {list(df.columns)}')
                return
            
            # Contador de items procesados
            agregados = 0
            actualizados = 0
            errores = []
            
            for index, row in df.iterrows():
                try:
                    # Obtener datos básicos
                    nombre_base = str(row[columnas_encontradas['nombre']]).strip()
                    if not nombre_base or nombre_base.lower() == 'nan':
                        continue
                    
                    stock = float(row[columnas_encontradas['stock']]) if pd.notna(row[columnas_encontradas['stock']]) else 0
                    
                    # Obtener datos opcionales
                    familia = str(row[columnas_encontradas.get('familia', 'nombre')]).strip() if 'familia' in columnas_encontradas else nombre_base
                    medida = str(row[columnas_encontradas.get('medida', 'medida')]).strip() if 'medida' in columnas_encontradas else 'N/A'
                    caracteristica = str(row[columnas_encontradas.get('caracteristica', 'caracteristica')]).strip() if 'caracteristica' in columnas_encontradas else 'N/A'
                    
                    costo_kg_usd = float(row[columnas_encontradas['costo']]) if 'costo' in columnas_encontradas and pd.notna(row[columnas_encontradas['costo']]) else 1.0
                    valor_dolar = float(row[columnas_encontradas['dolar']]) if 'dolar' in columnas_encontradas and pd.notna(row[columnas_encontradas['dolar']]) else 1000.0
                    stock_minimo = float(row[columnas_encontradas['stock_minimo']]) if 'stock_minimo' in columnas_encontradas and pd.notna(row[columnas_encontradas['stock_minimo']]) else 0
                    
                    # Crear nombre completo en formato "Familia - Medida - Característica"
                    if medida != 'N/A' and caracteristica != 'N/A':
                        nombre_completo = f"{familia} - {medida} - {caracteristica}"
                    else:
                        nombre_completo = nombre_base
                    
                    if stock <= 0:
                        continue
                    
                    # Verificar si la materia prima ya existe
                    existe = False
                    for i, mp in enumerate(self.stock_mp):
                        mp_nombre = mp.get('nombre') or mp.get('Nombre', '')
                        if mp_nombre.lower() == nombre_completo.lower():
                            # Actualizar materia prima existente
                            self.stock_mp[i] = {
                                'nombre': nombre_completo,
                                'kg': stock,
                                'costo_kilo_usd': costo_kg_usd,
                                'valor_dolar': valor_dolar,
                                'stock_minimo': stock_minimo
                            }
                            existe = True
                            actualizados += 1
                            break
                    
                    if not existe:
                        # Agregar nueva materia prima
                        nueva_mp = {
                            'nombre': nombre_completo,
                            'kg': stock,
                            'costo_kilo_usd': costo_kg_usd,
                            'valor_dolar': valor_dolar,
                            'stock_minimo': stock_minimo
                        }
                        self.stock_mp.append(nueva_mp)
                        agregados += 1
                        
                except Exception as e:
                    errores.append(f'Fila {index + 2}: {str(e)}')
            
            # Actualizar tabla y guardar
            self.actualizar_tabla_materia_prima()
            self.save_state()
            
            # Actualizar precios de materiales en productos
            precios = self._mapa_precios_mp()
            for p in self.productos:
                p.actualizar_precios_materiales(precios)
            self.actualizar_tabla_productos()
            
            # Mensaje de resultado
            mensaje = f'📊 Importación de Materia Prima Completada\n\n'
            mensaje += f'✅ Nuevas materias primas: {agregados}\n'
            mensaje += f'🔄 Materias primas actualizadas: {actualizados}\n'
            if errores:
                mensaje += f'\n⚠️ Errores encontrados ({len(errores)}):\n'
                mensaje += '\n'.join(errores[:3])  # Mostrar solo los primeros 3 errores
                if len(errores) > 3:
                    mensaje += f'\n... y {len(errores) - 3} errores más'
            else:
                mensaje += '\n🎉 ¡Importación exitosa sin errores!'
            
            messagebox.showinfo('Importación Completada', mensaje)
            
        except Exception as e:
            messagebox.showerror('Error', f'Error al importar archivo Excel: {str(e)}')

    # -------- Stock: Productos de Reventa --------
    def agregar_producto_reventa(self):
        familia = self.pr_familia_var.get().strip()
        medida = self.pr_medida_var.get().strip()
        caracteristica = self.pr_caracteristica_var.get().strip()
        cantidad_str = self.pr_cantidad_var.get().strip()
        costo_str = self.pr_costo_unitario_var.get().strip()
        otros_costos_str = self.pr_otros_costos_var.get().strip()
        moneda = self.pr_moneda_var.get()
        valor_dolar_str = self.pr_valor_dolar_var.get().strip()
        
        # Validar que se completen los campos mínimos requeridos
        if not (familia and cantidad_str and costo_str):
            messagebox.showerror('Error', 'Complete Familia, cantidad y costo unitario.')
            return
        
        # Usar valores por defecto para campos opcionales
        if not medida:
            medida = 'Sin Medida'
        if not caracteristica:
            caracteristica = 'Sin Característica'
        
        # Crear el nombre combinado para el producto
        nombre = f"{familia} - {medida} - {caracteristica}"
        
        try:
            cantidad = int(cantidad_str)
            costo_unitario = float(costo_str)
            otros_costos = float(otros_costos_str) if otros_costos_str else 0.0
        except Exception:
            messagebox.showerror('Error', 'Cantidad debe ser entero y costos deben ser numéricos.')
            return
        
        if cantidad <= 0 or costo_unitario <= 0:
            messagebox.showerror('Error', 'Cantidad y costo deben ser mayores a 0.')
            return
        
        if otros_costos < 0:
            messagebox.showerror('Error', 'Otros costos no pueden ser negativos.')
            return
        
        # Si es USD, validar valor del dólar
        valor_dolar = 1.0
        if moneda == 'USD':
            if not valor_dolar_str:
                messagebox.showerror('Error', 'Debe ingresar el valor del dólar para productos en USD.')
                return
            try:
                valor_dolar = float(valor_dolar_str)
                if valor_dolar <= 0:
                    messagebox.showerror('Error', 'El valor del dólar debe ser mayor a 0.')
                    return
            except Exception:
                messagebox.showerror('Error', 'El valor del dólar debe ser numérico.')
                return
        
        # Calcular costo unitario (mantener en moneda original, no convertir)
        costo_unitario_final = costo_unitario + otros_costos
        
        # Detectar si estamos en modo edición
        if self.editando_pr_index is not None:
            # Sobrescribir el producto de reventa en la posición guardada
            self.productos_reventa[self.editando_pr_index] = {
                'nombre': nombre,
                'cantidad': cantidad,
                'costo_unitario': costo_unitario,
                'otros_costos': otros_costos,
                'moneda': moneda,
                'valor_dolar': valor_dolar if moneda == 'USD' else None,
                'costo_unitario_final': costo_unitario_final
            }
            self.editando_pr_index = None  # Limpiar el índice de edición
        else:
            # Agregar o actualizar producto de reventa
            found = False
            for pr in self.productos_reventa:
                if pr['nombre'].lower() == nombre.lower():
                    pr['cantidad'] += cantidad
                    pr['costo_unitario'] = costo_unitario
                    pr['otros_costos'] = otros_costos
                    pr['moneda'] = moneda
                    pr['valor_dolar'] = valor_dolar if moneda == 'USD' else None
                    pr['costo_unitario_final'] = costo_unitario_final
                    found = True
                    break
            
            if not found:
                self.productos_reventa.append({
                    'nombre': nombre,
                    'cantidad': cantidad,
                    'costo_unitario': costo_unitario,
                    'otros_costos': otros_costos,
                    'moneda': moneda,
                    'valor_dolar': valor_dolar if moneda == 'USD' else None,
                    'costo_unitario_final': costo_unitario_final
                })
        
        # Limpiar formulario
        self.pr_familia_var.set('')
        self.pr_medida_var.set('')
        self.pr_caracteristica_var.set('')
        self.pr_cantidad_var.set('')
        self.pr_costo_unitario_var.set('')
        self.pr_otros_costos_var.set('')
        self.pr_valor_dolar_var.set('')
        self.pr_moneda_var.set('ARS')
        self._actualizar_campos_moneda_pr()
        
        self.actualizar_tabla_stock_prod()
        self.save_state()
        messagebox.showinfo('Éxito', f'Producto de reventa "{nombre}" agregado correctamente.')
    
    def editar_producto_reventa(self):
        if not hasattr(self, 'tree_stock_prod'):
            return
        selection = self.tree_stock_prod.selection()
        if not selection:
            messagebox.showwarning('Advertencia', 'Seleccione un producto de reventa para editar.')
            return
        
        item = selection[0]
        index = self.tree_stock_prod.index(item)
        
        if index >= len(self.productos_reventa):
            messagebox.showerror('Error', 'Índice inválido.')
            return
        
        pr = self.productos_reventa[index]
        
        # Cambiar a la pestaña de productos de reventa
        self.stock_tabs.select(1)  # 1 = índice de la pestaña "Productos de Reventa"
        
        # Cargar datos en el formulario
        # Extraer componentes del nombre (formato: "Familia - Medida - Característica")
        nombre_completo = pr['nombre']
        nombre_parts = nombre_completo.split(' - ')
        if len(nombre_parts) >= 3:
            self.pr_familia_var.set(nombre_parts[0])
            self.pr_medida_var.set(nombre_parts[1])
            self.pr_caracteristica_var.set(' - '.join(nombre_parts[2:]))
        else:
            # Fallback para nombres antiguos
            self.pr_familia_var.set(nombre_completo)
            self.pr_medida_var.set('')
            self.pr_caracteristica_var.set('')
        
        self.pr_cantidad_var.set(str(pr['cantidad']))
        self.pr_costo_unitario_var.set(str(pr['costo_unitario']))
        self.pr_otros_costos_var.set(str(pr.get('otros_costos', 0)))
        self.pr_moneda_var.set(pr.get('moneda', 'ARS'))
        if pr.get('moneda') == 'USD' and pr.get('valor_dolar'):
            self.pr_valor_dolar_var.set(str(pr['valor_dolar']))
        else:
            self.pr_valor_dolar_var.set('')
        
        # Actualizar visibilidad de campos
        self._actualizar_campos_moneda_pr()
        
        # Guardar el índice del producto de reventa en edición (NO eliminarlo)
        self.editando_pr_index = index
        messagebox.showinfo('Modo Edición', f'Producto de reventa "{pr["nombre"]}" cargado para edición.\nModifique los datos y presione "Agregar Producto de Reventa" para guardar los cambios.')
    
    def eliminar_producto_reventa(self):
        if not hasattr(self, 'tree_stock_prod'):
            return
        selection = self.tree_stock_prod.selection()
        if not selection:
            messagebox.showwarning('Advertencia', 'Seleccione un producto de reventa para eliminar.')
            return
        
        if messagebox.askyesno('Confirmar', '¿Eliminar producto de reventa seleccionado?'):
            item = selection[0]
            index = self.tree_stock_prod.index(item)
            self.productos_reventa.pop(index)
            self.actualizar_tabla_stock_prod()
            self.save_state()
    
    def exportar_plantilla_stock(self):
        """Exportar plantilla completa con Materia Prima y Productos de Reventa"""
        try:
            from tkinter import filedialog
            from openpyxl.styles import Font, PatternFill, Alignment
            
            filepath = filedialog.asksaveasfilename(
                defaultextension='.xlsx',
                filetypes=[('Excel files', '*.xlsx')],
                initialfile='plantilla_stock.xlsx'
            )
            if not filepath:
                return
            
            # Crear plantilla de Materia Prima con ejemplos
            df_mp = pd.DataFrame({
                'Familia': ['Ejemplo: Polietileno', 'Ejemplo: Resina PVC'],
                'Medida': ['1000 kg', '500 kg'],
                'Característica': ['Alta densidad', 'Transparente'],
                'Stock (kg)': [500.0, 250.0],
                'Costo/kg USD': [1.5, 2.0],
                'Valor dólar': [1000, 1000],
                'Stock Mínimo': [100.0, 50.0]
            })
            
            # Crear plantilla de Productos de Reventa con ejemplos de ARS y USD
            df_pr = pd.DataFrame({
                'Familia': ['Ejemplo: Tornillos', 'Ejemplo: Tuercas', 'Ejemplo: Válvulas', 'Ejemplo: Filtros'],
                'Medida': ['M8', 'M10', '1/2 pulgada', '10 micrones'],
                'Característica': ['Acero inoxidable', 'Acero galvanizado', 'Bronce', 'Filtro HEPA'],
                'Stock': [1000, 500, 100, 50],
                'Costo Unitario': [50.0, 75.0, 25.0, 15.0],  # En la moneda especificada
                'Costo Total': [50000.0, 37500.0, 2500.0, 750.0],  # En la moneda especificada
                'Moneda': ['ARS', 'ARS', 'USD', 'USD'],  # IMPORTANTE: Especificar ARS o USD
                'Valor dólar': [1000, 1000, 1000, 1000]  # IMPORTANTE: Valor del dólar cuando Moneda = USD
            })
            
            # Escribir ambas hojas en un solo archivo con formato
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df_mp.to_excel(writer, index=False, sheet_name='Materia Prima')
                df_pr.to_excel(writer, index=False, sheet_name='Productos Reventa')
                
                wb = writer.book
                
                # Formatear ambas hojas
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    
                    # Estilo para encabezados
                    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                    header_font = Font(bold=True, color='FFFFFF', size=11)
                    
                    # Aplicar estilo a encabezados
                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Ajustar ancho de columnas
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(cell.value)
                            except:
                                pass
                        
                        adjusted_width = (max_length + 2) * 1.2
                        ws.column_dimensions[column_letter].width = adjusted_width
            
            messagebox.showinfo(
                '📋 Plantilla Exportada', 
                f'✅ Plantilla creada en:\n{filepath}\n\n'
                f'📁 El archivo contiene DOS HOJAS:\n\n'
                f'📄 Hoja 1: "Materia Prima"\n'
                f'   • Familia, Medida, Característica\n'
                f'   • Stock (kg), Costo/kg USD\n'
                f'   • Valor dólar, Stock Mínimo\n\n'
                f'📄 Hoja 2: "Productos Reventa"\n'
                f'   • Familia, Medida, Característica\n'
                f'   • Stock, Costo Unitario, Costo Total\n'
                f'   • Moneda (ARS/USD), Valor dólar\n\n'
                f'💰 MANEJO DE MONEDAS:\n'
                f'• Moneda: Especifica "ARS" o "USD"\n'
                f'• Valor dólar: Obligatorio cuando Moneda = "USD"\n'
                f'• Costos: En la moneda especificada\n'
                f'• Conversión: Automática a ARS\n\n'
                f'💡 INSTRUCCIONES:\n'
                f'1. Edita las hojas según tus necesidades\n'
                f'2. Usa "📥 Importar Materia Prima" para la hoja 1\n'
                f'3. Usa "📥 Importar Stock" para la hoja 2\n'
                f'4. ¡Puedes usar el MISMO archivo para ambas importaciones!'
            )
        except Exception as e:
            messagebox.showerror('Error', f'Error al exportar plantilla: {str(e)}')
    
    def importar_stock_excel(self):
        """Importar productos de reventa desde archivo Excel con formato flexible"""
        try:
            from tkinter import filedialog
            filepath = filedialog.askopenfilename(
                title='Seleccionar archivo Excel de Stock',
                filetypes=[('Excel files', '*.xlsx *.xls')]
            )
            if not filepath:
                return
            
            # Leer el archivo Excel - buscar hoja "Productos Reventa"
            try:
                # Intentar leer de la hoja "Productos Reventa" primero
                try:
                    df = pd.read_excel(filepath, sheet_name='Productos Reventa')
                except:
                    # Si no existe, intentar con la primera hoja
                    df = pd.read_excel(filepath, sheet_name=0)
                    messagebox.showwarning('Aviso', 
                        'No se encontró la hoja "Productos Reventa". Se está usando la primera hoja del archivo.\n\n'
                        'Para un archivo con múltiples hojas, nombre las hojas como:\n'
                        '- "Materia Prima" para materias primas\n'
                        '- "Productos Reventa" para productos de reventa')
            except Exception as e:
                messagebox.showerror('Error', f'No se pudo leer el archivo Excel: {str(e)}')
                return
            
            # Mostrar las columnas disponibles para debug
            print(f"Columnas disponibles en el Excel: {list(df.columns)}")
            
            # Mapeo de columnas posibles (flexible)
            columnas_posibles = {
                'nombre': ['Nombre', 'nombre', 'Producto', 'producto', 'Descripción', 'descripcion'],
                'familia': ['Familia', 'familia', 'Categoría', 'categoria', 'Tipo', 'tipo'],
                'medida': ['Medida', 'medida', 'Tamaño', 'tamaño', 'Unidad', 'unidad', 'Size', 'size'],
                'caracteristica': ['Característica', 'caracteristica', 'Descripción', 'descripcion', 'Especificación', 'especificacion', 'Spec', 'spec'],
                'stock': ['Stock', 'stock', 'Cantidad', 'cantidad', 'Unidades', 'unidades', 'Quantity', 'quantity'],
                'costo_unitario': ['Costo Unitario', 'costo_unitario', 'Precio Unitario', 'precio_unitario', 'Costo por unidad', 'Unit Price', 'unit_price', 'Costo Unit', 'costo_unit'],
                'costo_total': ['Costo Total', 'costo_total', 'Costo Total Compra', 'Total', 'total', 'Importe', 'importe', 'Total Cost', 'total_cost'],
                'costo_usd': ['Costo en USD', 'costo_usd', 'Precio USD', 'precio_usd', 'USD Price', 'usd_price', 'Costo USD', 'costo_usd'],
                'dolar': ['Valor dólar', 'Dólar', 'USD', 'dolar', 'usd', 'Moneda', 'Exchange Rate', 'exchange_rate', 'Dollar', 'dollar'],
                'moneda': ['Moneda', 'moneda', 'Currency', 'currency', 'Tipo Moneda', 'tipo_moneda']
            }
            
            # Encontrar las columnas correctas
            columnas_encontradas = {}
            for tipo, opciones in columnas_posibles.items():
                for opcion in opciones:
                    if opcion in df.columns:
                        columnas_encontradas[tipo] = opcion
                        break
            
            # Si no hay columna 'nombre', intentar crear el nombre usando otras columnas
            if 'nombre' not in columnas_encontradas:
                if 'familia' in columnas_encontradas:
                    columnas_encontradas['nombre'] = columnas_encontradas['familia']
                elif 'medida' in columnas_encontradas:
                    columnas_encontradas['nombre'] = columnas_encontradas['medida']
                elif 'caracteristica' in columnas_encontradas:
                    columnas_encontradas['nombre'] = columnas_encontradas['caracteristica']
                else:
                    # Si no hay ninguna columna identificadora, usar la primera columna
                    columnas_encontradas['nombre'] = df.columns[0]
            
            # Validar columnas mínimas requeridas
            if 'stock' not in columnas_encontradas:
                messagebox.showerror('Error', 
                    f'El archivo debe contener al menos una columna de cantidad:\n'
                    f'- Stock, Cantidad, Unidades, etc.\n\n'
                    f'Columnas encontradas: {list(df.columns)}')
                return
            
            # Contador de productos procesados
            agregados = 0
            actualizados = 0
            errores = []
            
            for index, row in df.iterrows():
                try:
                    # Obtener datos básicos
                    nombre_base = str(row[columnas_encontradas['nombre']]).strip()
                    if not nombre_base or nombre_base.lower() == 'nan':
                        continue
                    
                    stock = int(row[columnas_encontradas['stock']]) if pd.notna(row[columnas_encontradas['stock']]) else 0
                    
                    # Obtener datos opcionales con validación
                    familia = str(row[columnas_encontradas.get('familia', 'nombre')]).strip() if 'familia' in columnas_encontradas and pd.notna(row[columnas_encontradas.get('familia', 'nombre')]) else nombre_base
                    if familia.lower() in ['nan', 'none', '']:
                        familia = nombre_base
                    
                    medida = str(row[columnas_encontradas.get('medida', 'medida')]).strip() if 'medida' in columnas_encontradas and pd.notna(row[columnas_encontradas.get('medida', 'medida')]) else 'N/A'
                    if medida.lower() in ['nan', 'none', '']:
                        medida = 'N/A'
                    
                    caracteristica = str(row[columnas_encontradas.get('caracteristica', 'caracteristica')]).strip() if 'caracteristica' in columnas_encontradas and pd.notna(row[columnas_encontradas.get('caracteristica', 'caracteristica')]) else 'N/A'
                    if caracteristica.lower() in ['nan', 'none', '']:
                        caracteristica = 'N/A'
                    
                    # Obtener costos con validación
                    try:
                        costo_unitario = float(row[columnas_encontradas['costo_unitario']]) if 'costo_unitario' in columnas_encontradas and pd.notna(row[columnas_encontradas['costo_unitario']]) else 0
                    except (ValueError, TypeError):
                        costo_unitario = 0
                    
                    try:
                        costo_total = float(row[columnas_encontradas['costo_total']]) if 'costo_total' in columnas_encontradas and pd.notna(row[columnas_encontradas['costo_total']]) else 0
                    except (ValueError, TypeError):
                        costo_total = 0
                    
                    try:
                        costo_usd = float(row[columnas_encontradas['costo_usd']]) if 'costo_usd' in columnas_encontradas and pd.notna(row[columnas_encontradas['costo_usd']]) else 0
                    except (ValueError, TypeError):
                        costo_usd = 0
                    
                    try:
                        valor_dolar = float(row[columnas_encontradas['dolar']]) if 'dolar' in columnas_encontradas and pd.notna(row[columnas_encontradas['dolar']]) else 1000.0
                    except (ValueError, TypeError):
                        valor_dolar = 1000.0
                    
                    moneda_raw = str(row[columnas_encontradas['moneda']]).strip().upper() if 'moneda' in columnas_encontradas and pd.notna(row[columnas_encontradas['moneda']]) else 'ARS'
                    moneda = 'USD' if moneda_raw in ['USD', 'DOLAR', 'DOLLAR', '$'] else 'ARS'
                    
                    # Crear nombre completo en formato "Familia - Medida - Característica"
                    if medida != 'N/A' and caracteristica != 'N/A':
                        nombre_completo = f"{familia} - {medida} - {caracteristica}"
                    else:
                        nombre_completo = nombre_base
                    
                    if stock <= 0:
                        continue
                    
                    # Validar que al menos haya un costo
                    if costo_unitario == 0 and costo_total == 0 and costo_usd == 0:
                        print(f"Advertencia: Producto '{nombre_completo}' sin costos, asignando costo unitario = 1.0")
                        costo_unitario = 1.0
                    
                    # Calcular costos si no están especificados
                    if costo_total == 0 and costo_unitario > 0:
                        costo_total = costo_unitario * stock
                    elif costo_unitario == 0 and costo_total > 0 and stock > 0:
                        costo_unitario = costo_total / stock
                    
                    # Calcular otros costos (diferencia entre costo total y costo unitario * stock)
                    otros_costos = 0.0
                    if costo_total > 0 and costo_unitario > 0 and stock > 0:
                        otros_costos = max(0, (costo_total - (costo_unitario * stock)) / stock)
                    
                    # Determinar moneda y valor del dólar
                    if costo_usd > 0:
                        # Si hay columna específica de costo en USD, usarla
                        costo_unitario = costo_usd
                        moneda = 'USD'
                        valor_dolar_actual = valor_dolar
                    elif moneda == 'USD':
                        # Si la moneda es USD, usar el costo unitario como USD
                        valor_dolar_actual = valor_dolar
                    else:
                        # Si es ARS, valor del dólar = 1
                        valor_dolar_actual = 1.0
                    
                    # Calcular costo unitario final (mantener en moneda original)
                    costo_unitario_final = costo_unitario + otros_costos
                    
                    # Verificar si el producto ya existe
                    existe = False
                    for i, pr in enumerate(self.productos_reventa):
                        if pr['nombre'].lower() == nombre_completo.lower():
                            # Actualizar producto existente - sumar cantidad
                            self.productos_reventa[i]['cantidad'] += stock
                            self.productos_reventa[i]['costo_unitario'] = costo_unitario
                            self.productos_reventa[i]['otros_costos'] = otros_costos
                            self.productos_reventa[i]['moneda'] = moneda
                            self.productos_reventa[i]['valor_dolar'] = valor_dolar_actual if moneda == 'USD' else None
                            self.productos_reventa[i]['costo_unitario_final'] = costo_unitario_final
                            
                            existe = True
                            actualizados += 1
                            break
                    
                    if not existe:
                        # Agregar nuevo producto con la estructura correcta
                        nuevo_producto = {
                            'nombre': nombre_completo,
                            'cantidad': stock,
                            'costo_unitario': costo_unitario,
                            'otros_costos': otros_costos,
                            'moneda': moneda,
                            'valor_dolar': valor_dolar_actual if moneda == 'USD' else None,
                            'costo_unitario_final': costo_unitario_final
                        }
                        
                        self.productos_reventa.append(nuevo_producto)
                        agregados += 1
                        
                except Exception as e:
                    errores.append(f'Fila {index + 2}: {str(e)}')
            
            # Actualizar tabla y guardar
            self.actualizar_tabla_stock_prod()
            self.save_state()
            
            # Mensaje de resultado
            mensaje = f'📦 Importación de Productos de Reventa Completada\n\n'
            mensaje += f'✅ Nuevos productos: {agregados}\n'
            mensaje += f'🔄 Productos actualizados: {actualizados}\n'
            if errores:
                mensaje += f'\n⚠️ Errores encontrados ({len(errores)}):\n'
                mensaje += '\n'.join(errores[:3])  # Mostrar solo los primeros 3 errores
                if len(errores) > 3:
                    mensaje += f'\n... y {len(errores) - 3} errores más'
            else:
                mensaje += '\n🎉 ¡Importación exitosa sin errores!'
            
            messagebox.showinfo('Importación Completada', mensaje)
            
        except Exception as e:
            messagebox.showerror('Error', f'Error al importar archivo Excel: {str(e)}')
    
    def actualizar_tabla_stock_prod(self):
        """Actualizar tabla de Productos de Reventa"""
        if not hasattr(self, 'tree_stock_prod'):
            return
        for item in self.tree_stock_prod.get_children():
            self.tree_stock_prod.delete(item)
        
        for pr in self.productos_reventa:
            # Extraer componentes del nombre (formato: "Familia - Medida - Característica")
            nombre_completo = pr['nombre']
            nombre_parts = nombre_completo.split(' - ')
            if len(nombre_parts) >= 3:
                familia = nombre_parts[0]
                medida = nombre_parts[1]
                caracteristica = ' - '.join(nombre_parts[2:])
            else:
                # Fallback para nombres antiguos
                familia = nombre_completo
                medida = ''
                caracteristica = ''
            
            cantidad = pr['cantidad']
            moneda = pr.get('moneda', 'ARS')
            costo_unitario = pr.get('costo_unitario', 0)
            otros_costos = pr.get('otros_costos', 0)
            costo_unitario_final = pr.get('costo_unitario_final', costo_unitario + otros_costos)
            
            # Mostrar costo con moneda
            if moneda == 'USD':
                costo_display = f'${costo_unitario_final:.2f} (USD)'
            else:
                costo_display = f'${costo_unitario_final:.2f}'
            
            # Costo total (mantener en la misma moneda)
            costo_total = cantidad * costo_unitario_final
            
            self.tree_stock_prod.insert('', 'end', values=(
                familia,
                medida,
                caracteristica,
                cantidad,
                costo_display,
                f'${costo_total:.2f}' + (' (USD)' if moneda == 'USD' else '')
            ))
        
        # Refrescar combo de ventas cuando cambie el stock
        try:
            self._refresh_combo_venta_productos()
        except Exception:
            pass

    def actualizar_tabla_stock_fabricados(self):
        """Actualizar tabla de Productos Fabricados"""
        if not hasattr(self, 'tree_stock_fabricados'):
            return
        for item in self.tree_stock_fabricados.get_children():
            self.tree_stock_fabricados.delete(item)
        
        for s in self.stock_prod:
            # Extraer componentes del nombre (formato: "Familia - Medida - Característica")
            nombre_completo = s.get('nombre')
            nombre_parts = nombre_completo.split(' - ')
            if len(nombre_parts) >= 3:
                familia = nombre_parts[0]
                medida = nombre_parts[1]
                caracteristica = ' - '.join(nombre_parts[2:])
            else:
                # Fallback para nombres antiguos
                familia = nombre_completo
                medida = ''
                caracteristica = ''
            
            cantidad = int(s.get('cantidad', 0))
            peso_unidad = float(s.get('peso_unidad', 0))
            total_kg = cantidad * peso_unidad
            costo_unit_total = s.get('costo_unit_total')
            costo_text = f'${float(costo_unit_total):.2f}' if costo_unit_total is not None else '-'
            
            # Calcular costo total
            costo_total = cantidad * float(costo_unit_total) if costo_unit_total is not None else 0
            costo_total_text = f'${costo_total:.2f}' if costo_unit_total is not None else '-'
            
            self.tree_stock_fabricados.insert('', 'end', values=(
                familia,
                medida,
                caracteristica,
                cantidad, 
                f'{peso_unidad:.2f}', 
                f'{total_kg:.2f}', 
                costo_text,
                costo_total_text
            ))
    
    def eliminar_stock_fabricado(self):
        """Eliminar productos fabricados"""
        if not hasattr(self, 'tree_stock_fabricados'):
            return
        selection = self.tree_stock_fabricados.selection()
        if not selection:
            messagebox.showwarning('Advertencia', 'Seleccione un producto fabricado para eliminar.')
            return
        if messagebox.askyesno('Confirmar', '¿Eliminar registro de stock del producto fabricado?'):
            item = selection[0]
            index = self.tree_stock_fabricados.index(item)
            self.stock_prod.pop(index)
            self.actualizar_tabla_stock_fabricados()
            self.save_state()
    
    def eliminar_stock_producto(self):
        """Eliminar productos fabricados (stock_prod) - función legacy - redirige a la nueva"""
        self.eliminar_stock_fabricado()

    # ==================== FUNCIONES DE CONTROL DE INGRESOS Y EGRESOS ====================
    
    def calcular_metricas_stock(self):
        """Calcula las métricas de ingresos, egresos y saldos para todas las categorías"""
        if not hasattr(self, 'label_ingreso_mp'):
            return
            
        # Calcular métricas de Materia Prima
        total_ingreso_mp = sum(float(log.get('cantidad', 0)) for log in self.log_ingresos_mp)
        total_egreso_mp = sum(float(log.get('cantidad', 0)) for log in self.log_egresos_mp)
        saldo_mp = total_ingreso_mp - total_egreso_mp
        
        # Actualizar labels de MP
        self.label_ingreso_mp.config(text=f'Ingresos: {total_ingreso_mp:.1f} kg')
        self.label_egreso_mp.config(text=f'Egresos: {total_egreso_mp:.1f} kg')
        self.label_saldo_mp.config(text=f'Saldo: {saldo_mp:.1f} kg')
        
        # Calcular métricas de Productos de Reventa
        total_ingreso_pr = sum(float(log.get('cantidad', 0)) for log in self.log_ingresos_pr)
        total_egreso_pr = sum(float(log.get('cantidad', 0)) for log in self.log_egresos_pr)
        saldo_pr = total_ingreso_pr - total_egreso_pr
        
        # Actualizar labels de PR
        self.label_ingreso_pr.config(text=f'Ingresos: {total_ingreso_pr:.0f} u')
        self.label_egreso_pr.config(text=f'Egresos: {total_egreso_pr:.0f} u')
        self.label_saldo_pr.config(text=f'Saldo: {saldo_pr:.0f} u')
        
        # Calcular métricas de Productos Fabricados
        total_ingreso_fab = sum(float(log.get('cantidad', 0)) for log in self.log_ingresos_fab)
        total_egreso_fab = sum(float(log.get('cantidad', 0)) for log in self.log_egresos_fab)
        saldo_fab = total_ingreso_fab - total_egreso_fab
        
        # Actualizar labels de FAB
        self.label_ingreso_fab.config(text=f'Ingresos: {total_ingreso_fab:.0f} u')
        self.label_egreso_fab.config(text=f'Egresos: {total_egreso_fab:.0f} u')
        self.label_saldo_fab.config(text=f'Saldo: {saldo_fab:.0f} u')
        
        # Actualizar colores de las flechas según el saldo
        self._actualizar_colores_flechas(saldo_mp, saldo_pr, saldo_fab)
    
    def _actualizar_colores_flechas(self, saldo_mp, saldo_pr, saldo_fab):
        """Actualiza los colores de las flechas según el saldo"""
        # Colores: verde para positivo, rojo para negativo, azul para cero
        
        # Materia Prima
        if saldo_mp > 0:
            self.arrow_saldo_mp.config(fg='#27AE60')  # Verde
        elif saldo_mp < 0:
            self.arrow_saldo_mp.config(fg='#E74C3C')  # Rojo
        else:
            self.arrow_saldo_mp.config(fg='#3498DB')  # Azul
            
        # Productos de Reventa
        if saldo_pr > 0:
            self.arrow_saldo_pr.config(fg='#27AE60')  # Verde
        elif saldo_pr < 0:
            self.arrow_saldo_pr.config(fg='#E74C3C')  # Rojo
        else:
            self.arrow_saldo_pr.config(fg='#3498DB')  # Azul
            
        # Productos Fabricados
        if saldo_fab > 0:
            self.arrow_saldo_fab.config(fg='#27AE60')  # Verde
        elif saldo_fab < 0:
            self.arrow_saldo_fab.config(fg='#E74C3C')  # Rojo
        else:
            self.arrow_saldo_fab.config(fg='#3498DB')  # Azul
    
    def registrar_movimiento_manual(self):
        """Registra un movimiento manual de ingreso o egreso"""
        familia = self.control_familia_var.get().strip()
        medida = self.control_medida_var.get().strip()
        caracteristica = self.control_caracteristica_var.get().strip()
        cantidad_str = self.control_cantidad_var.get().strip()
        motivo = self.control_motivo_var.get().strip()
        tipo = self.control_tipo_var.get()
        
        # Validaciones
        if not (familia and medida and caracteristica and cantidad_str):
            messagebox.showerror('Error', 'Por favor complete Familia, Medida, Característica y Cantidad.')
            return
            
        try:
            cantidad = float(cantidad_str)
            if cantidad <= 0:
                messagebox.showerror('Error', 'La cantidad debe ser mayor a 0.')
                return
        except ValueError:
            messagebox.showerror('Error', 'La cantidad debe ser un número válido.')
            return
            
        if not motivo:
            motivo = 'Movimiento manual'
            
        # Crear registro del movimiento
        from datetime import datetime
        movimiento = {
            'fecha': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'tipo': 'Ingreso' if tipo == 'ingreso' else 'Egreso',
            'familia': familia,
            'medida': medida,
            'caracteristica': caracteristica,
            'cantidad': cantidad,
            'motivo': motivo
        }
        
        # Determinar a qué lista agregar según el tipo de producto
        # Por simplicidad, asumimos que es materia prima (kg)
        # En una implementación más avanzada, se podría seleccionar el tipo
        
        if tipo == 'ingreso':
            self.log_ingresos_mp.append(movimiento)
        else:
            self.log_egresos_mp.append(movimiento)
            
        # Limpiar formulario
        self.control_familia_var.set('')
        self.control_medida_var.set('')
        self.control_caracteristica_var.set('')
        self.control_cantidad_var.set('')
        self.control_motivo_var.set('')
        
        # Actualizar métricas y log
        self.calcular_metricas_stock()
        self.actualizar_log_movimientos()
        self.save_state()
        
        messagebox.showinfo('Éxito', f'Movimiento de {tipo} registrado correctamente.')
    
    def actualizar_log_movimientos(self):
        """Actualiza la tabla de log de movimientos"""
        if not hasattr(self, 'tree_log_movimientos'):
            return
            
        # Limpiar tabla
        for item in self.tree_log_movimientos.get_children():
            self.tree_log_movimientos.delete(item)
            
        # Combinar todos los movimientos
        todos_movimientos = []
        
        # Agregar ingresos
        for movimiento in self.log_ingresos_mp + self.log_ingresos_pr + self.log_ingresos_fab:
            movimiento['tipo'] = 'Ingreso'
            todos_movimientos.append(movimiento)
            
        # Agregar egresos
        for movimiento in self.log_egresos_mp + self.log_egresos_pr + self.log_egresos_fab:
            movimiento['tipo'] = 'Egreso'
            todos_movimientos.append(movimiento)
            
        # Ordenar por fecha (más reciente primero)
        todos_movimientos.sort(key=lambda x: x['fecha'], reverse=True)
        
        # Insertar en la tabla
        for movimiento in todos_movimientos:
            self.tree_log_movimientos.insert('', 'end', values=(
                movimiento['fecha'],
                movimiento['tipo'],
                movimiento['familia'],
                movimiento['medida'],
                movimiento['caracteristica'],
                movimiento['cantidad'],
                movimiento['motivo']
            ))
    
    def registrar_ingreso_automatico(self, familia, medida, caracteristica, cantidad, motivo="Ingreso automático"):
        """Registra automáticamente un ingreso (usado cuando se agrega stock)"""
        from datetime import datetime
        movimiento = {
            'fecha': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'tipo': 'Ingreso',
            'familia': familia,
            'medida': medida,
            'caracteristica': caracteristica,
            'cantidad': cantidad,
            'motivo': motivo
        }
        
        # Determinar tipo de producto y agregar a la lista correspondiente
        # Por simplicidad, asumimos que es materia prima
        self.log_ingresos_mp.append(movimiento)
        self.calcular_metricas_stock()
        self.actualizar_log_movimientos()
    
    def registrar_egreso_automatico(self, familia, medida, caracteristica, cantidad, motivo="Egreso automático"):
        """Registra automáticamente un egreso (usado cuando se consume stock)"""
        from datetime import datetime
        movimiento = {
            'fecha': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'tipo': 'Egreso',
            'familia': familia,
            'medida': medida,
            'caracteristica': caracteristica,
            'cantidad': cantidad,
            'motivo': motivo
        }
        
        # Determinar tipo de producto y agregar a la lista correspondiente
        # Por simplicidad, asumimos que es materia prima
        self.log_egresos_mp.append(movimiento)
        self.calcular_metricas_stock()
        self.actualizar_log_movimientos()

    def registrar_ingreso_automatico_productos(self, familia, medida, caracteristica, cantidad, motivo="Ingreso automático"):
        """Registra automáticamente un ingreso de productos de reventa"""
        from datetime import datetime
        movimiento = {
            'fecha': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'tipo': 'Ingreso',
            'familia': familia,
            'medida': medida,
            'caracteristica': caracteristica,
            'cantidad': cantidad,
            'motivo': motivo
        }
        
        # Agregar a la lista de ingresos de productos de reventa
        if not hasattr(self, 'log_ingresos_pr'):
            self.log_ingresos_pr = []
        self.log_ingresos_pr.append(movimiento)
        self.calcular_metricas_stock()
        self.actualizar_log_movimientos()

    def registrar_egreso_automatico_productos(self, familia, medida, caracteristica, cantidad, motivo="Egreso automático"):
        """Registra automáticamente un egreso de productos de reventa"""
        from datetime import datetime
        movimiento = {
            'fecha': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'tipo': 'Egreso',
            'familia': familia,
            'medida': medida,
            'caracteristica': caracteristica,
            'cantidad': cantidad,
            'motivo': motivo
        }
        
        # Agregar a la lista de egresos de productos de reventa
        if not hasattr(self, 'log_egresos_pr'):
            self.log_egresos_pr = []
        self.log_egresos_pr.append(movimiento)
        self.calcular_metricas_stock()
        self.actualizar_log_movimientos()

    def registrar_ingreso_automatico_fabricados(self, familia, medida, caracteristica, cantidad, motivo="Ingreso automático"):
        """Registra automáticamente un ingreso de productos fabricados"""
        from datetime import datetime
        movimiento = {
            'fecha': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'tipo': 'Ingreso',
            'familia': familia,
            'medida': medida,
            'caracteristica': caracteristica,
            'cantidad': cantidad,
            'motivo': motivo
        }
        
        # Agregar a la lista de ingresos de productos fabricados
        if not hasattr(self, 'log_ingresos_fab'):
            self.log_ingresos_fab = []
        self.log_ingresos_fab.append(movimiento)
        self.calcular_metricas_stock()
        self.actualizar_log_movimientos()

    def obtener_ultimo_movimiento_material(self, familia, medida, caracteristica):
        """Obtiene el último movimiento de un material específico"""
        # Buscar en ingresos y egresos de materia prima
        nombre_material = f"{familia} - {medida} - {caracteristica}"
        
        # Combinar todos los movimientos de materia prima
        todos_movimientos = []
        if hasattr(self, 'log_ingresos_mp'):
            for mov in self.log_ingresos_mp:
                mov_nombre = f"{mov.get('familia', '')} - {mov.get('medida', '')} - {mov.get('caracteristica', '')}"
                if mov_nombre.lower() == nombre_material.lower():
                    todos_movimientos.append({
                        'fecha': mov.get('fecha', ''),
                        'tipo': 'Ingreso',
                        'cantidad': mov.get('cantidad', 0)
                    })
        
        if hasattr(self, 'log_egresos_mp'):
            for mov in self.log_egresos_mp:
                mov_nombre = f"{mov.get('familia', '')} - {mov.get('medida', '')} - {mov.get('caracteristica', '')}"
                if mov_nombre.lower() == nombre_material.lower():
                    todos_movimientos.append({
                        'fecha': mov.get('fecha', ''),
                        'tipo': 'Egreso',
                        'cantidad': mov.get('cantidad', 0)
                    })
        
        # Ordenar por fecha (más reciente primero)
        todos_movimientos.sort(key=lambda x: x['fecha'], reverse=True)
        
        # Retornar el último movimiento
        if todos_movimientos:
            ultimo = todos_movimientos[0]
            if ultimo['tipo'] == 'Ingreso':
                return f"↑ +{ultimo['cantidad']:.1f} kg"
            else:
                return f"↓ -{ultimo['cantidad']:.1f} kg"
        else:
            return "Sin movimientos"

    def verificar_alertas_stock_minimo(self):
        """Verifica y muestra alertas cuando el stock está bajo el mínimo"""
        alertas = []
        
        for s in self.stock_mp:
            nombre_completo = s.get('nombre') if 'nombre' in s else s.get('Nombre', '')
            stock_actual = float(s.get('kg') if 'kg' in s else s.get('Stock (kg)', 0))
            stock_minimo = float(s.get('stock_minimo', 0))
            
            if stock_minimo > 0 and stock_actual < stock_minimo:
                # Extraer componentes del nombre
                nombre_parts = nombre_completo.split(' - ')
                if len(nombre_parts) >= 3:
                    familia = nombre_parts[0]
                    medida = nombre_parts[1]
                    caracteristica = ' - '.join(nombre_parts[2:])
                else:
                    familia = nombre_completo
                    medida = ""
                    caracteristica = ""
                
                alertas.append({
                    'material': nombre_completo,
                    'stock_actual': stock_actual,
                    'stock_minimo': stock_minimo,
                    'diferencia': stock_minimo - stock_actual
                })
        
        if alertas:
            mensaje = "⚠️ ALERTA: Stock bajo el mínimo detectado:\n\n"
            for alerta in alertas:
                mensaje += f"• {alerta['material']}\n"
                mensaje += f"  Stock actual: {alerta['stock_actual']:.1f} kg\n"
                mensaje += f"  Stock mínimo: {alerta['stock_minimo']:.1f} kg\n"
                mensaje += f"  Faltan: {alerta['diferencia']:.1f} kg\n\n"
            
            mensaje += "Se recomienda realizar una compra para reponer el stock."
            messagebox.showwarning("Alerta de Stock Mínimo", mensaje)
        
        return alertas

    def marcar_completado(self):
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning('Advertencia', 'Seleccione un producto para marcar como completado.')
            return
        item = selection[0]
        index = self.tree.index(item)
        producto = self.productos[index]
        cantidad = int(producto.cantidad_fabricar or 0)
        if cantidad <= 0:
            messagebox.showwarning('Advertencia', 'La cantidad a fabricar debe ser mayor a 0.')
            return
        if not messagebox.askyesno('Confirmar', f'Agregar {cantidad} un. de "{producto.nombre}" a stock de productos y descontar MP?'):
            return
        # Verificar stock suficiente por material
        faltantes = []
        for m in (producto.materiales or []):
            nombre_mat = (m.get('nombre') or '').strip()
            kg_por_unidad = float(m.get('kg_por_unidad', 0))
            requerido = kg_por_unidad * cantidad
            disponible = 0.0
            encontrado = False
            material_encontrado = None
            
            # Normalizar nombre buscado (quitar espacios múltiples, convertir a minúsculas)
            nombre_mat_normalizado = ' '.join(nombre_mat.lower().split())
            
            for s in self.stock_mp:
                actual_nombre = (s.get('nombre') or s.get('Nombre') or '').strip()
                # Normalizar nombre del stock
                actual_nombre_normalizado = ' '.join(actual_nombre.lower().split())
                
                # Comparación exacta normalizada
                if actual_nombre_normalizado == nombre_mat_normalizado:
                    disponible = float(s.get('kg') if 'kg' in s else s.get('Stock (kg)', 0))
                    encontrado = True
                    material_encontrado = actual_nombre
                    break
            
            # Si no encontró coincidencia exacta, intentar coincidencia parcial
            if not encontrado:
                for s in self.stock_mp:
                    actual_nombre = (s.get('nombre') or s.get('Nombre') or '').strip()
                    actual_nombre_normalizado = ' '.join(actual_nombre.lower().split())
                    
                    # Extraer la familia (primera parte antes de " - ")
                    partes_nombre_mat = nombre_mat_normalizado.split(' - ')[0].strip()
                    partes_actual = actual_nombre_normalizado.split(' - ')[0].strip()
                    
                    # Comparar familias o búsqueda parcial
                    if partes_nombre_mat == partes_actual or \
                       nombre_mat_normalizado in actual_nombre_normalizado or \
                       actual_nombre_normalizado in nombre_mat_normalizado:
                        disponible = float(s.get('kg') if 'kg' in s else s.get('Stock (kg)', 0))
                        encontrado = True
                        material_encontrado = actual_nombre
                        break
            
            if not encontrado:
                faltantes.append(f'{nombre_mat}: necesita {requerido:.2f} kg, disponible {disponible:.2f} kg (MATERIAL NO ENCONTRADO)')
            elif disponible + 1e-9 < requerido:
                faltantes.append(f'{nombre_mat}: necesita {requerido:.2f} kg, disponible {disponible:.2f} kg (encontrado como: {material_encontrado})')
        if faltantes:
            messagebox.showerror('Stock insuficiente', 'No hay stock suficiente para:\n' + "\n".join(faltantes))
            return
        # Descontar stock por material y registrar egresos
        for m in (producto.materiales or []):
            nombre_mat = (m.get('nombre') or '').strip()
            kg_por_unidad = float(m.get('kg_por_unidad', 0))
            requerido = kg_por_unidad * cantidad
            
            # Extraer componentes del material para el registro
            nombre_parts = nombre_mat.split(' - ')
            if len(nombre_parts) >= 3:
                familia_mat = nombre_parts[0]
                medida_mat = nombre_parts[1]
                caracteristica_mat = ' - '.join(nombre_parts[2:])
            else:
                familia_mat = nombre_mat
                medida_mat = "N/A"
                caracteristica_mat = "N/A"
            
            # Usar la misma lógica de búsqueda flexible
            nombre_mat_normalizado = ' '.join(nombre_mat.lower().split())
            encontrado_para_descontar = False
            material_descontado = None
            
            # Primero intentar búsqueda exacta
            for s in self.stock_mp:
                actual_nombre = (s.get('nombre') or s.get('Nombre') or '').strip()
                actual_nombre_normalizado = ' '.join(actual_nombre.lower().split())
                
                if actual_nombre_normalizado == nombre_mat_normalizado:
                    # Obtener el valor actual de kg (manejar ambos formatos)
                    actual_kg = 0.0
                    if 'kg' in s:
                        actual_kg = float(s.get('kg', 0))
                    elif 'Stock (kg)' in s:
                        actual_kg = float(s.get('Stock (kg)', 0))
                    
                    # Actualizar el stock - asegurar formato consistente
                    nuevo_kg = max(0.0, actual_kg - requerido)
                    s['nombre'] = actual_nombre
                    s['kg'] = nuevo_kg
                    # Limpiar formatos antiguos
                    s.pop('Stock (kg)', None)
                    s.pop('Nombre', None)
                    
                    # Registrar egreso de materia prima
                    self.registrar_egreso_automatico(familia_mat, medida_mat, caracteristica_mat, requerido, f"Consumo para producción: {producto.nombre}")
                    encontrado_para_descontar = True
                    material_descontado = actual_nombre
                    break
            
            # Si no encontró coincidencia exacta, intentar coincidencia parcial
            if not encontrado_para_descontar:
                for s in self.stock_mp:
                    actual_nombre = (s.get('nombre') or s.get('Nombre') or '').strip()
                    actual_nombre_normalizado = ' '.join(actual_nombre.lower().split())
                    
                    partes_nombre_mat = nombre_mat_normalizado.split(' - ')[0].strip()
                    partes_actual = actual_nombre_normalizado.split(' - ')[0].strip()
                    
                    if partes_nombre_mat == partes_actual or \
                       nombre_mat_normalizado in actual_nombre_normalizado or \
                       actual_nombre_normalizado in nombre_mat_normalizado:
                        # Obtener el valor actual de kg (manejar ambos formatos)
                        actual_kg = 0.0
                        if 'kg' in s:
                            actual_kg = float(s.get('kg', 0))
                        elif 'Stock (kg)' in s:
                            actual_kg = float(s.get('Stock (kg)', 0))
                        
                        # Actualizar el stock - asegurar formato consistente
                        nuevo_kg = max(0.0, actual_kg - requerido)
                        s['nombre'] = actual_nombre
                        s['kg'] = nuevo_kg
                        # Limpiar formatos antiguos
                        s.pop('Stock (kg)', None)
                        s.pop('Nombre', None)
                        
                        # Registrar egreso de materia prima
                        self.registrar_egreso_automatico(familia_mat, medida_mat, caracteristica_mat, requerido, f"Consumo para producción: {producto.nombre}")
                        encontrado_para_descontar = True
                        material_descontado = actual_nombre
                        break
            
            # Verificar que se encontró y descontó el material
            if not encontrado_para_descontar:
                messagebox.showerror('Error', f'No se pudo encontrar el material "{nombre_mat}" en el stock para descontar {requerido:.2f} kg.\nPor favor verifique el stock de materia prima.')
                return
        self.actualizar_tabla_materia_prima()
        
        # Verificar alertas de stock mínimo después de consumir materia prima
        self.verificar_alertas_stock_minimo()
        
        # Agregar/sumar al stock de productos finalizados
        # Calcular costos antes de resetear valores del producto
        try:
            costo_mp_before = float(producto.costo_unitario_mp)
        except Exception:
            costo_mp_before = 0
        try:
            costo_mo_before = float(producto.costo_unitario_mano_obra)
        except Exception:
            costo_mo_before = 0
        costo_unit_total = costo_mp_before + costo_mo_before
        encontrado = False
        for s in self.stock_prod:
            if s['nombre'].strip().lower() == producto.nombre.strip().lower() and abs(float(s['peso_unidad']) - float(producto.peso_unidad)) < 1e-9:
                cant_actual = int(s.get('cantidad', 0))
                nueva_cant = cant_actual + cantidad
                costo_existente = float(s.get('costo_unit_total', costo_unit_total))
                s['costo_unit_total'] = ((costo_existente * cant_actual) + (costo_unit_total * cantidad)) / (nueva_cant if nueva_cant else 1)
                s['cantidad'] = nueva_cant
                encontrado = True
                break
        if not encontrado:
            self.stock_prod.append({'nombre': producto.nombre, 'cantidad': cantidad, 'peso_unidad': float(producto.peso_unidad), 'costo_unit_total': costo_unit_total})
        
        # Verificar si el producto existe en productos_reventa y reducir stock si existe
        producto_en_reventa = False
        cantidad_producida_restante = cantidad
        for pr in self.productos_reventa:
            nombre_pr = pr.get('nombre', '').strip()
            if nombre_pr.lower() == producto.nombre.strip().lower():
                cantidad_pr = int(pr.get('cantidad', 0))
                if cantidad_pr > 0:
                    # Reducir la cantidad en productos_reventa
                    cantidad_a_reducir = min(cantidad_producida_restante, cantidad_pr)
                    pr['cantidad'] = max(0, cantidad_pr - cantidad_a_reducir)
                    cantidad_producida_restante -= cantidad_a_reducir
                    producto_en_reventa = True
                    
                    # Si se redujo completamente, mostrar mensaje
                    if cantidad_a_reducir == cantidad_pr:
                        print(f"✓ Stock de reventa completado para '{producto.nombre}': se redujo {cantidad_a_reducir} unidades")
                    
                    # Si ya se usó toda la producción, no seguir buscando
                    if cantidad_producida_restante <= 0:
                        break
        
        # Si se redujo stock de reventa, mostrar resumen y actualizar
        if producto_en_reventa:
            cantidad_reducida = cantidad - cantidad_producida_restante
            if cantidad_reducida > 0:
                # Actualizar tabla de productos de reventa
                self.actualizar_tabla_stock_prod()
                print(f"ℹ Producción de '{producto.nombre}': {cantidad_reducida} unidades usadas para reducir stock de reventa")
        
        # Registrar ingreso de productos fabricados
        # Extraer componentes del producto para el registro
        nombre_parts = producto.nombre.split(' - ')
        if len(nombre_parts) >= 3:
            familia_prod = nombre_parts[0]
            medida_prod = nombre_parts[1]
            caracteristica_prod = ' - '.join(nombre_parts[2:])
        else:
            familia_prod = producto.nombre
            medida_prod = "N/A"
            caracteristica_prod = "N/A"
        
        self.registrar_ingreso_automatico_fabricados(familia_prod, medida_prod, caracteristica_prod, cantidad, f"Producción completada: {producto.nombre}")
        
        # Eliminar el producto de la lista ya que la tarea se completó
        self.productos.pop(index)
        
        self.actualizar_tabla_productos()
        self.actualizar_resumen_productos()
        self.actualizar_tabla_stock_prod()
        self.actualizar_tabla_stock_fabricados()
        self._refresh_combo_venta_productos()
        
        # Guardar el estado después de actualizar el stock
        self.save_state()
        # Registrar métrica
        try:
            import datetime as _dt
            fecha = _dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        except Exception:
            fecha = ''
        kg_consumidos = sum(float(m.get('kg_por_unidad', 0)) for m in (producto.materiales or [])) * cantidad
        # Recalcular costos unitarios para mostrar en métricas
        try:
            costo_mo = float(producto.costo_unitario_mano_obra or 0)
            costo_mp = float(producto.costo_unitario_mp or 0)
        except Exception:
            costo_mo = 0
            costo_mp = 0
        metrica = {
            'Fecha': fecha,
            'Tipo': 'Ingreso stock',
            'Producto': producto.nombre,
            'Cantidad (u)': cantidad,
            'Peso unidad (kg)': float(producto.peso_unidad),
            'Kg consumidos': kg_consumidos,
            'Costo unitario MP': round(costo_mp_before, 5),
            'Costo unitario MO': round(costo_mo_before, 5),
            'Costo total MP': round(costo_mp_before * cantidad, 5),
            'Precio de venta': producto.precio_venta if producto.precio_venta is not None else '',
            'Rentabilidad neta': round(producto.rentabilidad_neta, 5) if producto.rentabilidad_neta is not None else '',
            'Rentabilidad total': round((producto.rentabilidad_neta or 0) * cantidad, 5) if producto.rentabilidad_neta is not None else ''
        }
        self.metricas.append(metrica)
        self.actualizar_tabla_metricas()
        self.save_state()

    # -------- Utilidades de material (combobox Productos) --------
    def refresh_mp_combobox(self):
        if not hasattr(self, 'combo_mp'):
            return
        nombres = []
        for s in self.stock_mp:
            nombre = s.get('nombre') if 'nombre' in s else s.get('Nombre')
            if nombre:
                nombres.append(nombre)
        # valores únicos
        valores = sorted(list({n for n in nombres}))
        self.combo_mp['values'] = valores
        # Reset selección si ya no existe
        if self.sel_mp_var.get() not in valores:
            self.sel_mp_var.set('')

    def actualizar_tabla_materiales_form(self):
        if not hasattr(self, 'tree_materiales_prod'):
            return
        for item in self.tree_materiales_prod.get_children():
            self.tree_materiales_prod.delete(item)
        for m in self.materiales_producto_form:
            self.tree_materiales_prod.insert('', 'end', values=(m['nombre'], m['kg_por_unidad']))

    def agregar_material_a_producto_form(self):
        nombre = (self.sel_mp_var.get() or '').strip()
        kg_str = (self.kg_por_unidad_var.get() or '').strip()
        if not (nombre and kg_str):
            messagebox.showerror('Error', 'Seleccione material y complete kg por unidad.')
            return
        try:
            kg_val = float(kg_str)
        except Exception:
            messagebox.showerror('Error', 'Kg por unidad debe ser numérico.')
            return
        if kg_val <= 0:
            messagebox.showerror('Error', 'Kg por unidad debe ser mayor a 0.')
            return
        # Acumular si ya existe el material en la lista
        for m in self.materiales_producto_form:
            if m['nombre'].lower() == nombre.lower():
                m['kg_por_unidad'] = float(m.get('kg_por_unidad', 0)) + kg_val
                self.kg_por_unidad_var.set('')
                self.actualizar_tabla_materiales_form()
                return
        self.materiales_producto_form.append({'nombre': nombre, 'kg_por_unidad': kg_val})
        self.kg_por_unidad_var.set('')
        self.actualizar_tabla_materiales_form()

    # -------- Métricas --------
    def actualizar_tabla_metricas(self):
        # Producción / Ingreso stock
        if hasattr(self, 'tree_metricas_prod'):
            for item in self.tree_metricas_prod.get_children():
                self.tree_metricas_prod.delete(item)
            total_kg = 0
            prod_por_producto = {}
            for m in self.metricas:
                if m.get('Tipo') == 'Ingreso stock':
                    # Tomar valores guardados en la métrica si existen
                    costo_mp = float(m.get('Costo unitario MP', 0) or 0)
                    costo_mo = m.get('Costo unitario MO')
                    if costo_mo is None:
                        # Fallback: buscar en producto actual (puede ser 0 si ya se reseteó)
                        costo_mo = 0.0
                        for p in self.productos:
                            if p.nombre == m.get('Producto',''):
                                costo_mo = float(getattr(p, 'costo_unitario_mano_obra', 0) or 0)
                                break
                    else:
                        try:
                            costo_mo = float(costo_mo)
                        except Exception:
                            costo_mo = 0.0
                    costo_prod_unit = costo_mp + costo_mo
                    costo_total_mp = float(m.get('Costo total MP',0) or 0)
                    # Extraer componentes del nombre del producto
                    producto_nombre = m.get('Producto','')
                    nombre_parts = producto_nombre.split(' - ')
                    if len(nombre_parts) >= 3:
                        familia = nombre_parts[0]
                        medida = nombre_parts[1]
                        caracteristica = ' - '.join(nombre_parts[2:])
                    else:
                        # Fallback para nombres antiguos
                        familia = producto_nombre
                        medida = ''
                        caracteristica = ''
                    
                    self.tree_metricas_prod.insert('', 'end', values=(
                        m.get('Fecha',''), familia, medida, caracteristica, m.get('Cantidad (u)',0), 
                        round(float(m.get('Peso unidad (kg)',0) or 0), 5), 
                        round(float(m.get('Kg consumidos',0) or 0), 5),
                        round(costo_mp,5), round(costo_mo,5), round(costo_prod_unit,5), round(costo_total_mp,5), 
                        round(float(m.get('Precio de venta',0) or 0), 5) if m.get('Precio de venta') else '', 
                        round(float(m.get('Rentabilidad neta',0) or 0), 5) if m.get('Rentabilidad neta') else '', 
                        round(float(m.get('Rentabilidad total',0) or 0), 5) if m.get('Rentabilidad total') else ''
                    ))
                    try:
                        total_kg += float(m.get('Kg consumidos',0) or 0)
                        prod_por_producto[m.get('Producto','')] = prod_por_producto.get(m.get('Producto',''), 0) + float(m.get('Kg consumidos',0) or 0)
                    except Exception:
                        pass
            # actualizar texto y gráfico
            try:
                self.lbl_prod_text.config(text=f'Total kg consumidos: {total_kg:.2f}')
                if Figure is not None:
                    self.ax_prod.clear(); self.ax_prod.set_title('Kg consumidos')
                    if prod_por_producto:
                        self.ax_prod.bar(list(prod_por_producto.keys()), list(prod_por_producto.values()))
                    self.canvas_prod.draw()
            except Exception:
                pass
        # Ingresos MP
        if hasattr(self, 'tree_metricas_mp'):
            for item in self.tree_metricas_mp.get_children():
                self.tree_metricas_mp.delete(item)
            total_kg = 0
            mp_por_material = {}
            for m in self.metricas:
                if m.get('Tipo') == 'Ingreso MP':
                    self.tree_metricas_mp.insert('', 'end', values=(
                        m.get('Fecha',''), m.get('Producto',''), 
                        round(float(m.get('Kg consumidos',0) or 0), 5), 
                        round(float(m.get('Costo unitario MP',0) or 0), 5), 
                        round(float(m.get('Costo total MP',0) or 0), 5)
                    ))
                    try:
                        total_kg += float(m.get('Kg consumidos',0) or 0)
                        mp_por_material[m.get('Producto','')] = mp_por_material.get(m.get('Producto',''), 0) + float(m.get('Kg consumidos',0) or 0)
                    except Exception:
                        pass
            try:
                self.lbl_mp_text.config(text=f'Total kg MP: {total_kg:.2f}')
                if Figure is not None:
                    self.ax_mp.clear(); self.ax_mp.set_title('Kg MP ingresados')
                    if mp_por_material:
                        self.ax_mp.pie(list(mp_por_material.values()), labels=list(mp_por_material.keys()), autopct='%1.1f%%')
                    self.canvas_mp.draw()
            except Exception:
                pass
        # Ventas
        if hasattr(self, 'tree_metricas_sales'):
            for item in self.tree_metricas_sales.get_children():
                self.tree_metricas_sales.delete(item)
            total_ingreso = 0
            ventas_por_producto = {}
            for v in self.ventas:
                # Extraer componentes del nombre del producto
                producto_nombre = v.get('Producto','')
                nombre_parts = producto_nombre.split(' - ')
                if len(nombre_parts) >= 3:
                    familia = nombre_parts[0]
                    medida = nombre_parts[1]
                    caracteristica = ' - '.join(nombre_parts[2:])
                else:
                    # Fallback para nombres antiguos
                    familia = producto_nombre
                    medida = ''
                    caracteristica = ''
                
                self.tree_metricas_sales.insert('', 'end', values=(
                    v.get('Fecha',''), familia, medida, caracteristica, v.get('Cantidad (u)',0), 
                    round(float(v.get('Precio unitario',0) or 0), 5), 
                    round(float(v.get('Descuento (%)',0) or 0), 5), 
                    round(float(v.get('IIB (%)',0) or 0), 5),
                    round(float(v.get('Precio final',0) or 0), 5), 
                    round(float(v.get('Ingreso neto',0) or 0), 5), 
                    round(float(v.get('Ganancia total',0) or 0), 5)
                ))
                try:
                    total_ingreso += float(v.get('Ingreso neto',0) or 0)
                    ventas_por_producto[v.get('Producto','')] = ventas_por_producto.get(v.get('Producto',''), 0) + float(v.get('Ingreso neto',0) or 0)
                except Exception:
                    pass
            try:
                self.lbl_sales_text.config(text=f'Ingresos netos: {total_ingreso:.2f}')
                if Figure is not None:
                    self.ax_sales.clear(); self.ax_sales.set_title('Ingresos netos')
                    if ventas_por_producto:
                        self.ax_sales.plot(list(ventas_por_producto.keys()), list(ventas_por_producto.values()), marker='o')
                    self.canvas_sales.draw()
            except Exception:
                pass

    def eliminar_metrica(self):
        if not hasattr(self, 'tree_metricas'):
            return
        selection = self.tree_metricas.selection()
        if not selection:
            messagebox.showwarning('Advertencia', 'Seleccione una métrica para eliminar.')
            return
        if messagebox.askyesno('Confirmar', '¿Eliminar métrica seleccionada?'):
            item = selection[0]
            index = self.tree_metricas.index(item)
            self.metricas.pop(index)
            self.actualizar_tabla_metricas()
            self.save_state()

    # -------- Ventas --------
    def actualizar_tabla_ventas(self):
        if not hasattr(self, 'tree_ventas'):
            return
        for item in self.tree_ventas.get_children():
            self.tree_ventas.delete(item)
        for v in self.ventas:
            # Extraer componentes del nombre del producto
            producto_nombre = v.get('Producto','')
            nombre_parts = producto_nombre.split(' - ')
            if len(nombre_parts) >= 3:
                familia = nombre_parts[0]
                medida = nombre_parts[1]
                caracteristica = ' - '.join(nombre_parts[2:])
            else:
                # Fallback para nombres antiguos
                familia = producto_nombre
                medida = ''
                caracteristica = ''
            
            self.tree_ventas.insert('', 'end', values=(
                v.get('Fecha',''), familia, medida, caracteristica, v.get('Cantidad (u)',0), v.get('Precio unitario',0), v.get('Descuento (%)',0), v.get('IIB (%)',0),
                v.get('Precio final',0), v.get('Costo unitario',0), v.get('Ingreso bruto',0), v.get('Ingreso neto',0), v.get('Ganancia un.',0), v.get('Ganancia total',0), v.get('Stock antes',0), v.get('Stock después',0)
            ))

    def _refresh_combo_venta_productos(self):
        if not hasattr(self, 'combo_venta_prod'):
            return
        
        # Obtener tipo seleccionado
        tipo = self.venta_tipo_var.get() if hasattr(self, 'venta_tipo_var') else 'fabricado'
        
        # Mostrar 'Nombre (stock: N)'
        display_values = []
        self.venta_display_map = {}
        
        if tipo == 'fabricado':
            # Productos fabricados del stock_prod
            for s in self.stock_prod:
                nombre = s.get('nombre')
                cantidad = int(s.get('cantidad', 0))
                # Preferir costo unitario desde stock (costo_unit_total)
                cp = None
                try:
                    if s.get('costo_unit_total') is not None:
                        cp = float(s.get('costo_unit_total'))
                    else:
                        for p in self.productos:
                            if (p.nombre or '').strip().lower() == str(nombre).strip().lower():
                                cp = float(p.costo_unitario_mp) + float(p.costo_unitario_mano_obra)
                                break
                except Exception:
                    cp = None
                cp_txt = f", Costo: {cp:.2f}" if cp is not None else ''
                disp = f"{nombre} (stock: {cantidad}{cp_txt})"
                display_values.append(disp)
                self.venta_display_map[disp] = {'nombre': nombre, 'tipo': 'fabricado', 'costo': cp, 'disponible': cantidad}
        else:
            # Productos de reventa
            for pr in self.productos_reventa:
                nombre = pr.get('nombre')
                cantidad = pr.get('cantidad', 0)
                costo = pr.get('costo_unitario_final', pr.get('costo_unitario', 0))
                cp_txt = f", Costo: {costo:.2f}" if costo else ''
                disp = f"{nombre} (stock: {cantidad}{cp_txt})"
                display_values.append(disp)
                self.venta_display_map[disp] = {'nombre': nombre, 'tipo': 'reventa', 'costo': costo, 'disponible': cantidad}
        
        self.combo_venta_prod['values'] = display_values
        if self.venta_producto_var.get() not in display_values:
            # Seleccionar el primero disponible para evitar combo vacío
            if display_values:
                self.venta_producto_var.set(display_values[0])
                try:
                    self.venta_on_producto_change()
                except Exception:
                    pass
            else:
                self.venta_producto_var.set('')

    def registrar_venta(self):
        sel_val = (self.venta_producto_var.get() or '').strip()
        info = self.venta_display_map.get(sel_val, {})
        
        if isinstance(info, dict):
            prod_nombre = info.get('nombre', '')
            tipo_prod = info.get('tipo', 'fabricado')
        else:
            prod_nombre = info
            tipo_prod = 'fabricado'
            
        cantidad_str = (self.venta_cantidad_var.get() or '').strip()
        precio_str = (self.venta_precio_var.get() or '').strip()
        desc_str = (self.venta_descuento_var.get() or '').strip()
        if not (prod_nombre and cantidad_str and precio_str):
            messagebox.showerror('Error', 'Complete producto, cantidad y precio unitario.')
            return
        try:
            cantidad = int(float(cantidad_str))
            precio_unit = float(precio_str)
            descuento_pct = float(desc_str) if desc_str else 0.0
        except Exception:
            messagebox.showerror('Error', 'Cantidad, precio y descuento deben ser numéricos.')
            return
        if cantidad <= 0 or precio_unit < 0 or descuento_pct < 0:
            messagebox.showerror('Error', 'Valores inválidos.')
            return
        
        # Buscar en stock según tipo
        idx = None
        stock_list = self.stock_prod if tipo_prod == 'fabricado' else self.productos_reventa
        costo_key = 'costo_unit_total' if tipo_prod == 'fabricado' else 'costo_unitario_final'
        
        for i, s in enumerate(stock_list):
            if s.get('nombre','').strip().lower() == prod_nombre.lower():
                idx = i
                break
        if idx is None:
            messagebox.showerror('Error', f'Producto no encontrado en stock de {tipo_prod}.')
            return
        stock_actual = int(stock_list[idx].get('cantidad', 0))
        if cantidad > stock_actual:
            messagebox.showerror('Stock insuficiente', f'Disponible: {stock_actual}, solicitado: {cantidad}.')
            return
        
        # Descontar y registrar
        # Recalcular precio final por unidad (IIB primero, luego descuento)
        precio_neto_iib = precio_unit * (1 - (float(self.venta_iib_var.get() or 0) / 100))
        precio_final_un = precio_neto_iib * (1 - descuento_pct / 100)
        ingreso_bruto = precio_unit * cantidad
        ingreso_neto = precio_final_un * cantidad
        stock_list[idx]['cantidad'] = stock_actual - cantidad
        
        try:
            import datetime as _dt
            fecha = _dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        except Exception:
            fecha = ''
        # Costo unitario tomado del stock
        try:
            costo_unit = float(stock_list[idx].get(costo_key) or 0)
        except Exception:
            costo_unit = 0
        venta = {
            'Fecha': fecha,
            'Producto': prod_nombre,
            'Cantidad (u)': cantidad,
            'Precio unitario': precio_unit,
            'Descuento (%)': descuento_pct,
            'IIB (%)': float(self.venta_iib_var.get() or 0),
            'Precio final': round(precio_final_un, 5),
            'Costo unitario': round(costo_unit, 5),
            'Ingreso bruto': round(ingreso_bruto, 5),
            'Ingreso neto': round(ingreso_neto, 5),
            'Ganancia un.': round(precio_final_un - costo_unit, 5),
            'Ganancia total': round((precio_final_un - costo_unit) * cantidad, 5),
            'Stock antes': stock_actual,
            'Stock después': stock_actual - cantidad
        }
        self.ventas.append(venta)
        # Refrescar vistas y persistir
        self.actualizar_tabla_stock_prod()
        self.actualizar_tabla_stock_fabricados()
        self.actualizar_tabla_ventas()
        # Agregar resumen de revenue en métricas (fila sintética)
        self._recalcular_revenue_metricas()
        self.save_state()
        self._refresh_combo_venta_productos()

    # UI helpers de ventas
    def venta_on_producto_change(self):
        # Actualiza labels de costo y disponible al seleccionar producto
        sel_val = (self.venta_producto_var.get() or '').strip()
        info = self.venta_display_map.get(sel_val, {})
        
        if isinstance(info, dict):
            costo = info.get('costo', 0)
            disp = info.get('disponible', 0)
        else:
            # Compatibilidad con formato antiguo
            costo = 0
            disp = 0
        
        if hasattr(self, 'venta_costo_label'):
            self.venta_costo_label.config(text=f'Costo un.: {costo:.2f}' if costo else 'Costo un.: -')
        if hasattr(self, 'venta_disp_label'):
            self.venta_disp_label.config(text=f'Disponible: {disp}')
        # Recalcular totales
        self.venta_recalc_totales()

    def venta_recalc_totales(self):
        try:
            cant = float(self.venta_cantidad_var.get() or 0)
            precio = float(self.venta_precio_var.get() or 0)
            desc = float(self.venta_descuento_var.get() or 0)
            iib = float(self.venta_iib_var.get() or 0)
        except Exception:
            cant = 0
            precio = 0
            desc = 0
            iib = 0
        bruto = cant * precio
        # IIB se descuenta del precio unitario primero
        precio_neto_iib = precio * (1 - iib/100)
        # luego se aplica el descuento comercial (%)
        precio_desc = precio_neto_iib * (1 - desc/100)
        precio_final = precio_desc
        neto = cant * precio_final
        if hasattr(self, 'venta_total_bruto_lbl'):
            self.venta_total_bruto_lbl.config(text=f'Bruto: {bruto:.2f}')
        if hasattr(self, 'venta_total_neto_lbl'):
            self.venta_total_neto_lbl.config(text=f'Neto: {neto:.2f}')
        if hasattr(self, 'venta_preview_final_label'):
            self.venta_preview_final_label.config(text=f'Precio final: {precio_final:.2f}')
        # Ganancia = (precio_final - costo_unit_total) por unidad y total
        costo = 0
        sel_val = (self.venta_producto_var.get() or '').strip()
        nombre = self.venta_display_map.get(sel_val, '')
        for s in self.stock_prod:
            if s.get('nombre','') == nombre:
                try:
                    costo = float(s.get('costo_unit_total') or 0)
                except Exception:
                    costo = 0
                break
        ganancia_un = precio_final - costo
        ganancia_total = ganancia_un * cant
        if hasattr(self, 'venta_ganancia_un_lbl'):
            self.venta_ganancia_un_lbl.config(text=f'Ganancia un.: {ganancia_un:.2f}')
        if hasattr(self, 'venta_ganancia_total_lbl'):
            self.venta_ganancia_total_lbl.config(text=f'Ganancia total: {ganancia_total:.2f}')

    def venta_simular_descuento(self):
        try:
            precio = float(self.venta_precio_var.get() or 0)
            desc = float(self.venta_descuento_var.get() or 0)
            iib = float(self.venta_iib_var.get() or 0)
        except Exception:
            precio = 0
            desc = 0
            iib = 0
        # Primero descontar IIB, luego descuento comercial
        precio_desc = precio * (1 - iib/100)
        precio_desc = precio_desc * (1 - desc/100)
        if hasattr(self, 'venta_preview_desc_label'):
            self.venta_preview_desc_label.config(text=f'Precio con desc.: {precio_desc:.2f}')

    def eliminar_venta(self):
        if not hasattr(self, 'tree_ventas'):
            return
        selection = self.tree_ventas.selection()
        if not selection:
            messagebox.showwarning('Advertencia', 'Seleccione una venta para eliminar.')
            return
        if messagebox.askyesno('Confirmar', '¿Eliminar venta seleccionada?'):
            item = selection[0]
            index = self.tree_ventas.index(item)
            venta = self.ventas.pop(index)
            # Reponer stock
            for s in self.stock_prod:
                if s.get('nombre','') == venta.get('Producto',''):
                    s['cantidad'] = int(s.get('cantidad',0)) + int(venta.get('Cantidad (u)',0))
                    break
            self.actualizar_tabla_stock_prod()
            self.actualizar_tabla_ventas()
            self._recalcular_revenue_metricas()
            self.save_state()

    def _recalcular_revenue_metricas(self):
        # Quitar métricas de tipo revenue (marcadas con clave '__revenue')
        self.metricas = [m for m in self.metricas if not m.get('__revenue')]
        total_ingreso = sum(float(v.get('Ingreso neto',0)) for v in self.ventas)
        met = {
            '__revenue': True,
            'Fecha': '',
            'Tipo': 'Revenue',
            'Producto': 'REVENUE',
            'Cantidad (u)': '',
            'Peso unidad (kg)': '',
            'Kg consumidos': '',
            'Costo unitario MP': '',
            'Costo total MP': '',
            'Precio de venta': '',
            'Rentabilidad neta': '',
            'Rentabilidad total': round(total_ingreso, 5)
        }
        self.metricas.append(met)
        self.actualizar_tabla_metricas()

    def eliminar_material_de_producto_form(self):
        if not hasattr(self, 'tree_materiales_prod'):
            return
        selection = self.tree_materiales_prod.selection()
        if not selection:
            messagebox.showwarning('Advertencia', 'Seleccione un material para quitar.')
            return
        item = selection[0]
        index = self.tree_materiales_prod.index(item)
        if 0 <= index < len(self.materiales_producto_form):
            self.materiales_producto_form.pop(index)
            self.actualizar_tabla_materiales_form()

    
    def abrir_popup_agregar_material_costos(self):
        """Abrir pop-up para agregar material manualmente en costos"""
        # Crear ventana pop-up
        popup = tk.Toplevel(self.root)
        popup.title('Agregar Material')
        popup.geometry('450x400')
        popup.resizable(False, False)
        popup.configure(bg=STYLE_CONFIG['bg_color'])
        
        # Centrar la ventana
        popup.transient(self.root)
        popup.grab_set()
        
        # Centrar en pantalla
        popup.update_idletasks()
        x = (popup.winfo_screenwidth() // 2) - (popup.winfo_width() // 2)
        y = (popup.winfo_screenheight() // 2) - (popup.winfo_height() // 2)
        popup.geometry(f"+{x}+{y}")
        
        # Frame principal
        main_frame = tk.Frame(popup, bg=STYLE_CONFIG['bg_color'], padx=20, pady=20)
        main_frame.pack(fill='both', expand=True)
        
        # Título
        titulo = tk.Label(main_frame, text='Agregar Material Manual', 
                         font=('Segoe UI', 14, 'bold'), 
                         fg=STYLE_CONFIG['primary_color'], 
                         bg=STYLE_CONFIG['bg_color'])
        titulo.pack(pady=(0, 20))
        
        # Variables para los campos
        nombre_var = tk.StringVar()
        cantidad_var = tk.StringVar()
        precio_var = tk.StringVar()
        moneda_var = tk.StringVar(value='ARS')  # Por defecto en pesos
        
        # Frame para los campos
        campos_frame = tk.Frame(main_frame, bg=STYLE_CONFIG['bg_color'])
        campos_frame.pack(fill='x', pady=(0, 20))
        
        # Campo Nombre
        tk.Label(campos_frame, text='Nombre del Material:', 
                font=('Arial', 10, 'bold'), 
                fg=STYLE_CONFIG['text_color'], 
                bg=STYLE_CONFIG['bg_color']).pack(anchor='w', pady=(0, 5))
        entry_nombre = tk.Entry(campos_frame, textvariable=nombre_var, 
                               font=('Arial', 10), width=40,
                               relief='solid', borderwidth=1)
        entry_nombre.pack(fill='x', pady=(0, 15))
        entry_nombre.focus()
        
        # Campo Cantidad (Kg/unidad)
        tk.Label(campos_frame, text='Cantidad por Unidad (Kg):', 
                font=('Arial', 10, 'bold'), 
                fg=STYLE_CONFIG['text_color'], 
                bg=STYLE_CONFIG['bg_color']).pack(anchor='w', pady=(0, 5))
        entry_cantidad = tk.Entry(campos_frame, textvariable=cantidad_var, 
                                 font=('Arial', 10), width=40,
                                 relief='solid', borderwidth=1,
                                 validate='key', validatecommand=self.vcmd_decimal)
        entry_cantidad.pack(fill='x', pady=(0, 15))
        
        # Frame para precio y moneda
        precio_frame = tk.Frame(campos_frame, bg=STYLE_CONFIG['bg_color'])
        precio_frame.pack(fill='x', pady=(0, 20))
        
        # Campo Precio por Kilo
        tk.Label(precio_frame, text='Precio por Kilo:', 
                font=('Arial', 10, 'bold'), 
                fg=STYLE_CONFIG['text_color'], 
                bg=STYLE_CONFIG['bg_color']).pack(anchor='w', pady=(0, 5))
        
        # Frame para precio y moneda en la misma línea
        precio_input_frame = tk.Frame(precio_frame, bg=STYLE_CONFIG['bg_color'])
        precio_input_frame.pack(fill='x')
        
        entry_precio = tk.Entry(precio_input_frame, textvariable=precio_var, 
                               font=('Arial', 10), width=25,
                               relief='solid', borderwidth=1,
                               validate='key', validatecommand=self.vcmd_decimal)
        entry_precio.pack(side='left', fill='x', expand=True, padx=(0, 10))
        
        # Radio buttons para moneda
        moneda_frame = tk.Frame(precio_input_frame, bg=STYLE_CONFIG['bg_color'])
        moneda_frame.pack(side='right')
        
        tk.Radiobutton(moneda_frame, text='ARS', variable=moneda_var, value='ARS',
                      font=('Arial', 9), bg=STYLE_CONFIG['bg_color'],
                      fg=STYLE_CONFIG['text_color']).pack(side='left', padx=(0, 10))
        
        tk.Radiobutton(moneda_frame, text='USD', variable=moneda_var, value='USD',
                      font=('Arial', 9), bg=STYLE_CONFIG['bg_color'],
                      fg=STYLE_CONFIG['text_color']).pack(side='left')
        
        # Frame para botones
        botones_frame = tk.Frame(main_frame, bg=STYLE_CONFIG['bg_color'])
        botones_frame.pack(fill='x')
        
        def guardar_material():
            """Guardar el material ingresado"""
            nombre = nombre_var.get().strip()
            cantidad_str = cantidad_var.get().strip()
            precio_str = precio_var.get().strip()
            moneda = moneda_var.get()
            
            # Validaciones
            if not nombre:
                messagebox.showerror('Error', 'El nombre del material es obligatorio.')
                return
            
            if not cantidad_str:
                messagebox.showerror('Error', 'La cantidad por unidad es obligatoria.')
                return
                
            if not precio_str:
                messagebox.showerror('Error', 'El precio por kilo es obligatorio.')
                return
            
            try:
                cantidad = float(cantidad_str)
                precio = float(precio_str)
                
                if cantidad <= 0:
                    messagebox.showerror('Error', 'La cantidad debe ser mayor a 0.')
                    return
                    
                if precio <= 0:
                    messagebox.showerror('Error', 'El precio debe ser mayor a 0.')
                    return
                    
            except ValueError:
                messagebox.showerror('Error', 'La cantidad y el precio deben ser números válidos.')
                return
            
            # Si es USD, verificar que se haya ingresado el precio del dólar
            precio_final = precio
            if moneda == 'USD':
                dolar_str = self.costos_dolar_var.get().strip()
                if not dolar_str:
                    messagebox.showerror('Error', 'Debe ingresar el precio del dólar para materiales en USD.')
                    return
                try:
                    precio_dolar = float(dolar_str)
                    if precio_dolar <= 0:
                        messagebox.showerror('Error', 'El precio del dólar debe ser mayor a 0.')
                        return
                    precio_final = precio * precio_dolar  # Convertir USD a pesos
                except ValueError:
                    messagebox.showerror('Error', 'El precio del dólar debe ser un número válido.')
                    return
            
            # Agregar material a la lista
            self.materiales_costos_form.append({
                'nombre': nombre, 
                'kg_por_unidad': cantidad,
                'precio_manual': precio_final,  # Precio final en pesos
                'precio_original': precio,  # Precio original ingresado
                'moneda_original': moneda  # Moneda original
            })
            
            # Actualizar tabla
            self.actualizar_tabla_materiales_costos()
            
            # Cerrar pop-up
            popup.destroy()
            messagebox.showinfo('Éxito', f'Material "{nombre}" agregado correctamente.')
        
        def cancelar():
            """Cancelar y cerrar pop-up"""
            popup.destroy()
        
        # Botones
        btn_guardar = tk.Button(botones_frame, text='Guardar Material', 
                               command=guardar_material,
                               font=('Arial', 10, 'bold'),
                               bg=STYLE_CONFIG['success_color'], 
                               fg='white',
                               relief='flat', borderwidth=0,
                               padx=20, pady=8)
        btn_guardar.pack(side='right', padx=(10, 0))
        
        btn_cancelar = tk.Button(botones_frame, text='Cancelar', 
                                command=cancelar,
                                font=('Arial', 10, 'bold'),
                                bg=STYLE_CONFIG['secondary_color'], 
                                fg='white',
                                relief='flat', borderwidth=0,
                                padx=20, pady=8)
        btn_cancelar.pack(side='right')
        
        # Bind Enter para guardar
        popup.bind('<Return>', lambda e: guardar_material())

    def abrir_popup_editar_material_costos(self, index, material_original):
        """Abrir pop-up para editar material existente"""
        popup = tk.Toplevel(self.root)
        popup.title('Editar Material')
        popup.geometry('450x400')
        popup.resizable(False, False)
        popup.configure(bg=STYLE_CONFIG['bg_color'])
        
        # Centrar ventana
        popup.transient(self.root)
        popup.grab_set()
        
        # Frame principal
        main_frame = tk.Frame(popup, bg=STYLE_CONFIG['bg_color'], padx=20, pady=20)
        main_frame.pack(fill='both', expand=True)
        
        # Título
        titulo = tk.Label(main_frame, text='Editar Material', 
                         font=('Segoe UI', 14, 'bold'), 
                         fg=STYLE_CONFIG['primary_color'], 
                         bg=STYLE_CONFIG['bg_color'])
        titulo.pack(pady=(0, 20))
        
        # Variables para los campos (cargar valores existentes)
        nombre_var = tk.StringVar(value=material_original.get('nombre', ''))
        cantidad_var = tk.StringVar(value=str(material_original.get('kg_por_unidad', 0)))
        precio_var = tk.StringVar(value=str(material_original.get('precio_manual', 0)))
        moneda_var = tk.StringVar(value=material_original.get('moneda_original', 'ARS'))
        
        # Campos del formulario
        campos_frame = tk.Frame(main_frame, bg=STYLE_CONFIG['bg_color'])
        campos_frame.pack(fill='x', pady=10)
        
        # Nombre del material
        tk.Label(campos_frame, text='Nombre del Material:', font=('Arial', 10), fg=STYLE_CONFIG['text_color'], bg=STYLE_CONFIG['bg_color']).pack(anchor='w')
        nombre_entry = tk.Entry(campos_frame, textvariable=nombre_var, font=('Arial', 10), width=40)
        nombre_entry.pack(fill='x', pady=(5, 10))
        
        # Cantidad por unidad
        tk.Label(campos_frame, text='Cantidad por Unidad (kg):', font=('Arial', 10), fg=STYLE_CONFIG['text_color'], bg=STYLE_CONFIG['bg_color']).pack(anchor='w')
        cantidad_entry = tk.Entry(campos_frame, textvariable=cantidad_var, font=('Arial', 10), width=40)
        cantidad_entry.pack(fill='x', pady=(5, 10))
        
        # Moneda
        tk.Label(campos_frame, text='Moneda:', font=('Arial', 10), fg=STYLE_CONFIG['text_color'], bg=STYLE_CONFIG['bg_color']).pack(anchor='w')
        moneda_frame = tk.Frame(campos_frame, bg=STYLE_CONFIG['bg_color'])
        moneda_frame.pack(fill='x', pady=(5, 10))
        
        tk.Radiobutton(moneda_frame, text='ARS', variable=moneda_var, value='ARS', font=('Arial', 10), fg=STYLE_CONFIG['text_color'], bg=STYLE_CONFIG['bg_color']).pack(side='left', padx=(0, 20))
        tk.Radiobutton(moneda_frame, text='USD', variable=moneda_var, value='USD', font=('Arial', 10), fg=STYLE_CONFIG['text_color'], bg=STYLE_CONFIG['bg_color']).pack(side='left')
        
        # Precio por kg
        tk.Label(campos_frame, text='Precio por kg:', font=('Arial', 10), fg=STYLE_CONFIG['text_color'], bg=STYLE_CONFIG['bg_color']).pack(anchor='w')
        precio_entry = tk.Entry(campos_frame, textvariable=precio_var, font=('Arial', 10), width=40)
        precio_entry.pack(fill='x', pady=(5, 10))
        
        # Botones
        botones_frame = tk.Frame(main_frame, bg=STYLE_CONFIG['bg_color'])
        botones_frame.pack(fill='x', pady=(20, 0))
        
        def guardar_material():
            nombre = nombre_var.get().strip()
            cantidad_str = cantidad_var.get().strip()
            precio_str = precio_var.get().strip()
            moneda = moneda_var.get()
            
            if not nombre:
                messagebox.showerror('Error', 'Ingrese el nombre del material.')
                return
            
            if not cantidad_str:
                messagebox.showerror('Error', 'Ingrese la cantidad por unidad.')
                return
            
            if not precio_str:
                messagebox.showerror('Error', 'Ingrese el precio por kg.')
                return
            
            try:
                cantidad = float(cantidad_str)
                precio = float(precio_str)
                
                if cantidad < 0:
                    messagebox.showerror('Error', 'La cantidad debe ser mayor o igual a 0.')
                    return
                
                if precio < 0:
                    messagebox.showerror('Error', 'El precio debe ser mayor o igual a 0.')
                    return
                
                # Calcular precio final según moneda
                if moneda == 'USD':
                    # Si es USD, verificar que se haya ingresado el precio del dólar
                    dolar_str = self.costos_dolar_var.get().strip()
                    if not dolar_str:
                        messagebox.showerror('Error', 'Debe ingresar el precio del dólar para materiales en USD.')
                        return
                    try:
                        precio_dolar = float(dolar_str)
                        if precio_dolar <= 0:
                            messagebox.showerror('Error', 'El precio del dólar debe ser mayor a 0.')
                            return
                        precio_final = precio * precio_dolar  # Convertir USD a pesos
                    except ValueError:
                        messagebox.showerror('Error', 'El precio del dólar debe ser un número válido.')
                        return
                else:
                    precio_final = precio  # Ya está en pesos
                
                # Actualizar el material en la lista
                self.materiales_costos_form[index] = {
                    'nombre': nombre,
                    'kg_por_unidad': cantidad,
                    'precio_manual': precio_final,
                    'precio_original': precio,
                    'moneda_original': moneda
                }
                
                self.actualizar_tabla_materiales_costos()
                popup.destroy()
                messagebox.showinfo('Material actualizado', f'Material "{nombre}" actualizado correctamente.')
                
            except ValueError:
                messagebox.showerror('Error', 'La cantidad y el precio deben ser números válidos.')
                return
        
        btn_guardar = ttk.Button(botones_frame, text='Guardar', command=guardar_material, style='Action.TButton')
        btn_guardar.pack(side='left')
        
        btn_cancelar = ttk.Button(botones_frame, text='Cancelar', command=popup.destroy, style='Secondary.TButton')
        btn_cancelar.pack(side='right')
        
        # Bind Enter para guardar
        popup.bind('<Return>', lambda e: guardar_material())
    
    def editar_material_costos(self, event=None):
        """Editar material seleccionado de la tabla de materiales en costos"""
        if not hasattr(self, 'tree_materiales_costos'):
            return
        
        # Si se llama desde doble clic, usar el item del evento
        if event:
            item = event.widget.selection()[0] if event.widget.selection() else None
            if not item:
                return
            index = event.widget.index(item)
        else:
            # Si se llama desde botón, usar la selección actual
            selection = self.tree_materiales_costos.selection()
            if not selection:
                messagebox.showwarning('Advertencia', 'Seleccione un material para editar.')
                return
            item = selection[0]
            index = self.tree_materiales_costos.index(item)
        
        if 0 <= index < len(self.materiales_costos_form):
            material_original = self.materiales_costos_form[index]
            self.abrir_popup_editar_material_costos(index, material_original)
        else:
            messagebox.showerror('Error', 'Índice de material inválido.')
    
    def eliminar_material_costos(self):
        """Eliminar material seleccionado de la tabla de materiales en costos"""
        if not hasattr(self, 'tree_materiales_costos'):
            return
        
        selection = self.tree_materiales_costos.selection()
        if not selection:
            messagebox.showwarning('Advertencia', 'Seleccione un material para eliminar.')
            return
        
        item = selection[0]
        index = self.tree_materiales_costos.index(item)
        
        if 0 <= index < len(self.materiales_costos_form):
            material_eliminado = self.materiales_costos_form.pop(index)
            self.actualizar_tabla_materiales_costos()
            messagebox.showinfo('Material eliminado', f'Material "{material_eliminado["nombre"]}" eliminado correctamente.')
        else:
            messagebox.showerror('Error', 'Índice de material inválido.')
    
    def agregar_material_costos(self):
        """Abrir pop-up para agregar material manualmente"""
        self.abrir_popup_agregar_material_costos()
    
    def actualizar_tabla_materiales_costos(self):
        """Actualizar tabla de materiales en formulario de costos"""
        for item in self.tree_materiales_costos.get_children():
            self.tree_materiales_costos.delete(item)
        
        peso_total = 0.0
        mapa_precios = self._mapa_precios_mp()
        
        for mat in self.materiales_costos_form:
            nombre = mat['nombre']
            kg = mat['kg_por_unidad']
            peso_total += kg
            
            # Verificar si tiene precio manual
            if 'precio_manual' in mat:
                # Usar precio manual ingresado (en pesos)
                costo_pesos = mat['precio_manual']
                costo_total = kg * costo_pesos
                moneda_original = mat.get('moneda_original', 'ARS')
                precio_original = mat.get('precio_original', costo_pesos)
                
                # Mostrar precio original con su moneda
                if moneda_original == 'USD':
                    precio_mostrar = f'{precio_original:.2f} USD'
                else:
                    precio_mostrar = f'{precio_original:.2f} ARS'
            else:
                # Usar precio del stock (comportamiento anterior)
                precios = mapa_precios.get(nombre, {})
                costo_usd = precios.get('costo_kilo_usd', 0)
                valor_dolar = precios.get('valor_dolar', 0)
                costo_total = kg * costo_usd * valor_dolar
                precio_mostrar = f'{costo_usd:.2f} USD'
                moneda_original = 'USD'
            
            self.tree_materiales_costos.insert('', 'end', values=(
                nombre,
                f'{kg:.3f}',
                precio_mostrar,
                moneda_original,
                f'${costo_total:.2f}'
            ))
        
        self.lbl_peso_total_costos.config(text=f'Peso total/unidad: {peso_total:.3f} kg')
    
    def quitar_material_costos(self):
        """Quitar material seleccionado de la lista temporal de costos"""
        selection = self.tree_materiales_costos.selection()
        if not selection:
            messagebox.showwarning('Advertencia', 'Seleccione un material para quitar.')
            return
        item = selection[0]
        index = self.tree_materiales_costos.index(item)
        if 0 <= index < len(self.materiales_costos_form):
            self.materiales_costos_form.pop(index)
            self.actualizar_tabla_materiales_costos()
    
    def editar_iibb_item(self, event):
        """Abrir pop-up para editar el IIBB del item seleccionado"""
        # Obtener el item seleccionado
        selection = self.tree_costos.selection()
        if not selection:
            messagebox.showwarning('Advertencia', 'Por favor seleccione un item para editar su IIBB.')
            return
        
        # Obtener el índice del item
        item = selection[0]
        index = self.tree_costos.index(item)
        
        if 0 <= index < len(self.items_costos):
            producto = self.items_costos[index]
            
            # Crear pop-up para editar IIBB
            popup = tk.Toplevel(self.root)
            popup.title(f'Editar IIBB - {producto.nombre}')
            popup.geometry('400x200')
            popup.resizable(False, False)
            
            # Centrar el pop-up
            popup.transient(self.root)
            popup.grab_set()
            
            # Frame principal
            main_frame = tk.Frame(popup, bg=STYLE_CONFIG['bg_color'], padx=20, pady=20)
            main_frame.pack(fill='both', expand=True)
            
            # Título
            title_label = tk.Label(main_frame, text=f'Editar IIBB para TODOS los items\n(Seleccionado: {producto.nombre})', 
                                 font=('Arial', 12, 'bold'), fg=STYLE_CONFIG['primary_color'], 
                                 bg=STYLE_CONFIG['bg_color'], justify='center')
            title_label.pack(pady=(0, 20))
            
            # Frame para el input
            input_frame = tk.Frame(main_frame, bg=STYLE_CONFIG['bg_color'])
            input_frame.pack(pady=10)
            
            # Label y Entry para IIBB
            tk.Label(input_frame, text='🏛️ IIBB %:', font=('Arial', 10, 'bold'), 
                    fg=STYLE_CONFIG['text_color'], bg=STYLE_CONFIG['bg_color']).pack(side='left', padx=(0, 10))
            
            iibb_var = tk.StringVar(value=str(producto.iibb_porcentaje))
            iibb_entry = tk.Entry(input_frame, textvariable=iibb_var, width=15, font=('Arial', 10), 
                                relief='solid', borderwidth=1, validate='key', validatecommand=self.vcmd_decimal)
            iibb_entry.pack(side='left')
            iibb_entry.focus()
            iibb_entry.select_range(0, tk.END)
            
            # Nota informativa
            info_label = tk.Label(main_frame, text='⚠️ Este cambio se aplicará a TODOS los items de la planilla', 
                                font=('Arial', 9), fg=STYLE_CONFIG['warning_color'], 
                                bg=STYLE_CONFIG['bg_color'], justify='center')
            info_label.pack(pady=(10, 0))
            
            # Frame para botones
            btn_frame = tk.Frame(main_frame, bg=STYLE_CONFIG['bg_color'])
            btn_frame.pack(pady=20)
            
            def guardar_iibb():
                """Guardar el nuevo porcentaje de IIBB para TODOS los items"""
                try:
                    nuevo_porcentaje = float(iibb_var.get())
                    if nuevo_porcentaje < 0 or nuevo_porcentaje > 100:
                        messagebox.showerror('Error', 'El porcentaje de IIBB debe estar entre 0 y 100.')
                        return
                    
                    # Confirmar que se aplicará a todos los items
                    if messagebox.askyesno('Confirmar Cambio Masivo', 
                                         f'¿Está seguro de aplicar {nuevo_porcentaje}% de IIBB a TODOS los {len(self.items_costos)} items de la planilla?'):
                        
                        # Actualizar TODOS los productos
                        items_actualizados = 0
                        for item in self.items_costos:
                            # Actualizar el porcentaje de IIBB del producto
                            item.iibb_porcentaje = nuevo_porcentaje
                            
                            # Recalcular IIBB unitario y total
                            if item.precio_venta and item.precio_venta > 0:
                                item.iibb_unitario = item.precio_venta * nuevo_porcentaje / 100
                                item.total_iibb = item.iibb_unitario * item.cantidad_fabricar if item.cantidad_fabricar else 0
                            else:
                                item.iibb_unitario = 0
                                item.total_iibb = 0
                            
                            # Recalcular rentabilidad neta
                            if item.precio_venta is not None:
                                item.rentabilidad_neta = item.precio_venta - item.costo_base_unitario - item.iibb_unitario
                                item.rentabilidad_neta_total = item.rentabilidad_neta * item.cantidad_fabricar if item.cantidad_fabricar else None
                            
                            items_actualizados += 1
                        
                        # Actualizar tabla y guardar
                        self.actualizar_tabla_costos()
                        self.save_state()
                        
                        # Cerrar pop-up
                        popup.destroy()
                        messagebox.showinfo('Actualización Completada', 
                                           f'Se actualizó el IIBB a {nuevo_porcentaje}% en {items_actualizados} items de la planilla.')
                    
                except ValueError:
                    messagebox.showerror('Error', 'Por favor ingrese un número válido para el porcentaje de IIBB.')
            
            def cancelar():
                """Cancelar y cerrar pop-up"""
                popup.destroy()
            
            # Botones
            ttk.Button(btn_frame, text='💾 Guardar', command=guardar_iibb, 
                      style='Action.TButton').pack(side='left', padx=5)
            ttk.Button(btn_frame, text='❌ Cancelar', command=cancelar, 
                      style='Danger.TButton').pack(side='left', padx=5)
            
            # Configurar Enter para guardar
            popup.bind('<Return>', lambda e: guardar_iibb())
            popup.bind('<Escape>', lambda e: cancelar())

    def limpiar_formulario_costos(self, mostrar_alerta=True):
        """Limpiar todos los campos del formulario de costos"""
        self.costos_familia_var.set('')
        self.costos_medida_var.set('')
        self.costos_caracteristica_var.set('')
        self.costos_venta_var.set('')
        self.costos_venta_moneda_var.set('ARS')
        self.costos_iibb_porcentaje_var.set('')
        self.costos_descuento_var.set('')
        self.costos_cant_fab_var.set('')
        self.costos_cant_hora_var.set('')
        self.costos_dolar_var.set('')
        self.materiales_costos_form = []
        self.actualizar_tabla_materiales_costos()
        if mostrar_alerta:
            messagebox.showinfo('Formulario limpio', 'Todos los campos han sido limpiados.')
    
    def _obtener_valor_seguro(self, valor, valor_default=''):
        """Obtener valor de forma segura, usando valor por defecto si está vacío o es NaN"""
        import pandas as pd
        
        if pd.isna(valor) or valor == '' or valor is None or str(valor).lower() == 'nan':
            return valor_default
        return str(valor)
    
    def _obtener_numero_seguro(self, valor, valor_default=0):
        """Obtener número de forma segura, usando valor por defecto si está vacío o es NaN"""
        import pandas as pd
        
        if pd.isna(valor) or valor == '' or valor is None or str(valor).lower() == 'nan':
            return valor_default
        
        try:
            return float(valor)
        except (ValueError, TypeError):
            return valor_default

    def _obtener_valor_dolar_usado_producto(self, producto):
        """Obtener el valor del dólar que se usó originalmente para este producto"""
        # Si el producto tiene materiales en USD, buscar el valor del dólar usado
        for material in producto.materiales or []:
            if material.get('moneda_original') == 'USD':
                # Buscar en el mapa de precios el valor del dólar usado
                nombre_material = material.get('nombre')
                if nombre_material in self._mapa_precios_mp():
                    precio_info = self._mapa_precios_mp()[nombre_material]
                    return precio_info.get('valor_dolar', 0)
        
        # Si no encontramos materiales en USD, usar el valor actual del formulario
        return self._obtener_valor_dolar_costos()
    
    def _obtener_valor_dolar_costos(self, precio_dolar_planilla=None):
        """Obtener el valor del dólar del formulario de costos o de la planilla"""
        # Si se pasa precio de la planilla (importación), usarlo directamente
        if precio_dolar_planilla is not None:
            return precio_dolar_planilla if precio_dolar_planilla > 0 else 0.0
        
        # Si no se pasa precio de la planilla (entrada manual), usar el del formulario
        try:
            valor_str = self.costos_dolar_var.get().strip()
            if not valor_str:
                return 0.0
            valor = float(valor_str)
            return valor if valor > 0 else 0.0
        except (ValueError, AttributeError):
            return 0.0

    def _detectar_moneda_material(self, precio):
        """Detectar automáticamente si un precio de material está en USD o ARS basándose en el rango de valores"""
        try:
            precio_num = float(precio)
            if precio_num == 0:
                return 'ARS'  # Precio 0 se considera ARS por defecto
            
            # Si el precio está en un rango típico de USD (generalmente menor a 100)
            # y no es un precio típico de ARS (mayor a 500)
            if precio_num < 100 and precio_num > 0:
                return 'USD'
            elif precio_num >= 500:
                return 'ARS'
            else:
                # Para valores intermedios, usar heurística más específica
                if precio_num < 50:
                    return 'USD'
                else:
                    return 'ARS'
        except (ValueError, TypeError):
            return 'ARS'

    def _detectar_columnas_materiales(self, df):
        """Detectar automáticamente las columnas de materiales en diferentes formatos"""
        materiales_detectados = {}
        
        # Buscar columnas que contengan "Material" o patrones similares
        columnas_materiales = []
        for col in df.columns:
            col_lower = col.lower()
            if any(keyword in col_lower for keyword in ['material', 'materia', 'mp']):
                columnas_materiales.append(col)
        
        # Si no encontramos columnas con "Material", buscar por contenido
        if not columnas_materiales:
            # Buscar columnas que contengan valores típicos de materiales (precios bajos en USD)
            for col in df.columns:
                try:
                    # Verificar si la columna contiene valores numéricos en rango de materiales
                    valores_numericos = []
                    for val in df[col].dropna():
                        try:
                            num_val = float(str(val).replace(',', '.'))
                            if 0.1 <= num_val <= 50:  # Rango típico de materiales en USD
                                valores_numericos.append(num_val)
                        except:
                            continue
                    
                    # Si más del 50% de los valores están en el rango de materiales
                    if len(valores_numericos) > len(df) * 0.3:
                        columnas_materiales.append(col)
                except:
                    continue
        
        # Procesar las columnas detectadas
        for col in columnas_materiales:
            # Intentar diferentes formatos
            if '_' in col:
                # Formato: Material_1_Precio, Material_1_Cantidad, etc.
                parts = col.split('_')
                if len(parts) >= 3:
                    material_num = parts[1]
                    material_type = '_'.join(parts[2:])
                    
                    if material_num not in materiales_detectados:
                        materiales_detectados[material_num] = {}
                    materiales_detectados[material_num][material_type] = col
            else:
                # Formato simple: buscar columnas relacionadas
                col_lower = col.lower()
                if any(keyword in col_lower for keyword in ['precio', 'costo', 'valor']):
                    # Asumir que es Material_1_Precio
                    if '1' not in materiales_detectados:
                        materiales_detectados['1'] = {}
                    materiales_detectados['1']['Precio'] = col
                elif any(keyword in col_lower for keyword in ['cantidad', 'kg', 'peso']):
                    # Asumir que es Material_1_Cantidad
                    if '1' not in materiales_detectados:
                        materiales_detectados['1'] = {}
                    materiales_detectados['1']['Cantidad'] = col
        
        return materiales_detectados

    def _parsear_precio_complejo(self, valor):
        """Parsear precios con formatos complejos como '430 u$s 11.40', '7,2 USD', etc."""
        import pandas as pd
        
        if pd.isna(valor) or str(valor).strip() == '' or str(valor).lower() == 'nan':
            return 0, 'ARS'
        
        valor_str = str(valor).strip()
        
        # Caso 1: Formato "430 u$s 11.40" (cantidad + precio en USD)
        if 'u$s' in valor_str.lower():
            try:
                # Dividir por 'u$s' y tomar la segunda parte como precio
                parts = valor_str.split('u$s')
                if len(parts) >= 2:
                    precio_str = parts[1].strip()
                    # Limpiar y convertir el precio
                    precio_str = precio_str.replace(',', '.')  # Convertir comas a puntos
                    precio = float(precio_str)
                    return precio, 'USD'
            except (ValueError, IndexError):
                pass
        
        # Caso 2: Formato "7,2 USD" o "7.2 USD"
        if valor_str.upper().endswith(' USD'):
            try:
                precio_str = valor_str[:-4].strip()  # Quitar " USD"
                precio_str = precio_str.replace(',', '.')  # Convertir comas a puntos
                precio = float(precio_str)
                return precio, 'USD'
            except ValueError:
                pass
        
        # Caso 3: Formato "1345,33 ARS" o "1345.33 ARS"
        if valor_str.upper().endswith(' ARS'):
            try:
                precio_str = valor_str[:-4].strip()  # Quitar " ARS"
                precio_str = precio_str.replace(',', '.')  # Convertir comas a puntos
                precio = float(precio_str)
                return precio, 'ARS'
            except ValueError:
                pass
        
        # Caso 4: Solo número (asumir ARS por defecto)
        try:
            precio_str = valor_str.replace(',', '.')  # Convertir comas a puntos
            precio = float(precio_str)
            return precio, 'ARS'
        except ValueError:
            pass
        
        # Si no se puede parsear, retornar 0
        return 0, 'ARS'

    def importar_datos_costos(self):
        """Importar datos desde archivo Excel para la planilla de costos"""
        from tkinter import filedialog
        import pandas as pd
        import datetime
        
        # Configurar logging
        log_filename = f"importacion_costos_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(log_filename, encoding='utf-8'),
                logging.StreamHandler()
            ]
        )
        
        logging.info("=== INICIO DE IMPORTACION DE COSTOS ===")
        
        # Abrir diálogo para seleccionar archivo
        archivo = filedialog.askopenfilename(
            title="Seleccionar archivo de datos",
            filetypes=[
                ("Archivos Excel", "*.xlsx *.xls"),
                ("Archivos CSV", "*.csv"),
                ("Todos los archivos", "*.*")
            ]
        )
        
        if not archivo:
            return
        
        try:
            # Leer archivo Excel
            if archivo.endswith(('.xlsx', '.xls')):
                df = pd.read_excel(archivo)
            elif archivo.endswith('.csv'):
                df = pd.read_csv(archivo)
            else:
                messagebox.showerror('Error', 'Formato de archivo no soportado.')
                return
            
            if df.empty:
                messagebox.showerror('Error', 'El archivo está vacío.')
                return
            
            # Verificar columnas requeridas
            columnas_requeridas = ['Familia', 'Medida', 'Característica', 'Precio_Venta', 'Moneda_Precio', 
                                 'Cantidad_Fabricar', 'Cantidad_Hora', 'IIBB_Porcentaje', 'Precio_Dolar']
            
            columnas_faltantes = [col for col in columnas_requeridas if col not in df.columns]
            if columnas_faltantes:
                messagebox.showerror('Error', f'Faltan las siguientes columnas: {", ".join(columnas_faltantes)}')
                return
            
            # Procesar cada fila
            items_importados = 0
            items_reemplazados = 0
            errores = []
            logging.info(f"Total de filas a procesar: {len(df)}")
            
            for index, row in df.iterrows():
                try:
                    logging.info(f"Procesando fila {index + 1}: {row.get('Familia', 'Sin familia')}")
                    
                    # Limpiar formulario antes de cargar nuevos datos (sin alerta)
                    self.limpiar_formulario_costos(mostrar_alerta=False)
                    
                    # Cargar datos del producto (usando valores seguros)
                    self.costos_familia_var.set(self._obtener_valor_seguro(row.get('Familia', '')))
                    self.costos_medida_var.set(self._obtener_valor_seguro(row.get('Medida', '')))
                    self.costos_caracteristica_var.set(self._obtener_valor_seguro(row.get('Característica', '')))
                    
                    # Parsear precio de venta con formato complejo
                    precio_venta_raw = row.get('Precio_Venta', '')
                    precio_venta, moneda_precio = self._parsear_precio_complejo(precio_venta_raw)
                    logging.info(f"Precio raw: '{precio_venta_raw}' -> Parseado: {precio_venta} {moneda_precio}")
                    self.costos_venta_var.set(str(precio_venta))
                    self.costos_venta_moneda_var.set(moneda_precio)
                    
                    self.costos_cant_fab_var.set(str(self._obtener_numero_seguro(row.get('Cantidad_Fabricar', 0))))
                    self.costos_cant_hora_var.set(str(self._obtener_numero_seguro(row.get('Cantidad_Hora', 0))))
                    self.costos_iibb_porcentaje_var.set(str(self._obtener_numero_seguro(row.get('IIBB_Porcentaje', 0))))
                    self.costos_dolar_var.set(str(self._obtener_numero_seguro(row.get('Precio_Dolar', 0))))
                    
                    # Procesar materiales usando detección automática de columnas
                    materiales_detectados = self._detectar_columnas_materiales(df)
                    logging.info(f"Columnas de materiales detectadas: {materiales_detectados}")
                    
                    if materiales_detectados:
                        # Procesar cada material detectado
                        for material_num, material_info in materiales_detectados.items():
                            # Obtener datos del material desde las columnas detectadas
                            nombre_col = material_info.get('Nombre')
                            cantidad_col = material_info.get('Cantidad')
                            precio_col = material_info.get('Precio')
                            moneda_col = material_info.get('Moneda')
                            
                            # Si no hay columna de nombre, usar un nombre genérico
                            if nombre_col and pd.notna(row.get(nombre_col)):
                                nombre_material = str(row[nombre_col]).strip()
                            else:
                                nombre_material = f"Material {material_num}"
                            
                            # Obtener cantidad (usar 1 como default si no hay columna)
                            if cantidad_col and pd.notna(row.get(cantidad_col)):
                                cantidad = self._obtener_numero_seguro(row[cantidad_col], 1)
                            else:
                                cantidad = 1.0  # Default: 1 kg por unidad
                            
                            # Obtener precio
                            if precio_col and pd.notna(row.get(precio_col)):
                                precio_original = self._obtener_numero_seguro(row[precio_col], 0)
                            else:
                                precio_original = 0
                            
                            # Solo agregar si tiene precio válido
                            if precio_original > 0:
                                # Detectar moneda
                                if moneda_col and pd.notna(row.get(moneda_col)):
                                    moneda_original = str(row[moneda_col]).strip().upper()
                                else:
                                    moneda_original = self._detectar_moneda_material(precio_original)
                                
                                material = {
                                    'nombre': nombre_material,
                                    'kg_por_unidad': cantidad,
                                    'precio_original': precio_original,
                                    'moneda_original': moneda_original
                                }
                                
                                # Convertir precio a pesos si es necesario
                                if material['moneda_original'] == 'USD':
                                    precio_dolar = self._obtener_numero_seguro(row.get('Precio_Dolar', 0))
                                    if precio_dolar > 0:
                                        material['precio_manual'] = material['precio_original'] * precio_dolar
                                        logging.info(f"Material {nombre_material}: ${precio_original} USD convertido a ${material['precio_manual']} ARS")
                                    else:
                                        material['precio_manual'] = material['precio_original']
                                        logging.warning(f"Material {nombre_material}: precio en USD pero no hay valor del dólar")
                                else:
                                    material['precio_manual'] = material['precio_original']
                                
                                self.materiales_costos_form.append(material)
                                logging.info(f"Material agregado: {nombre_material} - {cantidad}kg - ${precio_original} {moneda_original}")
                    
                    # Si no se detectaron materiales automáticamente, usar el método anterior
                    if not materiales_detectados:
                        columnas_materiales = [col for col in df.columns if col.startswith('Material_')]
                        if columnas_materiales:
                            # Método anterior para compatibilidad
                            materiales_data = {}
                            for col in columnas_materiales:
                                valor = row[col]
                                if pd.notna(valor) and str(valor).strip() != '' and str(valor).lower() != 'nan':
                                    parts = col.split('_')
                                    if len(parts) >= 3:
                                        material_num = parts[1]
                                        material_type = '_'.join(parts[2:])
                                        
                                        if material_num not in materiales_data:
                                            materiales_data[material_num] = {}
                                        materiales_data[material_num][material_type] = valor
                            
                            # Procesar materiales del método anterior
                            for material_num, material_info in materiales_data.items():
                                nombre_material = self._obtener_valor_seguro(material_info.get('Nombre', ''))
                                if nombre_material:
                                    precio_original = self._obtener_numero_seguro(material_info.get('Precio', 0))
                                    moneda_explicita = self._obtener_valor_seguro(material_info.get('Moneda', ''))
                                    
                                    if moneda_explicita:
                                        moneda_original = moneda_explicita
                                    else:
                                        moneda_original = self._detectar_moneda_material(precio_original)
                                    
                                    material = {
                                        'nombre': nombre_material,
                                        'kg_por_unidad': self._obtener_numero_seguro(material_info.get('Cantidad', 1)),
                                        'precio_original': precio_original,
                                        'moneda_original': moneda_original
                                    }
                                    
                                    if material['moneda_original'] == 'USD':
                                        precio_dolar = self._obtener_numero_seguro(row.get('Precio_Dolar', 0))
                                        if precio_dolar > 0:
                                            material['precio_manual'] = material['precio_original'] * precio_dolar
                                        else:
                                            material['precio_manual'] = material['precio_original']
                                    else:
                                        material['precio_manual'] = material['precio_original']
                                    
                                    self.materiales_costos_form.append(material)
                    
                    # Actualizar tabla de materiales
                    self.actualizar_tabla_materiales_costos()
                    
                    # Agregar item a la planilla (sin mostrar alertas)
                    logging.info(f"Intentando agregar item {index + 1} a la planilla")
                    precio_dolar_planilla = self._obtener_numero_seguro(row.get('Precio_Dolar', 0))
                    logging.info(f"Item {index + 1}: Usando precio del dólar de la planilla = {precio_dolar_planilla}")
                    fue_reemplazo = self.agregar_item_costos_silencioso(precio_dolar_planilla)
                    if fue_reemplazo:
                        items_reemplazados += 1
                        logging.info(f"Item {index + 1} reemplazado exitosamente")
                    else:
                        items_importados += 1
                        logging.info(f"Item {index + 1} agregado exitosamente")
                    
                except Exception as e:
                    error_msg = f'Fila {index + 1}: {str(e)}'
                    errores.append(error_msg)
                    logging.error(f"Error en fila {index + 1}: {str(e)}")
                    continue
            
            # Mostrar solo un mensaje final
            logging.info(f"=== RESUMEN DE IMPORTACION ===")
            logging.info(f"Items nuevos agregados: {items_importados}")
            logging.info(f"Items duplicados reemplazados: {items_reemplazados}")
            logging.info(f"Errores encontrados: {len(errores)}")
            
            # Siempre guardar los datos en data.json después de la importación
            logging.info("Guardando datos en data.json")
            self.save_state()
            
            if items_importados > 0 or items_reemplazados > 0:
                # Actualizar la tabla en la interfaz
                logging.info("Actualizando tabla de costos en la interfaz")
                self.actualizar_tabla_costos()
                
                total_procesados = items_importados + items_reemplazados
                mensaje_resumen = f'Se procesaron {total_procesados} items:\n'
                mensaje_resumen += f'• {items_importados} items nuevos agregados\n'
                mensaje_resumen += f'• {items_reemplazados} items duplicados reemplazados\n\n'
                
                if errores:
                    logging.warning(f"Importación completada con {len(errores)} errores")
                    mensaje_resumen += f'Se encontraron {len(errores)} errores que fueron omitidos.\n\n'
                    mensaje_resumen += f'Ver log: {log_filename}'
                    messagebox.showinfo('Importación Completada', mensaje_resumen)
                else:
                    logging.info("Importación exitosa sin errores")
                    messagebox.showinfo('Importación Exitosa', mensaje_resumen)
            else:
                logging.error("No se pudo importar ningún item")
                messagebox.showerror('Error', 'No se pudo importar ningún item. Verifique el formato del archivo.')
            
        except Exception as e:
            logging.error(f"Error general en importación: {str(e)}")
            messagebox.showerror('Error', f'Error al importar archivo: {str(e)}')
    
    def agregar_item_costos_silencioso(self, precio_dolar_planilla=None):
        """Agregar un item a la planilla de costos sin mostrar alertas (para importación)"""
        
        familia = self.costos_familia_var.get().strip()
        medida = self.costos_medida_var.get().strip()
        caracteristica = self.costos_caracteristica_var.get().strip()
        venta = self.costos_venta_var.get()
        venta_moneda = self.costos_venta_moneda_var.get()
        iibb_porcentaje = self.costos_iibb_porcentaje_var.get()
        cant_fab = self.costos_cant_fab_var.get()
        cant_hora = self.costos_cant_hora_var.get()
        
        logging.info(f"Procesando item: {familia} - {medida} - {caracteristica}")
        
        # Calcular peso total de materiales
        peso_total = sum(float(m.get('kg_por_unidad', 0)) for m in self.materiales_costos_form)
        
        # Validar que se completen los campos mínimos requeridos
        if not familia:
            logging.error("Familia vacía - item rechazado")
            raise ValueError('Complete al menos la Familia.')
        
        # Usar valores por defecto para campos que pueden estar vacíos
        medida = medida if medida else 'Sin Medida'
        caracteristica = caracteristica if caracteristica else 'Sin Característica'
        
        # Usar valores por defecto para campos numéricos si están vacíos
        cant_fab = cant_fab if cant_fab else '0'
        cant_hora = cant_hora if cant_hora else '0'
        
        logging.info(f"Campos procesados - Familia: {familia}, Medida: {medida}, Característica: {caracteristica}")
        
        # Procesar precio de venta
        precio_venta_final = 0  # Valor por defecto si no hay precio
        if venta:
            try:
                precio_venta = float(venta)
                if precio_venta < 0:  # Permitir precio 0, solo rechazar precios negativos
                    raise ValueError('El precio de venta no puede ser negativo.')
                
                # Si es USD, convertir a pesos usando el precio del dólar de la planilla
                if venta_moneda == 'USD':
                    precio_dolar_actual = precio_dolar_planilla if precio_dolar_planilla and precio_dolar_planilla > 0 else 0
                    if precio_dolar_actual > 0:
                        precio_venta_final = precio_venta * precio_dolar_actual
                        logging.info(f"Precio en USD convertido: ${precio_venta} USD × {precio_dolar_actual} = ${precio_venta_final} ARS")
                    else:
                        precio_venta_final = precio_venta  # Mantener en USD si no hay precio del dólar
                        print(f"Precio en USD pero no hay precio del dólar disponible: ${precio_venta} USD")
                else:
                    precio_venta_final = precio_venta  # Ya está en pesos
                    print(f"Precio en ARS: ${precio_venta_final}")
            except ValueError as e:
                raise e
        else:
            print(f"No hay precio de venta para {nombre}, usando precio por defecto: $0")
        
        # Crear el nombre combinado para el producto
        nombre = f"{familia} - {medida} - {caracteristica}"
        
        # Si no hay materiales, agregar un material por defecto
        if not self.materiales_costos_form:
            material_por_defecto = {
                'nombre': 'Material por defecto',
                'kg_por_unidad': 0.0,
                'precio_original': 0.0,
                'moneda_original': 'ARS',
                'precio_manual': 0.0
            }
            self.materiales_costos_form.append(material_por_defecto)
        
        # Obtener media valor hora efectivo empleados
        valor_hora_ajustado_promedio = 0
        if self.empleados:
            valor_hora_ajustado_promedio = sum(e.valor_hora_efectivo for e in self.empleados) / len(self.empleados)
        
        try:
            logging.info(f"Creando producto: {nombre}")
            # Obtener valor del dólar (priorizando el de la planilla)
            valor_dolar_costos = self._obtener_valor_dolar_costos(precio_dolar_planilla)
            logging.info(f"Producto {nombre}: Usando precio del dólar = {valor_dolar_costos}")
            
            prod = Producto(
                nombre,
                peso_total,
                precio_venta_final if precio_venta_final else None,
                cant_fab,
                cant_hora,
                valor_hora_ajustado_promedio,
                materiales=list(self.materiales_costos_form),
                precios_materiales=self._mapa_precios_mp(),
                iibb_porcentaje=iibb_porcentaje,
                moneda_precio=venta_moneda
            )
            
            # Actualizar precios de materiales con el valor del dólar (priorizando el de la planilla)
            prod.actualizar_precios_materiales(self._mapa_precios_mp(), valor_dolar_costos)
            
            logging.info(f"Producto creado exitosamente: {nombre}")
        except Exception as e:
            logging.error(f"Error al crear producto {nombre}: {str(e)}")
            raise ValueError(f'Error en los datos: {e}')
        
        # Verificar si ya existe un item con la misma familia, medida y característica
        item_duplicado_index = None
        for i, item_existente in enumerate(self.items_costos):
            # Extraer familia, medida y característica del nombre existente
            partes_existente = item_existente.nombre.split(' - ')
            if len(partes_existente) >= 3:
                familia_existente = partes_existente[0]
                medida_existente = partes_existente[1]
                caracteristica_existente = ' - '.join(partes_existente[2:])  # En caso de que característica tenga guiones
                
                if (familia_existente == familia and 
                    medida_existente == medida and 
                    caracteristica_existente == caracteristica):
                    item_duplicado_index = i
                    break
        
        if item_duplicado_index is not None:
            # Reemplazar item duplicado
            self.items_costos[item_duplicado_index] = prod
            logging.info(f"Item duplicado reemplazado en posición {item_duplicado_index}: {nombre}")
            return True  # Indica que fue un reemplazo
        else:
            # Agregar nuevo item
            self.items_costos.append(prod)
            logging.info(f"Item agregado a la lista de costos. Total items: {len(self.items_costos)}")
            return False  # Indica que fue un nuevo item
        
        # Actualizar tabla y totales
        self.actualizar_tabla_costos()
        
        # Limpiar formulario (sin alerta para importación)
        self.limpiar_formulario_costos(mostrar_alerta=False)
    
    def agregar_item_costos(self):
        """Agregar un item a la planilla de costos (sin afectar stock)"""
        familia = self.costos_familia_var.get().strip()
        medida = self.costos_medida_var.get().strip()
        caracteristica = self.costos_caracteristica_var.get().strip()
        venta = self.costos_venta_var.get()
        venta_moneda = self.costos_venta_moneda_var.get()
        iibb_porcentaje = self.costos_iibb_porcentaje_var.get()
        descuento_str = self.costos_descuento_var.get()
        cant_fab = self.costos_cant_fab_var.get()
        cant_hora = self.costos_cant_hora_var.get()
        
        # Procesar descuento
        descuento_pct = 0.0
        if descuento_str:
            try:
                descuento_pct = float(descuento_str)
                if descuento_pct < 0:
                    descuento_pct = 0.0
                elif descuento_pct > 100:
                    descuento_pct = 100.0
            except ValueError:
                descuento_pct = 0.0
        
        # Calcular peso total de materiales
        peso_total = sum(float(m.get('kg_por_unidad', 0)) for m in self.materiales_costos_form)
        
        # Validar que se completen los campos mínimos requeridos
        if not familia:
            messagebox.showerror('Error', 'Complete al menos la Familia.')
            return
        
        # Usar valores por defecto para campos opcionales
        if not medida:
            medida = 'Sin Medida'
        if not caracteristica:
            caracteristica = 'Sin Característica'
        if not cant_fab:
            cant_fab = '0'
        if not cant_hora:
            cant_hora = '0'
        
        # Procesar precio de venta
        precio_venta_final = 0  # Valor por defecto si no hay precio
        if venta:
            try:
                precio_venta = float(venta)
                if precio_venta <= 0:
                    messagebox.showerror('Error', 'El precio de venta debe ser mayor a 0.')
                    return
                
                # Si es USD, convertir a pesos usando el precio del dólar del formulario
                if venta_moneda == 'USD':
                    valor_dolar_form = self._obtener_valor_dolar_costos()
                    if valor_dolar_form > 0:
                        precio_venta_final = precio_venta * valor_dolar_form
                        logging.info(f"Precio en USD convertido: ${precio_venta} USD × {valor_dolar_form} = ${precio_venta_final} ARS")
                    else:
                        precio_venta_final = precio_venta  # Mantener en USD si no hay precio del dólar
                        print(f"Precio en USD pero no hay precio del dólar disponible: ${precio_venta} USD")
                else:
                    precio_venta_final = precio_venta  # Ya está en pesos
                    print(f"Precio en ARS: ${precio_venta_final}")
                    
            except ValueError:
                messagebox.showerror('Error', 'El precio de venta debe ser un número válido.')
                return
        else:
            # Si no hay precio de venta, usar 0 como valor por defecto
            precio_venta_final = 0
        
        # Crear el nombre combinado para el producto
        nombre = f"{familia} - {medida} - {caracteristica}"
        
        # Los materiales son opcionales - si no hay materiales, usar valores por defecto
        if not self.materiales_costos_form:
            # Agregar un material por defecto con costo 0
            self.materiales_costos_form = [{
                'nombre': 'Material por defecto',
                'kg_por_unidad': 0.0,
                'precio_manual': 0.0,
                'precio_original': 0.0,
                'moneda_original': 'ARS'
            }]
        
        # Obtener media valor hora efectivo empleados
        valor_hora_ajustado_promedio = 0
        if self.empleados:
            valor_hora_ajustado_promedio = sum(e.valor_hora_efectivo for e in self.empleados) / len(self.empleados)
        
        try:
            # Obtener valor del dólar del formulario de costos (para entrada manual)
            # Si estamos editando un producto que tiene USD, usar su valor del dólar específico
            if self.editando_costos_index is not None:
                producto_original = self.items_costos[self.editando_costos_index]
                moneda_original = getattr(producto_original, 'moneda_precio', 'ARS')
                if moneda_original == 'USD':
                    valor_dolar_costos = self._obtener_valor_dolar_usado_producto(producto_original)
                else:
                    valor_dolar_costos = self._obtener_valor_dolar_costos()
            else:
                valor_dolar_costos = self._obtener_valor_dolar_costos()
            
            prod = Producto(
                nombre,
                peso_total,
                precio_venta_final if precio_venta_final else None,
                cant_fab,
                cant_hora,
                valor_hora_ajustado_promedio,
                materiales=list(self.materiales_costos_form),
                precios_materiales=self._mapa_precios_mp(),
                iibb_porcentaje=iibb_porcentaje,
                moneda_precio=venta_moneda
            )
            
            # Agregar descuento como atributo del producto
            prod.descuento_porcentaje = descuento_pct
            
            # Actualizar precios de materiales con el valor del dólar del formulario
            prod.actualizar_precios_materiales(self._mapa_precios_mp(), valor_dolar_costos)
        except Exception as e:
            messagebox.showerror('Error', f'Error en los datos: {e}')
            return
        
        # Detectar si estamos en modo edición
        if self.editando_costos_index is not None:
            # Sobrescribir el item en la posición guardada
            self.items_costos[self.editando_costos_index] = prod
            self.editando_costos_index = None  # Limpiar el índice de edición
        else:
            # Agregar nuevo item
            self.items_costos.append(prod)
        
        self.actualizar_tabla_costos()
        self.save_state()
        
        # Mostrar mensaje de éxito
        messagebox.showinfo('Éxito', f'Item "{nombre}" agregado a la planilla de costos.')
        
        # Limpiar formulario
        self.costos_familia_var.set('')
        self.costos_medida_var.set('')
        self.costos_caracteristica_var.set('')
        self.costos_venta_var.set('')
        self.costos_cant_fab_var.set('')
        self.costos_cant_hora_var.set('')
        self.costos_iibb_porcentaje_var.set('')
        self.costos_descuento_var.set('')
        self.costos_dolar_var.set('')
        self.costos_venta_moneda_var.set('ARS')
        self.materiales_costos_form = []
        self.actualizar_tabla_materiales_costos()
        
    def manejar_doble_clic_costos(self, event):
        """Manejar doble clic en la tabla de costos - editar item o IIBB según la columna"""
        # Obtener el item y la columna donde se hizo clic
        item = event.widget.identify('item', event.x, event.y)
        column = event.widget.identify('column', event.x, event.y)
        
        if not item:
            return
        
        # Si se hizo clic en la columna de IIBB %, editar IIBB
        if column == '#7':  # Columna "IIBB %"
            self.editar_iibb_item(event)
        else:
            # Para cualquier otra columna, editar el item completo
            self.editar_item_costos(event)
    
    def editar_item_costos(self, event=None):
        """Editar item seleccionado de la planilla de costos"""
        # Si se llama desde doble clic, usar el item del evento
        if event:
            item = event.widget.selection()[0] if event.widget.selection() else None
            if not item:
                return
            index = event.widget.index(item)
        else:
            # Si se llama desde botón, usar la selección actual
            selection = self.tree_costos.selection()
            if not selection:
                messagebox.showwarning('Advertencia', 'Seleccione un item para editar.')
                return
            item = selection[0]
            index = self.tree_costos.index(item)
        
        if 0 <= index < len(self.items_costos):
            producto = self.items_costos[index]
            
            # Llenar los campos con los datos del item seleccionado
            # Extraer componentes del nombre (formato: "Familia - Medida - Característica")
            nombre_parts = producto.nombre.split(' - ')
            if len(nombre_parts) >= 3:
                self.costos_familia_var.set(nombre_parts[0])
                self.costos_medida_var.set(nombre_parts[1])
                self.costos_caracteristica_var.set(' - '.join(nombre_parts[2:]))
            else:
                # Fallback para nombres antiguos
                self.costos_familia_var.set(producto.nombre)
                self.costos_medida_var.set('')
                self.costos_caracteristica_var.set('')
            
            # Cargar precio de venta con su moneda original
            if producto.precio_venta:
                # Si el producto tiene moneda_precio guardada, usarla
                moneda_original = getattr(producto, 'moneda_precio', 'ARS')
                self.costos_venta_moneda_var.set(moneda_original)
                
                # Si es USD, mostrar el precio original en USD, no el convertido
                if moneda_original == 'USD':
                    # Necesitamos calcular el precio original en USD
                    # Buscar el valor del dólar usado para este producto
                    valor_dolar_usado = self._obtener_valor_dolar_usado_producto(producto)
                    if valor_dolar_usado > 0:
                        precio_usd_original = producto.precio_venta / valor_dolar_usado
                        self.costos_venta_var.set(str(precio_usd_original))
                    else:
                        self.costos_venta_var.set(str(producto.precio_venta))
                else:
                    self.costos_venta_var.set(str(producto.precio_venta))
            else:
                self.costos_venta_var.set('')
                self.costos_venta_moneda_var.set('ARS')
            self.costos_cant_fab_var.set(str(producto.cantidad_fabricar))
            self.costos_cant_hora_var.set(str(producto.cantidad_por_hora))
            
            # Cargar IIBB porcentaje
            if producto.iibb_porcentaje:
                self.costos_iibb_porcentaje_var.set(str(producto.iibb_porcentaje))
            else:
                self.costos_iibb_porcentaje_var.set('')
            
            # Cargar descuento
            descuento = getattr(producto, 'descuento_porcentaje', 0.0)
            if descuento and descuento > 0:
                self.costos_descuento_var.set(str(descuento))
            else:
                self.costos_descuento_var.set('')
            
            # Cargar materiales
            self.materiales_costos_form = list(producto.materiales or [])
            self.actualizar_tabla_materiales_costos()
            
            # Si el producto tiene USD, cargar el valor del dólar específico usado
            moneda_original = getattr(producto, 'moneda_precio', 'ARS')
            if moneda_original == 'USD':
                valor_dolar_usado = self._obtener_valor_dolar_usado_producto(producto)
                if valor_dolar_usado > 0:
                    self.costos_dolar_var.set(str(valor_dolar_usado))
            
            # Guardar el índice del item en edición (NO eliminarlo)
            self.editando_costos_index = index
            messagebox.showinfo('Modo Edición', f'Item "{producto.nombre}" cargado para edición.\nModifique los datos y presione "Actualizar Planilla" para guardar los cambios.')
    
    def eliminar_item_costos(self):
        """Eliminar items seleccionados de la planilla de costos (permite selección múltiple)"""
        selection = self.tree_costos.selection()
        if not selection:
            messagebox.showwarning('Advertencia', 'Seleccione uno o más items para eliminar.')
            return
        
        # Obtener índices de los items seleccionados (ordenados de mayor a menor para eliminar correctamente)
        indices = []
        nombres = []
        for item in selection:
            index = self.tree_costos.index(item)
            if 0 <= index < len(self.items_costos):
                indices.append(index)
                nombres.append(self.items_costos[index].nombre)
        
        if not indices:
            messagebox.showwarning('Advertencia', 'No hay items válidos seleccionados.')
            return
        
        # Ordenar índices de mayor a menor para eliminar correctamente
        indices.sort(reverse=True)
        
        # Confirmar eliminación
        if len(indices) == 1:
            mensaje = f'¿Está seguro de eliminar el item "{nombres[0]}"?'
        else:
            mensaje = f'¿Está seguro de eliminar los {len(indices)} items seleccionados?\n\nItems a eliminar:\n' + '\n'.join(f'• {nombre}' for nombre in nombres)
        
        respuesta = messagebox.askyesno('Confirmar eliminación', mensaje)
        if respuesta:
            # Eliminar items (de mayor a menor índice para no afectar los índices de los siguientes)
            for index in indices:
                self.items_costos.pop(index)
            
            self.actualizar_tabla_costos()
            self.save_state()
            
            if len(indices) == 1:
                messagebox.showinfo('Éxito', f'Item "{nombres[0]}" eliminado de la planilla.')
            else:
                messagebox.showinfo('Éxito', f'{len(indices)} items eliminados de la planilla.')
    
    def enviar_a_produccion(self):
        """Enviar item seleccionado de la planilla de costos a la pestaña de productos"""
        selection = self.tree_costos.selection()
        if not selection:
            messagebox.showwarning('Advertencia', 'Por favor seleccione un item para enviar a producción.')
            return
        
        item = selection[0]
        index = self.tree_costos.index(item)
        self.transferir_costos_a_produccion(index)
    
    def transferir_costos_a_produccion(self, index):
        """Transferir un item de la planilla de costos a la pestaña de productos"""
        if 0 <= index < len(self.items_costos):
            item_costos = self.items_costos[index]
            
            # Verificar si el producto ya existe en producción
            existe = False
            for prod in self.productos:
                if prod.nombre.lower() == item_costos.nombre.lower():
                    existe = True
                    break
            
            if existe:
                respuesta = messagebox.askyesno(
                    'Producto existente', 
                    f'El producto "{item_costos.nombre}" ya existe en Producción.\n\n¿Desea sobrescribirlo con los datos de Costos?'
                )
                if not respuesta:
                    return
                
                # Eliminar el producto existente
                self.productos = [p for p in self.productos if p.nombre.lower() != item_costos.nombre.lower()]
            
            # Obtener valor hora promedio actualizado
            valor_hora_ajustado_promedio = 0
            if self.empleados:
                valor_hora_ajustado_promedio = sum(e.valor_hora_efectivo for e in self.empleados) / len(self.empleados)
            
            # Crear una copia del producto para la pestaña de producción
            nuevo_producto = Producto(
                item_costos.nombre,
                item_costos.peso_unidad,
                item_costos.precio_venta,
                item_costos.cantidad_fabricar,
                item_costos.cantidad_por_hora,
                valor_hora_ajustado_promedio,
                materiales=list(item_costos.materiales or []),
                precios_materiales=self._mapa_precios_mp()
            )
            
            # Agregar a productos
            self.productos.append(nuevo_producto)
            
            # Actualizar vista y guardar
            self.actualizar_tabla_productos()
            self.actualizar_resumen_productos()
            self.save_state()
            
            # Cambiar a la pestaña de productos
            self.tabs.select(0)  # Pestaña de productos es la primera
            
            messagebox.showinfo(
                'Éxito', 
                f'¡Producto "{item_costos.nombre}" agregado a Producción!\n\nLa pestaña de Producción se ha abierto automáticamente.'
            )
    
    def actualizar_tabla_costos(self):
        """Actualizar la tabla de la planilla de costos y totales"""
        # Limpiar tabla
        for item in self.tree_costos.get_children():
            self.tree_costos.delete(item)
        
        # NO recalcular costos de productos existentes - mantener sus precios originales
        # Solo mostrar los datos tal como están guardados
        
        # Totales generales
        total_mp = 0.0
        total_mo = 0.0
        total_iibb = 0.0
        total_produccion = 0.0
        total_ingreso = 0.0
        total_rentabilidad = 0.0
        margenes = []
        
        # Iterar por cada item de la planilla
        for prod in self.items_costos:
            # Extraer componentes del nombre (formato: "Familia - Medida - Característica")
            nombre_parts = prod.nombre.split(' - ')
            if len(nombre_parts) >= 3:
                familia = nombre_parts[0]
                medida = nombre_parts[1]
                caracteristica = ' - '.join(nombre_parts[2:])
            else:
                # Fallback para nombres antiguos
                familia = prod.nombre
                medida = ''
                caracteristica = ''
            peso_unidad = prod.peso_unidad
            costo_mp_unidad = prod.costo_unitario_mp
            costo_mo_unidad = prod.costo_unitario_mano_obra
            iibb_unidad = prod.iibb_unitario
            costo_total_unidad = costo_mp_unidad + costo_mo_unidad + iibb_unidad
            
            # Mostrar precio de venta tal como está guardado (ya convertido a pesos)
            precio_venta = prod.precio_venta if prod.precio_venta is not None else 0
            moneda_precio = getattr(prod, 'moneda_precio', 'ARS')
            
            # Los precios ya están en pesos (convertidos durante la importación o entrada manual)
            # No necesitamos convertir nuevamente
            
            # Obtener descuento del producto (si existe)
            descuento_pct = getattr(prod, 'descuento_porcentaje', 0.0)
            if descuento_pct is None:
                descuento_pct = 0.0
            try:
                descuento_pct = float(descuento_pct)
            except (ValueError, TypeError):
                descuento_pct = 0.0
            
            # Calcular margen bruto (sin descuento)
            margen_bruto = precio_venta - costo_total_unidad if precio_venta > 0 else 0
            margen_pct_bruto = (margen_bruto / precio_venta * 100) if precio_venta > 0 else 0
            
            # Calcular precio con descuento
            descuento_monto = precio_venta * (descuento_pct / 100) if precio_venta > 0 and descuento_pct > 0 else 0
            precio_con_descuento = precio_venta - descuento_monto if precio_venta > 0 else 0
            
            # Calcular margen neto (con descuento)
            margen_neto = precio_con_descuento - costo_total_unidad if precio_con_descuento > 0 else 0
            margen_pct_neto = (margen_neto / precio_con_descuento * 100) if precio_con_descuento > 0 else 0
            
            cant_fabricar = prod.cantidad_fabricar
            
            # Totales por producto
            costo_total_mp = prod.costo_total_mp
            costo_total_mo = prod.incidencia_mano_obra
            costo_total_iibb = prod.total_iibb
            costo_total_prod = costo_total_mp + costo_total_mo + costo_total_iibb
            ingreso_total = precio_con_descuento * cant_fabricar if precio_con_descuento and cant_fabricar else 0
            rentabilidad_total = ingreso_total - costo_total_prod
            
            # Acumular totales generales
            total_mp += costo_total_mp
            total_mo += costo_total_mo
            total_iibb += costo_total_iibb
            total_produccion += costo_total_prod
            total_ingreso += ingreso_total
            total_rentabilidad += rentabilidad_total
            
            if precio_venta > 0:
                margenes.append(margen_pct_bruto)
            
            # Formatear precio de venta (siempre en pesos)
            if precio_venta > 0:
                precio_formateado = f'${precio_venta:.2f}'
            else:
                precio_formateado = '-'
            
            # Formatear descuento
            if descuento_pct > 0:
                descuento_formateado = f'{descuento_pct:.1f}%'
            else:
                descuento_formateado = '-'
            
            # Insertar en la tabla con formato (incluyendo nuevas columnas)
            self.tree_costos.insert('', 'end', values=(
                familia,
                medida,
                caracteristica,
                f'{peso_unidad:.3f}',
                f'${costo_mp_unidad:.2f}',
                f'${costo_mo_unidad:.2f}',
                f'{prod.iibb_porcentaje:.1f}%',
                f'${iibb_unidad:.2f}',
                f'${costo_total_unidad:.2f}',
                precio_formateado,
                f'${margen_bruto:.2f}' if precio_venta else '-',
                f'{margen_pct_bruto:.1f}%' if precio_venta else '-',
                descuento_formateado,
                f'${margen_neto:.2f}' if precio_venta else '-',
                f'{margen_pct_neto:.1f}%' if precio_venta else '-'
            ))
        
        # Actualizar labels de totales
        self.lbl_total_mp_costos.config(text=f'Total Materia Prima: ${total_mp:,.2f}')
        self.lbl_total_mo_costos.config(text=f'Total Mano de Obra: ${total_mo:,.2f}')
        self.lbl_total_iibb_costos.config(text=f'Total IIBB: ${total_iibb:,.2f}')
        self.lbl_total_prod_costos.config(text=f'Total Producción: ${total_produccion:,.2f}')
        self.lbl_total_ingreso_costos.config(text=f'Total Ingresos: ${total_ingreso:,.2f}')
        self.lbl_total_rentabilidad_costos.config(text=f'Rentabilidad Total: ${total_rentabilidad:,.2f}')
        
        # Calcular margen promedio
        margen_promedio = sum(margenes) / len(margenes) if margenes else 0
        self.lbl_margen_promedio_costos.config(text=f'Margen Promedio: {margen_promedio:.1f}%')
        
        # Cambiar color de rentabilidad según sea positiva o negativa
        if total_rentabilidad >= 0:
            self.lbl_total_rentabilidad_costos.config(fg=STYLE_CONFIG['success_color'])
        else:
            self.lbl_total_rentabilidad_costos.config(fg=STYLE_CONFIG['danger_color'])
        
        # Actualizar gráfico de torta
        self._actualizar_grafico_costos(total_mp, total_mo, total_rentabilidad)
    
    def _actualizar_grafico_costos(self, total_mp, total_mo, total_rentabilidad):
        """Actualizar el gráfico de torta de distribución de costos"""
        if not Figure or not hasattr(self, 'ax_costos') or not hasattr(self, 'canvas_costos'):
            return
        
        # Limpiar el gráfico anterior
        self.ax_costos.clear()
        
        # Si no hay datos, mostrar mensaje
        if total_mp == 0 and total_mo == 0 and total_rentabilidad == 0:
            self.ax_costos.text(0.5, 0.5, 'Agregue items\npara ver el gráfico', 
                              ha='center', va='center', fontsize=12, 
                              color=STYLE_CONFIG['text_light'])
            self.ax_costos.set_xlim(0, 1)
            self.ax_costos.set_ylim(0, 1)
            self.ax_costos.axis('off')
        else:
            # Preparar datos para el gráfico
            labels = []
            values = []
            colors = []
            
            if total_mp > 0:
                labels.append('Materia Prima')
                values.append(total_mp)
                colors.append('#ef4444')  # Rojo
            
            if total_mo > 0:
                labels.append('Mano de Obra')
                values.append(total_mo)
                colors.append('#f59e0b')  # Naranja
            
            if total_rentabilidad > 0:
                labels.append('Rentabilidad')
                values.append(total_rentabilidad)
                colors.append('#10b981')  # Verde
            elif total_rentabilidad < 0:
                labels.append('Pérdida')
                values.append(abs(total_rentabilidad))
                colors.append('#dc2626')  # Rojo oscuro
            
            # Crear el gráfico de torta
            if values:
                wedges, texts, autotexts = self.ax_costos.pie(
                    values,
                    labels=labels,
                    colors=colors,
                    autopct='%1.1f%%',
                    startangle=90,
                    textprops={'fontsize': 10, 'weight': 'bold'}
                )
                
                # Configurar colores de texto de porcentajes
                for autotext in autotexts:
                    autotext.set_color('white')
        
        # Redibujar el canvas
        try:
            self.canvas_costos.draw()
            self.canvas_costos.flush_events()
        except Exception as e:
            print(f"Error al actualizar gráfico: {e}")

    def _refresh_combo_compra_materiales(self):
        """Función mantenida para compatibilidad - ya no es necesaria con campos de texto libre"""
        pass

    def _refresh_combo_compra_productos(self):
        """Función mantenida para compatibilidad - ya no es necesaria con campos de texto libre"""
        pass

    def registrar_compra_materia_prima(self):
        """Registrar una compra de materia prima"""
        try:
            material = self.compra_mp_material_var.get().strip()
            cantidad_str = self.compra_mp_cantidad_var.get().strip()
            precio_str = self.compra_mp_precio_var.get().strip()
            proveedor = self.compra_mp_proveedor_var.get().strip()
            fecha = self.compra_mp_fecha_var.get().strip()
            moneda = self.compra_mp_moneda_var.get()
            valor_dolar_str = self.compra_mp_valor_dolar_var.get().strip()

            # Validaciones
            if not material:
                messagebox.showerror("Error", "Debe ingresar un material")
                return

            if not cantidad_str:
                messagebox.showerror("Error", "Debe ingresar una cantidad")
                return

            if not precio_str:
                messagebox.showerror("Error", "Debe ingresar un precio")
                return

            if not proveedor:
                messagebox.showerror("Error", "Debe ingresar un proveedor")
                return

            if not fecha:
                messagebox.showerror("Error", "Debe ingresar una fecha")
                return

            # Si es USD, validar valor del dólar
            valor_dolar = 1.0
            if moneda == 'USD':
                if not valor_dolar_str:
                    messagebox.showerror("Error", "Debe ingresar el valor del dólar para compras en USD")
                    return
                try:
                    valor_dolar = float(valor_dolar_str)
                    if valor_dolar <= 0:
                        messagebox.showerror("Error", "El valor del dólar debe ser mayor a 0")
                        return
                except ValueError:
                    messagebox.showerror("Error", "El valor del dólar debe ser numérico")
                return

            try:
                cantidad = float(cantidad_str)
                precio = float(precio_str)
            except ValueError:
                messagebox.showerror("Error", "La cantidad y precio deben ser números válidos")
                return

            if cantidad <= 0:
                messagebox.showerror("Error", "La cantidad debe ser mayor a 0")
                return

            if precio <= 0:
                messagebox.showerror("Error", "El precio debe ser mayor a 0")
                return

            # Convertir precio a ARS si es USD (precio unitario en ARS)
            precio_final = precio * valor_dolar if moneda == 'USD' else precio
            
            # Calcular total: precio_unitario × valor_dolar × cantidad
            total_compra = precio * valor_dolar * cantidad

            # Construir nombre completo con formato estándar
            familia = self.compra_mp_familia_var.get().strip()
            medida = self.compra_mp_medida_var.get().strip()
            caracteristica = self.compra_mp_caracteristica_var.get().strip()
            
            # Crear nombre completo en formato "Familia - Medida - Característica"
            # Si no hay campos opcionales, usar el material como familia
            if not familia:
                familia = material
            if not medida:
                medida = "N/A"
            if not caracteristica:
                caracteristica = "N/A"
            
            nombre_completo = f"{familia} - {medida} - {caracteristica}"

            # Buscar si el material ya existe en stock
            material_encontrado = False
            for mat in self.stock_mp:
                nombre_existente = mat.get('nombre') or mat.get('Nombre', '')
                if nombre_existente.lower() == nombre_completo.lower():
                    # Actualizar cantidad existente
                    kg_actual = float(mat.get('kg') or mat.get('Stock (kg)', 0))
                    mat['kg'] = kg_actual + cantidad
                    mat['nombre'] = nombre_completo
                    material_encontrado = True
                    break

            # Si no existe, agregarlo
            if not material_encontrado:
                self.stock_mp.append({
                    'nombre': nombre_completo,
                    'kg': cantidad,
                    'costo_kilo_usd': precio_final,
                    'valor_dolar': valor_dolar if moneda == 'USD' else 1.0
                })

            # Registrar ingreso automático en el log
            self.registrar_ingreso_automatico(familia, medida, caracteristica, cantidad, f"Compra de materia prima - Proveedor: {proveedor}")

            # Agregar a logs de compras: precio_unitario × valor_dolar × cantidad
            self.agregar_log_compra_mp(fecha, material, cantidad, precio_final, proveedor, total_compra)

            # Guardar datos
            self.save_state()

            # Actualizar tablas de stock
            self.actualizar_tabla_materiales_form()
            self.actualizar_tabla_materia_prima()

            messagebox.showinfo("Éxito", f"Compra de {cantidad} kg de {material} registrada correctamente")
            self.limpiar_formulario_compra_mp()

        except Exception as e:
            messagebox.showerror("Error", f"Error al registrar compra: {str(e)}")

    def registrar_compra_producto(self):
        """Registrar una compra de producto"""
        try:
            producto = self.compra_prod_producto_var.get().strip()
            cantidad_str = self.compra_prod_cantidad_var.get().strip()
            precio_str = self.compra_prod_precio_var.get().strip()
            proveedor = self.compra_prod_proveedor_var.get().strip()
            fecha = self.compra_prod_fecha_var.get().strip()
            moneda = self.compra_prod_moneda_var.get()
            valor_dolar_str = self.compra_prod_valor_dolar_var.get().strip()

            # Validaciones
            if not producto:
                messagebox.showerror("Error", "Debe ingresar un producto")
                return

            if not cantidad_str:
                messagebox.showerror("Error", "Debe ingresar una cantidad")
                return

            if not precio_str:
                messagebox.showerror("Error", "Debe ingresar un precio")
                return

            if not proveedor:
                messagebox.showerror("Error", "Debe ingresar un proveedor")
                return

            if not fecha:
                messagebox.showerror("Error", "Debe ingresar una fecha")
                return

            # Si es USD, validar valor del dólar
            valor_dolar = 1.0
            if moneda == 'USD':
                if not valor_dolar_str:
                    messagebox.showerror("Error", "Debe ingresar el valor del dólar para compras en USD")
                    return
                try:
                    valor_dolar = float(valor_dolar_str)
                    if valor_dolar <= 0:
                        messagebox.showerror("Error", "El valor del dólar debe ser mayor a 0")
                        return
                except ValueError:
                    messagebox.showerror("Error", "El valor del dólar debe ser numérico")
                return

            try:
                cantidad = int(float(cantidad_str))
                precio = float(precio_str)
            except ValueError:
                messagebox.showerror("Error", "La cantidad y precio deben ser números válidos")
                return

            if cantidad <= 0:
                messagebox.showerror("Error", "La cantidad debe ser mayor a 0")
                return

            if precio <= 0:
                messagebox.showerror("Error", "El precio debe ser mayor a 0")
                return

            # Convertir precio a ARS si es USD (precio unitario en ARS)
            precio_final = precio * valor_dolar if moneda == 'USD' else precio
            
            # Calcular total: precio_unitario × valor_dolar × cantidad
            total_compra = precio * valor_dolar * cantidad

            # Construir nombre completo con formato estándar
            familia = self.compra_prod_familia_var.get().strip()
            medida = self.compra_prod_medida_var.get().strip()
            caracteristica = self.compra_prod_caracteristica_var.get().strip()
            
            # Crear nombre completo en formato "Familia - Medida - Característica"
            # Si no hay campos opcionales, usar el producto como familia
            if not familia:
                familia = producto
            if not medida:
                medida = "N/A"
            if not caracteristica:
                caracteristica = "N/A"
            
            nombre_completo = f"{familia} - {medida} - {caracteristica}"

            # Buscar si el producto ya existe en stock
            producto_encontrado = False
            for prod in self.productos_reventa:
                nombre_existente = prod.get('nombre', '')
                if nombre_existente.lower() == nombre_completo.lower():
                    # Actualizar cantidad existente
                    cantidad_actual = int(prod.get('cantidad', 0))
                    prod['cantidad'] = cantidad_actual + cantidad
                    prod['nombre'] = nombre_completo
                    prod['costo_unitario'] = precio_final
                    prod['moneda'] = moneda
                    prod['valor_dolar'] = valor_dolar if moneda == 'USD' else None
                    prod['costo_unitario_final'] = precio_final
                    producto_encontrado = True
                    break

            # Si no existe, agregarlo
            if not producto_encontrado:
                self.productos_reventa.append({
                    'nombre': nombre_completo,
                    'cantidad': cantidad,
                    'costo_unitario': precio_final,
                    'otros_costos': 0,
                    'moneda': moneda,
                    'valor_dolar': valor_dolar if moneda == 'USD' else None,
                    'costo_unitario_final': precio_final,
                    'costo_total': total_compra
                })

            # Registrar ingreso automático en el log para productos de reventa
            self.registrar_ingreso_automatico_productos(familia, medida, caracteristica, cantidad, f"Compra de producto - Proveedor: {proveedor}")

            # Agregar a logs de compras: precio_unitario × valor_dolar × cantidad
            self.agregar_log_compra_prod(fecha, producto, cantidad, precio_final, proveedor, total_compra)

            # Guardar datos
            self.save_state()

            # Actualizar tablas de stock
            self.actualizar_tabla_productos()
            self.actualizar_tabla_stock_prod()

            messagebox.showinfo("Éxito", f"Compra de {cantidad} unidades de {producto} registrada correctamente")
            self.limpiar_formulario_compra_prod()

        except Exception as e:
            messagebox.showerror("Error", f"Error al registrar compra: {str(e)}")

    def limpiar_formulario_compra_mp(self):
        """Limpiar formulario de compra de materia prima"""
        self.compra_mp_material_var.set('')
        self.compra_mp_familia_var.set('')
        self.compra_mp_medida_var.set('')
        self.compra_mp_caracteristica_var.set('')
        self.compra_mp_cantidad_var.set('')
        self.compra_mp_precio_var.set('')
        self.compra_mp_proveedor_var.set('')
        self.compra_mp_fecha_var.set('')
        self.compra_mp_moneda_var.set('ARS')
        self.compra_mp_valor_dolar_var.set('')
        self._actualizar_campos_moneda_compra_mp()

    def limpiar_formulario_compra_prod(self):
        """Limpiar formulario de compra de productos"""
        self.compra_prod_producto_var.set('')
        self.compra_prod_familia_var.set('')
        self.compra_prod_medida_var.set('')
        self.compra_prod_caracteristica_var.set('')
        self.compra_prod_cantidad_var.set('')
        self.compra_prod_precio_var.set('')
        self.compra_prod_proveedor_var.set('')
        self.compra_prod_fecha_var.set('')
        self.compra_prod_moneda_var.set('ARS')
        self.compra_prod_valor_dolar_var.set('')
        self._actualizar_campos_moneda_compra_prod()


    def cargar_logs_compras_a_tablas(self):
        """Cargar logs de compras desde las variables a las tablas"""
        try:
            # Cargar logs de compras de materia prima
            if hasattr(self, 'tree_logs_compra_mp') and hasattr(self, 'logs_compra_mp'):
                # Limpiar tabla
                for item in self.tree_logs_compra_mp.get_children():
                    self.tree_logs_compra_mp.delete(item)
                
                # Cargar logs
                for log in self.logs_compra_mp:
                    self.tree_logs_compra_mp.insert('', 'end', values=(
                        log.get('fecha', ''),
                        log.get('material', ''),
                        f"{log.get('cantidad', 0):.1f}",
                        f"${log.get('precio', 0):.2f}",
                        log.get('proveedor', ''),
                        f"${log.get('total', 0):.2f}"
                    ))
            
            # Cargar logs de compras de productos
            if hasattr(self, 'tree_logs_compra_prod') and hasattr(self, 'logs_compra_prod'):
                # Limpiar tabla
                for item in self.tree_logs_compra_prod.get_children():
                    self.tree_logs_compra_prod.delete(item)
                
                # Cargar logs
                for log in self.logs_compra_prod:
                    self.tree_logs_compra_prod.insert('', 'end', values=(
                        log.get('fecha', ''),
                        log.get('producto', ''),
                        f"{log.get('cantidad', 0)}",
                        f"${log.get('precio', 0):.2f}",
                        log.get('proveedor', ''),
                        f"${log.get('total', 0):.2f}"
                    ))
                    
        except Exception as e:
            print(f"Error al cargar logs de compras a las tablas: {e}")
    
    def agregar_log_compra_mp(self, fecha, material, cantidad, precio, proveedor, total):
        """Agregar un registro al log de compras de materia prima"""
        try:
            if not hasattr(self, 'tree_logs_compra_mp'):
                return
            
            # Crear el registro del log
            log_entry = {
                'fecha': fecha,
                'material': material,
                'cantidad': cantidad,
                'precio': precio,
                'proveedor': proveedor,
                'total': total
            }
            
            # Agregar a la lista de logs
            self.logs_compra_mp.insert(0, log_entry)  # Insertar al inicio
            
            # Mantener solo los últimos 100 registros
            if len(self.logs_compra_mp) > 100:
                self.logs_compra_mp = self.logs_compra_mp[:100]
            
            # Insertar en la tabla de logs
            self.tree_logs_compra_mp.insert('', 0, values=(
                fecha,
                material,
                f"{cantidad:.1f}",
                f"${precio:.2f}",
                proveedor,
                f"${total:.2f}"
            ))
            
            # Mantener solo los últimos 100 registros en la tabla
            items = self.tree_logs_compra_mp.get_children()
            if len(items) > 100:
                self.tree_logs_compra_mp.delete(items[-1])
                
        except Exception as e:
            print(f"Error al agregar log de compra MP: {e}")

    def agregar_log_compra_prod(self, fecha, producto, cantidad, precio, proveedor, total):
        """Agregar un registro al log de compras de productos"""
        try:
            if not hasattr(self, 'tree_logs_compra_prod'):
                return
            
            # Crear el registro del log
            log_entry = {
                'fecha': fecha,
                'producto': producto,
                'cantidad': cantidad,
                'precio': precio,
                'proveedor': proveedor,
                'total': total
            }
            
            # Agregar a la lista de logs
            self.logs_compra_prod.insert(0, log_entry)  # Insertar al inicio
            
            # Mantener solo los últimos 100 registros
            if len(self.logs_compra_prod) > 100:
                self.logs_compra_prod = self.logs_compra_prod[:100]
            
            # Insertar en la tabla de logs
            self.tree_logs_compra_prod.insert('', 0, values=(
                fecha,
                producto,
                f"{cantidad}",
                f"${precio:.2f}",
                proveedor,
                f"${total:.2f}"
            ))
            
            # Mantener solo los últimos 100 registros en la tabla
            items = self.tree_logs_compra_prod.get_children()
            if len(items) > 100:
                self.tree_logs_compra_prod.delete(items[-1])
                
        except Exception as e:
            print(f"Error al agregar log de compra productos: {e}")

if __name__ == '__main__':
    root = tk.Tk()
    app = App(root)
    root.mainloop()
