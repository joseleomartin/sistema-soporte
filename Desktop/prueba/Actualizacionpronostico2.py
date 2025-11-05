from concurrent.futures import ThreadPoolExecutor
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.keys import Keys
from lxml import etree
import pandas as pd
import numpy as np
import time
from openpyxl import load_workbook
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import re
import xlwings as xw
import gc
from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive
import os
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
import threading
from concurrent.futures import ThreadPoolExecutor
import time
import requests


lock = threading.Lock()


def setup_drive():
    SCOPES = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
    SERVICE_ACCOUNT_FILE = 'C:/Users/relim/Desktop/prueba/durable-binder-432321-q4-5f1ef9e64ea1.json'
    
    creds = service_account.Credentials.from_service_account_file(
        SERVICE_ACCOUNT_FILE, scopes=SCOPES)
    
    service = build('drive', 'v3', credentials=creds)
    return service

def upload_and_get_shared_link(service, file_path):
     with lock:
        file_metadata = {
            'name': os.path.basename(file_path),
            'mimeType': 'application/vnd.google-apps.spreadsheet'
        }
        media = MediaFileUpload(file_path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        file = service.files().create(body=file_metadata, media_body=media, fields='id').execute()
        file_id = file.get('id')
        
        # Cambiar permisos para compartir el archivo
        permissions = {
            'type': 'anyone',
            'role': 'writer'
        }
        service.permissions().create(fileId=file_id, body=permissions).execute()
        
        # Obtener enlace compartido
        file_info = service.files().get(fileId=file_id, fields='webViewLink').execute()
        shared_link = file_info.get('webViewLink')
        print(f"Enlace compartido generado: {shared_link}")
        return shared_link


def setup_driver():
    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--disable-extensions")
    chrome_options.add_argument("--disable-popup-blocking")
    chrome_options.add_argument("--blink-settings=imagesEnabled=false")
    chrome_options.add_argument('--ignore-certificate-errors')
    chrome_options.add_argument('--disable-web-security')
    chrome_options.add_argument('--allow-running-insecure-content')
    chrome_options.add_argument("--no-sandbox")
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
    return driver       

def login(driver):
    driver.get("https://www.marketscreener.com/login/")
    username_field = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, '/html/body/div[1]/div/div/main/div/div/div/div/form/div[1]/input'))
    )
    password_field = driver.find_element(By.XPATH, '/html/body/div[1]/div/div/main/div/div/div/div/form/div[2]/input')
    username_field.send_keys("leojosemartin@gmail.com")
    password_field.send_keys("Papaleo111.")
    password_field.send_keys(Keys.RETURN)
    time.sleep(2)  # Espera a que el inicio de sesión se complete

def capture_table(driver, url, xpath, header_xpath=None):
    print(f"Paso 1: Navegando a la URL: {url}")
    driver.get(url)
    time.sleep(2)
    print("Paso 2: Desplazándose hasta el final de la página")
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    print(f"Paso 3: Esperando la presencia del elemento con XPath: {xpath}")
    WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, xpath)))
    print("Paso 4: Extrayendo el contenido HTML de la página")
    html = driver.page_source
    dom = etree.HTML(html)
    
    # Extraer encabezados si se proporciona un header_xpath
    headers = []
    if header_xpath:
        header_elements = dom.xpath(header_xpath)
        headers = [''.join(header.xpath(".//text()")).strip() for header in header_elements]

    print(f"Paso 6: Buscando filas en la tabla con el XPath: {xpath}")
    rows = dom.xpath(xpath)
    if not rows:
        print(f"No se encontraron filas con el XPath: {xpath}")
        return []

    table_data = []
    for row in rows:
        cells = row.xpath(".//td | .//th")
        row_data = [''.join(cell.xpath(".//text()")).strip() for cell in cells]
        table_data.append(row_data)

    # Agregar encabezados al inicio de los datos de la tabla
    if headers:
        table_data.insert(0, headers)

    return table_data

EXCEPTIONS = {'0', "Diluted Shares Outstanding", "Basic Weighted Average Shares Outstanding", "Diluted Weighted Average Shares Outstanding", "Basic Weighted Average Shares Outstanding "
              "Diluted Weighted Average Shares Outstanding ","Basic Weighted Average Shares Outstanding ", "Payout Ratio", "American Depositary Receipts Ratio (ADR)",
                "American Depositary Receipts Ratio", "Effective Tax Rate - (Ratio)", "Effective Tax Rate", "Net EPS - Basic", "Basic EPS - Continuing Operations", "Net EPS", "Basic EPS", "Net EPS - Diluted", "Net EPS", "Diluted EPS - Continuing Operations", "Diluted EPS",
                "Normalized Basic EPS", "Normalized Diluted EPS",
                "Diluted Weighted Average Shares Outstanding ", "Basic Weighted Average Shares Outstanding ",
                "Diluted Shares Outstanding", "Diluted Weighted Average Shares Outstanding","Basic Weighted Average Shares Outstanding", "Fiscal period","Fiscal Period", "Fiscal Period: January", "Fiscal Period: February", "Fiscal Period: March", "Fiscal Period: April", "Fiscal Period: May", "Fiscal Period: June", "Fiscal Period: July", "Fiscal Period: August", "Fiscal Period: September", "Fiscal Period: October", "Fiscal Period: November", "Fiscal Period: December" }



def convert_value(value):
    if isinstance(value, str):
        value = value.strip()
        if value in EXCEPTIONS:
            return value
        if value == '-':
            return 0
        if value.endswith('M'):
            return int(float(value[:-1].replace(',', '')) * 1) # Multiplicar por 1 millón
        elif value.endswith('B'):
            return int(float(value[:-1].replace(',', '')) * 1_000)  # Multiplicar por 1 mil millones
        elif value.endswith('K'):
            return int(float(value[:-1].replace(',', '')) * 0.1)  # Multiplicar por 1 mil
        elif value.replace(',', '').lstrip('-').isdigit():  # Permitir números negativos
            return int(value.replace(',', ''))
        else:
            return value
    return value

def convert_dataframe(df):
    for col in df.columns:
        if col != 0:  # Aplicar conversión solo si la columna no es la columna 0
            df[col] = df[col].apply(convert_value)
    return df


def get_exchange_rate(from_currency, to_currency="USD"):
    """Obtiene la tasa de cambio desde la moneda de origen a la moneda de destino."""
    url = f"https://api.exchangerate-api.com/v4/latest/{from_currency}"
    response = requests.get(url)
    data = response.json()
    return data['rates'].get(to_currency, 1)  


def convert_to_usd(df, exchange_rate):
    # Convertir solo columnas numéricas después de aplicar map
    numeric_cols = df.columns[df.apply(lambda col: col.apply(lambda x: isinstance(x, (int, float)) or (isinstance(x, str) and x.replace(',', '').isdigit())).any())]

    # Multiplicar por la tasa de cambio solo si la fila no tiene un valor en la columna A que sea una excepción
    for index in df.index:
        if df.at[index, df.columns[0]] not in EXCEPTIONS:
            df.loc[index, numeric_cols] = df.loc[index, numeric_cols].apply(lambda x: x * exchange_rate if isinstance(x, (int, float)) else x)
    
    return df


def extract_currency_from_url(base_url):
    # Agregar el sufijo '-income-statement/' a la URL
    full_url = f"{base_url}-income-statement/"

    # Crear una nueva instancia del navegador
    driver = setup_driver()

    # Navegar a la URL completa
    driver.get(full_url)

    extracted_text = None  # Inicializar la variable para almacenar el texto

    try:
        # Utiliza un selector CSS para encontrar el elemento que contiene la moneda
        wait = WebDriverWait(driver, 20)  # Aumentar el tiempo de espera a 20 segundos
        element = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, ".js-selectedCurrency.px-10")))
        extracted_text = element.text.strip()  # Usar strip() para limpiar espacios extra
        print(f"Texto extraído desde el elemento en {full_url}: {extracted_text}")

    except Exception as e:
        print(f"Error inesperado: {e}")

    finally:
        driver.quit()  # Cerrar el navegador al finalizar

    return extracted_text  # Retornar el texto extraído


def tipo_empresa_from_xpath(driver, tipo_xpath):
    try:
        # Esperar a que el elemento esté presente
        empresa_element = WebDriverWait(driver, 2).until(
            EC.presence_of_element_located((By.XPATH, tipo_xpath))
        )
        # Extraer el texto del elemento
        empresa_text = empresa_element.get_attribute('textContent').strip()
        
        if empresa_text:
            print(f"Texto extraído del XPath: '{empresa_text}'")
        else:
            print("No se pudo extraer el texto, el resultado es vacío.")
        
        return empresa_text
    except Exception as e:
        print(f"Error al extraer el texto desde el XPath: {str(e)}")
        return "No encontrado"

def process_stock_code(stock_code, forecast_years, x2, drive):
    driver = setup_driver()
    login(driver)
    try:
        print(f"Processing stock code: {stock_code}")
        base_url = f"https://www.marketscreener.com/quote/stock/{stock_code}/finances"
        base_url2 = f"https://www.marketscreener.com/quote/stock/{stock_code}/valuation/"

        # Intentar capturar los datos de valoración
        try:
            valuation_data = capture_table(
                driver, f"{base_url2}",
                '//*[@id="valuationEnterpriseTable"]/tbody/tr',
                header_xpath='//*[@id="valuationEnterpriseTable_wrapper"]/div/div[1]/div/table/thead/tr/th'
            )
            if valuation_data is None or not valuation_data:
                valuation_data = []  # Asignar lista vacía si no hay datos
                print(f"No se encontraron datos de valoración para {stock_code}")
            else:
                print(f"Valuation data capturada para {stock_code}")
        except Exception as e:
            valuation_data = []
            print(f"No se pudo capturar valuation data para {stock_code}. Error: {e}")
                            
        # Capturar los datos financieros
        cash_flow_data = capture_table(
            driver, f"{base_url}-cash-flow-statement/",
            '//*[@id="horizontalFinancialTableN1_3"]/tbody/tr',
            header_xpath='//*[@id="horizontalFinancialTableN1_3_wrapper"]/div/div[1]/div/table/thead/tr/th'
        )   
        balance_sheet_data = capture_table(
            driver, f"{base_url}-balance-sheet/",
            '//*[@id="horizontalFinancialTableN1_2"]/tbody/tr',
            header_xpath='//*[@id="horizontalFinancialTableN1_2_wrapper"]/div/div[1]/div/table/thead/tr/th'
        )
        financial_data = capture_table(
            driver, f"{base_url}-income-statement/",
            '//*[@id="horizontalFinancialTableN1_1"]/tbody/tr',
            header_xpath='//*[@id="horizontalFinancialTableN1_1_wrapper"]/div/div[1]/div/table/thead/tr/th'
        )
        ratios_data = capture_table(
            driver, f"{base_url}-ratios/",
            '//*[@id="horizontalFinancialTableN1_5"]/tbody/tr',
            header_xpath='//*[@id="horizontalFinancialTableN1_5_wrapper"]/div/div[1]/div/table/thead/tr/th'
        )

        # Extraer el tipo de empresa desde el XPath específico
        tipo_xpath = "/html/body/div[1]/div/div[1]/main/div[1]/div[1]/div[2]/a[2]/h2"
        tipo_empresa = tipo_empresa_from_xpath(driver, tipo_xpath)

        currency = extract_currency_from_url(base_url)
        if not currency:
            currency = "USD"
            print(f"No se pudo detectar la moneda para {stock_code}, usando USD como predeterminado")


        # Agregar print para ver la moneda extraída
        print(f"Moneda detectada para {stock_code}: {currency}")

        # Obtener el tipo de cambio a dólares
        exchange_rate = get_exchange_rate(currency)
        print(f"Tasa de cambio a USD para {currency}: {exchange_rate}")

        stock_name_clean = re.sub(r'[^A-Za-z]+', ' ', stock_code).replace('-', ' ').strip()
        stock_codes_df = pd.read_excel(file_path, usecols=[0, 1], header=None)
        stock_name_from_file = stock_codes_df[stock_codes_df[0] == stock_code].iloc[0, 1] if not stock_codes_df[stock_codes_df[0] == stock_code].empty else ''


        data_data = [
            ["ticker", stock_name_from_file],
            ["nombre", stock_name_clean],
            ["empresa", tipo_empresa]
        ]

        # Convertir los datos financieros a DataFrames
        df_data= pd.DataFrame(data_data)
        df_cash_flow = pd.DataFrame(cash_flow_data)
        df_balance_sheet = pd.DataFrame(balance_sheet_data)
        df_financial = pd.DataFrame(financial_data)
        df_ratios = pd.DataFrame(ratios_data)
        df_valuation = pd.DataFrame(valuation_data)

        df_cash_flow = convert_dataframe(pd.DataFrame(cash_flow_data))
        df_balance_sheet = convert_dataframe(pd.DataFrame(balance_sheet_data))
        df_financial = convert_dataframe(pd.DataFrame(financial_data))
        df_ratios = convert_dataframe(pd.DataFrame(ratios_data))


        # Convertir los datos a dólares usando la tasa de cambio obtenida
        df_cash_flow = convert_to_usd(df_cash_flow, exchange_rate)
        df_balance_sheet = convert_to_usd(df_balance_sheet, exchange_rate)
        df_financial = convert_to_usd(df_financial, exchange_rate)
        df_ratios = convert_to_usd(df_ratios, exchange_rate)
    finally:
        driver.quit()
        try:
            os.system("taskkill /im chrome.exe /f")  # En Windows
            # os.system("pkill -f chrome")  # En Linux/Unix
        except Exception as e:
            print(f"Error al intentar matar procesos de Chrome: {e}")


    excel_path = rf"C:\Users\relim\Desktop\prueba\Resultados\{stock_name_from_file} - {stock_name_clean}.xlsx"
    with pd.ExcelWriter(excel_path, engine='xlsxwriter') as writer:
        df_data.to_excel(writer, sheet_name='Data', index=False, header=False)
        df_cash_flow.to_excel(writer, sheet_name='Cash Flow Statement', index=False, header=False)
        df_balance_sheet.to_excel(writer, sheet_name='Balance Sheet', index=False, header=False)
        df_financial.to_excel(writer, sheet_name='Financial', index=False, header=False)
        df_ratios.to_excel(writer, sheet_name='Ratios', index=False, header=False)
        df_valuation.to_excel(writer, sheet_name='Valuation', index=False, header=False)

    def extract_values(df, search_str):
        for idx, row in df.iterrows():
            if any(search_str in str(cell) for cell in row):
                return row[1:].values
        print(f"No se encontró '{search_str}' en la tabla.")
        return []

    fiscal_period = extract_values(df_cash_flow, 'Fiscal Period')

    # Extraemos los demás valores como antes
    unlevered_free_cash_flow = extract_values(df_cash_flow, 'Unlevered Free Cash Flow')
    total_debt = extract_values(df_balance_sheet, 'Net Debt')
    diluted_shares = extract_values(df_financial, 'Diluted Weighted Average Shares Outstanding')
    cash_equivalents = extract_values(df_balance_sheet, 'Cash and Equivalents')

    # Rellenamos con NaN para que todas las listas tengan la misma longitud
    max_length = max(len(fiscal_period), len(unlevered_free_cash_flow), len(total_debt), len(diluted_shares), len(cash_equivalents))
    fiscal_period = np.pad(fiscal_period, (0, max_length - len(fiscal_period)), constant_values=np.nan)
    unlevered_free_cash_flow = np.pad(unlevered_free_cash_flow, (0, max_length - len(unlevered_free_cash_flow)), constant_values=np.nan)
    total_debt = np.pad(total_debt, (0, max_length - len(total_debt)), constant_values=np.nan)
    diluted_shares = np.pad(diluted_shares, (0, max_length - len(diluted_shares)), constant_values=np.nan)
    cash_equivalents = np.pad(cash_equivalents, (0, max_length - len(cash_equivalents)), constant_values=np.nan)

    # Creamos el DataFrame incluyendo el Fiscal Period
    df_dcf = pd.DataFrame({
    'Fiscal Period': fiscal_period,
    'Unlevered Free Cash Flow': unlevered_free_cash_flow,
    'Total Debt': total_debt,
    'Diluted Shares Outstanding': diluted_shares,
    })

    # Transponer el DataFrame
    df_dcf = df_dcf.set_index('Fiscal Period').T

    # Obtener el último año fiscal
    last_year = int(fiscal_period[-1])
    new_years = [last_year + i for i in range(1, forecast_years + 1)]

    # Agregar los nuevos años como columnas
    df_dcf = pd.concat([df_dcf, pd.DataFrame(columns=new_years)], axis=1)
    

    # Escribir en el archivo Excel
    with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a') as writer:
        df_dcf.to_excel(writer, sheet_name='DCF', startrow=1, index=True, header=True)

    # Cargar el libro de trabajo y actualizar una celda
    wb = load_workbook(excel_path)
    ws = wb['DCF']
    ws['X2'] = x2

    # Calcular el índice de la última columna de flujo de caja libre no apalancado
    unlevered_last_col = len(df_dcf.columns) - forecast_years
    for i in range(forecast_years): 
        current_col = unlevered_last_col + 2 + i
        previous_col = unlevered_last_col + 1 + i

        if i == 0:
            formula = (
                f'=IF({ws.cell(row=3, column=unlevered_last_col + 1).coordinate}*'
                f'FORECAST({ws.cell(row=2, column=current_col).coordinate},'
                f'{ws.cell(row=3, column=2).coordinate}:{ws.cell(row=3, column=unlevered_last_col + 1).coordinate},'
                f'{ws.cell(row=2, column=2).coordinate}:{ws.cell(row=2, column=unlevered_last_col + 1).coordinate})>0,'
                f'FORECAST({ws.cell(row=2, column=current_col).coordinate},'
                f'{ws.cell(row=3, column=2).coordinate}:{ws.cell(row=3, column=unlevered_last_col + 1).coordinate},'
                f'{ws.cell(row=2, column=2).coordinate}:{ws.cell(row=2, column=unlevered_last_col + 1).coordinate}),'
                f'{ws.cell(row=3, column=unlevered_last_col + 1).coordinate})*$X$2'
            )
        else:
            formula = (
                f'=IF({ws.cell(row=3, column=unlevered_last_col + 1 + i).coordinate}*'
                f'FORECAST({ws.cell(row=2, column=current_col).coordinate},'
                f'{ws.cell(row=3, column=2).coordinate}:{ws.cell(row=3, column=current_col - 1).coordinate},'
                f'{ws.cell(row=2, column=2).coordinate}:{ws.cell(row=2, column=current_col - 1).coordinate})>0,'
                f'FORECAST({ws.cell(row=2, column=current_col).coordinate},'
                f'{ws.cell(row=3, column=2).coordinate}:{ws.cell(row=3, column=current_col - 1).coordinate},'
                f'{ws.cell(row=2, column=2).coordinate}:{ws.cell(row=2, column=current_col - 1).coordinate}),'
                f'{ws.cell(row=3, column=unlevered_last_col + 1 + i).coordinate})*$X$2'
            )
        ws.cell(row=3, column=current_col, value=formula)

    valor = ws.cell(row=3, column=current_col).value
    ws['K7'] = "TERMINAL VALUE"
    ws['L7'] = f'{valor} * (1 + 0.02) / (0.06 - 0.02)'
    terminal_value = ws['L7'].value

    start_col = unlevered_last_col + 2
    end_col = current_col
    terminal_col = end_col + 1

    def col_num_to_letter(col_num):
        col_letter = ""
        while col_num > 0:
            col_num, remainder = divmod(col_num - 1, 26)
            col_letter = chr(65 + remainder) + col_letter
        return col_letter

    start_col_letter = col_num_to_letter(start_col)
    end_col_letter = col_num_to_letter(end_col)
    terminal_col_letter = col_num_to_letter(terminal_col)
    next_start_col_letter = col_num_to_letter(start_col + 1)

    ws['K8'] = 'NPV of FCF'
    ws['L8'] = f'=NPV(0.06, {start_col_letter}3:{end_col_letter}3)'

    for col in range(start_col + 1, end_col + 1):
        col_letter = col_num_to_letter(col)
        ws[f'{col_letter}16'] = 0

    ws[f'{terminal_col_letter}16'] = terminal_value

    ws['K9'] = "NPV of TV"
    ws['L9'] = f'=NPV(0.06, {next_start_col_letter}16:{terminal_col_letter}16)'

    ws['K10'] = 'Total EV'
    ws['L10'] = '=L8 + L9'

    ws['K11'] = 'Net Debt'
    for col in reversed(range(2, ws.max_column + 1)):
        if isinstance(ws.cell(row=4, column=col).value, (int, float)):
            last_column_net_debt = col
            break
    ws['L11'] = ws.cell(row=4, column=last_column_net_debt).value

    ws['K12'] = 'Equity'
    ws['L12'] = f'=L10 - L11'

    ws['K13'] = 'Shares Outstanding'
    for col in reversed(range(2, ws.max_column + 1)):
        if isinstance(ws.cell(row=5, column=col).value, (int, float)):
            last_column_shares_outstanding = col
            break
    ws['L13'] = ws.cell(row=5, column=last_column_shares_outstanding).value

    ws['K14'] = 'Target Price'
    ws['L14'] = f'=L12/L13'

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if isinstance(cell.value, str) and ("#DIV/0!" in cell.value or "Error" in cell.value):
                cell.value = "Error"

    wb.save(excel_path)

    shared_link = upload_and_get_shared_link(drive, excel_path)

    scope = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
    creds = ServiceAccountCredentials.from_json_keyfile_name('C:/Users/relim/Desktop/prueba/durable-binder-432321-q4-5f1ef9e64ea1.json', scope)
    client = gspread.authorize(creds)
    sheet = client.open_by_key('193GrDtaNCWBkWwliitjSepbLlAcNlwQfUZp66WN6xyw').worksheet('Results')


    # Usar xlwings para abrir el archivo y calcular las fórmulas
    app = xw.App(visible=False)
    workbook = app.books.open(excel_path)
    workbook.save()  # Esto fuerza la evaluación de fórmulas
    workbook.close()
    app.quit()

    # Leer el archivo Excel para obtener el valor calculado
    df = pd.read_excel(excel_path, sheet_name='DCF', header=None)
    l14_value = df.at[13, 11]  # Filas y columnas están basadas en el índice 0

    # Verificar si l14_value es un error y asignar un valor manejable si lo es
    if isinstance(l14_value, str) and ("#DIV/0" in l14_value or "Error" in l14_value):
        l14_value = "Error en cálculo"  # Puedes asignar cualquier valor representativo
    elif pd.isna(l14_value):  # Si el valor es NaN (celda vacía o con error no detectado)
        l14_value = "Valor no disponible"


    try:
        # Buscar la fila correspondiente para el código de stock
        cell_list = sheet.col_values(1)  # Obtener todos los valores de la columna A
        if stock_code in cell_list:
            row_index = cell_list.index(stock_code) + 1  # Encuentra la fila existente
        else:
            # Si el código no está en la hoja, añade al final
            row_index = len(cell_list) + 1

        link = f'https://www.marketscreener.com/quote/stock/{stock_code}'
        formula_link = f'=HYPERLINK("{link}", "{stock_code}")'     
        
        # Actualizar la fila específica con los nuevos datos
        sheet.update_cell(row_index, 1, formula_link)
        sheet.update_cell(row_index, 2, stock_name_clean)
        sheet.update_cell(row_index, 3, l14_value)
        sheet.update_cell(row_index, 4, stock_name_from_file)
        sheet.update_cell(row_index, 8, shared_link)
        sheet.update_cell(row_index, 9, currency)
        sheet.update_cell(row_index, 10, exchange_rate)
        sheet.update_cell(row_index, 11, tipo_empresa)
        sheet.update_cell(row_index, 12, "11/11/2024")
        
        print(f"Datos actualizados: {stock_code}, {stock_name_clean}, {l14_value}")
    except Exception as e:
        print(f"Error al actualizar datos en Google Sheets: {e}")
    


def process_all_stock_codes(file_path, forecast_years, x2):
    stock_codes = pd.read_excel(file_path, usecols=[0], header=None).iloc[:, 0].tolist()
    drive = setup_drive() 

    with ThreadPoolExecutor(max_workers=3) as executor:
        futures = []
        for code in stock_codes:
            future = executor.submit(process_stock_code, code, forecast_years, x2, drive)
            futures.append(future)
            time.sleep(2)  # Retraso de 4 segundos entre el envío de tareas

        # Esperar a que se completen todas las tareas
        for future in futures:
            try:
                future.result()
            except Exception as e:
                print(f"Error en el procesamiento del código de stock: {e}")

    gc.collect()




file_path = r'C:\Users\relim\Desktop\prueba\StockCodes.xlsx'
forecast_years = 7
x2 = 1

process_all_stock_codes(file_path, forecast_years, x2)