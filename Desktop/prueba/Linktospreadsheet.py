from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.keys import Keys
from lxml import etree
import pandas as pd
import time
import numpy as np
from openpyxl import load_workbook
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import os
import re
import openpyxl
import xlwings as xw
import gspread

excel_path = r'C:\Users\relim\Desktop\prueba\StockCodes.xlsx'  # Reemplaza con la ruta de tu archivo Excel
df = pd.read_excel(excel_path, usecols=[0], header=None)
stock_codes = df[0].dropna().tolist()

# Solicita el año y x2 una sola vez
forecast_years = int(input("Introduce el año: "))
x2 = int(input("Introduce el valor de x2: "))

# Itera sobre cada stock_code
for stock_code in stock_codes:
    # Configuración de Selenium
    chrome_options = Options()
    chrome_options.add_argument("--headless")  # Ejecuta en modo headless (sin ventana)
    chrome_options.add_argument("--start-maximized")

    # Inicializa el WebDriver
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)

    try:
        # Iniciar sesión en el sitio web
        driver.get("https://www.marketscreener.com/login/")

        username_field = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '/html/body/div[1]/div/div/main/div/div/div/div/form/div[1]/input'))
        )
        password_field = driver.find_element(By.XPATH, '/html/body/div[1]/div/div/main/div/div/div/div/form/div[2]/input')

        username_field.send_keys("leojosemartin@gmail.com")
        password_field.send_keys("Papaleo111.")
        password_field.send_keys(Keys.RETURN)

        # Espera a que el inicio de sesión se complete
        time.sleep(10)

        # Función para capturar y procesar una tabla dada una URL y un XPath
        def capture_table(url, xpath):
            driver.get(url)
            time.sleep(5)  # Agregar tiempo de espera para que la página cargue
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, xpath)))

            # Capturar el HTML de la página después de cargar la tabla
            html = driver.page_source
            dom = etree.HTML(html)
            rows = dom.xpath(xpath)

            if not rows:
                print(f"No se encontraron filas con el XPath: {xpath}")
                return []

            table_data = []
            for row in rows:
                cells = row.xpath(".//td | .//th")
                row_data = [''.join(cell.xpath(".//text()")).strip() for cell in cells]
                table_data.append(row_data)

            return table_data

        # URLs con el código dinámico
        base_url = f"https://www.marketscreener.com/quote/stock/{stock_code}/finances"
        
        # Capturar las tablas
        cash_flow_data = capture_table(
            f"{base_url}-cash-flow-statement/",
            '//*[@id="horizontalFinancialTableN1_3"]/tbody/tr'
        )
        balance_sheet_data = capture_table(
            f"{base_url}-balance-sheet/",
            '//*[@id="horizontalFinancialTableN1_2"]/tbody/tr'
        )
        financial_data = capture_table(
            f"{base_url}-income-statement/",
            '//*[@id="horizontalFinancialTableN1_1"]/tbody/tr'
        )
        ratios_data = capture_table(
            f"{base_url}-ratios/",
            '//*[@id="horizontalFinancialTableN1_5"]/tbody/tr'
        )

    finally:
        # Cierra el navegador
        driver.quit()

    # Función para convertir valores de formatos como '1M' y '1.1B'
    def convert_value(value):
        if isinstance(value, str):
            value = value.strip()
            if value.endswith('M'):
                return int(float(value[:-1].replace(',', '')) * 1)
            elif value.endswith('B'):
                return int(float(value[:-1].replace(',', '')) * 1_000)
            elif value.replace(',', '').isdigit():
                return int(value.replace(',', ''))
            else:
                return value
        return value

    # Función para aplicar la conversión a un DataFrame
    def convert_dataframe(df):
        for col in df.columns:
            df[col] = df[col].apply(convert_value)
        return df

    # Crear el DataFrame con pandas
    df_cash_flow = pd.DataFrame(cash_flow_data)
    df_balance_sheet = pd.DataFrame(balance_sheet_data)
    df_financial = pd.DataFrame(financial_data)
    df_ratios = pd.DataFrame(ratios_data)

    # Convertir los datos de los DataFrames
    df_cash_flow = convert_dataframe(df_cash_flow)
    df_balance_sheet = convert_dataframe(df_balance_sheet)
    df_financial = convert_dataframe(df_financial)
    df_ratios = convert_dataframe(df_ratios)

    # Guardar los DataFrames en un archivo Excel, cada uno en una pestaña diferente
    excel_path = rf"C:\Users\relim\Desktop\prueba\{stock_code}.xlsx"
    with pd.ExcelWriter(excel_path, engine='xlsxwriter') as writer:
        df_cash_flow.to_excel(writer, sheet_name='Cash Flow Statement', index=False, header=False)
        df_balance_sheet.to_excel(writer, sheet_name='Balance Sheet', index=False, header=False)
        df_financial.to_excel(writer, sheet_name='Financial', index=False, header=False)
        df_ratios.to_excel(writer, sheet_name='Ratios', index=False, header=False)

    print(f"Tablas guardadas en '{excel_path}' en diferentes pestañas.")

    # Función para extraer los valores de Unlevered Free Cash Flow
    def extract_unlevered_free_cash_flow(df):
        search_str = 'Unlevered Free Cash Flow'
        for idx, row in df.iterrows():
            if any(search_str in str(cell) for cell in row):
                # Extraer el flujo de caja y los valores posteriores
                unlevered_free_cash_flow = row[1:].values
                return unlevered_free_cash_flow

        print(f"No se encontró '{search_str}' en la tabla.")
        return []

    # Función para extraer el Total Debt
    def extract_total_debt(df):
        search_str = 'Total Debt'
        for idx, row in df.iterrows():
            if any(search_str in str(cell) for cell in row):
                # Extraer el total de deuda y los valores posteriores
                total_debt = row[1:].values
                return total_debt

        print(f"No se encontró '{search_str}' en la tabla.")
        return []

    # Función para extraer el Diluted Weighted Average Shares Outstanding
    def extract_diluted_shares(df):
        search_str = 'Diluted Weighted Average Shares Outstanding'
        for idx, row in df.iterrows():
            if any(search_str in str(cell) for cell in row):
                # Extraer los shares y los valores posteriores
                diluted_shares = row[1:].values
                return diluted_shares

        print(f"No se encontró '{search_str}' en la tabla.")
        return []

    # Función para extraer el Cash and Equivalents
    def extract_cash_equivalents(df):
        search_str = 'Cash and Equivalents'
        for idx, row in df.iterrows():
            if any(search_str in str(cell) for cell in row):
                # Extraer el cash and equivalents y los valores posteriores
                cash_equivalents = row[1:].values
                return cash_equivalents

        print(f"No se encontró '{search_str}' en la tabla.")
        return []

    # Extraer los valores de Unlevered Free Cash Flow, Total Debt, Diluted Shares, y Cash and Equivalents
    unlevered_free_cash_flow = extract_unlevered_free_cash_flow(df_cash_flow)
    total_debt = extract_total_debt(df_balance_sheet)
    diluted_shares = extract_diluted_shares(df_financial)
    cash_equivalents = extract_cash_equivalents(df_balance_sheet)

    # Asegurarse de que todas las listas tengan la misma longitud
    max_length = max(len(unlevered_free_cash_flow), len(total_debt), len(diluted_shares), len(cash_equivalents))

    # Rellenar las listas más cortas con NaN para que todas tengan la misma longitud
    unlevered_free_cash_flow = np.pad(unlevered_free_cash_flow, (0, max_length - len(unlevered_free_cash_flow)), constant_values=np.nan)
    total_debt = np.pad(total_debt, (0, max_length - len(total_debt)), constant_values=np.nan)
    diluted_shares = np.pad(diluted_shares, (0, max_length - len(diluted_shares)), constant_values=np.nan)
    cash_equivalents = np.pad(cash_equivalents, (0, max_length - len(cash_equivalents)), constant_values=np.nan)

    # Crear un DataFrame para la pestaña DCF
    df_dcf = pd.DataFrame({
        'Unlevered Free Cash Flow': unlevered_free_cash_flow,
        'Total Debt': total_debt,
        'Diluted Shares Outstanding': diluted_shares,
        'Cash and Equivalents': cash_equivalents
    })

    # Transponer el DataFrame para que los datos sean horizontales
    df_dcf = df_dcf.T



    # Crear una lista de años desde el último año en la hoja
    last_year = int(df_dcf.columns[-1])
    new_years = [last_year + i for i in range(1, forecast_years + 1)]

    # Actualizar las fechas en el DataFrame
    df_dcf = df_dcf.reindex(columns=list(df_dcf.columns) + new_years)

    # Guardar el DataFrame en una nueva hoja usando openpyxl
    with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a') as writer:
        df_dcf.to_excel(writer, sheet_name='DCF', startrow=1, index=True, header=True)

    # Cargar el archivo Excel para modificar la fórmula
    wb = load_workbook(excel_path)
    ws = wb['DCF']

  
    ws['X2'] = x2

    # Encontrar la última columna con datos de Unlevered Free Cash Flow
    unlevered_last_col = len(df_dcf.columns) - forecast_years

    # Insertar la fórmula PRONOSTICO justo después del último dato de Unlevered Free Cash Flow
    for i in range(forecast_years):
        current_col = unlevered_last_col + 2 + i
        previous_col = unlevered_last_col + 1 + i

        if i == 0:
            formula = (
                f'=IF({ws.cell(row=3, column=unlevered_last_col + 1).coordinate}*'
                f'FORECAST({ws.cell(row=2, column=current_col).coordinate},'
                f'{ws.cell(row=3, column=2).coordinate}:{ws.cell(row=3, column=unlevered_last_col + 1).coordinate},'
                f'{ws.cell(row=2, column=2).coordinate}:{ws.cell(row=2, column=unlevered_last_col + 1).coordinate})>0,'
                f'FORECAST({ws.cell(row=2, column=current_col).coordinate},'
                f'{ws.cell(row=3, column=2).coordinate}:{ws.cell(row=3, column=unlevered_last_col + 1).coordinate},'
                f'{ws.cell(row=2, column=2).coordinate}:{ws.cell(row=2, column=unlevered_last_col + 1).coordinate}),'
                f'{ws.cell(row=3, column=unlevered_last_col + 1).coordinate})*$X$2'
            )
        else:
            formula = (
                f'=IF({ws.cell(row=3, column=unlevered_last_col + 1 + i).coordinate}*'
                f'FORECAST({ws.cell(row=2, column=current_col).coordinate},'
                f'{ws.cell(row=3, column=2).coordinate}:{ws.cell(row=3, column=current_col - 1).coordinate},'
                f'{ws.cell(row=2, column=2).coordinate}:{ws.cell(row=2, column=current_col - 1).coordinate})>0,'
                f'FORECAST({ws.cell(row=2, column=current_col).coordinate},'
                f'{ws.cell(row=3, column=2).coordinate}:{ws.cell(row=3, column=current_col - 1).coordinate},'
                f'{ws.cell(row=2, column=2).coordinate}:{ws.cell(row=2, column=current_col - 1).coordinate}),'
                f'{ws.cell(row=3, column=unlevered_last_col + 1 + i).coordinate})*$X$2'
            )
        ws.cell(row=3, column=current_col, value=formula)
    wb.save(excel_path)

    wb = load_workbook(excel_path)
    ws = wb['DCF']

    # Obtener el valor de la celda en la fila 3 y columna actual
    valor = ws.cell(row=3, column=current_col).value


    # Terminal value
    ws['K7'] = "TERMINAL VALUE"
    ws['L7'] = f'{valor} * (1 + 0.02) / (0.06 - 0.02)'
    terminal_value = ws['L7'].value



    start_col = unlevered_last_col + 2
    end_col = current_col
    terminal_col = end_col + 1

    # Construir las referencias de celdas en formato de columna
    start_col_letter = chr(64 + start_col)  # Convertir el número de columna a letra (por ejemplo, 5 -> E)
    end_col_letter = chr(64 + end_col)  
    terminal_col_letter = chr(64 + terminal_col)
    start_col_2_letter = chr(64 + terminal_col)
    next_start_col_letter = chr(64 + start_col + 1)


    #NPV del fcf
    ws['K8'] = 'Npv of fcf'
    ws['L8'] = f'=NPV(0.06, {start_col_letter}3:{end_col_letter}3)'

    #NPV del TV

    # Rellenar las celdas con 0 desde la columna de inicio hasta la columna final en la fila 9
    for col in range(start_col +1, end_col+1):  # Incluye la columna final
        col_letter = chr(64 + col)  # Convertir el número de columna a letra (ej., 5 -> E)
        ws[f'{col_letter}9'] = 0
        
    # Colocar el valor del Terminal Value en la celda correspondiente
    ws[f'{terminal_col_letter}9'] = terminal_value

    # Calcular el NPV del FCF
    ws['K8'] = 'Npv of fcf'
    ws['L8'] = f'=NPV(0.06, {start_col_letter}3:{end_col_letter}3)'

    # Calcular el NPV del Terminal Value
    ws['K9'] = "Npv of TV"
    ws['L9'] = f'=NPV(0.06, {next_start_col_letter}9:{terminal_col_letter}9)'

    #Total EV
    ws['k10'] = 'Total EV'
    ws['L10'] = f'=L8 + L9'

    # Net debt
    ws['K11'] = 'Net Debt'
    # Buscar la última columna con un valor numérico en la fila 4
    for col in reversed(range(2, ws.max_column + 1)):  # Recorre de derecha a izquierda
        if isinstance(ws.cell(row=4, column=col).value, (int, float)):
            last_column_net_debt = col
            break

    # Colocar el valor en la celda L11
    ws['L11'] = ws.cell(row=4, column=last_column_net_debt).value

    #Equity
    ws['k12'] = 'Equity'
    ws['L12'] = f'=L10 - L11'

    #Shares outstanding
    ws['K13'] = 'Shares Outstanding'
    for col in reversed(range(2, ws.max_column + 1)):  # Recorre de derecha a izquierda
        if isinstance(ws.cell(row=5, column=col).value, (int, float)):
            last_column_shares_outstanding = col
            break

    # Colocar el valor en la celda L13
    ws['L13'] = ws.cell(row=5, column=last_column_shares_outstanding).value

    # Target price
    ws['k14'] = 'Target Price'
    ws['L14'] = f'=L12/L13'

    wb.save(excel_path)

    from oauth2client.service_account import ServiceAccountCredentials
    import gspread

    # Define el alcance de la API
    scope = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]

    # Ruta al archivo JSON de credenciales
    creds = ServiceAccountCredentials.from_json_keyfile_name('C:/Users/relim/Desktop/prueba/durable-binder-432321-q4-5f1ef9e64ea1.json', scope)

    # Autenticación con Google Sheets
    client = gspread.authorize(creds)

    # Reemplaza con el ID correcto de tu hoja de cálculo
    spreadsheet_id = '193GrDtaNCWBkWwliitjSepbLlAcNlwQfUZp66WN6xyw'

    sheet_name = 'Results'  # Cambia esto por el nombre correcto de la hoja

    sheet = client.open_by_key(spreadsheet_id).worksheet(sheet_name)

    # Lee todos los valores en la columna A
    column_a_values = sheet.col_values(1)  # 1 es el índice de la columna A

    # Encuentra la primera celda vacía en la columna A
    if column_a_values:
        row_num = len(column_a_values) + 1  # La siguiente fila vacía
    else:
        row_num = 1  # Si la columna está vacía, coloca el stock_code en la primera fila

    # Escribe el stock_code en la primera celda vacía de la columna A
    cell_address = f'A{row_num}'
    try:
        # Enviar los datos en un formato adecuado (una lista dentro de una lista)
        sheet.update(cell_address, [[stock_code]])
        print(f'Stock code "{stock_code}" guardado en la celda {cell_address}')
    except gspread.exceptions.APIError as e:
        print(f"Error al actualizar la celda {cell_address}: {e}")



    import xlwings as xw

    # Ruta del archivo
    file_path = f'C:/Users/relim/Desktop/prueba/{stock_code}.xlsx'

    # Abrir el libro de trabajo con xlwings
    wb = xw.Book(file_path)

    # Seleccionar la hoja llamada 'DCF'
    ws = wb.sheets['DCF']

    # Leer el valor calculado de L14
    l14_value = ws.range('L14').value

    # Guardar los cambios y cerrar el libro si hiciste modificaciones
    wb.save(file_path)
    wb.close()

    from oauth2client.service_account import ServiceAccountCredentials
    import gspread

    # Define el alcance de la API
    scope = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]

    # Ruta al archivo JSON de credenciales
    creds = ServiceAccountCredentials.from_json_keyfile_name('C:/Users/relim/Desktop/prueba/durable-binder-432321-q4-5f1ef9e64ea1.json', scope)

    # Autenticación con Google Sheets
    client = gspread.authorize(creds)

    # Reemplaza con el ID correcto de tu hoja de cálculo
    spreadsheet_id = '193GrDtaNCWBkWwliitjSepbLlAcNlwQfUZp66WN6xyw'

    sheet_name = 'Results'  # Cambia esto por el nombre correcto de la hoja

    sheet = client.open_by_key(spreadsheet_id).worksheet(sheet_name)

    # Lee todos los valores en la columna C
    column_c_values = sheet.col_values(3)  # 3 es el índice de la columna C

    if column_c_values:
        row_num = len(column_c_values) + 1  # La siguiente fila vacía
    else:
        row_num = 1  # Si la columna está vacía, coloca el valor en la primera fila

    # Escribe el valor en la primera celda vacía de la columna C
    cell_address = f'C{row_num}'
    try:
        # Enviar los datos en un formato adecuado (una lista dentro de una lista)
        sheet.update(cell_address, [[l14_value]])
        print(f'Valor "{l14_value}" guardado en la celda {cell_address}')
    except gspread.exceptions.APIError as e:
        print(f"Error al actualizar la celda {cell_address}: {e}")

    from oauth2client.service_account import ServiceAccountCredentials
    import gspread
    import re

    # Define el alcance de la API
    scope = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]

    # Ruta al archivo JSON de credenciales
    creds = ServiceAccountCredentials.from_json_keyfile_name('C:/Users/relim/Desktop/prueba/durable-binder-432321-q4-5f1ef9e64ea1.json', scope)

    # Autenticación con Google Sheets
    client = gspread.authorize(creds)

    # Reemplaza con el ID correcto de tu hoja de cálculo
    spreadsheet_id = '193GrDtaNCWBkWwliitjSepbLlAcNlwQfUZp66WN6xyw'

    sheet_name = 'Results'  # Cambia esto por el nombre correcto de la hoja

    modified_stock_code = stock_code.replace('-', ' ')

    # Eliminar caracteres no alfabéticos, excepto espacios
    modified_stock_code = re.sub(r'[^a-zA-Z\s]', '', modified_stock_code)

    # Eliminar espacios múltiples
    modified_stock_code = " ".join(modified_stock_code.split())

    # Eliminar espacios múltiples
    modified_stock_code = " ".join(modified_stock_code.split())
    # Lee todos los valores en la columna B
    column_b_values = sheet.col_values(2)  # 2 es el índice de la columna B

    # Encuentra la primera celda vacía en la columna B
    if column_b_values:
        row_num = len(column_b_values) + 1  # La siguiente fila vacía
    else:
        row_num = 1  # Si la columna está vacía, coloca el stock_code modificado en la primera fila

    # Escribe el stock_code modificado en la primera celda vacía de la columna B
    cell_address_b = f'B{row_num}'
    try:
        sheet.update(cell_address_b, [[modified_stock_code]])
        print(f'Stock code "{modified_stock_code}" guardado en la celda {cell_address_b}')
    except gspread.exceptions.APIError as e:
        print(f"Error al actualizar la celda {cell_address_b}: {e}")

print("Proceso completado para todos los stock_codes.")