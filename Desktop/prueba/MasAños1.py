import os
import shutil
import openpyxl
import gspread
from google.oauth2 import service_account
from openpyxl.utils import get_column_letter
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload

# 📌 Ruta del archivo JSON con credenciales
SERVICE_ACCOUNT_FILE = 'C:/Users/relim/Desktop/prueba/durable-binder-432321-q4-5f1ef9e64ea1.json'

# 📌 Definir los permisos necesarios
SCOPES = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]

# 📌 Autenticación con Google Sheets API y Google Drive API
creds = service_account.Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
client = gspread.authorize(creds)

# 📌 ID de la hoja de cálculo principal donde se registrarán los enlaces
ID_HOJA_ACTUALIZACION = "193GrDtaNCWBkWwliitjSepbLlAcNlwQfUZp66WN6xyw"

# 📌 Cliente para Google Drive API
drive_service = build('drive', 'v3', credentials=creds)

def obtener_letra_columna(indice):
    """Convierte un índice numérico en una letra de columna de Excel (1 -> A, 2 -> B, etc.)."""
    return get_column_letter(indice)

def modificar_excel(ruta_origen, ruta_destino):
    """Modifica un archivo Excel según las reglas especificadas."""
    try:
        wb = openpyxl.load_workbook(ruta_origen)
        if "DCF" not in wb.sheetnames:
            print(f"⚠️ La hoja 'DCF' no está en {ruta_origen}. Archivo omitido.")
            return None
        
        hoja_dcf = wb["DCF"]

        # Borrar contenido de la columna X (24ª columna)
        for fila in hoja_dcf.iter_rows(min_row=1, max_row=hoja_dcf.max_row, min_col=24, max_col=24):
            for celda in fila:
                celda.value = None

        # Guardar temporalmente y recargar
        wb.save(ruta_destino)
        wb = openpyxl.load_workbook(ruta_destino)
        hoja_dcf = wb["DCF"]

        # Escribir "1" en X1 
        hoja_dcf['X1'] = 0.5

        # Reemplazar "$X$2" con "$X$1" en toda la hoja
        for fila in hoja_dcf.iter_rows():
            for celda in fila:
                if isinstance(celda.value, str) and "$X$2" in celda.value:
                    celda.value = celda.value.replace("$X$2", "$X$1")

        # Encontrar la última columna con datos en la fila 2
        ultima_columna = hoja_dcf.max_column
        while ultima_columna > 1 and hoja_dcf.cell(row=2, column=ultima_columna).value is None:
            ultima_columna -= 1

        ultimo_año = hoja_dcf.cell(row=2, column=ultima_columna).value or 11  # Asignar 11 si es None

        # Agregar 5 años adicionales en la fila 2
        for i in range(1, 6):
            hoja_dcf.cell(row=2, column=ultima_columna + i, value=ultimo_año + i)

        # Encontrar la última columna con datos en la fila 3 (para fórmulas)
        ultima_columna_formula = hoja_dcf.max_column
        while ultima_columna_formula > 1 and hoja_dcf.cell(row=3, column=ultima_columna_formula).value is None:
            ultima_columna_formula -= 1

        # Aplicar fórmulas en nuevas columnas
        for i in range(1, 6):
            col_actual = obtener_letra_columna(ultima_columna_formula + i)
            col_anterior = obtener_letra_columna(ultima_columna_formula + i - 1)
            col_inicio = obtener_letra_columna(2)

            formula = (
                f"=IF({col_anterior}3*FORECAST({col_actual}2,"
                f"{col_inicio}3:{col_anterior}3,"
                f"{col_inicio}2:{col_anterior}2)>0,"
                f"FORECAST({col_actual}2,"
                f"{col_inicio}3:{col_anterior}3,"
                f"{col_inicio}2:{col_anterior}2),"
                f"{col_anterior}3)*$X$1"
            )
            hoja_dcf.cell(row=3, column=ultima_columna_formula + i, value=formula)
        
        ultima_columna_con_ano = None
        for col in range(hoja_dcf.max_column, 1, -1):  # De derecha a izquierda
            if isinstance(hoja_dcf.cell(row=2, column=col).value, int):  # Si es un número (año)
                ultima_columna_con_ano = col
                break  # Encontramos la última columna con un año, podemos salir del bucle

        # Obtener la letra de la columna final correctamente
        col_fin = obtener_letra_columna(ultima_columna_con_ano)
        col_false = obtener_letra_columna(ultima_columna_con_ano - 1)
        col_inicio = obtener_letra_columna(2)  # Primera columna de años

        # Generar la fórmula correctamente con la columna final ajustada
        formula_7L = (
            f"=IF({col_fin}3*FORECAST({col_fin}2,{col_inicio}3:{col_fin}3,{col_inicio}2:{col_fin}2)>0,"
            f"FORECAST({col_fin}2,{col_inicio}3:{col_fin}3,{col_inicio}2:{col_fin}2),"
            f"{col_false}3)*$X$1 * (1+0.02)/(0.06-0.02)"
        )

        # Asignar la fórmula en L7
        hoja_dcf['L7'] = formula_7L
        # Guardar los cambios
        wb.save(ruta_destino)

        # Reemplazar números por años desde 2035 hacia atrás
        ultimo_año = 2035
        for i in range(ultima_columna + 5, 1, -1):
            hoja_dcf.cell(row=2, column=i, value=ultimo_año)
            ultimo_año -= 1

        # Buscar la primera columna con el año 2025 en la fila 2
        primer_ano_columna = None
        for col in range(2, hoja_dcf.max_column + 1):  # Empieza desde la columna 2 (B) hasta la última columna
            if hoja_dcf.cell(row=2, column=col).value == 2025:  # Verifica si la celda en fila 2 es 2025
                primer_ano_columna = col
                break  # Detenerse al encontrar 2025

        # Buscar el último año en la fila 2
        ultimo_ano_columna = None
        for col in range(hoja_dcf.max_column, 1, -1):  # Empieza desde la última columna hacia atrás
            if isinstance(hoja_dcf.cell(row=2, column=col).value, int):  # Verifica si es un número (año)
                ultimo_ano_columna = col
                break  # Detenerse al encontrar el último año

        # Obtener las letras de las columnas
        col_inicio = obtener_letra_columna(primer_ano_columna)  # Columna con el primer año (2025)
        col_fin = obtener_letra_columna(ultimo_ano_columna)    # Columna con el último año

        # Crear la fórmula para la celda L8, ajustada dinámicamente
        formula_vna = (
            f"=ARRAY_CONSTRAIN(ARRAYFORMULA(NPV (0.06;{col_inicio}3:{col_fin}3));1;1)"
        )

        # Asignar la fórmula en la celda L8
        hoja_dcf['L8'] = formula_vna

        num_columnas = ultimo_ano_columna - primer_ano_columna

        # Agregar ceros en la fila 16 hasta la primera fórmula
        for col in range(primer_ano_columna, ultimo_ano_columna):
            hoja_dcf.cell(row=16, column=col, value=0)  # Colocar ceros en cada celda de la fila 16

        # Obtener la fórmula que está en la celda L7
        formula_l7 = hoja_dcf['L7'].value

        # Aplicar la misma fórmula en la fila 16, columna correspondiente
        hoja_dcf.cell(row=16, column=ultimo_ano_columna, value=formula_l7)

        # Buscar la primera columna con 0 en la fila 16
        primer_0_columna = None
        for col in range(2, hoja_dcf.max_column + 1):  # Empieza desde la columna 2 (B) hasta la última columna
            if hoja_dcf.cell(row=16, column=col).value == 0:  # Verifica si es 0
                primer_0_columna = col
                break  # Detenerse al encontrar el primer 0

        # Buscar la última columna con un valor distinto de 0 en la fila 16 (la fórmula)
        ultimo_valor_columna = None
        for col in range(primer_0_columna, hoja_dcf.max_column + 1):  # Empieza desde el primer 0 encontrado
            if hoja_dcf.cell(row=16, column=col).value != 0:  # Verifica si el valor es distinto de 0
                ultimo_valor_columna = col
                break  # Detenerse al encontrar el primer valor distinto de 0

        # Obtener las letras de las columnas
        col_inicio = obtener_letra_columna(primer_0_columna)  # Columna con el primer 0
        col_fin = obtener_letra_columna(ultimo_valor_columna)  # Columna con el primer valor distinto de 0

        # Crear la fórmula para la celda L9, ajustada dinámicamente
        formula_vna = f"=ARRAY_CONSTRAIN(ARRAYFORMULA(NPV(0.06, {col_inicio}16:{col_fin}16)), 1, 1)"

        # Asignar la fórmula en la celda L9
        hoja_dcf['L9'] = formula_vna

        # Guardar los cambios finales
        wb.save(ruta_destino)
        print(f"✅ Archivo modificado y guardado: {ruta_destino}")
        return ruta_destino  # Retorna la ruta del archivo modificado

    except Exception as e:
        print(f"❌ Error al procesar {ruta_origen}: {e}")
        return None

    

def crear_google_sheet(nombre_archivo, ruta_excel):
    """
    Convierte el archivo Excel (ubicado en ruta_excel) a una hoja de cálculo de Google Sheets,
    la configura con acceso público y devuelve el enlace.
    """
    try:
        # Preparar el contenido del archivo Excel para subirlo y convertirlo.
        media = MediaFileUpload(
            ruta_excel,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            resumable=True
        )
        
        # Definir los metadatos para la conversión a Google Sheets
        sheet_metadata = {
            'name': nombre_archivo,
            'mimeType': 'application/vnd.google-apps.spreadsheet'
        }
        
        # Crear y convertir el archivo a Google Sheets
        sheet = drive_service.files().create(
            body=sheet_metadata,
            media_body=media,
            fields='id'
        ).execute()
        
        # Obtener el ID de la hoja de cálculo
        sheet_id = sheet.get('id')

        # Configurar permisos para que sea de acceso público (solo lectura)
        permiso_publico = {
            'type': 'anyone',
            'role': 'writer'
        }
        drive_service.permissions().create(fileId=sheet_id, body=permiso_publico).execute()

        # Generar el enlace de visualización
        enlace = f"https://docs.google.com/spreadsheets/d/{sheet_id}/view"
        print(f"📄 Hoja de cálculo creada y hecha pública: {enlace}")
        return enlace

    except Exception as e:
        print(f"❌ Error al crear la hoja de cálculo para {nombre_archivo}: {e}")
        return None

def actualizar_google_sheets(nombre_archivo, enlace_doc):
    """Sube los datos a Google Sheets en la hoja 'NuevosResultados'."""
    try:
        hoja = client.open_by_key(ID_HOJA_ACTUALIZACION).worksheet("NuevosResultados")

        # Buscar la primera fila vacía
        fila_vacia = len(hoja.col_values(1)) + 1

        # Insertar el nombre del archivo en la columna A y el enlace en la columna C
        hoja.update_cell(fila_vacia, 1, nombre_archivo)  # Columna A
        hoja.update_cell(fila_vacia, 3, enlace_doc)        # Columna C

        print(f"✅ Google Sheets actualizado en fila {fila_vacia}")

    except Exception as e:
        print(f"❌ Error al actualizar Google Sheets: {e}")

def procesar_archivos():
    """
    Procesa todos los archivos en la carpeta de origen, 
    modifica el Excel, lo guarda en 'F:\\ExcelsActualizados',
    lo convierte a Google Sheets (publicado) y actualiza la hoja 'NuevosResultados'.
    """
    carpeta_origen = r"F:\Excels"
    carpeta_destino = r"F:\ExcelsActualizados"

    os.makedirs(carpeta_destino, exist_ok=True)

    for archivo in os.listdir(carpeta_origen):
        if archivo.endswith(".xlsx"):
            ruta_origen = os.path.join(carpeta_origen, archivo)
            ruta_destino = os.path.join(carpeta_destino, archivo)

            # Copiar el archivo original a la carpeta de Excel actualizados
            shutil.copy2(ruta_origen, ruta_destino)
            archivo_modificado = modificar_excel(ruta_destino, ruta_destino)

            if archivo_modificado:
                # Convertir el Excel modificado a Google Sheets (convirtiendo y haciendo público)
                enlace_doc = crear_google_sheet(archivo, ruta_destino)
                if enlace_doc:
                    actualizar_google_sheets(archivo, enlace_doc)

if __name__ == "__main__":
    procesar_archivos()
