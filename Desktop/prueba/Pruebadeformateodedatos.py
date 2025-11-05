from concurrent.futures import ThreadPoolExecutor
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from lxml import etree
import pandas as pd
import numpy as np
import time
from openpyxl import load_workbook
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import re
import xlwings as xw
import gc
from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive
import os
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
import threading
from concurrent.futures import ThreadPoolExecutor
import time
from selenium.webdriver.edge.service import Service
from selenium.webdriver.edge.options import Options

lock = threading.Lock()


def setup_drive():
    SCOPES = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
    SERVICE_ACCOUNT_FILE = 'C:/Users/relim/Desktop/prueba/durable-binder-432321-q4-5f1ef9e64ea1.json'
    
    creds = service_account.Credentials.from_service_account_file(
        SERVICE_ACCOUNT_FILE, scopes=SCOPES)
    
    service = build('drive', 'v3', credentials=creds)
    return service

def upload_and_get_shared_link(service, file_path):
     with lock:
        file_metadata = {
            'name': os.path.basename(file_path),
            'mimeType': 'application/vnd.google-apps.spreadsheet'
        }
        media = MediaFileUpload(file_path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        file = service.files().create(body=file_metadata, media_body=media, fields='id').execute()
        file_id = file.get('id')
        
        # Cambiar permisos para compartir el archivo
        permissions = {
            'type': 'anyone',
            'role': 'writer'
        }
        service.permissions().create(fileId=file_id, body=permissions).execute()
        
        # Obtener enlace compartido
        file_info = service.files().get(fileId=file_id, fields='webViewLink').execute()
        shared_link = file_info.get('webViewLink')
        print(f"Enlace compartido generado: {shared_link}")
        return shared_link


def setup_driver():
    edge_options = Options()
    edge_options.add_argument("--headless")
    edge_options.add_argument("--start-maximized")
    edge_options.add_argument("--blink-settings=imagesEnabled=false")
    
    # Ruta al EdgeDriver descargado
    edge_driver_path = r'C:\Users\relim\Desktop\prueba\driver\msedgedriver.exe'
    
    driver = webdriver.Edge(service=Service(edge_driver_path), options=edge_options)
    return driver    

def login(driver):
    driver.get("https://www.marketscreener.com/login/")
    username_field = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, '/html/body/div[1]/div/div/main/div/div/div/div/form/div[1]/input'))
    )
    password_field = driver.find_element(By.XPATH, '/html/body/div[1]/div/div/main/div/div/div/div/form/div[2]/input')
    username_field.send_keys("leojosemartin@gmail.com")
    password_field.send_keys("Papaleo111.")
    password_field.send_keys(Keys.RETURN)
    time.sleep(2)  # Espera a que el inicio de sesión se complete

def capture_table(driver, url, xpath):
    driver.get(url)
    time.sleep(2)
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, xpath)))
    html = driver.page_source
    dom = etree.HTML(html)
    rows = dom.xpath(xpath)
    if not rows:
        print(f"No se encontraron filas con el XPath: {xpath}")
        return []
    table_data = []
    for row in rows:
        cells = row.xpath(".//td | .//th")
        row_data = [''.join(cell.xpath(".//text()")).strip() for cell in cells]
        table_data.append(row_data)
    return table_data

def convert_value(value):
        if isinstance(value, str):
            value = value.strip()
            if value == '-':
                return 0
            if value.endswith('M'):
                return int(float(value[:-1].replace(',', '')) * 1)
            elif value.endswith('B'):
                return int(float(value[:-1].replace(',', '')) * 1_000)
            elif value.endswith('K'):
                return int(float(value[:-1].replace(',', '')) * 0.1)
            elif value.replace(',', '').lstrip('-').isdigit():  # Permitir números negativos
                return int(value.replace(',', ''))
            else:
                return value
        return value

def convert_dataframe(df):
    for col in df.columns:
        df[col] = df[col].apply(convert_value)
    return df

def process_stock_code(stock_code, forecast_years, x2, drive):
    driver = setup_driver()
    login(driver)
    try:
        base_url = f"https://www.marketscreener.com/quote/stock/{stock_code}/finances"
        cash_flow_data = capture_table(
            driver, f"{base_url}-cash-flow-statement/",
            '//*[@id="horizontalFinancialTableN1_3"]/tbody/tr'
        )
        balance_sheet_data = capture_table(
            driver, f"{base_url}-balance-sheet/",
            '//*[@id="horizontalFinancialTableN1_2"]/tbody/tr'
        )
        financial_data = capture_table(
            driver, f"{base_url}-income-statement/",
            '//*[@id="horizontalFinancialTableN1_1"]/tbody/tr'
        )
        ratios_data = capture_table(
            driver, f"{base_url}-ratios/",
            '//*[@id="horizontalFinancialTableN1_5"]/tbody/tr'
        )
    finally:
        driver.quit()


    df_cash_flow = convert_dataframe(pd.DataFrame(cash_flow_data))
    df_balance_sheet = convert_dataframe(pd.DataFrame(balance_sheet_data))
    df_financial = convert_dataframe(pd.DataFrame(financial_data))
    df_ratios = convert_dataframe(pd.DataFrame(ratios_data))

    excel_path = rf"C:\Users\relim\Desktop\prueba\Resultados\{stock_code}.xlsx"
    with pd.ExcelWriter(excel_path, engine='xlsxwriter') as writer:
        df_cash_flow.to_excel(writer, sheet_name='Cash Flow Statement', index=False, header=False)
        df_balance_sheet.to_excel(writer, sheet_name='Balance Sheet', index=False, header=False)
        df_financial.to_excel(writer, sheet_name='Financial', index=False, header=False)
        df_ratios.to_excel(writer, sheet_name='Ratios', index=False, header=False)

    def extract_values(df, search_str):
        for idx, row in df.iterrows():
            if any(search_str in str(cell) for cell in row):
                return row[1:].values
        print(f"No se encontró '{search_str}' en la tabla.")
        return []

    unlevered_free_cash_flow = extract_values(df_cash_flow, 'Unlevered Free Cash Flow')
    total_debt = extract_values(df_balance_sheet, 'Net Debt')
    diluted_shares = extract_values(df_financial, 'Diluted Weighted Average Shares Outstanding')
    cash_equivalents = extract_values(df_balance_sheet, 'Cash and Equivalents')

    max_length = max(len(unlevered_free_cash_flow), len(total_debt), len(diluted_shares), len(cash_equivalents))
    unlevered_free_cash_flow = np.pad(unlevered_free_cash_flow, (0, max_length - len(unlevered_free_cash_flow)), constant_values=np.nan)
    total_debt = np.pad(total_debt, (0, max_length - len(total_debt)), constant_values=np.nan)
    diluted_shares = np.pad(diluted_shares, (0, max_length - len(diluted_shares)), constant_values=np.nan)
    cash_equivalents = np.pad(cash_equivalents, (0, max_length - len(cash_equivalents)), constant_values=np.nan)

    df_dcf = pd.DataFrame({
        'Unlevered Free Cash Flow': unlevered_free_cash_flow,
        'Total Debt': total_debt,
        'Diluted Shares Outstanding': diluted_shares,
        'Cash and Equivalents': cash_equivalents
    }).T

    last_year = int(df_dcf.columns[-1])
    new_years = [last_year + i for i in range(1, forecast_years + 1)]
    df_dcf = df_dcf.reindex(columns=list(df_dcf.columns) + new_years)

    with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a') as writer:
        df_dcf.to_excel(writer, sheet_name='DCF', startrow=1, index=True, header=True)

    wb = load_workbook(excel_path)
    ws = wb['DCF']
    ws['X2'] = x2

        
    unlevered_last_col = len(df_dcf.columns) - forecast_years
    for i in range(forecast_years):
        current_col = unlevered_last_col + 2 + i
        previous_col = unlevered_last_col + 1 + i

        if i == 0:
            formula = (
                f'=IF({ws.cell(row=3, column=unlevered_last_col + 1).coordinate}*'
                f'FORECAST({ws.cell(row=2, column=current_col).coordinate},'
                f'{ws.cell(row=3, column=2).coordinate}:{ws.cell(row=3, column=unlevered_last_col + 1).coordinate},'
                f'{ws.cell(row=2, column=2).coordinate}:{ws.cell(row=2, column=unlevered_last_col + 1).coordinate})>0,'
                f'FORECAST({ws.cell(row=2, column=current_col).coordinate},'
                f'{ws.cell(row=3, column=2).coordinate}:{ws.cell(row=3, column=unlevered_last_col + 1).coordinate},'
                f'{ws.cell(row=2, column=2).coordinate}:{ws.cell(row=2, column=unlevered_last_col + 1).coordinate}),'
                f'{ws.cell(row=3, column=unlevered_last_col + 1).coordinate})*$X$2'
            )
        else:
            formula = (
                f'=IF({ws.cell(row=3, column=unlevered_last_col + 1 + i).coordinate}*'
                f'FORECAST({ws.cell(row=2, column=current_col).coordinate},'
                f'{ws.cell(row=3, column=2).coordinate}:{ws.cell(row=3, column=current_col - 1).coordinate},'
                f'{ws.cell(row=2, column=2).coordinate}:{ws.cell(row=2, column=current_col - 1).coordinate})>0,'
                f'FORECAST({ws.cell(row=2, column=current_col).coordinate},'
                f'{ws.cell(row=3, column=2).coordinate}:{ws.cell(row=3, column=current_col - 1).coordinate},'
                f'{ws.cell(row=2, column=2).coordinate}:{ws.cell(row=2, column=current_col - 1).coordinate}),'
                f'{ws.cell(row=3, column=unlevered_last_col + 1 + i).coordinate})*$X$2'
            )
        ws.cell(row=3, column=current_col, value=formula)

    valor = ws.cell(row=3, column=current_col).value
    ws['K7'] = "TERMINAL VALUE"
    ws['L7'] = f'{valor} * (1 + 0.02) / (0.06 - 0.02)'
    terminal_value = ws['L7'].value

    start_col = unlevered_last_col + 2
    end_col = current_col
    terminal_col = end_col + 1

    def col_num_to_letter(col_num):
        col_letter = ""
        while col_num > 0:
            col_num, remainder = divmod(col_num - 1, 26)
            col_letter = chr(65 + remainder) + col_letter
        return col_letter

    start_col_letter = col_num_to_letter(start_col)
    end_col_letter = col_num_to_letter(end_col)
    terminal_col_letter = col_num_to_letter(terminal_col)
    next_start_col_letter = col_num_to_letter(start_col + 1)

    ws['K8'] = 'NPV of FCF'
    ws['L8'] = f'=NPV(0.06, {start_col_letter}3:{end_col_letter}3)'

    for col in range(start_col + 1, end_col + 1):
        col_letter = col_num_to_letter(col)
        ws[f'{col_letter}16'] = 0

    ws[f'{terminal_col_letter}16'] = terminal_value

    ws['K9'] = "NPV of TV"
    ws['L9'] = f'=NPV(0.06, {next_start_col_letter}16:{terminal_col_letter}16)'

    ws['K10'] = 'Total EV'
    ws['L10'] = '=L8 + L9'

    ws['K11'] = 'Net Debt'
    for col in reversed(range(2, ws.max_column + 1)):
        if isinstance(ws.cell(row=4, column=col).value, (int, float)):
            last_column_net_debt = col
            break
    ws['L11'] = ws.cell(row=4, column=last_column_net_debt).value

    ws['K12'] = 'Equity'
    ws['L12'] = f'=L10 - L11'

    ws['K13'] = 'Shares Outstanding'
    for col in reversed(range(2, ws.max_column + 1)):
        if isinstance(ws.cell(row=5, column=col).value, (int, float)):
            last_column_shares_outstanding = col
            break
    ws['L13'] = ws.cell(row=5, column=last_column_shares_outstanding).value

    ws['K14'] = 'Target Price'
    ws['L14'] = f'=L12/L13'

    wb.save(excel_path)

    shared_link = upload_and_get_shared_link(drive, excel_path)

    scope = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
    creds = ServiceAccountCredentials.from_json_keyfile_name('C:/Users/relim/Desktop/prueba/durable-binder-432321-q4-5f1ef9e64ea1.json', scope)
    client = gspread.authorize(creds)
    sheet = client.open_by_key('193GrDtaNCWBkWwliitjSepbLlAcNlwQfUZp66WN6xyw').worksheet('Results')

    stock_name_clean = re.sub(r'[^A-Za-z]+', ' ', stock_code).replace('-', ' ').strip()

    # Usar xlwings para abrir el archivo y calcular las fórmulas
    app = xw.App(visible=False)
    workbook = app.books.open(excel_path)
    workbook.save()  # Esto fuerza la evaluación de fórmulas
    workbook.close()
    app.quit()

    # Leer el archivo Excel para obtener el valor calculado
    df = pd.read_excel(excel_path, sheet_name='DCF', header=None)
    l14_value = df.at[13, 11]  # Filas y columnas están basadas en el índice 0

    stock_codes_df = pd.read_excel(file_path, usecols=[0, 1], header=None)
    stock_name_from_file = stock_codes_df[stock_codes_df[0] == stock_code].iloc[0, 1] if not stock_codes_df[stock_codes_df[0] == stock_code].empty else ''

    try:
        # Buscar la fila correspondiente para el código de stock
        cell_list = sheet.col_values(1)  # Obtener todos los valores de la columna A
        if stock_code in cell_list:
            row_index = cell_list.index(stock_code) + 1  # Encuentra la fila existente
        else:
            # Si el código no está en la hoja, añade al final
            row_index = len(cell_list) + 1

        link = f'https://www.marketscreener.com/quote/stock/{stock_code}'
        formula_link = f'=HYPERLINK("{link}", "{stock_code}")'     
        
        # Actualizar la fila específica con los nuevos datos
        sheet.update_cell(row_index, 1, formula_link)
        sheet.update_cell(row_index, 2, stock_name_clean)
        sheet.update_cell(row_index, 3, l14_value)
        sheet.update_cell(row_index, 4, stock_name_from_file)
        sheet.update_cell(row_index, 8, shared_link)
        
        print(f"Datos actualizados: {stock_code}, {stock_name_clean}, {l14_value}")
    except Exception as e:
        print(f"Error al actualizar datos en Google Sheets: {e}")
    


def process_all_stock_codes(file_path, forecast_years, x2):
    stock_codes = pd.read_excel(file_path, usecols=[0], header=None).iloc[:, 0].tolist()
    drive = setup_drive() 
    with ThreadPoolExecutor(max_workers=3) as executor:
        # Usar submit para enviar las tareas
        futures = [executor.submit(process_stock_code, code, forecast_years, x2, drive) for code in stock_codes]

        # Esperar a que se completen todas las tareas
        for future in futures:
            future.result()
            time.sleep(2)

    gc.collect()




file_path = r'C:\Users\relim\Desktop\prueba\StockCodes.xlsx'
forecast_years = 7
x2 = 1

process_all_stock_codes(file_path, forecast_years, x2)