import os
import shutil
import openpyxl
from openpyxl.utils import get_column_letter

def obtener_letra_columna(indice):
    return get_column_letter(indice)

def modificar_excel(ruta_origen, ruta_destino):
    wb = openpyxl.load_workbook(ruta_origen)
    if "DCF" not in wb.sheetnames:
        print(f"La hoja 'DCF' no está en {ruta_origen}")
        return
    
    hoja_dcf = wb["DCF"]
    
    # Borrar el contenido de la columna X (24ª columna)
    for fila in range(1, hoja_dcf.max_row + 1):
        hoja_dcf.cell(row=fila, column=24).value = None
    
    # Guardar el archivo y recargarlo
    wb.save(ruta_destino)
    wb = openpyxl.load_workbook(ruta_destino)
    hoja_dcf = wb["DCF"]

    # Agregar el valor 1 en la celda X1
    hoja_dcf['X1'] = 1
    
    # Cambiar el multiplicador $X$2 por $X$1 en todas las celdas de la hoja "DCF"
    for fila in hoja_dcf.iter_rows():
        for celda in fila:
            if celda.value and isinstance(celda.value, str) and "$X$2" in celda.value:
                celda.value = celda.value.replace("$X$2", "$X$1")

    # Encontrar la última columna con datos en la fila 2
    ultima_columna = hoja_dcf.max_column
    while hoja_dcf.cell(row=2, column=ultima_columna).value is None and ultima_columna > 1:
        ultima_columna -= 1
    
    ultimo_año = hoja_dcf.cell(row=2, column=ultima_columna).value
    if ultimo_año is None:
        ultimo_año = 11  # Asignar un valor predeterminado si es None
    
    # Agregar numeración continua en la fila 2, a partir del último número encontrado
    for i in range(ultima_columna + 1, ultima_columna + 6):
        hoja_dcf.cell(row=2, column=i, value=ultimo_año + (i - ultima_columna))
    
    # Aplicar la fórmula en la última columna con datos en la fila 3
    ultima_columna_formula = ultima_columna
    while hoja_dcf.cell(row=3, column=ultima_columna_formula).value is None and ultima_columna_formula > 1:
        ultima_columna_formula -= 1
    
    letra_col_formula = obtener_letra_columna(ultima_columna_formula)
    
    # Modificar las fórmulas con el multiplicador $X$1
    for i in range(ultima_columna_formula + 1, ultima_columna_formula + 6):
        letra_col_actual = obtener_letra_columna(i)
        letra_col_anterior = obtener_letra_columna(i - 1)
        letra_col_inicio = obtener_letra_columna(2)
        
        formula = (
            f"=IF({letra_col_anterior}3*FORECAST({letra_col_actual}2," 
            f"{letra_col_inicio}3:{letra_col_anterior}3," 
            f"{letra_col_inicio}2:{letra_col_anterior}2)>0," 
            f"FORECAST({letra_col_actual}2," 
            f"{letra_col_inicio}3:{letra_col_anterior}3," 
            f"{letra_col_inicio}2:{letra_col_anterior}2)," 
            f"{letra_col_anterior}3)*$X$1"
        )
        hoja_dcf.cell(row=3, column=i, value=formula)
    
    # Guardar el archivo modificado
    wb.save(ruta_destino)
    print(f"Archivo guardado: {ruta_destino}")

def procesar_archivos():
    carpeta_origen = r"F:\\Excels"
    carpeta_destino = r"F:\\ExcelsActualizados"
    
    if not os.path.exists(carpeta_destino):
        os.makedirs(carpeta_destino)
    
    for archivo in os.listdir(carpeta_origen):
        if archivo.endswith(".xlsx"):
            ruta_origen = os.path.join(carpeta_origen, archivo)
            ruta_destino = os.path.join(carpeta_destino, archivo)
            
            # Copiar archivo antes de modificarlo
            shutil.copy2(ruta_origen, ruta_destino)
            modificar_excel(ruta_destino, ruta_destino)

if __name__ == "__main__":
    procesar_archivos()
