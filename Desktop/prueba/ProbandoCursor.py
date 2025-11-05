from concurrent.futures import ThreadPoolExecutor
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.keys import Keys
from lxml import etree
import pandas as pd
import numpy as np
import time
from openpyxl import load_workbook
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import re
import xlwings as xw
import gc
from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive
import os
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
import threading
from concurrent.futures import ThreadPoolExecutor
import time
import requests
import shutil 
import tempfile
import yfinance as yf
import traceback


lock = threading.Lock()


def setup_drive():
    SCOPES = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
    SERVICE_ACCOUNT_FILE = 'C:/Users/relim/Desktop/prueba/durable-binder-432321-q4-5f1ef9e64ea1.json'
    
    creds = service_account.Credentials.from_service_account_file(
        SERVICE_ACCOUNT_FILE, scopes=SCOPES)
    
    service = build('drive', 'v3', credentials=creds)
    return service

def upload_and_get_shared_link(service, file_path):
     with lock:
        print(f"[LOG] Intentando subir el archivo: {file_path}")
        try:
            file_metadata = {
                'name': os.path.basename(file_path),
                'mimeType': 'application/vnd.google-apps.spreadsheet'
            }
            media = MediaFileUpload(file_path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            file = service.files().create(body=file_metadata, media_body=media, fields='id').execute()
            file_id = file.get('id')
            print(f"[LOG] Archivo subido, ID: {file_id}")
            
            # Cambiar permisos para compartir el archivo
            permissions = {
                'type': 'anyone',
                'role': 'writer'
            }
            service.permissions().create(fileId=file_id, body=permissions).execute()
            print(f"[LOG] Permisos cambiados a 'anyone' para el archivo: {file_id}")
            
            # Obtener enlace compartido
            file_info = service.files().get(fileId=file_id, fields='webViewLink').execute()
            shared_link = file_info.get('webViewLink')
            print(f"[LOG] Enlace compartido generado: {shared_link}")
            return shared_link
        except Exception as e:
            print(f"[ERROR] Error subiendo archivo a Google Drive: {e}")
            return None


def setup_driver():
    try:
        chrome_options = Options()
        chrome_options.add_argument("--headless=new")
        # User-Agent personalizado para simular un navegador real
        chrome_options.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36")
        # Flags anti-detección
        chrome_options.add_argument("--disable-blink-features=AutomationControlled")
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        chrome_options.add_experimental_option('useAutomationExtension', False)
        # Quitamos todas las opciones que puedan afectar la visibilidad
        chrome_options.add_argument("--start-maximized")  # Maximizar ventana
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--disable-extensions")
        chrome_options.add_argument("--disable-popup-blocking")
        chrome_options.add_argument("--blink-settings=imagesEnabled=false")
        chrome_options.add_argument('--ignore-certificate-errors')
        chrome_options.add_argument('--disable-web-security')
        chrome_options.add_argument('--allow-running-insecure-content')
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        
        # Configurar el servicio de Chrome
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=chrome_options)
        driver.set_page_load_timeout(30)
        return driver
    except Exception as e:
        print(f"Error al configurar el driver de Chrome: {e}")
        raise

def login(driver):
    try:
        print("   - Navegando a la página de login...")
        driver.get("https://www.marketscreener.com/login/")
        time.sleep(5)
        
        print("   - Buscando campos de login...")
        username_field = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.XPATH, '/html/body/div[1]/div/div/main/div/div/div/div/form/div[1]/input'))
        )
        password_field = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.XPATH, '/html/body/div[1]/div/div/main/div/div/div/div/form/div[2]/input'))
        )
        
        print("   - Ingresando credenciales...")
        username_field.clear()
        username_field.send_keys("leojosemartin@gmail.com")
        time.sleep(1)
        
        password_field.clear()
        password_field.send_keys("Papaleo111.")
        time.sleep(1)
        
        print("   - Enviando formulario...")
        password_field.send_keys(Keys.RETURN)
        time.sleep(5)
        
        # Verificar si el login fue exitoso
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//a[contains(@href, '/logout')]"))
        )
        print("   ✓ Login exitoso")
            
    except Exception as e:
        print(f"   ✗ Error durante el login: {str(e)}")
        raise

def capture_table(driver, url, xpath, header_xpath=None):
    try:
        print(f"\n   - Navegando a la URL...")
        driver.get(url)
        time.sleep(3)
        
        print(f"   - Desplazando página...")
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(2)
        
        print(f"   - Esperando elemento XPath: {xpath}")
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, xpath)))
        
        print(f"   - Extrayendo HTML...")
        html = driver.page_source
        dom = etree.HTML(html)
        
        headers = []
        if header_xpath:
            print(f"   - Extrayendo encabezados...")
            header_elements = dom.xpath(header_xpath)
            headers = [''.join(header.xpath(".//text()")).strip() for header in header_elements]
            print(f"   ✓ Encabezados encontrados: {len(headers)}")

        print(f"   - Buscando filas...")
        rows = dom.xpath(xpath)
        if not rows:
            print(f"   ⚠️ No se encontraron filas con el XPath: {xpath}")
            return []

        print(f"   - Procesando {len(rows)} filas...")
        table_data = []
        for row in rows:
            cells = row.xpath(".//td | .//th")
            row_data = [''.join(cell.xpath(".//text()")).strip() for cell in cells]
            table_data.append(row_data)

        if headers:
            table_data.insert(0, headers)

        print(f"   ✓ Datos capturados exitosamente: {len(table_data)} filas")
        return table_data

    except Exception as e:
        print(f"   ✗ Error al capturar datos: {str(e)}")
        return []


import os
import shutil

def cleanup_temp_folder(temp_folder):
    try:
        # Elimina todos los archivos y directorios en la carpeta Temp
        shutil.rmtree(temp_folder, ignore_errors=True)
        print(f"Carpeta temporal {temp_folder} limpiada exitosamente.")
    except Exception as e:
        print(f"Error al eliminar el contenido de la carpeta temporal: {e}")



EXCEPTIONS = {'0', "Diluted Shares Outstanding", "Basic Weighted Average Shares Outstanding", "Diluted Weighted Average Shares Outstanding", "Basic Weighted Average Shares Outstanding "
              "Diluted Weighted Average Shares Outstanding ","Basic Weighted Average Shares Outstanding ", "Payout Ratio", "American Depositary Receipts Ratio (ADR)",
                "American Depositary Receipts Ratio", "Effective Tax Rate - (Ratio)", "Effective Tax Rate", "Net EPS - Basic", "Basic EPS - Continuing Operations", "Net EPS", "Basic EPS", "Net EPS - Diluted", "Net EPS", "Diluted EPS - Continuing Operations", "Diluted EPS",
                "Normalized Basic EPS", "Normalized Diluted EPS",
                "Diluted Weighted Average Shares Outstanding ", "Basic Weighted Average Shares Outstanding ",
                "Diluted Shares Outstanding", "Diluted Weighted Average Shares Outstanding","Basic Weighted Average Shares Outstanding", "Fiscal period","Fiscal Period", "Fiscal Period: January", "Fiscal Period: February", "Fiscal Period: March", "Fiscal Period: April", "Fiscal Period: May", "Fiscal Period: June", "Fiscal Period: July", "Fiscal Period: August", "Fiscal Period: September", "Fiscal Period: October", "Fiscal Period: November", "Fiscal Period: December" }



def convert_value(value):
    if isinstance(value, str):
        value = value.strip()
        if value in EXCEPTIONS:
            return value
        if value == '-':
            return 0
        if value.endswith('M'):
            return int(float(value[:-1].replace(',', '')) * 1) # Multiplicar por 1 millón
        elif value.endswith('B'):
            return int(float(value[:-1].replace(',', '')) * 1_000)  # Multiplicar por 1 mil millones
        elif value.endswith('K'):
            return int(float(value[:-1].replace(',', '')) * 0.1)  # Multiplicar por 1 mil
        elif value.replace(',', '').lstrip('-').isdigit():  # Permitir números negativos
            return int(value.replace(',', ''))
        else:
            return value
    return value

def convert_dataframe(df):
    for col in df.columns:
        if col != 0:  # Aplicar conversión solo si la columna no es la columna 0
            df[col] = df[col].apply(convert_value)
    return df


def get_exchange_rate(from_currency, to_currency="USD"):
    """Obtiene la tasa de cambio desde la moneda de origen a la moneda de destino."""
    url = f"https://api.exchangerate-api.com/v4/latest/{from_currency}"
    response = requests.get(url)
    data = response.json()
    return data['rates'].get(to_currency, 1)  


def convert_to_usd(df, exchange_rate):
    # Convertir solo columnas numéricas después de aplicar map
    numeric_cols = df.columns[df.apply(lambda col: col.apply(lambda x: isinstance(x, (int, float)) or (isinstance(x, str) and x.replace(',', '').isdigit())).any())]

    # Multiplicar por la tasa de cambio solo si la fila no tiene un valor en la columna A que sea una excepción
    for index in df.index:
        if df.at[index, df.columns[0]] not in EXCEPTIONS:
            df.loc[index, numeric_cols] = df.loc[index, numeric_cols].apply(lambda x: x * exchange_rate if isinstance(x, (int, float)) else x)
    
    return df


def extract_currency_from_url(base_url):
    # Agregar el sufijo '-income-statement/' a la URL
    full_url = f"{base_url}-income-statement/"

    # Crear una nueva instancia del navegador
    driver = setup_driver()

    # Navegar a la URL completa
    driver.get(full_url)

    extracted_text = None  # Inicializar la variable para almacenar el texto

    try:
        # Utiliza un selector CSS para encontrar el elemento que contiene la moneda
        wait = WebDriverWait(driver, 20)  # Aumentar el tiempo de espera a 20 segundos
        element = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, ".js-selectedCurrency.px-10")))
        extracted_text = element.text.strip()  # Usar strip() para limpiar espacios extra
        print(f"Texto extraído desde el elemento en {full_url}: {extracted_text}")

    except Exception as e:
        print(f"Error inesperado: {e}")

    finally:
        driver.quit()  # Cerrar el navegador al finalizar

    return extracted_text  # Retornar el texto extraído


def tipo_empresa_from_xpath(driver, tipo_xpath):
    try:
        # Esperar a que el elemento esté presente
        empresa_element = WebDriverWait(driver, 2).until(
            EC.presence_of_element_located((By.XPATH, tipo_xpath))
        )
        # Extraer el texto del elemento
        empresa_text = empresa_element.get_attribute('textContent').strip()
        
        if empresa_text:
            print(f"Texto extraído del XPath: '{empresa_text}'")
        else:
            print("No se pudo extraer el texto, el resultado es vacío.")
        
        return empresa_text
    except Exception as e:
        print(f"Error al extraer el texto desde el XPath: {str(e)}")
        return "No encontrado"

def ticker_market(driver, ticker_xpath):
    try:
        # Esperar a que el elemento esté presente
        ticker_element = WebDriverWait(driver, 2).until(
            EC.presence_of_element_located((By.XPATH, ticker_xpath))
        )
        # Extraer el texto del elemento
        ticker_element = ticker_element.get_attribute('textContent').strip()
        
        if ticker_element:
            print(f"Texto extraído del XPath: '{ticker_element}'")
        else:
            print("No se pudo extraer el texto, el resultado es vacío.")
        
        return ticker_element
    except Exception as e:
        print(f"Error al extraer el texto desde el XPath: {str(e)}")
        return "No encontrado"

def process_stock_code(stock_code, forecast_years, x2, drive):
    driver = None
    try:
        print(f"\n{'='*50}")
        print(f"Iniciando procesamiento para {stock_code}")
        print(f"{'='*50}\n")
        
        print("1. Configurando el driver de Chrome...")
        driver = setup_driver()
        print("✓ Driver configurado exitosamente")
        
        print("\n2. Iniciando login...")
        # login(driver)
        # time.sleep(5)
        # print("✓ Login exitoso")

        base_url = f"https://www.marketscreener.com/quote/stock/{stock_code}/finances"
        base_url2 = f"https://www.marketscreener.com/quote/stock/{stock_code}/valuation/"

        print("\n3. Capturando datos de valoración...")
        valuation_data = []
        try:
            valuation_data = capture_table(
                driver, f"{base_url2}",
                '//*[@id="valuationEnterpriseTable"]/tbody/tr',
                header_xpath='//*[@id="valuationEnterpriseTable_wrapper"]/div/div[1]/div/table/thead/tr/th'
            )
            if not valuation_data:
                print("⚠️ No se encontraron datos de valoración")
            else:
                print(f"✓ Datos de valoración capturados: {len(valuation_data)} filas")
        except Exception as e:
            print(f"✗ Error al capturar valuation data: {str(e)}")

        print("\n4. Capturando datos financieros...")
        print("   - Cash Flow Statement")
        cash_flow_data = capture_table(
            driver, f"{base_url}-cash-flow-statement/",
            '//*[@id="horizontalFinancialTableN1_3"]/tbody/tr',
            header_xpath='//*[@id="horizontalFinancialTableN1_3_wrapper"]/div/div[1]/div/table/thead/tr/th'
        )
        print(f"   ✓ Cash Flow capturado: {len(cash_flow_data)} filas")
        
        print("   - Balance Sheet")
        balance_sheet_data = capture_table(
            driver, f"{base_url}-balance-sheet/",
            '//*[@id="horizontalFinancialTableN1_2"]/tbody/tr',
            header_xpath='//*[@id="horizontalFinancialTableN1_2_wrapper"]/div/div[1]/div/table/thead/tr/th'
        )
        print(f"   ✓ Balance Sheet capturado: {len(balance_sheet_data)} filas")
        
        print("   - Income Statement")
        financial_data = capture_table(
            driver, f"{base_url}-income-statement/",
            '//*[@id="horizontalFinancialTableN1_1"]/tbody/tr',
            header_xpath='//*[@id="horizontalFinancialTableN1_1_wrapper"]/div/div[1]/div/table/thead/tr/th'
        )
        print(f"   ✓ Income Statement capturado: {len(financial_data)} filas")
        
        print("   - Ratios")
        ratios_data = capture_table(
            driver, f"{base_url}-ratios/",
            '//*[@id="horizontalFinancialTableN1_5"]/tbody/tr',
            header_xpath='//*[@id="horizontalFinancialTableN1_5_wrapper"]/div/div[1]/div/table/thead/tr/th'
        )
        print(f"   ✓ Ratios capturados: {len(ratios_data)} filas")

        # Extraer el tipo de empresa desde el XPath específico
        tipo_xpath = "/html/body/div[1]/div/div[1]/main/div[1]/div[1]/div[2]/a[2]/h2"
        tipo_empresa = tipo_empresa_from_xpath(driver, tipo_xpath)

        ticker_xpath = "/html/body/div[1]/div/div[1]/main/div[1]/div/div[2]/h2[1]"
        ticker_empresa = ticker_market(driver, ticker_xpath)

        currency = extract_currency_from_url(base_url)
        if not currency:
            currency = "USD"
            print(f"No se pudo detectar la moneda para {stock_code}, usando USD como predeterminado")

        print(f"Moneda detectada para {stock_code}: {currency}")

        # Obtener el tipo de cambio a dólares
        exchange_rate = get_exchange_rate(currency)
        print(f"Tasa de cambio a USD para {currency}: {exchange_rate}")

        stock_name_clean = re.sub(r'[^A-Za-z]+', ' ', stock_code).replace('-', ' ').strip()
        stock_codes_df = pd.read_excel(file_path, usecols=[0, 1], header=None)
        stock_name_from_file = stock_codes_df[stock_codes_df[0] == stock_code].iloc[0, 1] if not stock_codes_df[stock_codes_df[0] == stock_code].empty else ''

        data_data = [
            ["ticker", ticker_empresa],
            ["nombre", stock_name_clean],
            ["empresa", tipo_empresa],
        ]

        ticker = yf.Ticker(stock_name_from_file) 
        print(ticker)

        # Convertir los datos financieros a DataFrames
        df_data = pd.DataFrame(data_data)
        df_cash_flow = pd.DataFrame(cash_flow_data)
        df_balance_sheet = pd.DataFrame(balance_sheet_data)
        df_financial = pd.DataFrame(financial_data)
        df_ratios = pd.DataFrame(ratios_data)
        df_valuation = pd.DataFrame(valuation_data)

        # Manejar error de yfinance
        try:
            df_yahoo_info = pd.DataFrame.from_dict(ticker.info, orient='index').T
            df_yahoo_info = df_yahoo_info.transpose()
        except Exception as e:
            print(f"   ⚠️ Error al obtener datos de Yahoo Finance para {stock_code}: {e}")
            df_yahoo_info = pd.DataFrame()

        df_cash_flow = convert_dataframe(pd.DataFrame(cash_flow_data))
        df_balance_sheet = convert_dataframe(pd.DataFrame(balance_sheet_data))
        df_financial = convert_dataframe(pd.DataFrame(financial_data))
        df_ratios = convert_dataframe(pd.DataFrame(ratios_data))

        # Convertir los datos a dólares usando la tasa de cambio obtenida
        df_cash_flow = convert_to_usd(df_cash_flow, exchange_rate)
        df_balance_sheet = convert_to_usd(df_balance_sheet, exchange_rate)
        df_financial = convert_to_usd(df_financial, exchange_rate)
        df_ratios = convert_to_usd(df_ratios, exchange_rate)

        # Si alguna tabla está vacía, rellenar con '-'
        def fill_if_empty(df, min_rows=2, min_cols=2):
            if df.empty:
                return pd.DataFrame([['-']*min_cols for _ in range(min_rows)])
            return df
        df_cash_flow = fill_if_empty(df_cash_flow)
        df_balance_sheet = fill_if_empty(df_balance_sheet)
        df_financial = fill_if_empty(df_financial)
        df_ratios = fill_if_empty(df_ratios)
        df_valuation = fill_if_empty(df_valuation)
        df_yahoo_info = fill_if_empty(df_yahoo_info)

        # Crear el archivo Excel antes de cerrar el navegador
        excel_path = rf"C:\Users\relim\Desktop\prueba\Resultados\{stock_name_from_file} - {stock_name_clean}.xlsx"
        with pd.ExcelWriter(excel_path, engine='xlsxwriter') as writer:
            df_data.to_excel(writer, sheet_name='Data', index=False, header=False)
            df_cash_flow.to_excel(writer, sheet_name='Cash Flow Statement', index=False, header=False)
            df_balance_sheet.to_excel(writer, sheet_name='Balance Sheet', index=False, header=False)
            df_financial.to_excel(writer, sheet_name='Financial', index=False, header=False)
            df_ratios.to_excel(writer, sheet_name='Ratios', index=False, header=False)
            df_valuation.to_excel(writer, sheet_name='Valuation', index=False, header=False)
            df_yahoo_info.to_excel(writer, sheet_name='Complementary', index=True, header=True)

        # Procesar el archivo Excel
        l14_value = process_excel_file(excel_path, df_cash_flow, df_balance_sheet, df_financial, df_ratios, df_valuation, forecast_years, x2)

        print(f"[DEBUG] Intentando subir el archivo a Google Drive: {excel_path}")
        shared_link = upload_and_get_shared_link(drive, excel_path)
        print(f"[DEBUG] Resultado de subida: {shared_link}")

        print(f"[DEBUG] Intentando actualizar Google Sheets para {stock_code}")
        # Definir valores por defecto para los campos faltantes
        country_value = "-"
        book_value = "-"
        pe_ratio = "-"
        market_cap = "-"
        dividend_ratio = "-"
        net_debt_growth = "-"
        total_assets_growth = "-"
        net_pp_growth = "-"
        update_google_sheets(stock_code, stock_name_clean, l14_value, ticker_empresa, shared_link, currency, exchange_rate, tipo_empresa, country_value, book_value, pe_ratio, market_cap, dividend_ratio, net_debt_growth, total_assets_growth, net_pp_growth)
        print(f"[DEBUG] Actualización de Google Sheets completada para {stock_code}")

    except Exception as e:
        print(f"\n✗ Error crítico procesando {stock_code}: {str(e)}")
        print(f"Stack trace: {traceback.format_exc()}")
        raise
    finally:
        print("\n5. Limpiando recursos...")
        if driver:
            try:
                driver.quit()
                print("✓ Driver cerrado exitosamente")
            except Exception as e:
                print(f"✗ Error al cerrar el driver: {str(e)}")
        
        try:
            print("   - Limpiando procesos de Chrome...")
            os.system("taskkill /im chrome.exe /f")
            os.system("taskkill /im chromedriver.exe /f")
            print("   ✓ Procesos de Chrome limpiados")
        except Exception as e:
            print(f"   ✗ Error al limpiar procesos de Chrome: {str(e)}")
        
        try:
            print("   - Limpiando carpeta temporal...")
            temp_folder = r"C:\Users\relim\AppData\Local\Temp"
            cleanup_temp_folder(temp_folder)
            print("   ✓ Carpeta temporal limpiada")
        except Exception as e:
            print(f"   ✗ Error al limpiar carpeta temporal: {str(e)}")
        
        print("   - Forzando recolección de basura...")
        gc.collect()
        print("✓ Recursos limpiados")
        print(f"\n{'='*50}")
        print(f"Procesamiento de {stock_code} finalizado")
        print(f"{'='*50}\n")


def process_excel_file(excel_path, df_cash_flow, df_balance_sheet, df_financial, df_ratios, df_valuation, forecast_years, x2):
    def extract_values(df, search_str):
        for idx, row in df.iterrows():
            if any(search_str in str(cell) for cell in row):
                return row[1:].values
        print(f"No se encontró '{search_str}' en la tabla.")
        return []

    fiscal_period = extract_values(df_cash_flow, 'Fiscal Period')
    unlevered_free_cash_flow = extract_values(df_cash_flow, 'Unlevered Free Cash Flow')
    total_debt = extract_values(df_balance_sheet, 'Net Debt')
    diluted_shares = extract_values(df_financial, 'Diluted Weighted Average Shares Outstanding')
    cash_equivalents = extract_values(df_balance_sheet, 'Cash and Equivalents')

    # Rellenar con NaN para que todas las listas tengan la misma longitud
    max_length = max(len(fiscal_period), len(unlevered_free_cash_flow), len(total_debt), len(diluted_shares), len(cash_equivalents))
    fiscal_period = np.pad(fiscal_period, (0, max_length - len(fiscal_period)), constant_values=np.nan)
    unlevered_free_cash_flow = np.pad(unlevered_free_cash_flow, (0, max_length - len(unlevered_free_cash_flow)), constant_values=np.nan)
    total_debt = np.pad(total_debt, (0, max_length - len(total_debt)), constant_values=np.nan)
    diluted_shares = np.pad(diluted_shares, (0, max_length - len(diluted_shares)), constant_values=np.nan)
    cash_equivalents = np.pad(cash_equivalents, (0, max_length - len(cash_equivalents)), constant_values=np.nan)

    # Crear DataFrame DCF
    df_dcf = pd.DataFrame({
        'Fiscal Period': fiscal_period,
        'Unlevered Free Cash Flow': unlevered_free_cash_flow,
        'Total Debt': total_debt,
        'Diluted Shares Outstanding': diluted_shares,
    })

    # Transponer el DataFrame
    df_dcf = df_dcf.set_index('Fiscal Period').T

    # Obtener el último año fiscal y agregar años futuros
    last_year = int(fiscal_period[-1])
    new_years = [last_year + i for i in range(1, forecast_years + 1)]
    df_dcf = pd.concat([df_dcf, pd.DataFrame(columns=new_years)], axis=1)

    # Escribir en el archivo Excel
    with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a') as writer:
        df_dcf.to_excel(writer, sheet_name='DCF', startrow=1, index=True, header=True)

    # Cargar el libro de trabajo y actualizar fórmulas
    wb = load_workbook(excel_path)
    ws = wb['DCF']
    ws['X2'] = x2

    # Calcular fórmulas para cada año
    unlevered_last_col = len(df_dcf.columns) - forecast_years
    for i in range(forecast_years):
        current_col = unlevered_last_col + 2 + i
        previous_col = unlevered_last_col + 1 + i

        if i == 0:
            formula = (
                f'=IF({ws.cell(row=3, column=unlevered_last_col + 1).coordinate}*'
                f'FORECAST({ws.cell(row=2, column=current_col).coordinate},'
                f'{ws.cell(row=3, column=2).coordinate}:{ws.cell(row=3, column=unlevered_last_col + 1).coordinate},'
                f'{ws.cell(row=2, column=2).coordinate}:{ws.cell(row=2, column=unlevered_last_col + 1).coordinate})>0,'
                f'FORECAST({ws.cell(row=2, column=current_col).coordinate},'
                f'{ws.cell(row=3, column=2).coordinate}:{ws.cell(row=3, column=unlevered_last_col + 1).coordinate},'
                f'{ws.cell(row=2, column=2).coordinate}:{ws.cell(row=2, column=unlevered_last_col + 1).coordinate}),'
                f'{ws.cell(row=3, column=unlevered_last_col + 1).coordinate})*$X$2'
            )
        else:
            formula = (
                f'=IF({ws.cell(row=3, column=unlevered_last_col + 1 + i).coordinate}*'
                f'FORECAST({ws.cell(row=2, column=current_col).coordinate},'
                f'{ws.cell(row=3, column=2).coordinate}:{ws.cell(row=3, column=current_col - 1).coordinate},'
                f'{ws.cell(row=2, column=2).coordinate}:{ws.cell(row=2, column=current_col - 1).coordinate})>0,'
                f'FORECAST({ws.cell(row=2, column=current_col).coordinate},'
                f'{ws.cell(row=3, column=2).coordinate}:{ws.cell(row=3, column=current_col - 1).coordinate},'
                f'{ws.cell(row=2, column=2).coordinate}:{ws.cell(row=2, column=current_col - 1).coordinate}),'
                f'{ws.cell(row=3, column=unlevered_last_col + 1 + i).coordinate})*$X$2'
            )
        ws.cell(row=3, column=current_col, value=formula)

    # Calcular terminal value y otros valores
    valor = ws.cell(row=3, column=current_col).value
    ws['K7'] = "TERMINAL VALUE"
    ws['L7'] = f'{valor} * (1 + 0.02) / (0.06 - 0.02)'
    terminal_value = ws['L7'].value

    start_col = unlevered_last_col + 3
    end_col = current_col
    terminal_col = end_col + 1

    def col_num_to_letter(col_num):
        col_letter = ""
        while col_num > 0:
            col_num, remainder = divmod(col_num - 1, 26)
            col_letter = chr(65 + remainder) + col_letter
        return col_letter

    start_col_letter = col_num_to_letter(start_col)
    end_col_letter = col_num_to_letter(end_col)
    terminal_col_letter = col_num_to_letter(terminal_col)
    next_start_col_letter = col_num_to_letter(start_col + 1)

    # Calcular NPV y otros valores
    ws['K8'] = 'NPV of FCF'
    ws['L8'] = f'=NPV(0.06, {start_col_letter}3:{end_col_letter}3)'

    for col in range(start_col + 1, end_col + 1):
        col_letter = col_num_to_letter(col)
        ws[f'{col_letter}16'] = 0

    ws[f'{terminal_col_letter}16'] = terminal_value

    ws['K9'] = "NPV of TV"
    ws['L9'] = f'=NPV(0.06, {next_start_col_letter}16:{terminal_col_letter}16)'

    ws['K10'] = 'Total EV'
    ws['L10'] = '=L8 + L9'

    ws['K11'] = 'Net Debt'
    for col in reversed(range(2, ws.max_column + 1)):
        if isinstance(ws.cell(row=4, column=col).value, (int, float)):
            last_column_net_debt = col
            break
    ws['L11'] = ws.cell(row=4, column=last_column_net_debt).value

    ws['K12'] = 'Equity'
    ws['L12'] = f'=L10 - L11'

    ws['K13'] = 'Shares Outstanding'
    for col in reversed(range(2, ws.max_column + 1)):
        if isinstance(ws.cell(row=5, column=col).value, (int, float)):
            last_column_shares_outstanding = col
            break
    ws['L13'] = ws.cell(row=5, column=last_column_shares_outstanding).value

    ws['K14'] = 'Target Price'
    ws['L14'] = f'=L12/L13'

    # Limpiar errores
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if isinstance(cell.value, str) and ("#DIV/0!" in cell.value or "Error" in cell.value):
                cell.value = "Error"

    wb.save(excel_path)

    # Forzar evaluación de fórmulas
    app = xw.App(visible=False)
    workbook = app.books.open(excel_path)
    workbook.save()
    workbook.close()
    app.quit()

    # Leer el valor calculado
    df = pd.read_excel(excel_path, sheet_name='DCF', header=None)
    l14_value = df.at[13, 11]

    if isinstance(l14_value, str) and ("#DIV/0" in l14_value or "Error" in l14_value):
        l14_value = "Error en cálculo"
    elif pd.isna(l14_value):
        l14_value = "Valor no disponible"

    return l14_value


def update_google_sheets(stock_code, stock_name_clean, l14_value, ticker_empresa, shared_link, currency, exchange_rate, tipo_empresa, country_value, book_value, pe_ratio, market_cap, dividend_ratio, net_debt_growth, total_assets_growth, net_pp_growth):
    try:
        print(f"[DEBUG] update_google_sheets llamado con:")
        print(f"  stock_code: {stock_code}")
        print(f"  stock_name_clean: {stock_name_clean}")
        print(f"  l14_value: {l14_value}")
        print(f"  ticker_empresa: {ticker_empresa}")
        print(f"  shared_link: {shared_link}")
        print(f"  currency: {currency}")
        print(f"  exchange_rate: {exchange_rate}")
        print(f"  tipo_empresa: {tipo_empresa}")
        print(f"  country_value: {country_value}")
        print(f"  book_value: {book_value}")
        print(f"  pe_ratio: {pe_ratio}")
        print(f"  market_cap: {market_cap}")
        print(f"  dividend_ratio: {dividend_ratio}")
        print(f"  net_debt_growth: {net_debt_growth}")
        print(f"  total_assets_growth: {total_assets_growth}")
        print(f"  net_pp_growth: {net_pp_growth}")
        scope = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
        creds = ServiceAccountCredentials.from_json_keyfile_name('C:/Users/relim/Desktop/prueba/durable-binder-432321-q4-5f1ef9e64ea1.json', scope)
        client = gspread.authorize(creds)
        sheet = client.open_by_key('193GrDtaNCWBkWwliitjSepbLlAcNlwQfUZp66WN6xyw').worksheet('Results')

        # Buscar la fila correspondiente para el código de stock
        cell_list = sheet.col_values(1)
        if stock_code in cell_list:
            row_index = cell_list.index(stock_code) + 1
        else:
            row_index = len(cell_list) + 1

        link = f'https://www.marketscreener.com/quote/stock/{stock_code}'
        formula_link = f'=HYPERLINK("{link}", "{stock_code}")'
        
        # Actualizar la fila específica con los nuevos datos
        print(f"[DEBUG] Actualizando fila {row_index} en Google Sheets...")
        sheet.update_cell(row_index, 1, formula_link)
        sheet.update_cell(row_index, 2, stock_name_clean)
        sheet.update_cell(row_index, 3, l14_value)
        sheet.update_cell(row_index, 4, ticker_empresa)
        sheet.update_cell(row_index, 5, shared_link)
        sheet.update_cell(row_index, 6, currency)
        sheet.update_cell(row_index, 7, exchange_rate)
        sheet.update_cell(row_index, 8, tipo_empresa)
        sheet.update_cell(row_index, 9, country_value)
        sheet.update_cell(row_index, 10, book_value)
        sheet.update_cell(row_index, 11, pe_ratio)
        sheet.update_cell(row_index, 12, market_cap)
        sheet.update_cell(row_index, 13, dividend_ratio)
        sheet.update_cell(row_index, 14, net_debt_growth)
        sheet.update_cell(row_index, 15, total_assets_growth)
        sheet.update_cell(row_index, 16, net_pp_growth)
        sheet.update_cell(row_index, 17, "16/05/2025")
        print(f"[DEBUG] Datos actualizados en Google Sheets para {stock_code}")
    except Exception as e:
        print(f"[ERROR] Error al actualizar datos en Google Sheets: {e}")


def process_all_stock_codes(file_path, forecast_years, x2):
    try:
        # Leer códigos de stock
        df = pd.read_excel(file_path)
        stock_codes = df['Ticker'].tolist()
        print(f"Se encontraron {len(stock_codes)} códigos de stock")

        # Configurar servicio de Google Drive
        scope = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
        creds = ServiceAccountCredentials.from_json_keyfile_name('C:/Users/relim/Desktop/prueba/durable-binder-432321-q4-5f1ef9e64ea1.json', scope)
        service = build('drive', 'v3', credentials=creds)
        print("Servicio de Google Drive configurado correctamente")

        # Procesar cada código de stock
        for stock_code in stock_codes:
            try:
                print(f"\nProcesando código: {stock_code}")
                process_stock_code(stock_code, forecast_years, x2, service)
                time.sleep(3)  # Esperar entre cada código
            except Exception as e:
                print(f"Error procesando {stock_code}: {e}")
                continue

    except Exception as e:
        print(f"Error general: {e}")
    finally:
        # Limpiar procesos de Chrome
        os.system("taskkill /f /im chrome.exe")
        os.system("taskkill /f /im chromedriver.exe")
        cleanup_temp_folder("C:\\Users\\relim\\AppData\\Local\\Temp")
        gc.collect()


file_path = r'C:\Users\relim\Desktop\prueba\StockCodes.xlsx'
forecast_years = 7
x2 = 1

process_all_stock_codes(file_path, forecast_years, x2)