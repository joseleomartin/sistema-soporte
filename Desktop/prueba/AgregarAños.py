import os
import shutil
import openpyxl
from openpyxl.utils import get_column_letter

def obtener_letra_columna(indice):
    return get_column_letter(indice)

def modificar_excel(ruta_origen, ruta_destino):
    wb = openpyxl.load_workbook(ruta_origen)
    if "DCF" not in wb.sheetnames:
        print(f"La hoja 'DCF' no está en {ruta_origen}")
        return
    
    hoja_dcf = wb["DCF"]
    
    # Encontrar la última columna con datos
    ultima_columna = hoja_dcf.max_column
    nueva_ultima_columna = ultima_columna + 5  # Agregar 5 nuevas columnas
    ultimo_año = hoja_dcf.cell(row=2, column=ultima_columna).value
    
    # Agregar numeración continua en la fila 2, omitiendo el 1
    for i in range(ultima_columna + 1, nueva_ultima_columna + 1):
        nuevo_valor = ultimo_año + (i - ultima_columna)
        if nuevo_valor != 1:
            hoja_dcf.cell(row=2, column=i, value=nuevo_valor)
    
    # Aplicar la fórmula corregida en la fila 3 (Unlevered Free Cash Flow)
    for i in range(ultima_columna + 1, nueva_ultima_columna + 1):
        letra_col_actual = obtener_letra_columna(i)
        letra_col_anterior = obtener_letra_columna(i - 1)
        letra_col_inicio = obtener_letra_columna(2)
        
        formula = (
            f"=IF({letra_col_anterior}3*FORECAST({letra_col_actual}2,"
            f"{letra_col_inicio}3:{letra_col_anterior}3,"
            f"{letra_col_inicio}2:{letra_col_anterior}2)>0,"
            f"FORECAST({letra_col_actual}2,"
            f"{letra_col_inicio}3:{letra_col_anterior}3,"
            f"{letra_col_inicio}2:{letra_col_anterior}2),"
            f"{letra_col_anterior}3)*$X$2"
        )
        hoja_dcf.cell(row=3, column=i, value=formula)
    
    # Guardar el archivo modificado
    wb.save(ruta_destino)
    print(f"Archivo guardado: {ruta_destino}")

def procesar_archivos():
    carpeta_origen = "F:\\Excels"
    carpeta_destino = "F:\\ExcelsCambiados"
    
    if not os.path.exists(carpeta_destino):
        os.makedirs(carpeta_destino)
    
    for archivo in os.listdir(carpeta_origen):
        if archivo.endswith(".xlsx"):
            ruta_origen = os.path.join(carpeta_origen, archivo)
            ruta_destino = os.path.join(carpeta_destino, archivo)
            
            # Copiar archivo antes de modificarlo
            shutil.copy2(ruta_origen, ruta_destino)
            modificar_excel(ruta_destino, ruta_destino)
            
if __name__ == "__main__":
    procesar_archivos()
