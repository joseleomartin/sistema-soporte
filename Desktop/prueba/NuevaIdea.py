from concurrent.futures import ThreadPoolExecutor
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.keys import Keys
from lxml import etree
import pandas as pd
import numpy as np
import time
from openpyxl import load_workbook
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import re
import xlwings as xw
import gc
from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive
import os
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
import threading
from concurrent.futures import ThreadPoolExecutor
import time
import requests
import shutil 
import tempfile
import yfinance as yf
import traceback


lock = threading.Lock()


def setup_drive():
    SCOPES = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
    SERVICE_ACCOUNT_FILE = 'C:/Users/relim/Desktop/prueba/durable-binder-432321-q4-5f1ef9e64ea1.json'
    
    creds = service_account.Credentials.from_service_account_file(
        SERVICE_ACCOUNT_FILE, scopes=SCOPES)
    
    service = build('drive', 'v3', credentials=creds)
    return service

def upload_and_get_shared_link(service, file_path):
     with lock:
        print(f"[LOG] Intentando subir el archivo: {file_path}")
        try:
            file_metadata = {
                'name': os.path.basename(file_path),
                'mimeType': 'application/vnd.google-apps.spreadsheet'
            }
            media = MediaFileUpload(file_path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            file = service.files().create(body=file_metadata, media_body=media, fields='id').execute()
            file_id = file.get('id')
            print(f"[LOG] Archivo subido, ID: {file_id}")
            
            # Cambiar permisos para compartir el archivo
            permissions = {
                'type': 'anyone',
                'role': 'writer'
            }
            service.permissions().create(fileId=file_id, body=permissions).execute()
            print(f"[LOG] Permisos cambiados a 'anyone' para el archivo: {file_id}")
            
            # Obtener enlace compartido
            file_info = service.files().get(fileId=file_id, fields='webViewLink').execute()
            shared_link = file_info.get('webViewLink')
            print(f"[LOG] Enlace compartido generado: {shared_link}")
            return shared_link
        except Exception as e:
            print(f"[ERROR] Error subiendo archivo a Google Drive: {e}")
            return None


def setup_driver():
    try:
        chrome_options = Options()
        chrome_options.add_argument("--headless=new")
        # User-Agent personalizado para simular un navegador real
        chrome_options.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36")
        # Flags anti-detección
        chrome_options.add_argument("--disable-blink-features=AutomationControlled")
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        chrome_options.add_experimental_option('useAutomationExtension', False)
        # Quitamos todas las opciones que puedan afectar la visibilidad
        chrome_options.add_argument("--start-maximized")  # Maximizar ventana
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--disable-extensions")
        chrome_options.add_argument("--disable-popup-blocking")
        chrome_options.add_argument("--blink-settings=imagesEnabled=false")
        chrome_options.add_argument('--ignore-certificate-errors')
        chrome_options.add_argument('--disable-web-security')
        chrome_options.add_argument('--allow-running-insecure-content')
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        
        # Configurar el servicio de Chrome
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=chrome_options)
        driver.set_page_load_timeout(30)
        return driver
    except Exception as e:
        print(f"Error al configurar el driver de Chrome: {e}")
        raise

def capture_table(driver, url, xpath, header_xpath=None):
    try:
        print(f"\n   - Navegando a la URL...")
        driver.get(url)
        time.sleep(3)
        
        print(f"   - Desplazando página...")
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(2)
        
        print(f"   - Esperando elemento XPath: {xpath}")
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, xpath)))
        
        print(f"   - Extrayendo HTML...")
        html = driver.page_source
        dom = etree.HTML(html)
        
        headers = []
        if header_xpath:
            print(f"   - Extrayendo encabezados...")
            header_elements = dom.xpath(header_xpath)
            headers = [''.join(header.xpath(".//text()")).strip() for header in header_elements]
            print(f"   ✓ Encabezados encontrados: {len(headers)}")

        print(f"   - Buscando filas...")
        rows = dom.xpath(xpath)
        if not rows:
            print(f"   ⚠️ No se encontraron filas con el XPath: {xpath}")
            return []

        print(f"   - Procesando {len(rows)} filas...")
        table_data = []
        
        # Primero, obtener los nombres de las filas (primera columna)
        row_names = []
        for row in rows:
            first_cell = row.xpath(".//td[1] | .//th[1]")
            if first_cell:
                row_name = ''.join(first_cell[0].xpath(".//text()")).strip()
                row_names.append(row_name)
        
        # Luego, para cada fila, obtener los valores en orden correcto
        for row in rows:
            cells = row.xpath(".//td | .//th")
            # Asegurarse de que los valores estén en el orden correcto
            row_data = []
            for cell in cells:
                value = ''.join(cell.xpath(".//text()")).strip()
                row_data.append(value)
            table_data.append(row_data)

        # Asegurarse de que los datos estén alineados correctamente
        if headers:
            # Verificar que los headers coincidan con el número de columnas en los datos
            if len(headers) != len(table_data[0]):
                print(f"   ⚠️ Número de headers ({len(headers)}) no coincide con número de columnas ({len(table_data[0])})")
                # Ajustar headers si es necesario
                if len(headers) > len(table_data[0]):
                    headers = headers[:len(table_data[0])]
                else:
                    headers.extend([''] * (len(table_data[0]) - len(headers)))
            
            # Insertar headers al principio
            table_data.insert(0, headers)

        print(f"   ✓ Datos capturados exitosamente: {len(table_data)} filas")
        return table_data

    except Exception as e:
        print(f"   ✗ Error al capturar datos: {str(e)}")
        return []


import os
import shutil

def cleanup_temp_folder(temp_folder):
    try:
        # Elimina todos los archivos y directorios en la carpeta Temp
        shutil.rmtree(temp_folder, ignore_errors=True)
        print(f"Carpeta temporal {temp_folder} limpiada exitosamente.")
    except Exception as e:
        print(f"Error al eliminar el contenido de la carpeta temporal: {e}")



EXCEPTIONS = {'0', "Diluted Shares Outstanding", "Basic Weighted Average Shares Outstanding", "Diluted Weighted Average Shares Outstanding", "Basic Weighted Average Shares Outstanding "
              "Diluted Weighted Average Shares Outstanding ","Basic Weighted Average Shares Outstanding ", "Payout Ratio", "American Depositary Receipts Ratio (ADR)",
                "American Depositary Receipts Ratio", "Effective Tax Rate - (Ratio)", "Effective Tax Rate", "Net EPS - Basic", "Basic EPS - Continuing Operations", "Net EPS", "Basic EPS", "Net EPS - Diluted", "Net EPS", "Diluted EPS - Continuing Operations", "Diluted EPS",
                "Normalized Basic EPS", "Normalized Diluted EPS",
                "Diluted Weighted Average Shares Outstanding ", "Basic Weighted Average Shares Outstanding ",
                "Diluted Shares Outstanding", "Diluted Weighted Average Shares Outstanding","Basic Weighted Average Shares Outstanding", "Fiscal period","Fiscal Period", "Fiscal Period: January", "Fiscal Period: February", "Fiscal Period: March", "Fiscal Period: April", "Fiscal Period: May", "Fiscal Period: June", "Fiscal Period: July", "Fiscal Period: August", "Fiscal Period: September", "Fiscal Period: October", "Fiscal Period: November", "Fiscal Period: December" }



def convert_dataframe(df):
    for col in df.columns:
        if col != df.columns[0]:  # Aplicar conversión solo si la columna no es la primera (índice/nombres)
            df[col] = df[col].apply(convert_value)
    return df

def convert_value(value):
    if isinstance(value, str):
        value = value.strip()
        if value in EXCEPTIONS:
            return value
        if value == '-':
            return 0
        if value.endswith('M'):
            # Intentar convertir a número antes de multiplicar
            try:
                num_value = float(value[:-1].replace(',', ''))
                return int(num_value * 1_000_000) # Multiplicar por 1 millón
            except ValueError:
                return value # Si falla la conversión, retornar el valor original
        elif value.endswith('B'):
             # Intentar convertir a número antes de multiplicar
            try:
                num_value = float(value[:-1].replace(',', ''))
                return int(num_value * 1_000_000_000)  # Multiplicar por 1 mil millones
            except ValueError:
                return value
        elif value.endswith('K'):
             # Intentar convertir a número antes de multiplicar
            try:
                num_value = float(value[:-1].replace(',', ''))
                return int(num_value * 1_000)  # Multiplicar por 1 mil (ajustado a 1000)
            except ValueError:
                return value
        elif value.replace(',', '').lstrip('-').isdigit() or (value.replace(',', '').lstrip('-').replace('.', '', 1).isdigit() and '.' in value):
             # Permitir números negativos y decimales
            try:
                 # Intentar convertir a float primero para manejar decimales, luego a int si es posible sin pérdida
                num_value = float(value.replace(',', ''))
                return int(num_value) if num_value.is_integer() else num_value
            except ValueError:
                return value
        else:
            return value # Retornar el valor original si no se puede convertir
    return value

def get_exchange_rate(from_currency, to_currency="USD"):
    """Obtiene la tasa de cambio desde la moneda de origen a la moneda de destino."""
    url = f"https://api.exchangerate-api.com/v4/latest/{from_currency}"
    response = requests.get(url)
    data = response.json()
    return data['rates'].get(to_currency, 1)  


def convert_to_usd(df, exchange_rate):
    # Convertir solo columnas numéricas después de aplicar map
    numeric_cols = df.columns[df.apply(lambda col: col.apply(lambda x: isinstance(x, (int, float)) or (isinstance(x, str) and x.replace(',', '').isdigit())).any())]

    # Multiplicar por la tasa de cambio solo si la fila no tiene un valor en la columna A que sea una excepción
    for index in df.index:
        if df.at[index, df.columns[0]] not in EXCEPTIONS:
            df.loc[index, numeric_cols] = df.loc[index, numeric_cols].apply(lambda x: x * exchange_rate if isinstance(x, (int, float)) else x)
    
    return df


def extract_currency_from_url(base_url):
    # Agregar el sufijo '-income-statement/' a la URL
    full_url = f"{base_url}-income-statement/"

    # Crear una nueva instancia del navegador
    driver = setup_driver()

    # Navegar a la URL completa
    driver.get(full_url)

    extracted_text = None  # Inicializar la variable para almacenar el texto

    try:
        # Utiliza un selector CSS para encontrar el elemento que contiene la moneda
        wait = WebDriverWait(driver, 20)  # Aumentar el tiempo de espera a 20 segundos
        element = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, ".js-selectedCurrency.px-10")))
        extracted_text = element.text.strip()  # Usar strip() para limpiar espacios extra
        print(f"Texto extraído desde el elemento en {full_url}: {extracted_text}")

    except Exception as e:
        print(f"Error inesperado: {e}")

    finally:
        driver.quit()  # Cerrar el navegador al finalizar

    return extracted_text  # Retornar el texto extraído


def tipo_empresa_from_xpath(driver, tipo_xpath):
    try:
        # Esperar a que el elemento esté presente
        empresa_element = WebDriverWait(driver, 2).until(
            EC.presence_of_element_located((By.XPATH, tipo_xpath))
        )
        # Extraer el texto del elemento
        empresa_text = empresa_element.get_attribute('textContent').strip()
        
        if empresa_text:
            print(f"Texto extraído del XPath: '{empresa_text}'")
        else:
            print("No se pudo extraer el texto, el resultado es vacío.")
        
        return empresa_text
    except Exception as e:
        print(f"Error al extraer el texto desde el XPath: {str(e)}")
        return "No encontrado"

def ticker_market(driver, ticker_xpath):
    try:
        # Esperar a que el elemento esté presente
        ticker_element = WebDriverWait(driver, 2).until(
            EC.presence_of_element_located((By.XPATH, ticker_xpath))
        )
        # Extraer el texto del elemento
        ticker_element = ticker_element.get_attribute('textContent').strip()
        
        if ticker_element:
            print(f"Texto extraído del XPath: '{ticker_element}'")
        else:
            print("No se pudo extraer el texto, el resultado es vacío.")
        
        return ticker_element
    except Exception as e:
        print(f"Error al extraer el texto desde el XPath: {str(e)}")
        return "No encontrado"

def process_stock_code(stock_code, forecast_years, x2, drive):
    driver = None
    try:
        print(f"\n{'='*50}")
        print(f"Iniciando procesamiento para {stock_code}")
        print(f"{'='*50}\n")

        print("1. Configurando el driver de Chrome...")
        driver = setup_driver()
        print("✓ Driver configurado exitosamente")

        base_url = f"https://www.marketscreener.com/quote/stock/{stock_code}/finances"
        base_url2 = f"https://www.marketscreener.com/quote/stock/{stock_code}/valuation/"

        # --- Conectar a Google Sheets y obtener enlace histórico --- #
        print("\n3. Obteniendo enlace histórico desde Google Sheets...")
        historical_sheet_link = None
        client = None
        try:
            scope = ["https://www.googleapis.com/auth/spreadsheets"]
            creds = ServiceAccountCredentials.from_json_keyfile_name('C:/Users/relim/Desktop/prueba/durable-binder-432321-q4-5f1ef9e64ea1.json', scope)
            client = gspread.authorize(creds)

            # Abrir el Google Sheet principal
            main_sheet_id = '1snwyzqO13lf_D9bXdFXAwCEc_-q53S0jb7ZAGPOOSyQ' # ID de tu hoja principal
            main_sheet = client.open_by_key(main_sheet_id)

            # Seleccionar la hoja "Calculado"
            calculado_sheet = main_sheet.worksheet('Calculado')

            # Extraer el ticker corto (parte antes del primer guion)
            short_ticker = stock_code.split('-')[0]
            print(f"   - Buscando ticker corto '{short_ticker}' en la hoja 'Calculado'...")
            cell = calculado_sheet.find(short_ticker, in_column=1) # Buscar en la primera columna (columna A)

            if cell:
                row_index = cell.row
                # Obtener el link del sub-enlace de la tercera columna (columna C, index 2)
                historical_sheet_link = calculado_sheet.cell(row_index, 3).value
                print(f"✓ Enlace histórico encontrado para {stock_code}: {historical_sheet_link}")
            else:
                print(f"⚠️ Ticker {stock_code} no encontrado en la hoja 'Calculado'. No se combinarán datos históricos.")
        except Exception as e:
            print(f"✗ Error al acceder o procesar la hoja principal de Google Sheets para obtener enlace histórico: {str(e)}")
        # --- Fin de obtención de enlace histórico ---

        # --- Extraer tipo de empresa, ticker, moneda y tasa de cambio --- #
        print("\n4. Extrayendo información básica y tasa de cambio...")

        # Extraer el tipo de empresa desde el XPath específico
        tipo_xpath = "/html/body/div[1]/div/div[1]/main/div[1]/div[1]/div[2]/a[2]/h2"
        tipo_empresa = tipo_empresa_from_xpath(driver, tipo_xpath)

        ticker_xpath = "/html/body/div[1]/div/div[1]/main/div[1]/div/div[2]/h2[1]"
        ticker_empresa = ticker_market(driver, ticker_xpath)

        currency = extract_currency_from_url(base_url)
        if not currency:
            currency = "USD"
            print(f"No se pudo detectar la moneda para {stock_code}, usando USD como predeterminado")

        print(f"   - Moneda detectada para {stock_code}: {currency}")

        # Obtener el tipo de cambio a dólares
        exchange_rate = get_exchange_rate(currency)
        print(f"   - Tasa de cambio a USD para {currency}: {exchange_rate}")

        stock_name_clean = re.sub(r'[^A-Za-z]+', ' ', stock_code).replace('-', ' ').strip()
        stock_codes_df = pd.read_excel(file_path, usecols=[0, 1], header=None)
        stock_name_from_file = stock_codes_df[stock_codes_df[0] == stock_code].iloc[0, 1] if not stock_codes_df[stock_codes_df[0] == stock_code].empty else ''

        data_data = [
            ["ticker", ticker_empresa],
            ["nombre", stock_name_clean],
            ["empresa", tipo_empresa],
        ]

        # Convertir df_data a DataFrame aquí
        df_data = pd.DataFrame(data_data)

        # Obtener datos básicos de Yahoo Finance (para Complementary)
        scraped_yahoo_info_df = pd.DataFrame()
        try:
             ticker = yf.Ticker(stock_name_from_file)
             scraped_yahoo_info_df = pd.DataFrame.from_dict(ticker.info, orient='index').T
             print(f"   ✓ Datos básicos de Yahoo Finance obtenidos.")
        except Exception as e:
             print(f"   ✗ Error al obtener datos básicos de Yahoo Finance para {stock_code}: {e}")
        # --- Fin de extracción de información básica y tasa de cambio ---

        # --- Nueva función auxiliar para obtener y combinar DataFrames --- #
        def get_combined_dataframe(sheet_name, driver, historical_sheet_link, client, base_url, base_url2, exchange_rate):
            print(f"\n   - Procesando hoja: {sheet_name}")
            scraped_data = []
            historical_df = pd.DataFrame()

            # 1. Primero obtener datos históricos (si el enlace existe)
            if historical_sheet_link and client:
                try:
                    match = re.search(r'/spreadsheets/d/([a-zA-Z0-9-_]+)', historical_sheet_link)
                    if match:
                        historical_sheet_id = match.group(1)
                        historical_sheet = client.open_by_key(historical_sheet_id)

                        print(f"      - Leyendo hoja histórica: {sheet_name}...")
                        worksheet = historical_sheet.worksheet(sheet_name)
                        data = worksheet.get_all_values()
                        if data:
                            historical_df = pd.DataFrame(data)
                            if len(historical_df) > 1:
                                historical_df.columns = historical_df.iloc[0] # Primera fila como encabezado
                                historical_df = historical_df[1:].reset_index(drop=True) # Resto son datos
                            else:
                                 # Si solo hay una fila, tratarla como datos sin encabezado claro
                                 historical_df = historical_df.reset_index(drop=True)

                            # Ensure first column is string for merging/indexing
                            if not historical_df.empty:
                                historical_df.iloc[:, 0] = historical_df.iloc[:, 0].astype(str)

                            print(f"      ✓ Hoja histórica '{sheet_name}' leída exitosamente.")
                        else:
                            print(f"      ⚠️ La hoja histórica '{sheet_name}' está vacía.")
                            historical_df = pd.DataFrame()

                except Exception as e:
                    print(f"      ✗ Error al leer la hoja histórica '{sheet_name}': {str(e)}")
                    historical_df = pd.DataFrame()

            # 2. Luego scrapear datos recientes
            try:
                print(f"      - Scrapeando datos recientes para {sheet_name}...")
                if sheet_name == 'Valuation':
                    scraped_data = capture_table(
                        driver, f"{base_url2}",
                        '//*[@id="valuationEnterpriseTable"]/tbody/tr',
                        header_xpath='//*[@id="valuationEnterpriseTable_wrapper"]/div/div[1]/div/table/thead/tr/th'
                    )
                elif sheet_name == 'Cash Flow Statement':
                    scraped_data = capture_table(
                        driver, f"{base_url}-cash-flow-statement/",
                        '//*[@id="horizontalFinancialTableN1_3"]/tbody/tr',
                        header_xpath='//*[@id="horizontalFinancialTableN1_3_wrapper"]/div/div[1]/div/table/thead/tr/th'
                    )
                elif sheet_name == 'Balance Sheet':
                    scraped_data = capture_table(
                        driver, f"{base_url}-balance-sheet/",
                        '//*[@id="horizontalFinancialTableN1_2"]/tbody/tr',
                        header_xpath='//*[@id="horizontalFinancialTableN1_2_wrapper"]/div/div[1]/div/table/thead/tr/th'
                    )
                elif sheet_name == 'Financial':
                    scraped_data = capture_table(
                        driver, f"{base_url}-income-statement/",
                        '//*[@id="horizontalFinancialTableN1_1"]/tbody/tr',
                        header_xpath='//*[@id="horizontalFinancialTableN1_1_wrapper"]/div/div[1]/div/table/thead/tr/th'
                    )
                elif sheet_name == 'Ratios':
                    scraped_data = capture_table(
                        driver, f"{base_url}-ratios/",
                        '//*[@id="horizontalFinancialTableN1_5"]/tbody/tr',
                        header_xpath='//*[@id="horizontalFinancialTableN1_5_wrapper"]/div/div[1]/div/table/thead/tr/th'
                    )
                print(f"      ✓ Datos recientes para {sheet_name} scrapeados: {len(scraped_data)} filas")

            except Exception as e:
                print(f"      ✗ Error al scrapear datos recientes para {sheet_name}: {str(e)}")
                scraped_data = []

            # Convertir datos scrapeados a DataFrame y aplicar conversiones
            scraped_df = pd.DataFrame(scraped_data)
            if not scraped_df.empty:
                scraped_df = convert_dataframe(scraped_df)
                scraped_df = convert_to_usd(scraped_df, exchange_rate)

            # 3. Combinar datos (históricos primero, luego scrapeados)
            print(f"      - Combinando datos para {sheet_name}...")

            if historical_df.empty and scraped_df.empty:
                combined_df = pd.DataFrame()
            elif historical_df.empty:
                combined_df = scraped_df
            elif scraped_df.empty:
                combined_df = historical_df
            else:
                # Usar la primera columna como clave de alineación (nombres de métricas)
                alignment_key_name = historical_df.columns[0] if not historical_df.empty else scraped_df.columns[0]

                # Asegurarse de que ambos DataFrames usen el mismo nombre para la columna clave
                historical_df.rename(columns={historical_df.columns[0]: alignment_key_name}, inplace=True)
                scraped_df.rename(columns={scraped_df.columns[0]: alignment_key_name}, inplace=True)

                # Establecer la clave de alineación como índice
                historical_df_indexed = historical_df.set_index(alignment_key_name)
                scraped_df_indexed = scraped_df.set_index(alignment_key_name)

                # Obtener todos los nombres de métricas únicos de ambos DFs
                all_metrics = historical_df_indexed.index.union(scraped_df_indexed.index)

                # Obtener todas las columnas (años) únicas de ambos DFs
                all_columns = historical_df_indexed.columns.union(scraped_df_indexed.columns)

                # Función para detectar si una columna tiene valores reales (no todos son nan)
                def has_real_values(df, col):
                    try:
                        return df[col].notna().any()
                    except KeyError:
                        return False

                # Reordenar columnas: primero las que tienen valores reales, luego las que son nan
                real_columns = []
                nan_columns = []
                
                for col in all_columns:
                    try:
                        if has_real_values(scraped_df_indexed, col) or has_real_values(historical_df_indexed, col):
                            real_columns.append(col)
                        else:
                            nan_columns.append(col)
                    except Exception as e:
                        print(f"      ⚠️ Error al procesar columna {col}: {str(e)}")
                        nan_columns.append(col)

                all_columns_sorted = real_columns + nan_columns

                # Crear un nuevo DataFrame para la combinación con el índice y columnas correctas
                combined_df_indexed = pd.DataFrame(index=all_metrics, columns=all_columns_sorted)

                # Llenar el DataFrame combinado iterando por métrica y año
                for metric in all_metrics:
                    for col in all_columns_sorted:
                        try:
                            hist_val = historical_df_indexed.loc[metric, col] if metric in historical_df_indexed.index and col in historical_df_indexed.columns else np.nan
                            scrap_val = scraped_df_indexed.loc[metric, col] if metric in scraped_df_indexed.index and col in scraped_df_indexed.columns else np.nan

                            # Si hay un valor scrapeado y no es nan, usarlo
                            if pd.notna(scrap_val):
                                combined_df_indexed.loc[metric, col] = scrap_val
                            # Si no hay valor scrapeado o es nan, usar el histórico
                            elif pd.notna(hist_val):
                                combined_df_indexed.loc[metric, col] = hist_val
                            else:
                                combined_df_indexed.loc[metric, col] = np.nan
                        except Exception as e:
                            print(f"      ⚠️ Error al combinar valores para métrica {metric}, columna {col}: {str(e)}")
                            combined_df_indexed.loc[metric, col] = np.nan

                # Resetear el índice
                combined_df = combined_df_indexed.reset_index()
                combined_df.rename(columns={'index': alignment_key_name}, inplace=True)

            print(f"      ✓ Datos para {sheet_name} combinados.")
            return combined_df
        # --- Fin de la función auxiliar --- #

        # --- Obtener DataFrames combinados para cada sección --- #
        print("\n5. Obteniendo y combinando datos para cada sección...")

        # Valuation
        df_valuation = get_combined_dataframe('Valuation', driver, historical_sheet_link, client, base_url, base_url2, exchange_rate)

        # Cash Flow Statement
        df_cash_flow = get_combined_dataframe('Cash Flow Statement', driver, historical_sheet_link, client, base_url, base_url2, exchange_rate)

        # Balance Sheet
        df_balance_sheet = get_combined_dataframe('Balance Sheet', driver, historical_sheet_link, client, base_url, base_url2, exchange_rate)

        # Financial (Income Statement)
        df_financial = get_combined_dataframe('Financial', driver, historical_sheet_link, client, base_url, base_url2, exchange_rate)

        # Ratios
        df_ratios = get_combined_dataframe('Ratios', driver, historical_sheet_link, client, base_url, base_url2, exchange_rate)

        # Complementary (Yahoo Finance data) - scrapeada por separado y combinada aquí
        print("\n   - Procesando hoja: Complementary")
        historical_complementary_df = pd.DataFrame()
        if historical_sheet_link and client:
             try:
                 match = re.search(r'/spreadsheets/d/([a-zA-Z0-9-_]+)', historical_sheet_link)
                 if match:
                     historical_sheet_id = match.group(1)
                     historical_sheet = client.open_by_key(historical_sheet_id)
                     print(f"      - Leyendo hoja histórica: Complementary...")
                     worksheet = historical_sheet.worksheet('Complementary')
                     data = worksheet.get_all_values()
                     if data:
                          # Para Complementary, la primera fila es encabezado, la primera columna es la clave
                          historical_complementary_df = pd.DataFrame(data[1:], columns=data[0])
                          # Convertir la primera columna a string y establecer como índice temporal para la combinación
                          historical_complementary_df.iloc[:, 0] = historical_complementary_df.iloc[:, 0].astype(str)
                          historical_complementary_df = historical_complementary_df.set_index(historical_complementary_df.columns[0])
                          print(f"      ✓ Hoja histórica 'Complementary' leída exitosamente.")
                     else:
                          print(f"      ⚠️ La hoja histórica 'Complementary' está vacía.")
             except Exception as e:
                 print(f"      ✗ Error al leer la hoja histórica 'Complementary': {str(e)}")

        # Scrapear datos de Yahoo Finance (Complementary)
        scraped_yahoo_info_df = pd.DataFrame()
        try:
             ticker = yf.Ticker(stock_name_from_file)
             scraped_yahoo_info_df = pd.DataFrame.from_dict(ticker.info, orient='index').T
             # Convertir la primera columna a string y establecer como índice temporal para la combinación
             scraped_yahoo_info_df.iloc[:, 0] = scraped_yahoo_info_df.iloc[:, 0].astype(str)
             scraped_yahoo_info_df = scraped_yahoo_info_df.set_index(scraped_yahoo_info_df.columns[0])
             print(f"      ✓ Datos recientes de Yahoo Finance scrapeados.")
        except Exception as e:
             print(f"      ✗ Error al obtener datos de Yahoo Finance para {stock_code}: {e}")

        # Combinar datos Complementary
        print(f"      - Combinando datos para Complementary...")
        if historical_complementary_df.empty:
            df_yahoo_info = scraped_yahoo_info_df
        elif scraped_yahoo_info_df.empty:
            df_yahoo_info = historical_complementary_df
        else:
             # Combinar horizontalmente los DataFrames indexados. Preferir scraped_yahoo_info_df
             df_yahoo_info = scraped_yahoo_info_df.combine_first(historical_complementary_df)

        # Resetear el índice para que la primera columna vuelva a ser una columna normal para Complementary
        df_yahoo_info = df_yahoo_info.reset_index()
        # Renombrar la primera columna a 'Item' o similar si no tiene nombre
        if df_yahoo_info.columns[0] == 'index':
             df_yahoo_info.rename(columns={'index': 'Item'}, inplace=True)

        print(f"   ✓ Datos para Complementary combinados.")

        # --- Fin de obtención de DataFrames combinados --- #

        # Si alguna tabla combinada está vacía, rellenar con '-'
        def fill_if_empty(df, min_rows=2, min_cols=2):
            if df.empty:
                return pd.DataFrame([['-']*min_cols for _ in range(min_rows)])
            # Asegurarse de que todos los valores sean string para evitar errores al escribir en excel si hay tipos mixtos
            return df.astype(str) # Convertir todo a string aquí para seguridad

        df_valuation = fill_if_empty(df_valuation, min_rows=1)
        df_cash_flow = fill_if_empty(df_cash_flow, min_rows=1)
        df_balance_sheet = fill_if_empty(df_balance_sheet, min_rows=1)
        df_financial = fill_if_empty(df_financial, min_rows=1)
        df_ratios = fill_if_empty(df_ratios, min_rows=1)
        df_yahoo_info = fill_if_empty(df_yahoo_info, min_rows=1)

        # Crear el archivo Excel usando los DataFrames combinados
        excel_path = rf"C:\Users\relim\Desktop\prueba\Resultados\{stock_name_from_file} - {stock_name_clean}.xlsx"
        print(f"\n6. Creando archivo Excel: {excel_path}")
        with pd.ExcelWriter(excel_path, engine='xlsxwriter') as writer:
            df_data.to_excel(writer, sheet_name='Data', index=False, header=False)
            df_cash_flow.to_excel(writer, sheet_name='Cash Flow Statement', index=False, header=False)
            df_balance_sheet.to_excel(writer, sheet_name='Balance Sheet', index=False, header=False)
            df_financial.to_excel(writer, sheet_name='Financial', index=False, header=False)
            df_ratios.to_excel(writer, sheet_name='Ratios', index=False, header=False)
            df_valuation.to_excel(writer, sheet_name='Valuation', index=False, header=False)
            df_yahoo_info.to_excel(writer, sheet_name='Complementary', index=False, header=True) # Complementary no necesita index=True si ya lo reseteamos
        print(f"✓ Archivo Excel creado exitosamente.")

        # Procesar el archivo Excel para DCF (ahora trabaja con datos ya combinados)
        print("\n7. Procesando archivo Excel para cálculo DCF...")
        l14_value = process_excel_file(excel_path, df_cash_flow, df_balance_sheet, df_financial, df_ratios, df_valuation, forecast_years, x2)
        print(f"✓ Cálculo DCF completado. L14 Value: {l14_value}")

        print(f"\n8. Subiendo archivo a Google Drive: {excel_path}")
        shared_link = upload_and_get_shared_link(drive, excel_path)
        print(f"✓ Archivo subido. Enlace compartido: {shared_link}")

        print(f"\n9. Actualizando Google Sheets principal para {stock_code}...")
        # Obtener valores adicionales del DataFrame df_yahoo_info (que ya está combinado y llenado)
        country_value = "-"
        book_value = "-"
        pe_ratio = "-"
        market_cap = "-"
        dividend_ratio = "-"
        net_debt_growth = "-"
        total_assets_growth = "-"
        net_pp_growth = "-"

        # Intentar obtener valores de df_yahoo_info
        if not df_yahoo_info.empty and 'Item' in df_yahoo_info.columns:
            # Convertir a dict para fácil acceso por el nombre del item
            yahoo_info_dict = df_yahoo_info.set_index('Item').to_dict('index')

            # Asumiendo que los valores están en la primera columna de datos después del 'Item'
            country_value = yahoo_info_dict.get('country', {}).get(df_yahoo_info.columns[1], '-') if len(df_yahoo_info.columns) > 1 else '-'
            book_value = yahoo_info_dict.get('bookValue', {}).get(df_yahoo_info.columns[1], '-') if len(df_yahoo_info.columns) > 1 else '-'
            pe_ratio = yahoo_info_dict.get('trailingPE', {}).get(df_yahoo_info.columns[1], '-') if len(df_yahoo_info.columns) > 1 else '-'
            market_cap = yahoo_info_dict.get('marketCap', {}).get(df_yahoo_info.columns[1], '-') if len(df_yahoo_info.columns) > 1 else '-'
            dividend_ratio = yahoo_info_dict.get('dividendYield', {}).get(df_yahoo_info.columns[1], '-') if len(df_yahoo_info.columns) > 1 else '-'

            # También podemos intentar obtener algunas tasas de crecimiento si existen en Complementary (aunque el scrapeo inicial de Yahoo Finance no las da)
            # Habría que adaptar si estas tasas de crecimiento provienen de otro lugar o se calculan después.
            # Por ahora, las dejamos como placeholders o intentamos buscarlas si se agregan a Complementary.


        update_google_sheets(stock_code, stock_name_clean, l14_value, ticker_empresa, shared_link, currency, exchange_rate, tipo_empresa, country_value, book_value, pe_ratio, market_cap, dividend_ratio, net_debt_growth, total_assets_growth, net_pp_growth)
        print(f"✓ Google Sheets principal actualizado para {stock_code}")

    except Exception as e:
        print(f"\n✗ Error crítico procesando {stock_code}: {str(e)}")
        print(f"Stack trace: {traceback.format_exc()}")
        # No raise the exception here to allow the loop in process_all_stock_codes to continue
    finally:
        print("\n10. Limpiando recursos...") # Changed step number
        if driver:
            try:
                driver.quit()
                print("✓ Driver cerrado exitosamente")
            except Exception as e:
                print(f"✗ Error al cerrar el driver: {str(e)}")

        try:
            print("   - Limpiando procesos de Chrome...")
            os.system("taskkill /im chrome.exe /f")
            os.system("taskkill /im chromedriver.exe /f")
            print("   ✓ Procesos de Chrome limpiados")
        except Exception as e:
            print(f"   ✗ Error al limpiar procesos de Chrome: {str(e)}")

        try:
            print("   - Limpiando carpeta temporal...")
            temp_folder = r"C:\Users\relim\AppData\Local\Temp"
            cleanup_temp_folder(temp_folder)
            print("   ✓ Carpeta temporal limpiada")
        except Exception as e:
            print(f"   ✗ Error al limpiar carpeta temporal: {str(e)}")

        print("   - Forzando recolección de basura...")
        gc.collect()
        print("✓ Recursos limpiados")
        print(f"\n{'='*50}")
        print(f"Procesamiento de {stock_code} finalizado")
        print(f"{'='*50}\n")


def process_excel_file(excel_path, df_cash_flow, df_balance_sheet, df_financial, df_ratios, df_valuation, forecast_years, x2):
    def find_row_data(df, search_str):
        # Ensure all values in the DataFrame are strings for searching
        df_str = df.astype(str)
        for idx, row in df_str.iterrows():
            if any(search_str.lower() in str(cell).lower() for cell in row):
                # Return original values from the non-string converted DataFrame (excluding the first column)
                return df.iloc[idx, 1:].values.tolist()
        print(f"No se encontró '{search_str}' en la tabla.")
        return []

    def find_row_headers(df, search_str):
         # Ensure all values in the DataFrame are strings for searching
        df_str = df.astype(str)
        for idx, row in df_str.iterrows():
            if any(search_str.lower() in str(cell).lower() for cell in row):
                # Return original values from the non-string converted DataFrame (excluding the first column)
                return df.iloc[idx, 1:].values.tolist()
        print(f"No se encontró la fila de encabezados para '{search_str}' en la tabla.")
        return []


    # Encontrar la fila de encabezados de periodo fiscal
    fiscal_period_headers = find_row_headers(df_cash_flow, 'Fiscal Period')

    # Encontrar las filas de datos para las métricas clave
    unlevered_free_cash_flow_data = find_row_data(df_cash_flow, 'Unlevered Free Cash Flow')
    total_debt_data = find_row_data(df_balance_sheet, 'Net Debt')
    diluted_shares_data = find_row_data(df_financial, 'Diluted Weighted Average Shares Outstanding')
    # cash_equivalents_data = find_row_data(df_balance_sheet, 'Cash and Equivalents') # No se usa directamente en DCF por ahora

    # Alinear datos con los encabezados de periodo fiscal
    # Crear un diccionario temporal para alinear cada columna de datos con su encabezado correspondiente
    aligned_data = {}
    # Usar la longitud de los encabezados como referencia principal
    data_length = len(fiscal_period_headers)

    # Asegurarse de que todas las listas de datos tengan al menos la misma longitud que los encabezados
    # Si son más cortas, rellenar con NaN. Si son más largas, truncar (esto no debería pasar si la extracción es correcta)
    unlevered_free_cash_flow_data = unlevered_free_cash_flow_data[:data_length] + [np.nan] * max(0, data_length - len(unlevered_free_cash_flow_data))
    total_debt_data = total_debt_data[:data_length] + [np.nan] * max(0, data_length - len(total_debt_data))
    diluted_shares_data = diluted_shares_data[:data_length] + [np.nan] * max(0, data_length - len(diluted_shares_data))


    # Crear el DataFrame DCF inicial alineado por columnas
    df_dcf = pd.DataFrame({
        'Fiscal Period': fiscal_period_headers,
        'Unlevered Free Cash Flow': unlevered_free_cash_flow_data,
        'Total Debt': total_debt_data,
        'Diluted Shares Outstanding': diluted_shares_data,
    })

    # Transponer el DataFrame
    df_dcf = df_dcf.set_index('Fiscal Period').T

    # Limpiar los nombres de columna para quedarnos solo con años o periodos relevantes
    # Esto reemplaza la lógica anterior de fiscal_years = [...]
    df_dcf.columns = [str(col).strip() for col in df_dcf.columns]
    # Eliminar columnas que son 'nan' o vacías si existen como nombres de columna después de transponer
    df_dcf = df_dcf.rename(columns=lambda x: x.replace('nan', '').strip()) # Clean 'nan' strings
    df_dcf = df_dcf.loc[:, df_dcf.columns.astype(bool)] # Remove empty column names

    # Asegurarse de que las columnas (años/periodos) sean únicas y ordenadas
    # Esto es crucial para evitar la duplicación visual en Excel
    # Convertir a numérico si es posible para ordenar correctamente
    try:
        ordered_columns = sorted(df_dcf.columns, key=lambda x: int(x) if str(x).isdigit() else float('inf'))
        df_dcf = df_dcf[ordered_columns]
    except ValueError:
         # If columns are not strictly numeric, keep the current order after cleaning
         pass # Or implement a different sorting logic if needed


    # Obtener el último año fiscal (ahora de las columnas del DataFrame limpiado) y agregar años futuros
    l14_value = "Error en cálculo DCF"

    if len(df_dcf.columns) > 0:
        try:
            # El último año es la última columna en el DataFrame ordenado
            last_year_str = df_dcf.columns[-1]
            last_year = int(last_year_str)
            new_years = [str(last_year + i) for i in range(1, forecast_years + 1)]
            # Agregar las nuevas columnas al DataFrame DCF
            for year in new_years:
                df_dcf[year] = np.nan # Agregar columnas para años futuros, inicializadas con NaN

            # Escribir en el archivo Excel
            with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a') as writer:
                # Escribir el DataFrame DCF, con los años fiscales como encabezados
                df_dcf.to_excel(writer, sheet_name='DCF', startrow=1, index=True, header=True)

            # Cargar el libro de trabajo y actualizar fórmulas
            wb = load_workbook(excel_path)
            ws = wb['DCF']
            ws['X2'] = x2 # Asegurarse de que X2 se mantenga o se posicione correctamente

            # Calcular fórmulas para cada año
            # Las columnas de datos históricos están en df_dcf.columns hasta antes de los new_years
            historical_years_count = len(df_dcf.columns) - forecast_years
            start_data_col_index = 2 # Columna B (después del índice)
            last_data_col_index = start_data_col_index + historical_years_count - 1
            start_forecast_col_index = last_data_col_index + 1 # Columna donde empieza el primer año pronosticado

            unlevered_row = 3 # Fila 'Unlevered Free Cash Flow' en Excel
            fiscal_period_row = 2 # Fila de los años en Excel


            for i in range(forecast_years):
                current_col_index = start_forecast_col_index + i
                current_col_letter = col_num_to_letter(current_col_index)

                if i == 0:
                    # Para el primer año de pronóstico, usar todos los años históricos disponibles
                    forecast_range_data = f'{col_num_to_letter(start_data_col_index)}{unlevered_row}:{col_num_to_letter(last_data_col_index)}{unlevered_row}'
                    forecast_range_headers = f'{col_num_to_letter(start_data_col_index)}{fiscal_period_row}:{col_num_to_letter(last_data_col_index)}{fiscal_period_row}'
                    formula = (
                        f'=IF({ws.cell(row=unlevered_row, column=last_data_col_index).coordinate}*'
                        f'FORECAST({ws.cell(row=fiscal_period_row, column=current_col_index).coordinate},'
                        f'{forecast_range_data},'
                        f'{forecast_range_headers})>0,'
                        f'FORECAST({ws.cell(row=fiscal_period_row, column=current_col_index).coordinate},'
                        f'{forecast_range_data},'
                        f'{forecast_range_headers}),'
                        f'{ws.cell(row=unlevered_row, column=last_data_col_index).coordinate})*$X$2'
                    )
                else:
                    # Para años futuros, usar los años históricos y los años pronosticados previamente
                    previous_forecast_col_index = start_forecast_col_index + i - 1
                    forecast_range_data = f'{col_num_to_letter(start_data_col_index)}{unlevered_row}:{col_num_to_letter(previous_forecast_col_index)}{unlevered_row}'
                    forecast_range_headers = f'{col_num_to_letter(start_data_col_index)}{fiscal_period_row}:{col_num_to_letter(previous_forecast_col_index)}{fiscal_period_row}'
                    formula = (
                         f'=IF({ws.cell(row=unlevered_row, column=previous_forecast_col_index).coordinate}*'
                         f'FORECAST({ws.cell(row=fiscal_period_row, column=current_col_index).coordinate},'
                         f'{forecast_range_data},'
                         f'{forecast_range_headers})>0,'
                         f'FORECAST({ws.cell(row=fiscal_period_row, column=current_col_index).coordinate},'
                         f'{forecast_range_data},'
                         f'{forecast_range_headers}),'
                         f'{ws.cell(row=unlevered_row, column=previous_forecast_col_index).coordinate})*$X$2'
                    )
                ws.cell(row=unlevered_row, column=current_col_index, value=formula)


            # Calcular terminal value y otros valores
            last_forecast_col_index = start_forecast_col_index + forecast_years - 1
            last_forecast_col_letter = col_num_to_letter(last_forecast_col_index)

            ws['K7'] = "TERMINAL VALUE"
            ws['L7'] = f'={last_forecast_col_letter}{unlevered_row} * (1 + 0.02) / (0.06 - 0.02)'

            npv_tv_row = 16
            # Limpiar y configurar fila 16 para NPV of TV
            for col_index in range(start_data_col_index, last_forecast_col_index + 2): # Incluir la columna después del último año pronosticado
                 ws.cell(row=npv_tv_row, column=col_index).value = None

            # Poner 0 en las columnas de años históricos para la fila 16
            for col_index in range(start_data_col_index, last_data_col_index + 1):
                ws.cell(row=npv_tv_row, column=col_index).value = 0

            # Poner el valor del Terminal Value en la columna después del último año pronosticado
            terminal_value_col_index = last_forecast_col_index + 1
            terminal_value_col_letter = col_num_to_letter(terminal_value_col_index)
            ws.cell(row=npv_tv_row, column=terminal_value_col_index).value = f'=L7' # Referencia a la celda del Terminal Value


            ws['K8'] = 'NPV of FCF'
            start_npv_fcf_col_letter = col_num_to_letter(start_forecast_col_index)
            end_npv_fcf_col_letter = col_num_to_letter(last_forecast_col_index)
            ws['L8'] = f'=NPV(0.06, {start_npv_fcf_col_letter}{unlevered_row}:{end_npv_fcf_col_letter}{unlevered_row})'


            ws['K9'] = "NPV of TV"
            start_npv_tv_col_letter = col_num_to_letter(start_forecast_col_index)
            end_npv_tv_col_letter = col_num_to_letter(terminal_value_col_index)
            ws['L9'] = f'=NPV(0.06, {start_npv_tv_col_letter}{npv_tv_row}:{end_npv_tv_col_letter}{npv_tv_row})'


            ws['K10'] = 'Total EV'
            ws['L10'] = '=L8 + L9'

            ws['K11'] = 'Net Debt'
            last_column_net_debt = None
            for col in reversed(range(start_data_col_index, ws.max_column + 1)):
                cell_value = ws.cell(row=4, column=col).value
                if isinstance(cell_value, (int, float)):
                    last_column_net_debt = col
                    break
            if last_column_net_debt is not None:
                ws['L11'] = ws.cell(row=4, column=last_column_net_debt).value
            else:
                 ws['L11'] = "Valor no disponible"


            ws['K12'] = 'Equity'
            ws['L12'] = f'=L10 - L11'

            ws['K13'] = 'Shares Outstanding'
            last_column_shares_outstanding = None
            for col in reversed(range(start_data_col_index, ws.max_column + 1)):
                 cell_value = ws.cell(row=5, column=col).value
                 if isinstance(cell_value, (int, float)):
                     last_column_shares_outstanding = col
                     break
            if last_column_shares_outstanding is not None:
                ws['L13'] = ws.cell(row=5, column=last_column_shares_outstanding).value
            else:
                ws['L13'] = "Valor no disponible"


            ws['K14'] = 'Target Price'
            if isinstance(ws['L13'].value, (int, float)) and ws['L13'].value != 0:
                 ws['L14'] = f'=L12/L13'
            else:
                 ws['L14'] = "Error: División por cero o valor no disponible"


            # Limpiar errores
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    if isinstance(cell.value, str) and ("#DIV/0!" in cell.value or "Error" in cell.value):
                        cell.value = "Error"

            wb.save(excel_path)

            # Forzar evaluación de fórmulas
            try:
                app = xw.App(visible=False, add_book=False) # Add add_book=False
                workbook = app.books.open(excel_path)
                workbook.save()
                workbook.close()
                app.quit()
            except Exception as excel_e:
                 print(f"✗ Error al forzar la evaluación de fórmulas con xlwings: {str(excel_e)}")


            # Leer el valor calculado
            df = pd.read_excel(excel_path, sheet_name='DCF', header=None)
            l14_value = df.at[13, 11]

            if isinstance(l14_value, str) and ("#DIV/0" in l14_value or "Error" in l14_value or "Valor no disponible" in l14_value):
                l14_value = "Error en cálculo"
            elif pd.isna(l14_value):
                l14_value = "Valor no disponible"

        except Exception as e:
            print(f"✗ Error durante el cálculo del DCF para {excel_path}: {str(e)}")
            l14_value = "Error en cálculo DCF"

    else:
        print("⚠️ No se encontraron años fiscales válidos para el cálculo del DCF.")
        l14_value = "Error en cálculo DCF (No hay años fiscales)"


    return l14_value

def col_num_to_letter(col_num):
    """Helper function to convert column number (1-based) to Excel column letter."""
    col_letter = ""
    while col_num > 0:
        col_num, remainder = divmod(col_num - 1, 26)
        col_letter = chr(65 + remainder) + col_letter
    return col_letter

def update_google_sheets(stock_code, stock_name_clean, l14_value, ticker_empresa, shared_link, currency, exchange_rate, tipo_empresa, country_value, book_value, pe_ratio, market_cap, dividend_ratio, net_debt_growth, total_assets_growth, net_pp_growth):
    try:
        print(f"[DEBUG] update_google_sheets llamado con:")
        print(f"  stock_code: {stock_code}")
        print(f"  stock_name_clean: {stock_name_clean}")
        print(f"  l14_value: {l14_value}")
        print(f"  ticker_empresa: {ticker_empresa}")
        print(f"  shared_link: {shared_link}")
        print(f"  currency: {currency}")
        print(f"  exchange_rate: {exchange_rate}")
        print(f"  tipo_empresa: {tipo_empresa}")
        print(f"  country_value: {country_value}")
        print(f"  book_value: {book_value}")
        print(f"  pe_ratio: {pe_ratio}")
        print(f"  market_cap: {market_cap}")
        print(f"  dividend_ratio: {dividend_ratio}")
        print(f"  net_debt_growth: {net_debt_growth}")
        print(f"  total_assets_growth: {total_assets_growth}")
        print(f"  net_pp_growth: {net_pp_growth}")
        scope = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
        creds = ServiceAccountCredentials.from_json_keyfile_name('C:/Users/relim/Desktop/prueba/durable-binder-432321-q4-5f1ef9e64ea1.json', scope)
        client = gspread.authorize(creds)
        sheet = client.open_by_key('1snwyzqO13lf_D9bXdFXAwCEc_-q53S0jb7ZAGPOOSyQ').worksheet('Results')

        # Buscar la fila correspondiente para el código de stock
        cell_list = sheet.col_values(1)
        if stock_code in cell_list:
            row_index = cell_list.index(stock_code) + 1
        else:
            row_index = len(cell_list) + 1

        link = f'https://www.marketscreener.com/quote/stock/{stock_code}'
        formula_link = f'=HYPERLINK("{link}", "{stock_code}")'
        
        # Actualizar la fila específica con los nuevos datos
        print(f"[DEBUG] Actualizando fila {row_index} en Google Sheets...")
        sheet.update_cell(row_index, 1, formula_link)
        sheet.update_cell(row_index, 2, stock_name_clean)
        sheet.update_cell(row_index, 3, l14_value)
        sheet.update_cell(row_index, 4, ticker_empresa)
        sheet.update_cell(row_index, 5, shared_link)
        sheet.update_cell(row_index, 6, currency)
        sheet.update_cell(row_index, 7, exchange_rate)
        sheet.update_cell(row_index, 8, tipo_empresa)
        sheet.update_cell(row_index, 9, country_value)
        sheet.update_cell(row_index, 10, book_value)
        sheet.update_cell(row_index, 11, pe_ratio)
        sheet.update_cell(row_index, 12, market_cap)
        sheet.update_cell(row_index, 13, dividend_ratio)
        sheet.update_cell(row_index, 14, net_debt_growth)
        sheet.update_cell(row_index, 15, total_assets_growth)
        sheet.update_cell(row_index, 16, net_pp_growth)
        sheet.update_cell(row_index, 17, "16/05/2025")
        print(f"[DEBUG] Datos actualizados en Google Sheets para {stock_code}")
    except Exception as e:
        print(f"[ERROR] Error al actualizar datos en Google Sheets: {e}")


def process_all_stock_codes(file_path, forecast_years, x2):
    try:
        print("\n=== INICIO DEL PROGRAMA ===")

        # Configurar servicio de Google Sheets y Drive
        scope = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
        creds = ServiceAccountCredentials.from_json_keyfile_name('C:/Users/relim/Desktop/prueba/durable-binder-432321-q4-5f1ef9e64ea1.json', scope)
        client = gspread.authorize(creds)
        drive_service = build('drive', 'v3', credentials=creds)

        # Abrir el Google Sheet principal
        main_sheet_id = '1snwyzqO13lf_D9bXdFXAwCEc_-q53S0jb7ZAGPOOSyQ' # ID de tu hoja principal
        main_sheet = client.open_by_key(main_sheet_id)

        # Leer códigos de stock y enlaces de la hoja "ScrapNuevo"
        print(f"Leyendo códigos de stock y enlaces de la hoja 'ScrapNuevo'...")
        scrap_nuevo_sheet = main_sheet.worksheet('ScrapNuevo')
        # Asumiendo que la columna A tiene el Ticker y la columna C tiene el enlace al nuevo spreadsheet
        scrap_nuevo_data = scrap_nuevo_sheet.get_all_values()
        # Saltar la fila de encabezado si existe
        if len(scrap_nuevo_data) > 1:
            # Crear un diccionario {ticker: nuevo_sheet_link}
            stock_codes_and_new_links = {row[0]: row[2] for row in scrap_nuevo_data[1:] if len(row) > 2 and row[0].strip()}
            stock_codes = list(stock_codes_and_new_links.keys())
        else:
            stock_codes_and_new_links = {}
            stock_codes = []

        print(f"Se encontraron {len(stock_codes)} códigos de stock en 'ScrapNuevo'")
        if stock_codes:
             print(f"Códigos encontrados: {stock_codes[:5]}...")  # Mostrar los primeros 5 códigos

        # Leer enlaces históricos de la hoja "Calculado"
        print(f"\nLeyendo enlaces históricos de la hoja 'Calculado'...")
        calculado_sheet = main_sheet.worksheet('Calculado')
        # Asumiendo que la columna A tiene el Ticker y la columna C tiene el enlace al spreadsheet histórico
        calculado_data = calculado_sheet.get_all_values()
        # Saltar la fila de encabezado si existe
        if len(calculado_data) > 1:
             # Crear un diccionario {ticker: historical_sheet_link}
             historical_links = {row[0]: row[2] for row in calculado_data[1:] if len(row) > 2 and row[0].strip()}
        else:
             historical_links = {}

        print(f"Se encontraron {len(historical_links)} enlaces históricos en 'Calculado'")


        # Procesar cada código de stock encontrado en "ScrapNuevo"
        for stock_code in stock_codes:
            new_sheet_link = stock_codes_and_new_links.get(stock_code)
            historical_sheet_link = historical_links.get(stock_code, None) # Obtener enlace histórico, None si no existe

            # --- MODIFICACIÓN AQUÍ ---
            # Verificar si el enlace de datos nuevos es un string y parece un enlace de Google Sheets
            if not isinstance(new_sheet_link, str) or not re.search(r'/spreadsheets/d/', new_sheet_link):
                print(f"\n⚠️ Saltando {stock_code}: El enlace de datos nuevos en 'ScrapNuevo' no es un enlace de Google Sheets válido o está vacío: '{new_sheet_link}'.")
                continue
            # --- FIN MODIFICACIÓN ---

            try:
                print(f"\n{'='*50}")
                print(f"Iniciando procesamiento para: {stock_code}")
                print(f"{'='*50}")

                # Ahora process_stock_code trabajará con los enlaces a los Google Sheets
                process_stock_code_from_sheets(stock_code, new_sheet_link, historical_sheet_link, client, drive_service, forecast_years, x2, main_sheet)

                time.sleep(3)  # Esperar entre cada código
            except Exception as e:
                print(f"Error procesando {stock_code}: {e}")
                print(f"Stack trace: {traceback.format_exc()}")
                continue

    except Exception as e:
        print(f"Error general: {e}")
        print(f"Stack trace: {traceback.format_exc()}")
    finally:
        print("\n=== LIMPIEZA FINAL ===")
        # Limpiar procesos de Chrome
        os.system("taskkill /f /im chrome.exe")
        os.system("taskkill /f /im chromedriver.exe")
        cleanup_temp_folder("C:\\Users\\relim\\AppData\\Local\\Temp")
        gc.collect()
        print("=== PROGRAMA FINALIZADO ===")


# Necesitamos una nueva función process_stock_code que lea de sheets
def process_stock_code_from_sheets(stock_code, new_sheet_link, historical_sheet_link, client, drive_service, forecast_years, x2, main_sheet):
    print(f"[DEBUG] process_stock_code_from_sheets iniciado para {stock_code}")

    new_sheet_id = None
    historical_sheet_id = None

    # Extraer IDs de los enlaces
    match_new = re.search(r'/spreadsheets/d/([a-zA-Z0-9-_]+)', new_sheet_link)
    if match_new:
        new_sheet_id = match_new.group(1)
        print(f"[DEBUG] ID de hoja nueva extraído: {new_sheet_id}")
    else:
        print(f"[ERROR] No se pudo extraer el ID de la hoja nueva del enlace: {new_sheet_link}")
        return # Salir si no hay enlace nuevo válido

    if historical_sheet_link:
        match_hist = re.search(r'/spreadsheets/d/([a-zA-Z0-9-_]+)', historical_sheet_link)
        if match_hist:
            historical_sheet_id = match_hist.group(1)
            print(f"[DEBUG] ID de hoja histórica extraído: {historical_sheet_id}")
        else:
            print(f"[WARNING] No se pudo extraer el ID de la hoja histórica del enlace: {historical_sheet_link}")

    # Leer datos de las hojas
    sheets_to_process = ['Valuation', 'Cash Flow Statement', 'Balance Sheet', 'Financial', 'Ratios', 'Complementary']
    combined_dfs = {}
    stock_name_clean = "" # Debemos obtener el nombre y ticker de alguna parte, ¿quizás del nuevo sheet o del sheet principal?
    ticker_empresa = ""
    currency = "USD" # Asumir USD si no se puede obtener
    exchange_rate = 1.0 # Asumir 1.0 si no se puede obtener
    tipo_empresa = "No especificado"


    # Intentar obtener información básica del nuevo sheet si existe
    try:
        if new_sheet_id:
            new_sheet = client.open_by_key(new_sheet_id)
            # Intentar leer la hoja 'Data' si existe para obtener ticker, nombre, empresa
            try:
                data_sheet_new = new_sheet.worksheet('Data')
                data_values = data_sheet_new.get_all_values()
                data_dict = dict(data_values) # Convertir a diccionario {clave: valor}
                ticker_empresa = data_dict.get('ticker', '')
                stock_name_clean = data_dict.get('nombre', '')
                tipo_empresa = data_dict.get('empresa', 'No especificado')
                print(f"[DEBUG] Información básica obtenida del nuevo sheet: Ticker={ticker_empresa}, Nombre={stock_name_clean}, Empresa={tipo_empresa}")
            except gspread.WorksheetNotFound:
                print("[WARNING] Hoja 'Data' no encontrada en el nuevo spreadsheet.")
            except Exception as e:
                 print(f"[WARNING] Error al leer la hoja 'Data' del nuevo spreadsheet: {e}")

            # Intentar obtener currency y exchange rate si están en alguna parte del nuevo sheet
            # Esto puede requerir lógica adicional dependiendo de dónde se guarden estos datos
            # Por ahora, mantenemos la lógica de obtenerlos del scraping si es absolutamente necesario,
            # pero idealmente deberían estar en los datos guardados.
            # Si no se pueden obtener, usamos USD y 1.0 por defecto.

            # Si el nombre de la acción aún está vacío, intentar obtenerlo de la hoja principal (ScrapNuevo)
            if not stock_name_clean:
                 try:
                     scrap_nuevo_sheet = main_sheet.worksheet('ScrapNuevo')
                     cell = scrap_nuevo_sheet.find(stock_code, in_column=1) # Buscar en la primera columna (columna A)
                     if cell:
                          row_index = cell.row
                          stock_name_clean = scrap_nuevo_sheet.cell(row_index, 2).value # Asumiendo que el nombre está en la columna B
                          print(f"[DEBUG] Nombre obtenido de la hoja principal 'ScrapNuevo': {stock_name_clean}")
                 except Exception as e:
                      print(f"[WARNING] Error al obtener el nombre de la hoja principal 'ScrapNuevo': {e}")


        # Si el nombre sigue vacío, usar el stock_code
        if not stock_name_clean:
             stock_name_clean = stock_code
             print(f"[DEBUG] Usando stock_code como nombre: {stock_name_clean}")

         # Si el ticker sigue vacío, usar el stock_code
        if not ticker_empresa:
             ticker_empresa = stock_code
             print(f"[DEBUG] Usando stock_code como ticker: {ticker_empresa}")


    except Exception as e:
        print(f"[ERROR] Error al abrir o procesar el nuevo spreadsheet (para info básica): {e}")
        # Continuar con valores por defecto si falla

    # Combinar datos para cada hoja
    for sheet_name in sheets_to_process:
        print(f"\n   - Procesando hoja: {sheet_name} para combinar")
        new_df = pd.DataFrame()
        historical_df = pd.DataFrame()

        # Leer datos del nuevo spreadsheet
        try:
            if new_sheet_id:
                new_sheet = client.open_by_key(new_sheet_id)
                worksheet = new_sheet.worksheet(sheet_name)
                data = worksheet.get_all_values()
                if data:
                     new_df = pd.DataFrame(data[1:], columns=data[0]) # Primera fila como encabezado
                     # Ensure first column is string for merging/indexing
                     if not new_df.empty:
                         new_df.iloc[:, 0] = new_df.iloc[:, 0].astype(str)
                     print(f"      ✓ Hoja '{sheet_name}' leída del nuevo spreadsheet.")
                else:
                     print(f"      ⚠️ La hoja '{sheet_name}' en el nuevo spreadsheet está vacía.")
        except gspread.WorksheetNotFound:
            print(f"      ⚠️ Hoja '{sheet_name}' no encontrada en el nuevo spreadsheet.")
        except Exception as e:
            print(f"      ✗ Error al leer la hoja '{sheet_name}' del nuevo spreadsheet: {str(e)}")


        # Leer datos del spreadsheet histórico
        try:
            if historical_sheet_id:
                historical_sheet = client.open_by_key(historical_sheet_id)
                worksheet = historical_sheet.worksheet(sheet_name)
                data = worksheet.get_all_values()
                if data:
                     historical_df = pd.DataFrame(data[1:], columns=data[0]) # Primera fila como encabezado
                     # Ensure first column is string for merging/indexing
                     if not historical_df.empty:
                         historical_df.iloc[:, 0] = historical_df.iloc[:, 0].astype(str)

                     print(f"      ✓ Hoja '{sheet_name}' leída del spreadsheet histórico.")
                else:
                    print(f"      ⚠️ La hoja '{sheet_name}' en el spreadsheet histórico está vacía.")
        except gspread.WorksheetNotFound:
            print(f"      ⚠️ Hoja '{sheet_name}' no encontrada en el spreadsheet histórico.")
        except Exception as e:
            print(f"      ✗ Error al leer la hoja '{sheet_name}' del spreadsheet histórico: {str(e)}")

        # Combinar los DataFrames (new_df tiene prioridad)
        print(f"      - Combinando datos para {sheet_name}...")
        if new_df.empty and historical_df.empty:
            combined_df = pd.DataFrame()
            print(f"      ⚠️ Ambos DataFrames ({sheet_name}) están vacíos.")
        elif new_df.empty:
            combined_df = historical_df
            print(f"      ✓ Usando solo datos históricos para {sheet_name}.")
        elif historical_df.empty:
            combined_df = new_df
            print(f"      ✓ Usando solo datos nuevos para {sheet_name}.")
        else:
            # Usar la primera columna como clave de alineación (nombres de métricas)
            alignment_key_name = new_df.columns[0] if not new_df.empty else historical_df.columns[0]

            # Asegurarse de que ambos DataFrames usen el mismo nombre para la columna clave
            new_df.rename(columns={new_df.columns[0]: alignment_key_name}, inplace=True)
            historical_df.rename(columns={historical_df.columns[0]: alignment_key_name}, inplace=True)

            # Establecer la clave de alineación como índice
            new_df_indexed = new_df.set_index(alignment_key_name)
            historical_df_indexed = historical_df.set_index(alignment_key_name)

            # Obtener todas las métricas únicas y columnas únicas
            all_metrics = new_df_indexed.index.union(historical_df_indexed.index)
            all_columns = new_df_indexed.columns.union(historical_df_indexed.columns)

             # Ordenar columnas numéricamente si es posible, de lo contrario alfabéticamente
            try:
                # Convertir a entero para ordenar, si no es posible, usar float('inf') para enviarlos al final
                ordered_columns = sorted(all_columns, key=lambda x: int(str(x).strip()) if str(x).strip().isdigit() else float('inf'))
            except ValueError:
                # Si falla la conversión, ordenar alfabéticamente
                ordered_columns = sorted(all_columns)

            # Crear un nuevo DataFrame combinado con el índice y columnas correctas
            combined_df_indexed = pd.DataFrame(index=all_metrics, columns=ordered_columns)

            # Llenar el DataFrame combinado: new_df tiene prioridad
            for metric in all_metrics:
                for col in ordered_columns:
                    new_val = new_df_indexed.loc[metric, col] if metric in new_df_indexed.index and col in new_df_indexed.columns else np.nan
                    hist_val = historical_df_indexed.loc[metric, col] if metric in historical_df_indexed.index and col in historical_df_indexed.columns else np.nan

                    # Si hay un valor en el nuevo DF y no es NaN, usarlo
                    if pd.notna(new_val):
                        combined_df_indexed.loc[metric, col] = new_val
                    # Si no, usar el valor del histórico si no es NaN
                    elif pd.notna(hist_val):
                        combined_df_indexed.loc[metric, col] = hist_val
                    else:
                        # Si ambos son NaN, mantener NaN
                        combined_df_indexed.loc[metric, col] = np.nan

                # Resetear el índice
                combined_df = combined_df_indexed.reset_index()
                combined_df.rename(columns={'index': alignment_key_name}, inplace=True)

            print(f"      ✓ Datos para {sheet_name} combinados exitosamente.")


        # Asegurarse de que el DataFrame combinado no esté vacío antes de guardarlo
        # También convertir a string para evitar problemas de formato al escribir en Sheets
        if not combined_df.empty:
             # Asegurarse de que todos los valores sean string para evitar errores al escribir en Sheets
            combined_dfs[sheet_name] = combined_df.astype(str)
        else:
             # Si el DataFrame combinado está vacío, crear uno mínimo con guiones para que la hoja no esté completamente vacía
             print(f"      ⚠️ DataFrame combinado para {sheet_name} está vacío. Creando DataFrame de placeholder.")
             placeholder_df = pd.DataFrame([['-']*2 for _ in range(2)]) # Crear un DataFrame 2x2 con guiones
             if sheet_name == 'Complementary': # Complementary tiene encabezado
                  placeholder_df.columns = ['Item', 'Valor']
             combined_dfs[sheet_name] = placeholder_df.astype(str)


    # Crear un nuevo spreadsheet en Google Drive
    print(f"\nCreando nuevo spreadsheet combinado para {stock_code} en Google Drive...")
    combined_spreadsheet_title = f"{stock_name_clean} - Datos Combinados" # Título del nuevo spreadsheet
    # Puedes definir el ID de la carpeta "Results" si es fija, o buscarla.
    # Por ahora, crearemos en la raíz o en una carpeta si se especifica un ID.
    # Supongamos que la carpeta 'Results' existe y conocemos su ID (obtenerlo manualmente o buscarlo una vez)
    # Para simplificar, lo crearemos en la raíz por ahora, o puedes reemplazar 'root' con el ID de tu carpeta.
    # Para crear en una carpeta específica, necesitas el folder_id.
    # Ejemplo: folder_id = 'YOUR_RESULTS_FOLDER_ID'
    # file_metadata = {'name': combined_spreadsheet_title, 'mimeType': 'application/vnd.google-apps.spreadsheet', 'parents': [folder_id]}

    # Buscar la carpeta "Resultados" para obtener su ID
    results_folder_id = None
    try:
        response = drive_service.files().list(
            q="mimeType='application/vnd.google-apps.folder' and name='Resultados' and trashed=false",
            spaces='drive',
            fields='files(id, name)'
        ).execute()
        folders = response.get('files', [])
        if folders:
            results_folder_id = folders[0]['id']
            print(f"[DEBUG] Carpeta 'Resultados' encontrada con ID: {results_folder_id}")
        else:
            print("[WARNING] Carpeta 'Resultados' no encontrada. Creando spreadsheet en la raíz del Drive.")
    except Exception as e:
        print(f"[ERROR] Error buscando la carpeta 'Resultados': {e}. Creando spreadsheet en la raíz.")


    file_metadata = {'name': combined_spreadsheet_title, 'mimeType': 'application/vnd.google-apps.spreadsheet'}
    if results_folder_id:
        file_metadata['parents'] = [results_folder_id]


    combined_spreadsheet = drive_service.files().create(body=file_metadata, fields='id, webViewLink').execute()
    combined_spreadsheet_id = combined_spreadsheet.get('id')
    combined_spreadsheet_link = combined_spreadsheet.get('webViewLink')
    print(f"✓ Spreadsheet combinado creado. ID: {combined_spreadsheet_id}")
    print(f"✓ Enlace: {combined_spreadsheet_link}")

    # Escribir los DataFrames combinados en el nuevo spreadsheet
    print(f"\nEscribiendo datos combinados en el nuevo spreadsheet...")
    try:
        for sheet_name, df in combined_dfs.items():
            # Verificar si la hoja ya existe
            try:
                 worksheet = client.open_by_key(combined_spreadsheet_id).worksheet(sheet_name)
                 # Si existe, borrar el contenido actual (excepto la primera fila si es Complementary)
                 if sheet_name == 'Complementary':
                     worksheet.clear() # Esto borra todo, luego escribiremos el DF completo con encabezado
                     print(f"      Limpiando hoja existente '{sheet_name}' antes de escribir.")
                 else:
                      # Para otras hojas, borramos todo
                      worksheet.clear()
                      print(f"      Limpiando hoja existente '{sheet_name}' antes de escribir.")
            except gspread.WorksheetNotFound:
                # Si no existe, crear la hoja
                print(f"      Creando hoja '{sheet_name}' en el nuevo spreadsheet.")
                worksheet = client.open_by_key(combined_spreadsheet_id).add_worksheet(title=sheet_name, rows="1000", cols="50")
            except Exception as e:
                print(f"      ⚠️ Error al verificar o crear la hoja '{sheet_name}': {e}. Intentando crearla de nuevo.")
                try:
                     worksheet = client.open_by_key(combined_spreadsheet_id).add_worksheet(title=sheet_name, rows="1000", cols="50")
                     print(f"      Hoja '{sheet_name}' creada después del error.")
                except Exception as e_retry:
                     print(f"      ✗ Error crítico al crear la hoja '{sheet_name}' después del reintento: {e_retry}. Saltando esta hoja.")
                     continue # Saltar a la siguiente hoja si falla la creación

            # Escribir el DataFrame en la hoja
            # Convertir DataFrame a lista de listas, incluyendo el encabezado si es necesario
            if sheet_name == 'Complementary' and not df.empty:
                 values_to_write = [df.columns.values.tolist()] + df.values.tolist()
            else:
                 # Para las otras hojas, escribimos sin encabezado de pandas (usamos el de la primera fila de datos)
                 values_to_write = df.values.tolist()

            if values_to_write:
                try:
                    worksheet.update(values_to_write)
                    print(f"      ✓ Datos escritos en la hoja '{sheet_name}'.")
                except Exception as e:
                    print(f"      ✗ Error al escribir datos en la hoja '{sheet_name}': {e}")
            else:
                 print(f"      ⚠️ No hay datos para escribir en la hoja '{sheet_name}'.")


    except Exception as e:
        print(f"[ERROR] Error al escribir DataFrames en el nuevo spreadsheet: {e}")


    # Procesar el nuevo spreadsheet para DCF
    print("\nProcesando spreadsheet combinado para cálculo DCF...")
    # Ahora process_excel_file leerá directamente del spreadsheet ID
    l14_value = process_google_sheet_for_dcf(combined_spreadsheet_id, client, forecast_years, x2)
    print(f"✓ Cálculo DCF completado. L14 Value: {l14_value}")

    # Actualizar Google Sheets principal
    print(f"\nActualizando Google Sheets principal ('Results') para {stock_code}...")

    # Obtener valores adicionales del DataFrame df_yahoo_info (ahora desde combined_dfs['Complementary'])
    country_value = "-"
    book_value = "-"
    pe_ratio = "-"
    market_cap = "-"
    dividend_ratio = "-"
    net_debt_growth = "-" # Estos pueden requerir cálculo o estar en otras pestañas
    total_assets_growth = "-"
    net_pp_growth = "-"

    if 'Complementary' in combined_dfs and not combined_dfs['Complementary'].empty:
        df_yahoo_info = combined_dfs['Complementary']
        # Convertir a dict para fácil acceso por el nombre del item, asegurando que la primera columna sea la clave
        if not df_yahoo_info.empty and df_yahoo_info.columns[0]:
             try:
                yahoo_info_dict = df_yahoo_info.set_index(df_yahoo_info.columns[0]).to_dict('index')

                # Asumiendo que los valores están en la primera columna de datos después del índice
                # Verificar si hay al menos 2 columnas antes de acceder a df_yahoo_info.columns[1]
                data_column_name = df_yahoo_info.columns[1] if len(df_yahoo_info.columns) > 1 else None

                if data_column_name:
                    country_value = yahoo_info_dict.get('country', {}).get(data_column_name, '-')
                    book_value = yahoo_info_dict.get('bookValue', {}).get(data_column_name, '-')
                    pe_ratio = yahoo_info_dict.get('trailingPE', {}).get(data_column_name, '-')
                    market_cap = yahoo_info_dict.get('marketCap', {}).get(data_column_name, '-')
                    dividend_ratio = yahoo_info_dict.get('dividendYield', {}).get(data_column_name, '-')
                    # Para los crecimientos, tendríamos que buscarlos en otras pestañas combinadas si es necesario
             except Exception as e:
                  print(f"[WARNING] Error al procesar df_yahoo_info combinado: {e}")
        else:
             print("[WARNING] df_yahoo_info combinado está vacío o no tiene una primera columna válida.")

    # También podrías intentar obtener currency y exchange_rate de alguna parte aquí si no se obtuvieron antes

    update_google_sheets_results(main_sheet, stock_code, stock_name_clean, l14_value, ticker_empresa, combined_spreadsheet_link, currency, exchange_rate, tipo_empresa, country_value, book_value, pe_ratio, market_cap, dividend_ratio, net_debt_growth, total_assets_growth, net_pp_growth)
    print(f"✓ Google Sheets principal actualizado para {stock_code}")
    print(f"[DEBUG] process_stock_code_from_sheets finalizado para {stock_code}")


# Nueva función para procesar DCF directamente en Google Sheets
def process_google_sheet_for_dcf(spreadsheet_id, client, forecast_years, x2):
    print(f"[DEBUG] process_google_sheet_for_dcf iniciado para spreadsheet ID: {spreadsheet_id}")
    l14_value = "Error en cálculo DCF"

    try:
        spreadsheet = client.open_by_key(spreadsheet_id)

        # Leer datos de las hojas combinadas necesarias para el DCF
        try:
            df_cash_flow = pd.DataFrame(spreadsheet.worksheet('Cash Flow Statement').get_all_values())
            df_balance_sheet = pd.DataFrame(spreadsheet.worksheet('Balance Sheet').get_all_values())
            df_financial = pd.DataFrame(spreadsheet.worksheet('Financial').get_all_values())
            # df_ratios = pd.DataFrame(spreadsheet.worksheet('Ratios').get_all_values()) # No se usa directamente en el cálculo DCF
            # df_valuation = pd.DataFrame(spreadsheet.worksheet('Valuation').get_all_values()) # No se usa directamente en el cálculo DCF

            # Asumiendo que la primera fila es el encabezado y la primera columna es la clave para estas hojas
            if not df_cash_flow.empty:
                df_cash_flow.columns = df_cash_flow.iloc[0]
                df_cash_flow = df_cash_flow[1:].reset_index(drop=True)
                df_cash_flow.iloc[:, 0] = df_cash_flow.iloc[:, 0].astype(str) # Ensure key column is string

            if not df_balance_sheet.empty:
                 df_balance_sheet.columns = df_balance_sheet.iloc[0]
                 df_balance_sheet = df_balance_sheet[1:].reset_index(drop=True)
                 df_balance_sheet.iloc[:, 0] = df_balance_sheet.iloc[:, 0].astype(str) # Ensure key column is string

            if not df_financial.empty:
                 df_financial.columns = df_financial.iloc[0]
                 df_financial = df_financial[1:].reset_index(drop=True)
                 df_financial.iloc[:, 0] = df_financial.iloc[:, 0].astype(str) # Ensure key column is string


        except gspread.WorksheetNotFound as e:
             print(f"[ERROR] Hoja necesaria para DCF no encontrada: {e}")
             return "Error: Hoja DCF no encontrada"
        except Exception as e:
             print(f"[ERROR] Error al leer datos para DCF desde sheets: {e}")
             return "Error: Lectura datos DCF"


        # --- Lógica de cálculo DCF (similar a process_excel_file pero con DataFrames leídos de Sheets) ---

        def find_row_data_from_df(df, search_str):
            if df.empty:
                return []
            df_str = df.astype(str) # Convertir a string para buscar
            for idx, row in df_str.iterrows():
                if any(search_str.lower() in str(cell).lower() for cell in row):
                    # Retornar valores originales (no string)
                    return df.iloc[idx, 1:].values.tolist()
            print(f"[WARNING] No se encontró '{search_str}' en la tabla para DCF.")
            return []

        # Encontrar la fila de encabezados de periodo fiscal
        fiscal_period_headers = find_row_data_from_df(df_cash_flow, 'Fiscal Period')

        # Encontrar las filas de datos para las métricas clave
        unlevered_free_cash_flow_data = find_row_data_from_df(df_cash_flow, 'Unlevered Free Cash Flow')
        total_debt_data = find_row_data_from_df(df_balance_sheet, 'Net Debt')
        diluted_shares_data = find_row_data_from_df(df_financial, 'Diluted Weighted Average Shares Outstanding')


        # Alinear datos con los encabezados de periodo fiscal
        # Usar la longitud de los encabezados como referencia principal
        data_length = len(fiscal_period_headers)

        # Asegurarse de que todas las listas de datos tengan al menos la misma longitud que los encabezados
        unlevered_free_cash_flow_data = unlevered_free_cash_flow_data[:data_length] + [np.nan] * max(0, data_length - len(unlevered_free_cash_flow_data))
        total_debt_data = total_debt_data[:data_length] + [np.nan] * max(0, data_length - len(total_debt_data))
        diluted_shares_data = diluted_shares_data[:data_length] + [np.nan] * max(0, data_length - len(diluted_shares_data))


        # Crear el DataFrame DCF inicial alineado por columnas
        df_dcf = pd.DataFrame({
            'Fiscal Period': fiscal_period_headers,
            'Unlevered Free Cash Flow': unlevered_free_cash_flow_data,
            'Total Debt': total_debt_data,
            'Diluted Shares Outstanding': diluted_shares_data,
        })

        # Transponer el DataFrame
        df_dcf = df_dcf.set_index('Fiscal Period').T

        # Limpiar los nombres de columna para quedarnos solo con años o periodos relevantes
        df_dcf.columns = [str(col).strip() for col in df_dcf.columns]
        df_dcf = df_dcf.rename(columns=lambda x: x.replace('nan', '').strip())
        df_dcf = df_dcf.loc[:, df_dcf.columns.astype(bool)]

        # Asegurarse de que las columnas (años/periodos) sean únicas y ordenadas
        # Esto es crucial para evitar la duplicación visual en Excel
        # Convertir a numérico si es posible para ordenar correctamente
        try:
            ordered_columns = sorted(df_dcf.columns, key=lambda x: int(x) if str(x).isdigit() else float('inf'))
            df_dcf = df_dcf[ordered_columns]
        except ValueError:
             pass # Mantener orden actual si no son solo números

        # Agregar años futuros al DataFrame DCF
        if len(df_dcf.columns) > 0:
            try:
                # El último año es la última columna en el DataFrame ordenado
                last_year_str = df_dcf.columns[-1]
                last_year = int(last_year_str)
                new_years = [str(last_year + i) for i in range(1, forecast_years + 1)]
                # Agregar las nuevas columnas al DataFrame DCF
                for year in new_years:
                    df_dcf[year] = np.nan # Agregar columnas para años futuros, inicializadas con NaN

            except Exception as e:
                print(f"[WARNING] No se pudo agregar años futuros al DCF DataFrame: {e}")
                # Continuar sin años futuros si falla


        # Escribir en la hoja DCF del spreadsheet combinado
        print(f"Escribiendo DataFrame DCF en el spreadsheet ID: {spreadsheet_id}")
        try:
            ws_dcf = spreadsheet.worksheet('DCF')
             # Si existe, borrar el contenido actual
            ws_dcf.clear()
            print("      Limpiando hoja existente 'DCF' antes de escribir.")
        except gspread.WorksheetNotFound:
            # Si no existe, crear la hoja
            print("      Creando hoja 'DCF' en el nuevo spreadsheet.")
            ws_dcf = spreadsheet.add_worksheet(title='DCF', rows="1000", cols="50")
        except Exception as e:
             print(f"[ERROR] Error al verificar o crear la hoja 'DCF': {e}. No se pudo continuar con el cálculo.")
             return "Error: No se pudo crear hoja DCF"


        # Escribir el DataFrame DCF
        # Escribir el encabezado (Fiscal Period y años)
        header_values = [['Fiscal Period'] + df_dcf.columns.tolist()]
        ws_dcf.update(header_values, range_name='A1')

        # Escribir el resto del DataFrame (índice y valores)
        data_values = df_dcf.reset_index().values.tolist()
        if data_values:
             ws_dcf.update(data_values, range_name='A2')
             print("      ✓ DataFrame DCF escrito en la hoja 'DCF'.")
        else:
             print("      ⚠️ DataFrame DCF está vacío. No se escribió data.")


        # Escribir valor de x2 (Hardcodeado en K2 según el código original, o L2?)
        # El código original lo pone en X2 de la hoja DCF. Mantengamos eso.
        ws_dcf.update_cell(2, 24, x2) # Celda X2 (col 24, row 2)
        print(f"      ✓ Valor X2 ({x2}) escrito en la celda X2 de la hoja 'DCF'.")


        # Calcular fórmulas en Google Sheets usando la API de gspread
        # Similar a la lógica de process_excel_file, pero construyendo fórmulas de Google Sheets
        # Las columnas de datos históricos están en df_dcf.columns hasta antes de los new_years
        historical_years_count = len(df_dcf.columns) - forecast_years
        start_data_col_index = 2 # Columna B (después del índice)
        last_data_col_index = start_data_col_index + historical_years_count - 1
        start_forecast_col_index = last_data_col_index + 1 # Columna donde empieza el primer año pronosticado

        unlevered_row = 3 # Fila 'Unlevered Free Cash Flow' en Excel
        fiscal_period_row = 1 # Fila de los años en Google Sheets (encabezado)


        for i in range(forecast_years):
            current_col_index = start_forecast_col_index + i
            current_col_letter = col_num_to_letter(current_col_index)

            if i == 0:
                # Para el primer año de pronóstico, usar todos los años históricos disponibles
                forecast_range_data = f'{col_num_to_letter(start_data_col_index)}{unlevered_row}:{col_num_to_letter(last_data_col_index)}{unlevered_row}'
                forecast_range_headers = f'{col_num_to_letter(start_data_col_index)}{fiscal_period_row}:{col_num_to_letter(last_data_col_index)}{fiscal_period_row}'
                formula = (
                    f'=IF({ws_dcf.cell(row=unlevered_row, column=last_data_col_index).coordinate}*'
                    f'FORECAST({ws_dcf.cell(row=fiscal_period_row, column=current_col_index).coordinate},'
                    f'{forecast_range_data},'
                    f'{forecast_range_headers})>0,'
                    f'FORECAST({ws_dcf.cell(row=fiscal_period_row, column=current_col_index).coordinate},'
                    f'{forecast_range_data},'
                    f'{forecast_range_headers}),'
                    f'{ws_dcf.cell(row=unlevered_row, column=last_data_col_index).coordinate})*$X$2'
                )
            else:
                # Para años futuros, usar los años históricos y los años pronosticados previamente
                previous_forecast_col_index = start_forecast_col_index + i - 1
                forecast_range_data = f'{col_num_to_letter(start_data_col_index)}{unlevered_row}:{col_num_to_letter(previous_forecast_col_index)}{unlevered_row}'
                forecast_range_headers = f'{col_num_to_letter(start_data_col_index)}{fiscal_period_row}:{col_num_to_letter(previous_forecast_col_index)}{fiscal_period_row}'
                formula = (
                     f'=IF({ws_dcf.cell(row=unlevered_row, column=previous_forecast_col_index).coordinate}*'
                     f'FORECAST({ws_dcf.cell(row=fiscal_period_row, column=current_col_index).coordinate},'
                     f'{forecast_range_data},'
                     f'{forecast_range_headers})>0,'
                     f'FORECAST({ws_dcf.cell(row=fiscal_period_row, column=current_col_index).coordinate},'
                     f'{forecast_range_data},'
                     f'{forecast_range_headers}),'
                     f'{ws_dcf.cell(row=unlevered_row, column=previous_forecast_col_index).coordinate})*$X$2'
                )
            ws_dcf.cell(row=unlevered_row, column=current_col_index, value=formula)


        # Calcular terminal value y otros valores
        last_forecast_col_index = start_forecast_col_index + forecast_years - 1
        last_forecast_col_letter = col_num_to_letter(last_forecast_col_index)

        # TERMINAL VALUE (K7 en excel original, L7 en Google Sheets?) -> Pongamos en K7 y L7 para simular el excel
        formulas_to_update.append({'range': 'K7', 'values': [['TERMINAL VALUE']]})
        # L7: =ÚLTIMO_FCF_PRONOSTICADO * (1 + 0.02) / (0.06 - 0.02)
        # Último FCF pronosticado está en la fila unlevered_row_sheet, última columna pronosticada (last_forecast_col_letter)
        formulas_to_update.append({
             'range': 'L7',
             'values': [[f'={last_forecast_col_letter}{unlevered_row_sheet} * (1 + 0.02) / (0.06 - 0.02)']]
        })

        # NPV of FCF (K8 y L8)
        formulas_to_update.append({'range': 'K8', 'values': [['NPV of FCF']]})
        start_npv_fcf_col_letter = col_num_to_letter(start_forecast_col_index)
        end_npv_fcf_col_letter = col_num_to_letter(last_forecast_col_index)
        ws_dcf.update_cell(8, 1, f'=NPV(0.06, {start_npv_fcf_col_letter}{unlevered_row_sheet}:{end_npv_fcf_col_letter}{unlevered_row_sheet})')


        # NPV of TV (K9 y L9)
        formulas_to_update.append({'range': 'K9', 'values': [['NPV of TV']]})
        start_npv_tv_col_letter = col_num_to_letter(start_forecast_col_index)
        end_npv_tv_col_letter = col_num_to_letter(terminal_value_col_index)
        ws_dcf.update_cell(9, 1, f'=NPV(0.06, {start_npv_tv_col_letter}{npv_tv_row_sheet}:{end_npv_tv_col_letter}{npv_tv_row_sheet})')


        # Total EV (K10 y L10)
        formulas_to_update.append({'range': 'K10', 'values': [['Total EV']]})
        formulas_to_update.append({'range': 'L10', 'values': [['=L8 + L9']]})

        # Net Debt (K11 y L11)
        formulas_to_update.append({'range': 'K11', 'values': [['Net Debt']]})
        last_total_debt_col_letter = col_num_to_letter(start_data_col_index + len(total_debt_data) -1) if len(total_debt_data) > 0 else None # Última columna con dato
        if last_total_debt_col_letter:
             formulas_to_update.append({
                 'range': 'L11',
                 'values': [[f'={last_total_debt_col_letter}{total_debt_row_sheet}']]
             })
        else:
             formulas_to_update.append({'range': 'L11', 'values': [["Valor no disponible"]]})
             print("[WARNING] No se encontraron datos de Net Debt para la fórmula L11.")


        # Equity (K12 y L12)
        formulas_to_update.append({'range': 'K12', 'values': [['Equity']]})
        formulas_to_update.append({'range': 'L12', 'values': [[f'=L10 - L11']]})

        # Shares Outstanding (K13 y L13)
        formulas_to_update.append({'range': 'K13', 'values': [['Shares Outstanding']]})
        last_total_shares_col_letter = col_num_to_letter(start_data_col_index + len(total_shares_data) - 1) if len(total_shares_data) > 0 else None
        if last_total_shares_col_letter:
             formulas_to_update.append({
                 'range': 'L13',
                 'values': [[f'={last_total_shares_col_letter}{total_shares_row_sheet}']]
             })
        else:
             formulas_to_update.append({'range': 'L13', 'values': [["Valor no disponible"]]})
             print("[WARNING] No se encontraron datos de Shares Outstanding para la fórmula L13.")


        # Target Price (K14 y L14)
        formulas_to_update.append({'range': 'K14', 'values': [['Target Price']]})
        # La fórmula es L12/L13, pero debemos manejar la división por cero o valor no disponible
        formulas_to_update.append({
            'range': 'L14',
            'values': [[f'=IF(ISNUMBER(L13), IF(L13<>0, L12/L13, "#DIV/0!"), "Valor no disponible")']]
        })

        # Realizar la actualización por lotes de las fórmulas
        if formulas_to_update:
            try:
                ws_dcf.batch_update(formulas_to_update)
                print("      ✓ Fórmulas DCF escritas en la hoja 'DCF'.")
            except Exception as e:
                print(f"      ✗ Error al escribir fórmulas en la hoja 'DCF': {e}")
        else:
             print("      ⚠️ No hay fórmulas para escribir en la hoja 'DCF'.")


        # Leer el valor calculado de L14 (celda con Target Price)
        try:
            l14_cell = ws_dcf.acell('L14') # Leer celda L14
            l14_value = l14_cell.value
            print(f"[DEBUG] Valor leído de L14: {l14_value}")
        except Exception as e:
            print(f"[ERROR] Error al leer el valor de L14: {e}")
            l14_value = "Error al leer L14"

    except Exception as e:
        print(f"[ERROR] Error general durante el cálculo del DCF en Google Sheets: {str(e)}")
        l14_value = "Error general DCF"

    print(f"[DEBUG] process_google_sheet_for_dcf finalizado. Retornando L14: {l14_value}")
    return l14_value


# Función para actualizar la hoja "Results" del spreadsheet principal
def update_google_sheets_results(main_sheet, stock_code, stock_name_clean, l14_value, ticker_empresa, shared_link, currency, exchange_rate, tipo_empresa, country_value, book_value, pe_ratio, market_cap, dividend_ratio, net_debt_growth, total_assets_growth, net_pp_growth):
     try:
        print(f"[DEBUG] update_google_sheets_results llamado para {stock_code}")

        sheet = main_sheet.worksheet('Results')

        # Buscar la fila correspondiente para el código de stock
        cell = sheet.find(stock_code, in_column=1)
        if cell:
            row_index = cell.row
            print(f"[DEBUG] Ticker {stock_code} encontrado en la fila {row_index}. Actualizando fila.")
        else:
            row_index = len(sheet.col_values(1)) + 1 # Añadir al final si no se encuentra
            print(f"[DEBUG] Ticker {stock_code} no encontrado. Añadiendo a la nueva fila {row_index}.")

        # Crear el hipervínculo para la columna A (Ticker)
        link_marketscreener = f'https://www.marketscreener.com/quote/stock/{stock_code}'
        formula_link_ticker = f'=HYPERLINK("{link_marketscreener}", "{stock_code}")'

        # Actualizar la fila específica con los nuevos datos
        # Definir los valores a actualizar en el orden de las columnas de la hoja "Results"
        # Columnas esperadas: A=Ticker (link), B=Nombre, C=L14 (Target Price), D=Ticker Empresa, E=Enlace Combinado, F=Moneda, G=Tasa Cambio, H=Tipo Empresa, I=País, J=Book Value, K=PE Ratio, L=Market Cap, M=Dividend Ratio, N=Net Debt Growth, O=Total Assets Growth, P=Net PP Growth, Q=Fecha Actualización

        values_to_update = [
            formula_link_ticker,        # Columna A: Ticker con enlace a Marketscreener
            stock_name_clean,           # Columna B: Nombre de la empresa
            str(l14_value),             # Columna C: Target Price (L14 del DCF) - asegurar que sea string
            ticker_empresa,             # Columna D: Ticker de la empresa
            shared_link,                # Columna E: Enlace al Spreadsheet Combinado en Drive
            currency,                   # Columna F: Moneda
            str(exchange_rate),         # Columna G: Tasa de Cambio a USD - asegurar que sea string
            tipo_empresa,               # Columna H: Tipo de Empresa
            country_value,              # Columna I: País (desde Complementary)
            str(book_value),            # Columna J: Book Value (desde Complementary) - asegurar que sea string
            str(pe_ratio),              # Columna K: PE Ratio (desde Complementary) - asegurar que sea string
            str(market_cap),            # Columna L: Market Cap (desde Complementary) - asegurar que sea string
            str(dividend_ratio),        # Columna M: Dividend Ratio (desde Complementary) - asegurar que sea string
            str(net_debt_growth),       # Columna N: Net Debt Growth (placeholder) - asegurar que sea string
            str(total_assets_growth),   # Columna O: Total Assets Growth (placeholder) - asegurar que sea string
            str(net_pp_growth),         # Columna P: Net PP Growth (placeholder) - asegurar que sea string
            "16/05/2025"                # Columna Q: Fecha Actualización (hardcoded por ahora)
        ]

        # Actualizar el rango de la fila completa
        # Determinar la columna final (Q es la 17ª columna)
        end_column_letter = col_num_to_letter(len(values_to_update)) # Calcular la letra de la última columna
        update_range = f'A{row_index}:{end_column_letter}{row_index}'

        try:
            sheet.update(values_to_update, range_name=update_range)
            print(f"✓ Fila {row_index} en 'Results' actualizada para {stock_code}")
        except Exception as e:
            print(f"[ERROR] Error al actualizar el rango {update_range} en Google Sheets: {e}")


     except Exception as e:
         print(f"[ERROR] Error general al actualizar datos en Google Sheets ('Results'): {e}")

# Reemplazar la llamada a process_all_stock_codes en el script principal
# con la nueva lógica que no requiere el archivo local como input inicial
# La ruta del archivo local 'StockCodes.xlsx' ya no es necesaria como input principal para process_all_stock_codes
# Pero la ruta del JSON de credenciales sí
# Las variables forecast_years y x2 sí son necesarias
# file_path = r'C:\Users\relim\Desktop\prueba\StockCodes.xlsx' # Ya no se usa para leer tickers
forecast_years = 7
x2 = 1

# Llamar a la nueva función principal que orquesta el proceso
process_all_stock_codes("dummy_path.xlsx", forecast_years, x2) # Pasamos un dummy_path ya que no se usa