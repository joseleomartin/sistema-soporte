from concurrent.futures import ThreadPoolExecutor
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.keys import Keys
from lxml import etree
import pandas as pd
import numpy as np
import time
from openpyxl import load_workbook
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import re
import xlwings as xw
import gc
from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive
import os
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
import statsmodels.api as sm
import threading
from concurrent.futures import ThreadPoolExecutor
import time
import requests
import shutil 
import tempfile
import yfinance as yf
from sklearn.linear_model import LinearRegression
import scipy.stats as stats  # Agregar al inicio del archivo si no está
import matplotlib.pyplot as plt
from sklearn.metrics import mean_absolute_error, mean_squared_error


lock = threading.Lock()


def setup_drive():
    SCOPES = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
    SERVICE_ACCOUNT_FILE = 'C:/Users/relim/Desktop/prueba/durable-binder-432321-q4-5f1ef9e64ea1.json'
    
    creds = service_account.Credentials.from_service_account_file(
        SERVICE_ACCOUNT_FILE, scopes=SCOPES)
    
    service = build('drive', 'v3', credentials=creds)
    return service

def upload_and_get_shared_link(service, file_path):
     with lock:
        file_metadata = {
            'name': os.path.basename(file_path),
            'mimeType': 'application/vnd.google-apps.spreadsheet'
        }
        media = MediaFileUpload(file_path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        file = service.files().create(body=file_metadata, media_body=media, fields='id').execute()
        file_id = file.get('id')
        
        # Cambiar permisos para compartir el archivo
        permissions = {
            'type': 'anyone',
            'role': 'writer'
        }
        service.permissions().create(fileId=file_id, body=permissions).execute()
        
        # Obtener enlace compartido
        file_info = service.files().get(fileId=file_id, fields='webViewLink').execute()
        shared_link = file_info.get('webViewLink')
        print(f"Enlace compartido generado: {shared_link}")
        return shared_link


def setup_driver():
    try:
        chrome_options = Options()
        chrome_options.add_argument("--headless=new")  # Usar la nueva sintaxis de headless
        # User-Agent personalizado para simular un navegador real
        chrome_options.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36")
        # Flags anti-detección
        chrome_options.add_argument("--disable-blink-features=AutomationControlled")
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation", "enable-logging"])
        chrome_options.add_experimental_option('useAutomationExtension', False)
        # Configuraciones adicionales para mejorar la estabilidad
        chrome_options.add_argument("--start-maximized")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--disable-extensions")
        chrome_options.add_argument("--disable-popup-blocking")
        chrome_options.add_argument("--blink-settings=imagesEnabled=false")
        chrome_options.add_argument('--ignore-certificate-errors')
        chrome_options.add_argument('--disable-web-security')
        chrome_options.add_argument('--allow-running-insecure-content')
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        # Silenciar logs de Chromium y evitar registro de push (GCM)
        chrome_options.add_argument("--disable-notifications")
        chrome_options.add_argument("--disable-features=BackgroundMode,PushMessaging,IdleDetection")
        chrome_options.add_argument("--disable-logging")
        chrome_options.add_argument("--log-level=3")
        
        # Configurar el servicio de Chrome
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=chrome_options)
        driver.set_page_load_timeout(30)
        return driver
    except Exception as e:
        print(f"Error al configurar el driver de Chrome: {e}")
        raise

def login(driver):
    print(f"[LOG] Navegando a página de login...")
    driver.set_page_load_timeout(30)
    driver.get("https://www.marketscreener.com/login/")
    print(f"[LOG] Página de login cargada, esperando campo de usuario (timeout: 15s)...")
    try:
        username_field = WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.XPATH, '/html/body/div[1]/div/div/main/div/div/div/div/form/div[1]/input'))
        )
        print(f"[LOG] Campo de usuario encontrado, buscando campo de contraseña...")
        password_field = driver.find_element(By.XPATH, '/html/body/div[1]/div/div/main/div/div/div/div/form/div[2]/input')
        print(f"[LOG] Campos encontrados, ingresando credenciales...")
        username_field.clear()
        username_field.send_keys("leojosemartin@gmail.com")
        password_field.clear()
        password_field.send_keys("Papaleo111.")
        print(f"[LOG] Credenciales ingresadas, enviando formulario...")
        password_field.send_keys(Keys.RETURN)
        print(f"[LOG] Formulario enviado, esperando confirmación de login (5s)...")
        time.sleep(5)  # Espera a que el inicio de sesión se complete
        
        # Verificar que el login fue exitoso (buscar elementos que solo aparecen después del login)
        try:
            WebDriverWait(driver, 5).until(
                lambda d: "login" not in d.current_url.lower() or d.current_url != "https://www.marketscreener.com/login/"
            )
            print(f"[LOG] Login completado - URL actual: {driver.current_url}")
        except:
            print(f"[LOG] ADVERTENCIA: No se pudo confirmar el login, pero continuando...")
    except Exception as e:
        print(f"[LOG] ERROR durante el proceso de login: {e}")
        raise

def capture_table(driver, url, xpath, header_xpath=None):
    print(f"Paso 1: Navegando a la URL: {url}")
    driver.get(url)
    time.sleep(2)
    print("Paso 2: Desplazándose hasta el final de la página")
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    print(f"Paso 3: Esperando la presencia del elemento con XPath: {xpath}")
    WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, xpath)))
    print("Paso 4: Extrayendo el contenido HTML de la página")
    html = driver.page_source
    dom = etree.HTML(html)
    
    # Extraer encabezados si se proporciona un header_xpath
    headers = []
    if header_xpath:
        header_elements = dom.xpath(header_xpath)
        headers = [''.join(header.xpath(".//text()")).strip() for header in header_elements]

    print(f"Paso 6: Buscando filas en la tabla con el XPath: {xpath}")
    rows = dom.xpath(xpath)
    if not rows:
        print(f"No se encontraron filas con el XPath: {xpath}")
        return []

    table_data = []
    for row in rows:
        cells = row.xpath(".//td | .//th")
        row_data = [''.join(cell.xpath(".//text()")).strip() for cell in cells]
        table_data.append(row_data)

    # Agregar encabezados al inicio de los datos de la tabla
    if headers:
        table_data.insert(0, headers)

    return table_data


import os
import shutil

def cleanup_temp_folder(temp_folder):
    try:
        # Elimina todos los archivos y directorios en la carpeta Temp
        shutil.rmtree(temp_folder, ignore_errors=True)
        print(f"Carpeta temporal {temp_folder} limpiada exitosamente.")
    except Exception as e:
        print(f"Error al eliminar el contenido de la carpeta temporal: {e}")



EXCEPTIONS = {'0', "Diluted Shares Outstanding", "Basic Weighted Average Shares Outstanding", "Diluted Weighted Average Shares Outstanding", "Basic Weighted Average Shares Outstanding "
              "Diluted Weighted Average Shares Outstanding ","Basic Weighted Average Shares Outstanding ", "Payout Ratio", "American Depositary Receipts Ratio (ADR)",
                "American Depositary Receipts Ratio", "Effective Tax Rate - (Ratio)", "Effective Tax Rate", "Net EPS - Basic", "Basic EPS - Continuing Operations", "Net EPS", "Basic EPS", "Net EPS - Diluted", "Net EPS", "Diluted EPS - Continuing Operations", "Diluted EPS",
                "Normalized Basic EPS", "Normalized Diluted EPS",
                "Diluted Weighted Average Shares Outstanding ", "Basic Weighted Average Shares Outstanding ",
                "Diluted Shares Outstanding", "Diluted Weighted Average Shares Outstanding","Basic Weighted Average Shares Outstanding", "Fiscal period","Fiscal Period", "Fiscal Period: January", "Fiscal Period: February", "Fiscal Period: March", "Fiscal Period: April", "Fiscal Period: May", "Fiscal Period: June", "Fiscal Period: July", "Fiscal Period: August", "Fiscal Period: September", "Fiscal Period: October", "Fiscal Period: November", "Fiscal Period: December" }



def convert_value(value):
    if isinstance(value, str):
        value = value.strip()
        if value in EXCEPTIONS:
            return value
        if value == '-':
            return 0
        if value.endswith('M'):
            return float(value[:-1].replace(',', ''))  # Ya está en millones
        elif value.endswith('B'):
            return float(value[:-1].replace(',', '')) * 1000  # 1B = 1000 millones
        elif value.endswith('K'):
            return float(value[:-1].replace(',', '')) / 1000  # 1K = 0.001 millones
        elif value.replace(',', '').lstrip('-').isdigit():  # Permitir números negativos
            return int(value.replace(',', ''))
        else:
            return value
    return value

def convert_dataframe(df):
    for col in df.columns:
        if col != 0:  # Aplicar conversión solo si la columna no es la columna 0
            df[col] = df[col].apply(convert_value)
    return df


def get_exchange_rate(from_currency, to_currency="USD"):
    """Obtiene la tasa de cambio desde la moneda de origen a la moneda de destino."""
    url = f"https://api.exchangerate-api.com/v4/latest/{from_currency}"
    response = requests.get(url)
    data = response.json()
    return data['rates'].get(to_currency, 1)  


def convert_to_usd(df, exchange_rate):
    # Convertir solo columnas numéricas después de aplicar map
    numeric_cols = df.columns[df.apply(lambda col: col.apply(lambda x: isinstance(x, (int, float)) or (isinstance(x, str) and x.replace(',', '').isdigit())).any())]

    # Multiplicar por la tasa de cambio solo si la fila no tiene un valor en la columna A que sea una excepción
    for index in df.index:
        if df.at[index, df.columns[0]] not in EXCEPTIONS:
            df.loc[index, numeric_cols] = df.loc[index, numeric_cols].apply(lambda x: x * exchange_rate if isinstance(x, (int, float)) else x)
    
    return df


def extract_currency_from_url(base_url):
    # Agregar el sufijo '-income-statement/' a la URL
    full_url = f"{base_url}-income-statement/"

    # Crear una nueva instancia del navegador
    driver = setup_driver()

    # Navegar a la URL completa
    driver.get(full_url)

    extracted_text = None  # Inicializar la variable para almacenar el texto

    try:
        # Utiliza un selector CSS para encontrar el elemento que contiene la moneda
        wait = WebDriverWait(driver, 20)  # Aumentar el tiempo de espera a 20 segundos
        element = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, ".js-selectedCurrency.px-10")))
        extracted_text = element.text.strip()  # Usar strip() para limpiar espacios extra
        print(f"Texto extraído desde el elemento en {full_url}: {extracted_text}")

    except Exception as e:
        print(f"Error inesperado: {e}")

    finally:
        driver.quit()  # Cerrar el navegador al finalizar

    return extracted_text  # Retornar el texto extraído


def tipo_empresa_from_xpath(driver, tipo_xpath):
    try:
        # Esperar a que el elemento esté presente
        empresa_element = WebDriverWait(driver, 2).until(
            EC.presence_of_element_located((By.XPATH, tipo_xpath))
        )
        # Extraer el texto del elemento
        empresa_text = empresa_element.get_attribute('textContent').strip()
        
        if empresa_text:
            print(f"Texto extraído del XPath: '{empresa_text}'")
        else:
            print("No se pudo extraer el texto, el resultado es vacío.")
        
        return empresa_text
    except Exception as e:
        print(f"Error al extraer el texto desde el XPath: {str(e)}")
        return "No encontrado"
    
    

def ticker_market(driver, ticker_xpath):
    try:
        # Esperar a que el elemento esté presente
        ticker_element = WebDriverWait(driver, 2).until(
            EC.presence_of_element_located((By.XPATH, ticker_xpath))
        )
        # Extraer el texto del elemento
        ticker_element = ticker_element.get_attribute('textContent').strip()
        
        if ticker_element:
            print(f"Texto extraído del XPath: '{ticker_element}'")
        else:
            print("No se pudo extraer el texto, el resultado es vacío.")
        
        return ticker_element
    except Exception as e:
        print(f"Error al extraer el texto desde el XPath: {str(e)}")
        return "No encontrado"

def process_stock_code(stock_code, forecast_years, x2, drive):
    print(f"[LOG] ===== INICIANDO PROCESO PARA STOCK CODE: {stock_code} =====")
    print(f"[LOG] Parámetros: forecast_years={forecast_years}, x2={x2}")
    
    # Inicializar variables por defecto para evitar errores si falla la captura
    cash_flow_data = []
    balance_sheet_data = []
    financial_data = []
    ratios_data = []
    valuation_data = []
    ticker_empresa = "No encontrado"
    tipo_empresa = "No encontrado"
    currency = "USD"
    exchange_rate = 1.0
    pais = "No encontrado"
    stock_name_clean = stock_code
    df_yahoo_info = pd.DataFrame()
    df_history = pd.DataFrame()
    
    # Limpiar procesos de Chrome anteriores antes de iniciar uno nuevo
    print(f"[LOG] Limpiando procesos de Chrome anteriores...")
    try:
        os.system("taskkill /im chrome.exe /f >nul 2>&1")
        time.sleep(2)  # Esperar a que se cierren los procesos
        print(f"[LOG] Procesos de Chrome anteriores limpiados")
    except:
        print(f"[LOG] No se pudieron limpiar procesos anteriores (puede que no existan)")
    
    driver = setup_driver()
    print(f"[LOG] Driver de Chrome configurado correctamente")
    
    try:
        print(f"[LOG] Procesando stock code: {stock_code}")
        
        # Realizar login antes de procesar los datos
        print(f"[LOG] Iniciando login en Marketscreener...")
        try:
            login(driver)
            print(f"[LOG] Login exitoso")
        except Exception as login_error:
            print(f"[LOG] ERROR en login: {login_error}")
            import traceback
            traceback.print_exc()
            raise  # Re-lanzar el error para que se maneje en el bloque except principal
        
        base_url = f"https://www.marketscreener.com/quote/stock/{stock_code}/finances"
        base_url2 = f"https://www.marketscreener.com/quote/stock/{stock_code}/valuation/"

        # Intentar capturar los datos de valoración
        try:
            valuation_data = capture_table(
                driver, f"{base_url2}",
                '//*[@id="valuationEnterpriseTable"]/tbody/tr',
                header_xpath='//*[@id="valuationEnterpriseTable_wrapper"]/div/div[1]/div/table/thead/tr/th'
            )
            if valuation_data is None or not valuation_data:
                valuation_data = []  # Asignar lista vacía si no hay datos
                print(f"No se encontraron datos de valoración para {stock_code}")
            else:
                print(f"Valuation data capturada para {stock_code}")
        except Exception as e:
            valuation_data = []
            print(f"No se pudo capturar valuation data para {stock_code}. Error: {e}")
                            
        # Capturar los datos financieros
        print(f"[LOG] Capturando datos de Cash Flow Statement...")
        cash_flow_data = capture_table(
            driver, f"{base_url}-cash-flow-statement/",
            '//*[@id="horizontalFinancialTableN1_3"]/tbody/tr',
            header_xpath='//*[@id="horizontalFinancialTableN1_3_wrapper"]/div/div[1]/div/table/thead/tr/th'
        )
        print(f"[LOG] Cash Flow Statement capturado: {len(cash_flow_data)} filas")
        
        print(f"[LOG] Capturando datos de Balance Sheet...")
        balance_sheet_data = capture_table(
            driver, f"{base_url}-balance-sheet/",
            '//*[@id="horizontalFinancialTableN1_2"]/tbody/tr',
            header_xpath='//*[@id="horizontalFinancialTableN1_2_wrapper"]/div/div[1]/div/table/thead/tr/th'
        )
        print(f"[LOG] Balance Sheet capturado: {len(balance_sheet_data)} filas")
        
        print(f"[LOG] Capturando datos de Income Statement...")
        financial_data = capture_table(
            driver, f"{base_url}-income-statement/",
            '//*[@id="horizontalFinancialTableN1_1"]/tbody/tr',
            header_xpath='//*[@id="horizontalFinancialTableN1_1_wrapper"]/div/div[1]/div/table/thead/tr/th'
        )
        print(f"[LOG] Income Statement capturado: {len(financial_data)} filas")
        
        print(f"[LOG] Capturando datos de Ratios...")
        ratios_data = capture_table(
            driver, f"{base_url}-ratios/",
            '//*[@id="horizontalFinancialTableN1_5"]/tbody/tr',
            header_xpath='//*[@id="horizontalFinancialTableN1_5_wrapper"]/div/div[1]/div/table/thead/tr/th'
        )
        print(f"[LOG] Ratios capturado: {len(ratios_data)} filas")

        # Extraer el tipo de empresa desde el XPath específico
        tipo_xpath = "/html/body/div[1]/div/div[1]/main/div[1]/div[1]/div[2]/a[2]/h2"
        tipo_empresa = tipo_empresa_from_xpath(driver, tipo_xpath)

        ticker_xpath = "/html/body/div[1]/div/div[1]/main/div[1]/div/div[2]/h2[1]"
        ticker_empresa = ticker_market(driver, ticker_xpath)

        currency = extract_currency_from_url(base_url)
        if not currency:
            currency = "USD"
            print(f"No se pudo detectar la moneda para {stock_code}, usando USD como predeterminado")

        print(f"Moneda detectada para {stock_code}: {currency}")

        exchange_rate = get_exchange_rate(currency)
        print(f"Tasa de cambio a USD para {currency}: {exchange_rate}")

        stock_name_clean = re.sub(r'[^A-Za-z]+', ' ', stock_code).replace('-', ' ').strip()
        stock_codes_df = pd.read_excel(file_path, usecols=[0, 1], header=None)
        stock_name_from_file = stock_codes_df[stock_codes_df[0] == stock_code].iloc[0, 1] if not stock_codes_df[stock_codes_df[0] == stock_code].empty else ''

        # Extraer el país desde el xpath proporcionado
        pais_xpath = "/html/body/div[1]/div/div[1]/main/div[1]/div[1]/div[2]/a[1]/h2/span/i"
        try:
            pais_element = WebDriverWait(driver, 2).until(
                EC.presence_of_element_located((By.XPATH, pais_xpath))
            )
            pais = pais_element.get_attribute('title')
            if pais:
                print(f"País extraído: {pais}")
            else:
                pais = "No encontrado"
                print("No se pudo extraer el país, el resultado es vacío.")
        except Exception as e:
            pais = "No encontrado"
            print(f"Error al extraer el país desde el XPath: {str(e)}")

        # --- Formateo de Ticker para Yahoo Finance ---
        yahoo_ticker = ticker_empresa
        if pais == "Hong Kong":
            yahoo_ticker = ticker_empresa.zfill(4) + ".HK"
        elif pais == "China":
            if ticker_empresa.startswith('6'):
                yahoo_ticker = ticker_empresa + ".SS"
            else:
                yahoo_ticker = ticker_empresa + ".SZ"
        elif pais == "India" or currency == "INR":
            yahoo_ticker = ticker_empresa + ".NS"
        elif pais == "Japón" or pais == "Japan" or currency == "JPY":
            yahoo_ticker = ticker_empresa + ".T"
        elif pais == "Tailandia" or pais == "Thailand" or currency == "THB":
            yahoo_ticker = ticker_empresa + ".BK"
        elif pais == "Pakistán" or pais == "Pakistan" or currency == "PKR":
            yahoo_ticker = ticker_empresa + ".KA"
        elif pais == "Estados Unidos" or pais == "United States" or currency == "USD":
            yahoo_ticker = ticker_empresa
        elif pais == "Reino Unido" or pais == "United Kingdom" or currency == "GBP":
            yahoo_ticker = ticker_empresa + ".L"
        elif pais == "Alemania" or pais == "Germany" or currency == "EUR":
            yahoo_ticker = ticker_empresa + ".DE"
        elif pais == "Francia" or pais == "France":
            yahoo_ticker = ticker_empresa + ".PA"
        elif pais == "Australia" or currency == "AUD":
            yahoo_ticker = ticker_empresa + ".AX"
        elif pais == "Canadá" or pais == "Canada" or currency == "CAD":
            yahoo_ticker = ticker_empresa + ".TO"
        elif pais == "Suiza" or pais == "Switzerland" or currency == "CHF":
            yahoo_ticker = ticker_empresa + ".SW"
        elif pais == "Italia" or pais == "Italy":
            yahoo_ticker = ticker_empresa + ".MI"
        elif pais == "España" or pais == "Spain":
            yahoo_ticker = ticker_empresa + ".MC"
        elif pais == "Brasil" or pais == "Brazil" or currency == "BRL":
            yahoo_ticker = ticker_empresa + ".SA"
        elif pais == "Rusia" or pais == "Russia" or currency == "RUB":
            yahoo_ticker = ticker_empresa + ".ME"
        elif pais == "Turquía" or pais == "Turkey" or currency == "TRY":
            yahoo_ticker = ticker_empresa + ".IS"
        elif pais == "Corea del Sur" or pais == "South Korea" or currency == "KRW":
            yahoo_ticker = ticker_empresa + ".KS"
        elif pais == "Taiwán" or pais == "Taiwan" or currency == "TWD":
            yahoo_ticker = ticker_empresa + ".TW"
        elif pais == "Singapur" or pais == "Singapore" or currency == "SGD":
            yahoo_ticker = ticker_empresa + ".SI"
        elif pais == "Sudáfrica" or pais == "South Africa" or currency == "ZAR":
            yahoo_ticker = ticker_empresa + ".JO"
        elif pais == "Malasia" or pais == "Malaysia" or currency == "MYR":
            yahoo_ticker = ticker_empresa + ".KL"
        elif pais == "Indonesia" or currency == "IDR":
            yahoo_ticker = ticker_empresa + ".JK"
        elif pais == "Filipinas" or pais == "Philippines" or currency == "PHP":
            yahoo_ticker = ticker_empresa + ".PS"
        elif pais == "México" or pais == "Mexico" or currency == "MXN":
            yahoo_ticker = ticker_empresa + ".MX"
        elif pais == "Emiratos Árabes Unidos" or pais == "United Arab Emirates" or currency == "AED":
            yahoo_ticker = ticker_empresa + ".AD"
        else:
            yahoo_ticker = ticker_empresa  # Por defecto, sin sufijo
        print(f"--- DEBUG: Ticker original '{ticker_empresa}', Ticker para Yahoo '{yahoo_ticker}' (basado en país '{pais}', moneda '{currency}') ---")
        # --- Fin Formateo de Ticker ---

        # Agregar datos generales a la lista para la hoja 'Data'
        data_data = [
            ["ticker", ticker_empresa],
            ["nombre", stock_name_clean],
            ["empresa", tipo_empresa],
            ["pais", pais],
        ]

        # --- BLOQUE YAHOO FINANCE ---
        max_retries = 3
        retry_delay = 20  # segundos
        df_yahoo_info = pd.DataFrame()
        df_history = pd.DataFrame()

        for attempt in range(max_retries):
            try:
                print(f"Intentando obtener datos de Yahoo Finance para {yahoo_ticker} (intento {attempt + 1}/{max_retries})")
                ticker = yf.Ticker(yahoo_ticker)
                
                # 1. Obtener info general
                info_dict = ticker.info
                # Crear DataFrame desde el diccionario, manejando el caso de que esté vacío
                if info_dict and 'symbol' in info_dict: # Verificar que la info no esté vacía
                    df_yahoo_info = pd.DataFrame([info_dict])
                else:
                    print(f"--- DEBUG: No se encontró información 'info' para {yahoo_ticker} en Yahoo Finance. ---")
                    df_yahoo_info = pd.DataFrame()
                print(f"Datos de 'info' obtenidos para {yahoo_ticker}")

                # 2. Obtener historial de precios
                try:
                    df_history = ticker.history(period="max")
                    if df_history.empty:
                        print(f"--- DEBUG: No se encontró historial 'history' para {yahoo_ticker}. El DataFrame está vacío.")
                    else:
                        print(f"Datos de 'history' obtenidos para {yahoo_ticker}")
                except Exception as hist_e:
                    print(f"No se pudieron obtener datos históricos: {hist_e}")
                    df_history = pd.DataFrame() # Dejar vacío si falla

                break # Éxito, salir del bucle de reintentos
            except Exception as e:
                if "429" in str(e) or "Invalid Crumb" in str(e):
                    print(f"Error de API en intento {attempt + 1}. Esperando {retry_delay} segundos...")
                    time.sleep(retry_delay)
                    retry_delay *= 2
                else:
                    print(f"Error al obtener datos de Yahoo Finance: {e}")
                    df_yahoo_info = pd.DataFrame()
                    df_history = pd.DataFrame()
                    break
        
        if df_yahoo_info.empty:
            print(f"Fallo final al obtener datos de 'info' para {yahoo_ticker}")
        # --- FIN BLOQUE YAHOO FINANCE ---

        # Convertir los datos financieros a DataFrames
        df_data= pd.DataFrame(data_data)
        df_cash_flow = pd.DataFrame(cash_flow_data)
        df_balance_sheet = pd.DataFrame(balance_sheet_data)
        df_financial = pd.DataFrame(financial_data)
        df_ratios = pd.DataFrame(ratios_data)
        df_valuation = pd.DataFrame(valuation_data)
        

        df_cash_flow = convert_dataframe(pd.DataFrame(cash_flow_data))
        df_balance_sheet = convert_dataframe(pd.DataFrame(balance_sheet_data))
        df_financial = convert_dataframe(pd.DataFrame(financial_data))
        df_ratios = convert_dataframe(pd.DataFrame(ratios_data))


        # Convertir los datos a dólares usando la tasa de cambio obtenida
        df_cash_flow = convert_to_usd(df_cash_flow, exchange_rate)
        df_balance_sheet = convert_to_usd(df_balance_sheet, exchange_rate)
        df_financial = convert_to_usd(df_financial, exchange_rate)
        df_ratios = convert_to_usd(df_ratios, exchange_rate)

        print(f"[LOG] Datos financieros convertidos a USD correctamente")
        
        # Cerrar el driver después de capturar todos los datos (ya no lo necesitamos)
        print(f"[LOG] Cerrando navegador después de capturar datos...")
        driver.quit()
        print(f"[LOG] Navegador cerrado correctamente")
        
        # Mata procesos de Chrome
        try:
            os.system("taskkill /im chrome.exe /f")  # En Windows
            print(f"[LOG] Procesos de Chrome finalizados")
        except:
            print(f"[LOG] No se pudieron finalizar procesos de Chrome (puede que ya estén cerrados)")

        # Limpia el perfil temporal
        temp_folder = r"C:\Users\relim\AppData\Local\Temp" 
        cleanup_temp_folder(temp_folder)

    except Exception as e:
        print(f"[LOG] ERROR durante captura de datos: {e}")
        import traceback
        print(f"[LOG] Traceback completo:")
        traceback.print_exc()
        print(f"[LOG] Intentando cerrar driver...")
        try:
            driver.quit()
            print(f"[LOG] Driver cerrado en bloque except")
        except Exception as driver_error:
            print(f"[LOG] Error al cerrar driver (puede que ya esté cerrado): {driver_error}")
        try:
            os.system("taskkill /im chrome.exe /f")
            print(f"[LOG] Procesos de Chrome finalizados en bloque except")
        except:
            print(f"[LOG] No se pudieron finalizar procesos de Chrome")
        print(f"[LOG] Continuando con procesamiento usando valores por defecto (datos pueden estar incompletos)")

    # DEBUG: Preparando para escribir en Excel
    print(f"[LOG] ===== PREPARANDO PARA ESCRIBIR EN EXCEL: {stock_code} =====")
    print(f"--- DEBUG: Preparando para escribir en Excel: {stock_code} ---")
    # Solo mostrar logs si los objetos existen y no están vacíos
    if 'df_history' in locals() and df_history is not None:
        print(f"--- DEBUG: df_history shape: {df_history.shape} ---")
        print(df_history.head(10))
    if 'df_yahoo_info' in locals() and df_yahoo_info is not None:
        print(f"--- DEBUG: df_yahoo_info shape: {df_yahoo_info.shape} ---")
        print(df_yahoo_info.head(5))
    if 'df_data' in locals() and df_data is not None:
        print(f"--- DEBUG: df_data shape: {df_data.shape} ---")
        print(df_data.head(5))
    if 'df_monthly' in locals() and df_monthly is not None:
        print(f"--- DEBUG: df_monthly shape: {df_monthly.shape} ---")
        print(df_monthly.head(5))
    if 'df_pred' in locals() and df_pred is not None:
        print(f"--- DEBUG: df_pred shape: {df_pred.shape} ---")
        print(df_pred.head(5))
    if 'metrics' in locals() and metrics is not None:
        print(f"--- DEBUG: Métricas calculadas: {metrics} ---")

    excel_path = rf"C:\Users\relim\Desktop\prueba\Resultados\{ticker_empresa} - {stock_name_clean}.xlsx"
    # DEBUG: Mostrar información de todos los DataFrames antes de guardar
    for name, df in [
        ("Data", df_data),
        ("Cash Flow Statement", df_cash_flow),
        ("Balance Sheet", df_balance_sheet),
        ("Financial", df_financial),
        ("Ratios", df_ratios),
        ("Valuation", df_valuation),
        ("Complementary", df_yahoo_info),
        ("Yahoo History", df_history)
    ]:
        print(f"--- DEBUG: {name} ---")
        print("Empty:", df.empty)
        print("Columns:", df.columns)
        print("Dtypes:", df.dtypes)
        print(df.head())

   
    col_conceptos = df_valuation.columns[0]

# Asegurar que hay suficientes filas
    if len(df_valuation) >= 25:
        # Extraer las filas necesarias
        fila0 = df_valuation.iloc[0]    # Fiscal Period: nombres de años
        fila1 = df_valuation.iloc[1]    # Capitalization
        fila25 = df_valuation.iloc[24]  # Nbr of stocks (in thousands)

        # Extraer los conceptos
        concepto1 = fila1[col_conceptos]
        concepto25 = fila25[col_conceptos]

        # Años y valores (omitimos la columna de concepto)
        años = fila0.drop(labels=col_conceptos).tolist()
        valores_fila1 = fila1.drop(labels=col_conceptos).tolist()
        valores_fila25 = fila25.drop(labels=col_conceptos).tolist()

        # Filtrar los años que tengan datos válidos en ambas métricas
        años_validos = []
        cap_filtrada = []
        stocks_filtrada = []

        for i, año in enumerate(años):
            if valores_fila1[i] != "-" and valores_fila25[i] != "-":
                años_validos.append(año)
                cap_filtrada.append(valores_fila1[i])
                stocks_filtrada.append(valores_fila25[i])
            else:
                print(f"--- INFO: Año {año} eliminado por contener '-' ---")

        # Calcular Price = Capitalization / Stocks * 1000
        prices = []
        for cap, stocks in zip(cap_filtrada, stocks_filtrada):
            try:
                cap_val = float(str(cap).replace(",", ""))
                stocks_val = float(str(stocks).replace(",", ""))
                price = (cap_val / stocks_val) * 1000
                prices.append(round(price, 4))
            except Exception as e:
                prices.append("-")

        # Calcular Year Returns %
        year_returns = [""]  # primer año no tiene retorno
        for i in range(1, len(prices)):
            try:
                prev = prices[i - 1]
                curr = prices[i]
                if isinstance(prev, (int, float)) and isinstance(curr, (int, float)) and curr != 0:
                    ret = (prev / curr) - 1  # tu lógica: precio 1 / precio 2 - 1
                    year_returns.append(round(ret, 4))  # en porcentaje con 4 decimales
                else:
                    year_returns.append("")
            except Exception as e:
                print(f"Error en cálculo de retorno año {i}: {e}")
                year_returns.append("")

        # Construir el DataFrame
        data = {
            "Fiscal Period": [concepto1, concepto25, "Price", "Year Returns %"]
        }
        for i, año in enumerate(años_validos):
            data[año] = [
                cap_filtrada[i],
                stocks_filtrada[i],
                prices[i],
                year_returns[i]
            ]
        
        df_linear_regression = pd.DataFrame(data)

        # --- Agregar una fila vacía ---
        empty_row = pd.DataFrame([[""] * len(df_linear_regression.columns)], columns=df_linear_regression.columns)
        df_linear_regression = pd.concat([df_linear_regression, empty_row], ignore_index=True)

        # --- Insertar título "Cash Flow Statement" ---
        titulo_row = pd.DataFrame([["Cash Flow Statement"] + [""] * (len(df_linear_regression.columns) - 1)], columns=df_linear_regression.columns)
        df_linear_regression = pd.concat([df_linear_regression, titulo_row], ignore_index=True)

        # --- Insertar contenido de df_cash_flow ---
        df_cash_flow_limpio = df_cash_flow.copy()
        ncols = len(df_linear_regression.columns)

        if df_cash_flow_limpio.shape[1] >= ncols:
            df_cash_flow_limpio = df_cash_flow_limpio.iloc[:, :ncols]
        else:
            for _ in range(ncols - df_cash_flow_limpio.shape[1]):
                df_cash_flow_limpio[f"col_{_}"] = ""

        df_cash_flow_limpio.columns = df_linear_regression.columns
        df_linear_regression = pd.concat([df_linear_regression, df_cash_flow_limpio], ignore_index=True)

        empty_rows = pd.DataFrame([[""] * len(df_linear_regression.columns) for _ in range(2)], columns=df_linear_regression.columns)
        df_linear_regression = pd.concat([df_linear_regression, empty_rows], ignore_index=True)

        row_year_returns = df_linear_regression[df_linear_regression["Fiscal Period"] == "Year Returns %"]

        if not row_year_returns.empty:
            # Pegamos la fila de nuevo debajo de los espacios
            df_linear_regression = pd.concat([df_linear_regression, row_year_returns], ignore_index=True)
        else:
            print("--- WARN: No se encontró la fila 'Year Returns %', no se pudo copiar ---")

        # --- Calcular Year Returns % para métricas adicionales ---
        metricas_a_calcular = [
            "Net Income",
            "Cash from Operations",
            "Capital Expenditure",
            "Cash from Investing",
            "Repurchase of Common Stock",
            "Unlevered Free Cash Flow"
        ]

        col_conceptos_cashflow = df_cash_flow.columns[0]

        for nombre_metrica in metricas_a_calcular:
            fila = df_cash_flow[df_cash_flow[col_conceptos_cashflow] == nombre_metrica]
            if fila.empty:
                print(f"--- WARN: No se encontró la métrica '{nombre_metrica}' en df_cash_flow ---")
                continue

            fila = fila.iloc[0]
            valores = fila.drop(labels=col_conceptos_cashflow).tolist()

            valores_float = []
            for v in valores:
                try:
                    valor_float = float(str(v).replace(",", "").replace("(", "-").replace(")", ""))
                    valores_float.append(valor_float)
                except:
                    valores_float.append(None)

            year_returns = [""]  # Primer año no tiene retorno
            for i in range(1, len(valores_float)):
                try:
                    prev = valores_float[i - 1]
                    curr = valores_float[i]
                    if isinstance(prev, (int, float)) and isinstance(curr, (int, float)):
                        if prev == 0 or prev is None or curr is None:
                            year_returns.append(0)
                        else:
                            ret = (curr / prev) - 1  # Variación relativa
                            year_returns.append(round(ret, 4))
                    else:
                        year_returns.append("")
                except Exception as e:
                    print(f"Error en cálculo de retorno año {i}: {e}")
                    year_returns.append("")

            fila_resultado = [f"{nombre_metrica} Year Returns %"] + year_returns
            while len(fila_resultado) < len(df_linear_regression.columns):
                fila_resultado.append("")
            fila_resultado = fila_resultado[:len(df_linear_regression.columns)]

            nueva_fila_df = pd.DataFrame([fila_resultado], columns=df_linear_regression.columns)
            df_linear_regression = pd.concat([df_linear_regression, nueva_fila_df], ignore_index=True)

        # --- FIN del loop de métricas ---

        # --- Ahora hacemos regresión lineal ---
            import numpy as np

            fila_returns = df_linear_regression[df_linear_regression["Fiscal Period"] == "Year Returns %"]
            if not fila_returns.empty:
                # Tomamos la última fila (si hay varias) y desde la columna 2 (índice 2) para saltar la columna vacía
                fila_returns = fila_returns.iloc[-1, 2:]
            else:
                print("--- ERROR: No se encontró la fila 'Year Returns %' para regresión ---")
                fila_returns = []

            metricas = metricas_a_calcular  # Reutilizamos la lista

            filas_metricas = []
            for nombre in metricas:
                fila = df_linear_regression[df_linear_regression["Fiscal Period"] == f"{nombre} Year Returns %"]
                if not fila.empty:
                    # También tomamos desde la columna 2 para ignorar primera columna vacía
                    filas_metricas.append(fila.iloc[0, 2:])
                else:
                    print(f"--- WARN: No se encontró '{nombre} Year Returns %' para regresión ---")

            if fila_returns is not None and len(filas_metricas) == len(metricas):
                try:
                    # Convertir a numpy arrays float, asignando np.nan donde no hay dato válido
                    y = np.array([float(x) if x not in ("", None) else np.nan for x in fila_returns], dtype=np.float64)
                    X_raw = np.array([[float(x) if x not in ("", None) else np.nan for x in fila] for fila in filas_metricas], dtype=np.float64)

                    X = X_raw.T  # transponer para tener filas = observaciones

                    # Filtrar solo filas sin NaN en y ni en X para regresión válida
                    mask = ~np.isnan(y) & ~np.isnan(X).any(axis=1)
                    y_clean = y[mask]
                    X_clean = X[mask]

                    # Añadir columna intercepto
                    X_final = np.column_stack((np.ones(X_clean.shape[0]), X_clean))

                    coeficientes, residuals, rank, s = np.linalg.lstsq(X_final, y_clean, rcond=None)

                    etiquetas = ["Intercept"] + metricas
                    valores = [round(c, 6) for c in coeficientes]
                    df_resultados = pd.DataFrame({
                        "Fiscal Period": etiquetas,
                        "Regresión Lineal Coef.": valores
                    })

                    # Asegurar que tenga las mismas columnas que df_linear_regression
                    while df_resultados.shape[1] < df_linear_regression.shape[1]:
                        df_resultados[f"col_{df_resultados.shape[1]}"] = ""

                    df_resultados.columns = df_linear_regression.columns

                    # Insertar fila vacía y luego resultados al final
                    empty_row = pd.DataFrame([[""] * len(df_linear_regression.columns)], columns=df_linear_regression.columns)
                    df_linear_regression = pd.concat([df_linear_regression, empty_row], ignore_index=True)
                    df_linear_regression = pd.concat([df_linear_regression, df_resultados], ignore_index=True)

                    print("--- DEBUG: Coeficientes agregados al DataFrame ---")
                    print(df_resultados)

                except Exception as e:
                    print(f"--- ERROR en regresión lineal: {e}")
            else:
                print("--- ERROR: Datos insuficientes para regresión lineal ---")


            
    else:
        print("df_valuation no tiene al menos 25 filas.")
        df_linear_regression = pd.DataFrame()

    


    # --- Paso 1: Preparar y limpiar df_valuation ---

    df_val_indexed = df_valuation.set_index(df_valuation.columns[0])
    df_val_t = df_val_indexed.T
    df_val_t.index.name = "Fiscal Period"
    df_regresion_valuation = df_val_t.reset_index()
    if "Fiscal Period" in df_regresion_valuation.columns:
        df_regresion_valuation = df_regresion_valuation.drop(columns=["Fiscal Period"])

    def limpiar_valor_todo(valor):
        if isinstance(valor, str):
            val = valor.replace("x", "").replace(",", "").strip()
            if val in ("", "-", "∞"):
                return 0
            if "%" in val:
                try:
                    return float(val.replace("%", "")) / 100
                except ValueError:
                    return None
            try:
                return float(val)
            except ValueError:
                return val
        return valor

    df_regresion_valuation.columns = (
        df_regresion_valuation.columns
        .str.strip()
        .str.replace('\n', ' ')
        .str.replace('\r', '')
        .str.replace(r'\s+', ' ', regex=True)
    )

    df_regresion_valuation = df_regresion_valuation.applymap(limpiar_valor_todo)

    columnas_principal = [
        df_regresion_valuation.columns[0],  # Fiscal Period
        "P/E ratio", "PBR", "PEG",
        "Capitalization / Revenue", "EV / Revenue", "EV / EBITDA", "EV / EBIT", "EV / FCF",
        "Reference price 2"
    ]
    col_fiscal = columnas_principal[0]

    tabla_prev_regresion = df_regresion_valuation[columnas_principal].copy()
    tabla_prev_regresion[col_fiscal] = pd.to_numeric(tabla_prev_regresion[col_fiscal], errors="coerce")

    metricas = columnas_principal[1:-1]
    y_col = columnas_principal[-1]

    df_filtrado = tabla_prev_regresion[(tabla_prev_regresion[col_fiscal] <= 2024)]
    df_valid = df_filtrado[metricas + [y_col]].dropna()

    X_raw = df_valid[metricas].astype(float).to_numpy()
    y = df_valid[y_col].astype(float).to_numpy()
    X_final = np.column_stack((np.ones(X_raw.shape[0]), X_raw))

    coeficientes, residuals, rank, s = np.linalg.lstsq(X_final, y, rcond=None)
    etiquetas = ["Intercept"] + metricas
    valores = [round(c, 6) for c in coeficientes]
    coef_dict = dict(zip(etiquetas, valores))

    # 2. Cálculo de errores por coeficiente
    try:
        n = X_final.shape[0]
        p = X_final.shape[1] - 1
        y_pred = X_final @ coeficientes
        residuos = y - y_pred
        sse = np.sum(residuos ** 2)
        mse = sse / (n - p - 1)
        var_b = mse * np.linalg.inv(X_final.T @ X_final).diagonal()
        std_err = np.sqrt(var_b)
        t_stats = coeficientes / std_err
        p_values = [2 * (1 - stats.t.cdf(np.abs(t), df=n - p - 1)) for t in t_stats]

        # Redondear para visualización
        std_err = [round(e, 6) for e in std_err]
        t_stats = [round(t, 6) for t in t_stats]
        p_values = [round(pv, 6) for pv in p_values]

    except Exception as e:
        std_err = ["N/A"] * len(coeficientes)
        t_stats = ["N/A"] * len(coeficientes)
        p_values = ["N/A"] * len(coeficientes)
        print(f"Advertencia: No se pudieron calcular errores de regresión: {e}")
        

    # 3. Construir DataFrames de resultados
    tabla_resultados = pd.DataFrame([valores], columns=etiquetas)
    fila_std = pd.DataFrame([std_err], columns=etiquetas)
    fila_tstat = pd.DataFrame([t_stats], columns=etiquetas)
    fila_pval = pd.DataFrame([p_values], columns=etiquetas)

    # Agregar nombre de fila como primera columna (opcional, para mayor claridad en Excel)
    tabla_resultados.insert(0, "Estadístico", ["Coeficiente"])
    fila_std.insert(0, "Estadístico", ["Std Error"])
    fila_tstat.insert(0, "Estadístico", ["t-stat"])
    fila_pval.insert(0, "Estadístico", ["p-value"])

    # Encabezado para la tabla de regresión (con la columna extra)
    encabezado_regresion = pd.DataFrame([[""] + etiquetas], columns=["Estadístico"] + etiquetas)
    fila_vacia_regresion = pd.DataFrame([[""] * (len(etiquetas) + 1)], columns=["Estadístico"] + etiquetas)

    # 4. Predicciones verticales (igual que antes)
    tabla_pred = tabla_prev_regresion[tabla_prev_regresion[col_fiscal] <= 2025].copy()
    targets = []
    for _, fila in tabla_pred.iterrows():
        anio = fila[col_fiscal]
        valores_metricas = fila[metricas].astype(float).to_numpy()
        predicho = coef_dict["Intercept"] + np.dot([coef_dict[m] for m in metricas], valores_metricas)
        actual = fila[y_col]
        diferencia = round((predicho - actual) / actual, 2) * 100
        targets.append({
            "Price Result": "Target",
            "Price": round(predicho, 2),
            "Difference": diferencia,
            "Year": anio
        })
        targets.append({
            "Price Result": "Current",
            "Price": actual,
            "Difference": "",
            "Year": anio
        })
    df_targets = pd.DataFrame(targets)
    df_pred_vertical = df_targets.pivot(index="Year", columns="Price Result", values="Price")
    df_pred_vertical["Difference"] = df_targets[df_targets["Price Result"] == "Target"].set_index("Year")["Difference"]
    df_pred_vertical = df_pred_vertical[["Current", "Target", "Difference"]]
    df_pred_vertical = df_pred_vertical.reset_index()

    target_price_2025 = None
    upside_pct_2025 = None

    try:
        fila_2025 = df_pred_vertical[df_pred_vertical["Year"] == 2025].iloc[0]

        current_2025 = fila_2025["Current"]
        target_price_2025 = fila_2025["Target"]
        upside_pct_2025 = round((target_price_2025 - current_2025) / current_2025 * 100, 2)

    except Exception as e:
        print(f"Advertencia: No se pudo calcular el target o upside para 2025: {e}")
        target_price_2025 = None
        upside_pct_2025 = None

    # Encabezados y filas vacías para el resto de la tabla
    def fila_encabezado(cols):
        return pd.DataFrame([cols], columns=cols)

    encabezado_principal = fila_encabezado(columnas_principal)
    fila_vacia_principal = pd.DataFrame([[""] * len(columnas_principal)], columns=columnas_principal)
    encabezado_predicciones_vertical = pd.DataFrame([df_pred_vertical.columns.tolist()], columns=df_pred_vertical.columns)
    fila_vacia_predicciones_vertical = pd.DataFrame([[""] * len(df_pred_vertical.columns)], columns=df_pred_vertical.columns)

    # 5. Concatenar todo en el orden deseado
    df_final = pd.concat([
        encabezado_principal,
        tabla_prev_regresion,
        fila_vacia_principal,
        encabezado_regresion,
        tabla_resultados,
        fila_std,
        fila_tstat,
        fila_pval,
        fila_vacia_regresion,
        encabezado_predicciones_vertical,
        df_pred_vertical,
        fila_vacia_predicciones_vertical
    ], ignore_index=True).fillna("")


    try:
        with pd.ExcelWriter(excel_path, engine='xlsxwriter') as writer:
            df_data.to_excel(writer, sheet_name='Data', index=False, header=False)
            df_cash_flow.to_excel(writer, sheet_name='Cash Flow Statement', index=False, header=False)
            df_balance_sheet.to_excel(writer, sheet_name='Balance Sheet', index=False, header=False)
            df_financial.to_excel(writer, sheet_name='Financial', index=False, header=False)
            df_ratios.to_excel(writer, sheet_name='Ratios', index=False, header=False)
            df_valuation.to_excel(writer, sheet_name='Valuation', index=False, header=False)
            df_linear_regression.to_excel(writer, sheet_name='Linear Regression', index=False)
            df_final.to_excel(writer, sheet_name='Regresion valuation', index=False, header=False)
            if not df_yahoo_info.empty:
                df_yahoo_info.T.reset_index().to_excel(writer, sheet_name='Complementary', index=False, header=["Campo", "Valor"])
            if not df_history.empty and len(df_history.index) > 0:
                # Eliminar la zona horaria del índice antes de guardar
                df_history.index = df_history.index.tz_localize(None)
                # Filtrar solo el último día de cada mes
                df_monthly = df_history.copy()
                df_monthly['Year'] = df_monthly.index.year
                df_monthly['Month'] = df_monthly.index.month
                df_monthly = df_monthly.groupby(['Year', 'Month']).tail(1)
                df_monthly = df_monthly.drop(['Year', 'Month'], axis=1)
                df_hist = df_monthly.reset_index()
                df_hist = df_hist.dropna(subset=['Close'])
                df_hist['Days'] = (df_hist['Date'] - df_hist['Date'].min()).dt.days
                X = df_hist['Days'].values.reshape(-1, 1)  # Usar array sin nombres para evitar warning
                # Regresión para cada columna y métricas
                preds = {}
                metrics = []
                future_days = np.arange(df_hist['Days'].max() + 30, df_hist['Days'].max() + 122*30, 30).reshape(-1, 1)
                future_dates = [df_hist['Date'].min() + pd.Timedelta(days=int(day)) for day in future_days.flatten()]
                for col in ['Open', 'High', 'Low', 'Close']:
                    y = df_hist[col].values  # Usar array sin nombres
                    model = LinearRegression()
                    model.fit(X, y)
                    preds[col] = model.predict(future_days)
                    y_pred_hist = model.predict(X)
                    mae = mean_absolute_error(y, y_pred_hist)
                    mse = mean_squared_error(y, y_pred_hist)
                    rmse = np.sqrt(mse)
                    accuracy = (1 - (mae / np.mean(np.abs(y)))) * 100 if np.mean(np.abs(y)) != 0 else 0
                    metrics.append({'Variable': col, 'MAE': mae, 'MSE': mse, 'RMSE': rmse, 'Accuracy (%)': accuracy})
                # Crear DataFrame de predicción
                df_pred = pd.DataFrame({
                    'Date': future_dates,
                    'Open': preds['Open'],
                    'High': preds['High'],
                    'Low': preds['Low'],
                    'Close': preds['Close'],
                    'Volume': [None]*len(future_dates),
                    'Dividends': [None]*len(future_dates),
                    'Stock Splits': [None]*len(future_dates)
                })
                # Asegurar que la columna Date esté presente y bien formateada
                df_monthly = df_monthly.copy()
                df_monthly['Date'] = df_monthly.index
                df_monthly = df_monthly.reset_index(drop=True)
                cols = ['Date', 'Open', 'High', 'Low', 'Close', 'Volume', 'Dividends', 'Stock Splits']
                df_monthly = df_monthly[cols]
                df_pred = df_pred[cols]
                # Concatenar los datos históricos y las predicciones, separadas por una fila vacía y encabezado
                df_concat = pd.concat([
                    df_monthly,
                    pd.DataFrame([['']*len(df_monthly.columns)], columns=df_monthly.columns),
                    pd.DataFrame([['Predicción Lineal'] + ['']*(len(df_monthly.columns)-1)], columns=df_monthly.columns),
                    df_pred
                ], ignore_index=True)
                # Crear DataFrame de métricas en formato tabla
                df_metrics = pd.DataFrame(metrics)
                df_metrics = df_metrics.set_index('Variable').T.reset_index()
                # Agregar una fila vacía y luego las métricas al final de la hoja
                df_concat = pd.concat([
                    df_concat,
                    pd.DataFrame([['']*len(df_concat.columns)], columns=df_concat.columns),
                    df_metrics
                ], ignore_index=True)
                # Guardar la hoja principal con métricas al final
                df_concat.to_excel(writer, sheet_name='Yahoo History', index=False, header=True)
                # Graficar
                plt.figure(figsize=(12,6))
                plt.plot(df_hist['Date'], df_hist['Close'], label='Histórico Close')
                plt.plot(future_dates, preds['Close'], label='Predicción Close', linestyle='--')
                plt.plot(df_hist['Date'], df_hist['Open'], label='Histórico Open')
                plt.plot(future_dates, preds['Open'], label='Predicción Open', linestyle='--')
                plt.plot(df_hist['Date'], df_hist['High'], label='Histórico High')
                plt.plot(future_dates, preds['High'], label='Predicción High', linestyle='--')
                plt.plot(df_hist['Date'], df_hist['Low'], label='Histórico Low')
                plt.plot(future_dates, preds['Low'], label='Predicción Low', linestyle='--')
                plt.xlabel('Fecha')
                plt.ylabel('Precio')
                plt.title('Regresión Lineal - Predicción Mensual (Open, High, Low, Close)')
                plt.legend()
                plt.tight_layout()
                plt.savefig(excel_path.replace('.xlsx', '_prediccion.png'))
                plt.close()
            else:
                print(f"--- AVISO: No hay datos históricos en Yahoo Finance para {yahoo_ticker} ---")
            print(f"--- DEBUG: Llamada a to_excel para la hoja 'Complementary' ejecutada para {stock_code}. ---")
        print(f"--- DEBUG: Archivo guardado correctamente en: {excel_path} ---")
    except Exception as e:
        print(f"--- ERROR al guardar el archivo Excel: {e} ---")
    print(f"--- DEBUG: Proceso de escritura en Excel finalizado para {stock_code}. ---")

    # --- REGRESIÓN LINEAL SIMPLE SOBRE CASH FLOW CON LIMPIEZA Y DEPURACIÓN ---
    def convertir_valor(val):
        if val is None or val == '' or val == '-':
            return np.nan
        val = str(val).replace(',', '').strip()
        try:
            if val.endswith('B'):
                return float(val[:-1]) * 1_000_000_000
            elif val.endswith('M'):
                return float(val[:-1]) * 1_000_000
            elif val.endswith('K'):
                return float(val[:-1]) * 1_000
            else:
                return float(val)
        except:
            return np.nan

    df_cf_reg = pd.DataFrame(cash_flow_data)
    if not df_cf_reg.empty:
        df_cf_reg.columns = df_cf_reg.iloc[0]
        df_cf_reg = df_cf_reg[1:]
        # Limpiar nombres de columnas (quitar espacios y convertir a string)
        df_cf_reg.columns = [str(col).strip() for col in df_cf_reg.columns]
        print(f"Nombres de columnas después de limpiar: {df_cf_reg.columns.tolist()}")
        # Imprimir una fila de ejemplo
        print(f"Fila de ejemplo: {df_cf_reg.iloc[0].to_dict()}")
        conceptos = df_cf_reg.iloc[:, 0].values
        anios = [col for col in df_cf_reg.columns[1:] if str(col).isdigit()]
        print(f"Años detectados: {anios}")
        print(f"Conceptos detectados: {len(conceptos)}")
        resultados = []
        for idx, row in df_cf_reg.iterrows():
            y = []
            x = []
            for idx_anio, anio in enumerate(anios):
                valor = row[anio] if anio in row else None
                v = convertir_valor(valor)
                if np.isfinite(v):
                    y.append(v)
                    x.append(idx_anio)  # Normalizar los años a 0, 1, 2...
            print(f"Concepto: {row.iloc[0]} | X (años): {x} | Y (valores): {y}")
            if len(y) > 1:
                X = np.array(x).reshape(-1, 1)
                Y = np.array(y)
                model = LinearRegression()
                model.fit(X, Y)
                y_pred = model.predict(X)
                score = model.score(X, Y)
                n = len(Y)
                p = 1  # Solo una variable X
                residuals = Y - y_pred
                sse = np.sum(residuals ** 2)
                sst = np.sum((Y - np.mean(Y)) ** 2)
                mse = sse / (n - p - 1)
                se_beta = np.sqrt(mse / np.sum((X.flatten() - np.mean(X.flatten())) ** 2))
                t_stat = model.coef_[0] / se_beta
                p_value = 2 * (1 - stats.t.cdf(np.abs(t_stat), df=n - p - 1))
                # Intervalo de confianza 95%
                t_crit = stats.t.ppf(0.975, df=n - p - 1)
                ci_lower = model.coef_[0] - t_crit * se_beta
                ci_upper = model.coef_[0] + t_crit * se_beta
                std_error_reg = np.sqrt(mse)
                # ANOVA
                msr = (sst - sse) / p
                f_stat = msr / mse if mse != 0 else float('inf')
                significance_f = 1 - stats.f.cdf(f_stat, p, n - p - 1)
                resultados.append({
                    'Concepto': row.iloc[0],
                    'Coeficiente': model.coef_[0],
                    'Error estándar': se_beta,
                    't-stat': t_stat,
                    'p-value': p_value,
                    'IC 95% inf': ci_lower,
                    'IC 95% sup': ci_upper,
                    'Intercepto': model.intercept_,
                    'R2': score,
                    'Std Error Reg': std_error_reg,
                    'Observaciones': n,
                    'F': f_stat,
                    'Significance F': significance_f,
                    'Modelo': f"Y = {model.coef_[0]:.4f} * X + {model.intercept_:.4f}"
                })
        if resultados:
            df_resultados = pd.DataFrame(resultados)
            df_resultados.to_excel(writer, sheet_name='Cash Flow Regresion', index=False)

            # Agregar tabla ANOVA al Excel después de la tabla de regresión
            # Usar los valores de la última regresión (puedes ajustar para varias si lo deseas)
            anova_data = []
            concepto = resultados[-1]['Concepto']
            ssr = sst - sse  # Suma de cuadrados de la regresión
            dfr = p
            msr = (sst - sse) / p
            dfe = n - p - 1
            mse = sse / dfe
            f_stat = msr / mse if mse != 0 else float('inf')
            significance_f = 1 - stats.f.cdf(f_stat, p, n - p - 1)
            anova_data.append(['Regresión', ssr, dfr, msr, f_stat, significance_f])
            anova_data.append(['Residual', sse, dfe, mse, '', ''])
            anova_data.append(['Total', sst, n - 1, '', '', ''])
            df_anova = pd.DataFrame(anova_data, columns=['Fuente', 'Suma de Cuadrados', 'Grados de Libertad', 'Media Cuadrática', 'F', 'Significance F'])
            # Escribir la tabla ANOVA debajo de la tabla de regresión
            startrow = len(df_resultados) + 3
            df_anova.to_excel(writer, sheet_name='Cash Flow Regresion', index=False, startrow=startrow)
        else:
            pd.DataFrame({'Mensaje': ['No se encontraron datos válidos para la regresión.']}).to_excel(writer, sheet_name='Cash Flow Regresion', index=False)
    # --- FIN REGRESIÓN ---

    def extract_values(df, search_str):
        for idx, row in df.iterrows():
            if any(search_str in str(cell) for cell in row):
                return row[1:].values
        print(f"No se encontró '{search_str}' en la tabla.")
        return []

    fiscal_period = extract_values(df_cash_flow, 'Fiscal Period')

    # Extraemos los demás valores como antes
    unlevered_free_cash_flow = extract_values(df_cash_flow, 'Unlevered Free Cash Flow')
    total_debt = extract_values(df_balance_sheet, 'Net Debt')
    diluted_shares = extract_values(df_financial, 'Diluted Weighted Average Shares Outstanding')
    cash_equivalents = extract_values(df_balance_sheet, 'Cash and Equivalents')

    # Rellenamos con NaN para que todas las listas tengan la misma longitud
    max_length = max(len(fiscal_period), len(unlevered_free_cash_flow), len(total_debt), len(diluted_shares), len(cash_equivalents))
    fiscal_period = np.pad(fiscal_period, (0, max_length - len(fiscal_period)), constant_values=np.nan)
    unlevered_free_cash_flow = np.pad(unlevered_free_cash_flow, (0, max_length - len(unlevered_free_cash_flow)), constant_values=np.nan)
    total_debt = np.pad(total_debt, (0, max_length - len(total_debt)), constant_values=np.nan)
    diluted_shares = np.pad(diluted_shares, (0, max_length - len(diluted_shares)), constant_values=np.nan)
    cash_equivalents = np.pad(cash_equivalents, (0, max_length - len(cash_equivalents)), constant_values=np.nan)

    # Creamos el DataFrame incluyendo el Fiscal Period
    df_dcf = pd.DataFrame({
    'Fiscal Period': fiscal_period,
    'Unlevered Free Cash Flow': unlevered_free_cash_flow,
    'Total Debt': total_debt,
    'Diluted Shares Outstanding': diluted_shares,
    })

    # Transponer el DataFrame
    df_dcf = df_dcf.set_index('Fiscal Period').T

    # Obtener el último año fiscal
    last_year = int(fiscal_period[-1])
    new_years = [last_year + i for i in range(1, forecast_years + 1)]

    # Agregar los nuevos años como columnas
    df_dcf = pd.concat([df_dcf, pd.DataFrame(columns=new_years)], axis=1)
    

    # Escribir en el archivo Excel
    with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a') as writer:
        df_dcf.to_excel(writer, sheet_name='DCF', startrow=1, index=True, header=True)

    # Cargar el libro de trabajo y actualizar una celda
    wb = load_workbook(excel_path)
    ws = wb['DCF']
    ws['X2'] = x2

    # Calcular el índice de la última columna de flujo de caja libre no apalancado
    unlevered_last_col = len(df_dcf.columns) - forecast_years
    for i in range(forecast_years): 
        current_col = unlevered_last_col + 2 + i
        previous_col = unlevered_last_col + 1 + i

        if i == 0:
            formula = (
                f'=IF({ws.cell(row=3, column=unlevered_last_col + 1).coordinate}*'
                f'FORECAST({ws.cell(row=2, column=current_col).coordinate},'
                f'{ws.cell(row=3, column=2).coordinate}:{ws.cell(row=3, column=unlevered_last_col + 1).coordinate},'
                f'{ws.cell(row=2, column=2).coordinate}:{ws.cell(row=2, column=unlevered_last_col + 1).coordinate})>0,'
                f'FORECAST({ws.cell(row=2, column=current_col).coordinate},'
                f'{ws.cell(row=3, column=2).coordinate}:{ws.cell(row=3, column=unlevered_last_col + 1).coordinate},'
                f'{ws.cell(row=2, column=2).coordinate}:{ws.cell(row=2, column=unlevered_last_col + 1).coordinate}),'
                f'{ws.cell(row=3, column=unlevered_last_col + 1).coordinate})*$X$2'
            )
        else:
            formula = (
                f'=IF({ws.cell(row=3, column=unlevered_last_col + 1 + i).coordinate}*'
                f'FORECAST({ws.cell(row=2, column=current_col).coordinate},'
                f'{ws.cell(row=3, column=2).coordinate}:{ws.cell(row=3, column=current_col - 1).coordinate},'
                f'{ws.cell(row=2, column=2).coordinate}:{ws.cell(row=2, column=current_col - 1).coordinate})>0,'
                f'FORECAST({ws.cell(row=2, column=current_col).coordinate},'
                f'{ws.cell(row=3, column=2).coordinate}:{ws.cell(row=3, column=current_col - 1).coordinate},'
                f'{ws.cell(row=2, column=2).coordinate}:{ws.cell(row=2, column=current_col - 1).coordinate}),'
                f'{ws.cell(row=3, column=unlevered_last_col + 1 + i).coordinate})*$X$2'
            )
        ws.cell(row=3, column=current_col, value=formula)

    valor = ws.cell(row=3, column=current_col).value
    ws['K7'] = "TERMINAL VALUE"
    ws['L7'] = f'{valor} * (1 + 0.02) / (0.06 - 0.02)'
    terminal_value = ws['L7'].value

    start_col = unlevered_last_col + 3
    end_col = current_col
    terminal_col = end_col + 1

    def col_num_to_letter(col_num):
        col_letter = ""
        while col_num > 0:
            col_num, remainder = divmod(col_num - 1, 26)
            col_letter = chr(65 + remainder) + col_letter
        return col_letter

    start_col_letter = col_num_to_letter(start_col)
    end_col_letter = col_num_to_letter(end_col)
    terminal_col_letter = col_num_to_letter(terminal_col)
    next_start_col_letter = col_num_to_letter(start_col + 1)

    ws['K8'] = 'NPV of FCF'
    ws['L8'] = f'=NPV(0.06, {start_col_letter}3:{end_col_letter}3)'

    for col in range(start_col + 1, end_col + 1):
        col_letter = col_num_to_letter(col)
        ws[f'{col_letter}16'] = 0

    ws[f'{terminal_col_letter}16'] = terminal_value

    ws['K9'] = "NPV of TV"
    ws['L9'] = f'=NPV(0.06, {next_start_col_letter}16:{terminal_col_letter}16)'

    ws['K10'] = 'Total EV'
    ws['L10'] = '=L8 + L9'

    ws['K11'] = 'Net Debt'
    for col in reversed(range(2, ws.max_column + 1)):
        if isinstance(ws.cell(row=4, column=col).value, (int, float)):
            last_column_net_debt = col
            break
    ws['L11'] = ws.cell(row=4, column=last_column_net_debt).value

    ws['K12'] = 'Equity'
    ws['L12'] = f'=L10 - L11'

    ws['K13'] = 'Shares Outstanding'
    for col in reversed(range(2, ws.max_column + 1)):
        if isinstance(ws.cell(row=5, column=col).value, (int, float)):
            last_column_shares_outstanding = col
            break
    ws['L13'] = ws.cell(row=5, column=last_column_shares_outstanding).value

    ws['K14'] = 'Target Price'
    ws['L14'] = f'=L12/L13'

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if isinstance(cell.value, str) and ("#DIV/0!" in cell.value or "Error" in cell.value):
                cell.value = "Error"

    # Crear la nueva hoja 'new DCF' copiando el contenido de la hoja DCF
    ws_new_dcf = wb.copy_worksheet(ws)
    ws_new_dcf.title = 'new DCF'
    
    # Agregar cálculos estadísticos en la nueva hoja
    # Average = AVERAGE(B3:K3) - Promedio de datos históricos
    ws_new_dcf['B10'] = '=AVERAGE(B3:K3)'
    
    # Std+ = B10 + STDEV(B3:K3) - Average + Desviación estándar
    ws_new_dcf['B11'] = '=B10+STDEV(B3:K3)'
    
    # Std- = B10 - STDEV(B3:K3) - Average - Desviación estándar  
    ws_new_dcf['B12'] = '=B10-STDEV(B3:K3)'
    
    # Agregar etiquetas para claridad
    ws_new_dcf['A10'] = 'Average'
    ws_new_dcf['A11'] = 'Std+'
    ws_new_dcf['A12'] = 'Std-'
    
    # Crear 3 casos completos con todos los datos copiados
    # Caso 1: Average - Copiar toda la estructura y reemplazar solo los pronósticos (con 3 espacios antes)
    # Copiar años (fila 2) incluyendo 2035
    for col in range(2, 23):  # Columnas B a V (hasta 2035)
        col_letter = chr(64 + col)  # B, C, D, etc.
        ws_new_dcf[f'{col_letter}18'] = ws_new_dcf[f'{col_letter}2'].value
    # Agregar 2036 (columna W) para Average
    ws_new_dcf['W18'] = 2036
    
    # Copiar Unlevered Free Cash Flow (fila 3) pero reemplazar pronósticos con Average
    for col in range(2, 12):  # Columnas B a K (datos históricos)
        col_letter = chr(64 + col)
        ws_new_dcf[f'{col_letter}19'] = ws_new_dcf[f'{col_letter}3'].value
    for col in range(12, 23):  # Columnas L a W (pronósticos con Average incluyendo 2035)
        col_letter = chr(64 + col)
        ws_new_dcf[f'{col_letter}19'] = '=$B$10'
    # Agregar dato de 2036 para Average
    ws_new_dcf['W19'] = '=$B$10'
    
    # Copiar Total Debt (fila 4)
    for col in range(2, 12):  # Columnas B a K
        col_letter = chr(64 + col)
        ws_new_dcf[f'{col_letter}20'] = ws_new_dcf[f'{col_letter}4'].value
    
    # Copiar Diluted Shares Outstanding (fila 5)
    for col in range(2, 12):  # Columnas B a K
        col_letter = chr(64 + col)
        ws_new_dcf[f'{col_letter}21'] = ws_new_dcf[f'{col_letter}5'].value
    
    # Agregar etiquetas para el caso Average
    ws_new_dcf['A18'] = 'Average'
    
    # Caso 2: Std+ - Crear sección completa para Std+ (con más espacio)
    # Copiar años (fila 2) incluyendo 2035
    for col in range(2, 23):  # Columnas B a V (hasta 2035)
        col_letter = chr(64 + col)
        ws_new_dcf[f'{col_letter}35'] = ws_new_dcf[f'{col_letter}2'].value
    # Agregar 2036 (columna W) para Std+
    ws_new_dcf['W35'] = 2036
    
    # Copiar Unlevered Free Cash Flow pero reemplazar pronósticos con Std+
    for col in range(2, 12):  # Columnas B a K (datos históricos)
        col_letter = chr(64 + col)
        ws_new_dcf[f'{col_letter}36'] = ws_new_dcf[f'{col_letter}3'].value
    for col in range(12, 23):  # Columnas L a W (pronósticos con Std+ incluyendo 2035)
        col_letter = chr(64 + col)
        ws_new_dcf[f'{col_letter}36'] = '=$B$11'
    # Agregar dato de 2036 para Std+
    ws_new_dcf['W36'] = '=$B$11'
    
    # Copiar Total Debt
    for col in range(2, 12):  # Columnas B a K
        col_letter = chr(64 + col)
        ws_new_dcf[f'{col_letter}37'] = ws_new_dcf[f'{col_letter}4'].value
    
    # Copiar Diluted Shares Outstanding
    for col in range(2, 12):  # Columnas B a K
        col_letter = chr(64 + col)
        ws_new_dcf[f'{col_letter}38'] = ws_new_dcf[f'{col_letter}5'].value
    
    # Agregar etiquetas para el caso Std+
    ws_new_dcf['A35'] = 'Std+'
    
    # Caso 3: Std- - Crear sección completa para Std- (con más espacio)
    # Copiar años (fila 2) incluyendo 2035
    for col in range(2, 23):  # Columnas B a V (hasta 2035)
        col_letter = chr(64 + col)
        ws_new_dcf[f'{col_letter}52'] = ws_new_dcf[f'{col_letter}2'].value
    # Agregar 2036 (columna W) para Std-
    ws_new_dcf['W52'] = 2036
    
    # Copiar Unlevered Free Cash Flow pero reemplazar pronósticos con Std-
    for col in range(2, 12):  # Columnas B a K (datos históricos)
        col_letter = chr(64 + col)
        ws_new_dcf[f'{col_letter}53'] = ws_new_dcf[f'{col_letter}3'].value
    for col in range(12, 23):  # Columnas L a W (pronósticos con Std- incluyendo 2035)
        col_letter = chr(64 + col)
        ws_new_dcf[f'{col_letter}53'] = '=$B$12'
    # Agregar dato de 2036 para Std-
    ws_new_dcf['W53'] = '=$B$12'
    
    # Asegurar que X2 tenga el mismo valor que en la hoja original
    ws_new_dcf['X2'] = x2
    
    # Agregar valores en X3 y X4
    ws_new_dcf['X3'] = 0.65  # 65%
    ws_new_dcf['X4'] = 0.175  # 15%
    
    # Agregar fórmula en X5 que combine los Target Prices de los 3 casos
    ws_new_dcf['X5'] = '=L63*X4+L46*X4+L29*X3'  # Std-*X4 + Std+*X4 + Average*X3
    
    # Agregar etiqueta "PRICE" en W5
    ws_new_dcf['W5'] = 'PRICE'
    
    # Agregar ceros después de los datos para cada caso (tantos como años de pronóstico)
    # Caso 1 (Average) - 10 ceros en L..U y Terminal Value en V
    for col in range(12, 23):  # Columnas L a V (10 ceros + Terminal en V)
        col_letter = chr(64 + col)
        if col == 22:  # Columna V - Terminal Value
            ws_new_dcf[f'{col_letter}30'] = '=L22'
        else:  # Columnas L a U - 10 ceros
            ws_new_dcf[f'{col_letter}30'] = 0
    
    # Caso 2 (Std+) - 10 ceros en L..U y Terminal Value en V
    for col in range(12, 23):  # Columnas L a V (10 ceros + Terminal en V)
        col_letter = chr(64 + col)
        if col == 22:  # Columna V - Terminal Value
            ws_new_dcf[f'{col_letter}47'] = '=L39'
        else:  # Columnas L a U - 10 ceros
            ws_new_dcf[f'{col_letter}47'] = 0
    
    # Caso 3 (Std-) - 10 ceros en L..U y Terminal Value en V
    for col in range(12, 23):  # Columnas L a V (10 ceros + Terminal en V)
        col_letter = chr(64 + col)
        if col == 22:  # Columna V - Terminal Value
            ws_new_dcf[f'{col_letter}64'] = '=L56'
        else:  # Columnas L a U - 10 ceros
            ws_new_dcf[f'{col_letter}64'] = 0

    # Calcular Terminal Value y otros valores para cada caso
    # Caso 1 (Average) - Terminal Value y cálculos (Filas 22-29)
    ws_new_dcf['K22'] = 'TERMINAL VALUE'
    ws_new_dcf['L22'] = '=IF(V19*FORECAST(W18,B19:V19,B18:V18)>0,FORECAST(W18,B19:V19,B18:V18),V19)*$X$2 * (1 + 0.02) / (0.06 - 0.02)'
    ws_new_dcf['K23'] = 'NPV of FCF'
    ws_new_dcf['L23'] = '=NPV(0.06, M19:W19)'
    ws_new_dcf['K24'] = 'NPV of TV'
    ws_new_dcf['L24'] = '=NPV(0.06, L30:V30)'
    ws_new_dcf['K25'] = 'Total EV'
    ws_new_dcf['L25'] = '=L23 + L24'
    ws_new_dcf['K26'] = 'Net Debt'
    ws_new_dcf['L26'] = f'={ws_new_dcf["K20"].value}'  # Usar el valor real de Net Debt de la empresa
    ws_new_dcf['K27'] = 'Equity'
    ws_new_dcf['L27'] = '=L25 - L26'
    ws_new_dcf['K28'] = 'Shares Outstanding'
    ws_new_dcf['L28'] = f'={ws_new_dcf["K21"].value}'  # Usar el valor real de Shares de la empresa
    ws_new_dcf['K29'] = 'Target Price'
    ws_new_dcf['L29'] = '=L27/L28'
    
    # Caso 2 (Std+) - Terminal Value y cálculos (Filas 39-46)
    ws_new_dcf['K39'] = 'TERMINAL VALUE'
    ws_new_dcf['L39'] = '=IF(V36*FORECAST(W35,B36:V36,B35:V35)>0,FORECAST(W35,B36:V36,B35:V35),V36)*$X$2 * (1 + 0.02) / (0.06 - 0.02)'
    ws_new_dcf['K40'] = 'NPV of FCF'
    ws_new_dcf['L40'] = '=NPV(0.06, M36:W36)'
    ws_new_dcf['K41'] = 'NPV of TV'
    ws_new_dcf['L41'] = '=NPV(0.06, L47:V47)'
    ws_new_dcf['K42'] = 'Total EV'
    ws_new_dcf['L42'] = '=L40 + L41'
    ws_new_dcf['K43'] = 'Net Debt'
    ws_new_dcf['L43'] = f'={ws_new_dcf["K20"].value}'  # Usar el valor real de Net Debt de la empresa
    ws_new_dcf['K44'] = 'Equity'
    ws_new_dcf['L44'] = '=L42 - L43'
    ws_new_dcf['K45'] = 'Shares Outstanding'
    ws_new_dcf['L45'] = f'={ws_new_dcf["K21"].value}'  # Usar el valor real de Shares de la empresa
    ws_new_dcf['K46'] = 'Target Price'
    ws_new_dcf['L46'] = '=L44/L45'
    
    # Caso 3 (Std-) - Terminal Value y cálculos (Filas 56-63)
    ws_new_dcf['K56'] = 'TERMINAL VALUE'
    ws_new_dcf['L56'] = '=IF(V53*FORECAST(W52,B53:V53,B52:V52)>0,FORECAST(W52,B53:V53,B52:V52),V53)*$X$2 * (1 + 0.02) / (0.06 - 0.02)'
    ws_new_dcf['K57'] = 'NPV of FCF'
    ws_new_dcf['L57'] = '=NPV(0.06, M53:W53)'
    ws_new_dcf['K58'] = 'NPV of TV'
    ws_new_dcf['L58'] = '=NPV(0.06, L64:V64)'
    ws_new_dcf['K59'] = 'Total EV'
    ws_new_dcf['L59'] = '=L57 + L58'
    ws_new_dcf['K60'] = 'Net Debt'
    ws_new_dcf['L60'] = f'={ws_new_dcf["K20"].value}'  # Usar el valor real de Net Debt de la empresa
    ws_new_dcf['K61'] = 'Equity'
    ws_new_dcf['L61'] = '=L59 - L60'
    ws_new_dcf['K62'] = 'Shares Outstanding'
    ws_new_dcf['L62'] = f'={ws_new_dcf["K21"].value}'  # Usar el valor real de Shares de la empresa
    ws_new_dcf['K63'] = 'Target Price'
    ws_new_dcf['L63'] = '=L61/L62'
    
    wb.save(excel_path)


    # Leer el valor de L14 directamente con xlwings para obtener el resultado calculado
    print(f"[LOG] Leyendo valores calculados de Excel...")
    import xlwings as xw
    print(f"[LOG] Abriendo Excel con xlwings: {excel_path}")
    app = xw.App(visible=False)
    app.display_alerts = False
    app.screen_updating = False
    wb = app.books.open(excel_path)
    print(f"[LOG] Excel abierto, accediendo a hoja 'DCF'...")
    ws = wb.sheets['DCF']
    l14_value = ws.range('L14').value
    print(f"[LOG] Valor leído de L14 (DCF VALUE): {l14_value}")
    
    # Leer el valor de X5 de la hoja 'new DCF'
    print(f"[LOG] Accediendo a hoja 'new DCF'...")
    try:
        ws_new_dcf = wb.sheets['new DCF']
        x5_value = ws_new_dcf.range('X5').value
        print(f"[LOG] Valor leído de X5 (DCF NEW): {x5_value}")
    except Exception as e:
        print(f"[LOG] Error al leer X5 de new DCF: {e}")
        x5_value = None
    
    wb.close()
    app.quit()
    print(f"[LOG] Excel cerrado correctamente")

    # Verificar si l14_value es un error y asignar un valor manejable si lo es
    if isinstance(l14_value, str) and ("#DIV/0" in l14_value or "Error" in l14_value):
        l14_value = "Error en cálculo"  # Puedes asignar cualquier valor representativo
    elif pd.isna(l14_value):  # Si el valor es NaN (celda vacía o con error no detectado)
        l14_value = "Valor no disponible"

    wb = load_workbook(excel_path)

    # Cargar la hoja 'Complementary' para extraer el valor de 'Open' y 'Country'
    ws_complementary = wb['Valuation']
    # Función para buscar valores en la hoja 'Complementary'
    def extract_value_from_complementary(search_str):
        search_str = search_str.lower()

        for row in ws_complementary.iter_rows(min_row=1, max_row=200):
            first_cell = row[0]
            first_cell_value = str(first_cell.value).strip().lower() if first_cell.value else ''

            if search_str in first_cell_value:
                numeric_values = []
                for cell in row:
                    val = cell.value
                    if val is None:
                        continue
                    # Intentar convertir a float
                    try:
                        num = float(str(val).strip())
                        numeric_values.append(num)
                    except ValueError:
                        continue  # No es número, ignorar

                if numeric_values:
                    return numeric_values[-1]  # último número de la fila
        return None

    # Obtener el valor de 'Open' de la hoja 'Complementary'
    open_value = extract_value_from_complementary("Reference price")
    print(f"Open value: {open_value}")

    # Obtener el valor de 'Country' de la hoja 'Complementary

    market_cap = extract_value_from_complementary('Capitalization')

    pe_ratio = extract_value_from_complementary('P/E ratio')
    print(f"Open value: {pe_ratio}")

  

    # Imprimir para depuración
    print(f"Open value: {open_value}")

    print(f"Target Price (L14): {l14_value}")

    # Calcular el 'Upside' si 'Open' y 'Target Price' están disponibles
    if open_value and isinstance(l14_value, (int, float)):
        upside = ((l14_value - open_value) / open_value) * 100
        print(f"Upside calculado: {upside:.2f}%")
    else:
        upside = None
        print("No se pudo calcular el upside debido a datos faltantes.")

    # Verificar si 'Data' ya existe en el archivo, si no, crearla
    if 'Data' not in wb.sheetnames:
        ws_data = wb.create_sheet('Data')
    else:
        ws_data = wb['Data']

    # Escribir los valores en la hoja 'Data' desde la fila 4
    start_row = 4

    # Datos a guardar
    ws_data[f"A{start_row}"] = "Open"
    ws_data[f"B{start_row}"] = open_value if open_value else "No encontrado"

    ws_data[f"A{start_row + 1}"] = "Target Price"
    ws_data[f"B{start_row + 1}"] = l14_value if l14_value else "No encontrado"

    ws_data[f"A{start_row + 2}"] = "Upside"
    ws_data[f"B{start_row + 2}"] = f"{upside:.2f}%" if upside is not None else "No calculado"

    # Guardar el país justo debajo de Upside
    ws_data[f"A{start_row + 3}"] = "País"
    ws_data[f"B{start_row + 3}"] = pais if pais else "No encontrado"

    ws_data[f"A{start_row + 5}"] = "Market Cap"
    ws_data[f"B{start_row + 5}"] = market_cap if market_cap else "No encontrado"

    ws_data[f"A{start_row + 7}"] = "Pe ratio"
    ws_data[f"B{start_row + 7}"] = pe_ratio if pe_ratio else "No encontrado"

    # Guardar el archivo actualizado
    wb.save(excel_path)


    ws_balance_sheet = wb['Balance Sheet']

    
    def extract_values_from_balance_sheet(search_str):
        values = []
        for row in ws_balance_sheet.iter_rows(min_row=1, max_row=200, min_col=1):
            if search_str.lower() in str(row[0].value).lower():
                values.extend([cell.value for cell in row])
        
        return values


   
    def calculate_growth_percentage(current, previous):
        if current is not None and previous not in (None, 0):  
            print(f"Calculando porcentaje de crecimiento entre: Current = {current}, Previous = {previous}")
            growth = ((current - previous) / previous) * 100
            print(f"Porcentaje de Crecimiento = (({current} - {previous}) / {previous}) * 100 = {growth}%")
            return growth
        else:
            print("No se puede calcular el porcentaje de crecimiento (división por cero o valor nulo).")
            return None

    
    net_debt_values = extract_values_from_balance_sheet('Net Debt')
    total_assets_values = extract_values_from_balance_sheet('Total Assets')
    net_pp_values = extract_values_from_balance_sheet('Net Property Plant and Equipment')

    
    net_debt_current, net_debt_previous = net_debt_values[-2], net_debt_values[-1] if len(net_debt_values) >= 2 else (None, None)
    total_assets_current, total_assets_previous = total_assets_values[-2], total_assets_values[-1] if len(total_assets_values) >= 2 else (None, None)
    net_pp_current, net_pp_previous = net_pp_values[-2], net_pp_values[-1] if len(net_pp_values) >= 2 else (None, None)

    # Inicializar todas las variables de crecimiento como None
    net_debt_growth = None
    total_assets_growth = None
    net_pp_growth = None
    
    # Calcular los porcentajes de crecimiento
    print(f"[LOG] Calculando porcentajes de crecimiento...")
    print(f"[LOG] Net Debt - Actual: {net_debt_current}, Previous: {net_debt_previous}")
    net_debt_growth = calculate_growth_percentage(net_debt_current, net_debt_previous)
    print(f"[LOG] Net Debt Growth calculado: {net_debt_growth}")
    
    print(f"[LOG] Total Assets - Actual: {total_assets_current}, Previous: {total_assets_previous}")
    total_assets_growth = calculate_growth_percentage(total_assets_current, total_assets_previous)
    print(f"[LOG] Total Assets Growth calculado: {total_assets_growth}")
    
    print(f"[LOG] Net PP - Actual: {net_pp_current}, Previous: {net_pp_previous}")
    net_pp_growth = calculate_growth_percentage(net_pp_current, net_pp_previous)
    print(f"[LOG] Net PP Growth calculado: {net_pp_growth}")

   
    ws_data = wb['Data']

    # Escribir los resultados en la hoja 'Data' en las filas 12, 13 y 14
    ws_data[f"A12"] = "Net Debt Growth"
    ws_data[f"B12"] = f"{net_debt_growth:.2f}%" if net_debt_growth is not None else "No calculado"

    ws_data[f"A13"] = "Total Assets Growth"
    ws_data[f"B13"] = f"{total_assets_growth:.2f}%" if total_assets_growth is not None else "No calculado"

    ws_data[f"A14"] = "Net Property, Plant and Equipment Growth"
    ws_data[f"B14"] = f"{net_pp_growth:.2f}%" if net_pp_growth is not None else "No calculado"

    # Guardar el archivo actualizado
    wb.save(excel_path)
    print("Porcentajes de crecimiento calculados y guardados en las filas 12, 13 y 14 de la hoja 'Data'.")




    # Cargar el archivo
    wb = load_workbook(excel_path)

    # Función para buscar EBITDA en la fila 20 y calcular el crecimiento
    def calculate_ebitda_growth():
        ws_valuation = wb["Valuation"]
        
        # Verificar los valores en la fila 20
        ebitda_values = []
        
        # Imprimir todos los valores de la fila 20 para depuración
        print(f"Valores en la fila 20: {[cell.value for cell in ws_valuation[20]]}")
        
        # Asegurarnos de que estamos buscando desde la columna B (índice 1) en adelante
        for cell in ws_valuation[20][1:]:  # Fila 20, desde la columna B (índice 1)
            value = cell.value
            if isinstance(value, str):  # Si el valor es un texto, intentar convertirlo a número
                try:
                    value = float(value.replace(",", ""))  # Eliminamos comas si las hay
                except ValueError:
                    value = None  # Si no se puede convertir, lo descartamos
            if isinstance(value, (int, float)):  # Asegurarnos de que es un valor numérico
                ebitda_values.append(value)
        
        # Ignorar el primer valor si es el '1' extraño al principio
        if ebitda_values and ebitda_values[0] == 1:
            ebitda_values = ebitda_values[1:]

        # Comprobar si se han encontrado suficientes valores para calcular el crecimiento
        if len(ebitda_values) >= 2:
            print(f"Valores de EBITDA encontrados en la fila 20: {ebitda_values}")
            return ebitda_values
        else:
            print("No se han encontrado suficientes valores de EBITDA en la fila 20.")
            return []

    # Extraer los valores de EBITDA
    ebitda_values = calculate_ebitda_growth()

    # Inicializar ebitda_growth como None
    ebitda_growth = None
    
    # Comprobar si se han encontrado suficientes valores
    if ebitda_values:
        # Calcular el crecimiento porcentual
        first_ebitda = ebitda_values[0]  # Primer valor de EBITDA
        last_ebitda = ebitda_values[-1]  # Último valor de EBITDA
        ebitda_growth = ((last_ebitda - first_ebitda) / first_ebitda) * 100  # Cálculo del crecimiento porcentual

        # Cargar la hoja 'Data' y guardar los resultados en las celdas A14 y B14
        ws_data = wb["Data"]
        ws_data["A15"] = "EBITDA Growth"
        ws_data["B15"] = f"{ebitda_growth:.2f}%"  # Guardar el crecimiento en B14

        # Guardar el archivo actualizado
        wb.save(excel_path)
        print("Cálculo de EBITDA Growth guardado en la hoja 'Data'.")
        
    else:
        print("No se han encontrado suficientes valores de EBITDA para calcular el crecimiento.")
        # Aún así, guardar "No calculado" en la hoja Data
        ws_data = wb["Data"]
        ws_data["A15"] = "EBITDA Growth"
        ws_data["B15"] = "No calculado"
        wb.save(excel_path)




    

    print(f"[LOG] Iniciando proceso de escritura a Google Sheets para {stock_code}")
    
    shared_link = upload_and_get_shared_link(drive, excel_path)
    print(f"[LOG] Enlace compartido generado: {shared_link}")

    print(f"[LOG] Configurando credenciales de Google Sheets...")
    scope = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
    creds = ServiceAccountCredentials.from_json_keyfile_name('C:/Users/relim/Desktop/prueba/durable-binder-432321-q4-5f1ef9e64ea1.json', scope)
    print(f"[LOG] Credenciales cargadas correctamente")
    
    print(f"[LOG] Autenticando con gspread...")
    client = gspread.authorize(creds)
    print(f"[LOG] Autenticación exitosa")
    
    print(f"[LOG] Abriendo spreadsheet con ID: 1D8p0lKGDf8h2fp6f3zdXb5M09mJtEExIfBKKINJurJk")
    spreadsheet = client.open_by_key('1D8p0lKGDf8h2fp6f3zdXb5M09mJtEExIfBKKINJurJk')
    print(f"[LOG] Spreadsheet abierto correctamente")
    
    print(f"[LOG] Accediendo a la hoja 'Results'...")
    sheet = spreadsheet.worksheet('Results')
    print(f"[LOG] Hoja 'Results' accedida correctamente")

    print(f"[LOG] Buscando fila para stock_code: {stock_code}")
    cell_list = sheet.col_values(1)  # Obtener todos los valores de la columna A
    print(f"[LOG] Total de filas encontradas en columna A: {len(cell_list)}")
    
    if stock_code in cell_list:
        row_index = cell_list.index(stock_code) + 1  # Encuentra la fila existente
        print(f"[LOG] Stock code encontrado en fila existente: {row_index}")
    else:
        # Si el código no está en la hoja, añade al final
        row_index = len(cell_list) + 1
        print(f"[LOG] Stock code no encontrado, se agregará en nueva fila: {row_index}")

    print(f"[LOG] Preparando datos para escritura en fila {row_index}...")
    
    # Preparar todos los valores para batch update
    link = f'https://www.marketscreener.com/quote/stock/{stock_code}'
    formula_link = f'=HYPERLINK("{link}", "{stock_code}")'
    
    # Preparar valores con conversiones
    print(f"[LOG] Preparando columna 3 (DCF NEW) con valor: {x5_value}")
    try:
        x5_numeric = float(x5_value) if x5_value is not None else None
        col3_value = round(x5_numeric, 2) if x5_numeric is not None else ""
    except (ValueError, TypeError):
        col3_value = x5_value if x5_value is not None else ""
    print(f"[LOG] Columna 3 preparada: {col3_value}")
    
    print(f"[LOG] Preparando columna 4 (DCF VALUE) con valor: {l14_value}")
    try:
        l14_numeric = float(l14_value) if l14_value is not None else None
        col4_value = round(l14_numeric, 2) if l14_numeric is not None else ""
    except (ValueError, TypeError):
        col4_value = l14_value if l14_value is not None else ""
    print(f"[LOG] Columna 4 preparada: {col4_value}")
    
    print(f"[LOG] Preparando columna 6 (UPSIDE REG) con valor: {upside_pct_2025}")
    try:
        upside_numeric = float(upside_pct_2025) if upside_pct_2025 is not None else None
        col6_value = upside_numeric / 100 if upside_numeric is not None else ""
    except (ValueError, TypeError):
        col6_value = upside_pct_2025 if upside_pct_2025 is not None else ""
    print(f"[LOG] Columna 6 preparada: {col6_value}")
    
    print(f"[LOG] Preparando columna 14 (NET DEBT GROWTH) con valor: {net_debt_growth}")
    try:
        net_debt_numeric = float(net_debt_growth) if net_debt_growth is not None else None
        col14_value = net_debt_numeric / 100 if net_debt_numeric is not None else ""
    except (ValueError, TypeError):
        col14_value = net_debt_growth if net_debt_growth is not None else ""
    print(f"[LOG] Columna 14 preparada: {col14_value}")
    
    print(f"[LOG] Preparando columna 15 (ASSETS GROWTH) con valor: {total_assets_growth}")
    try:
        total_assets_numeric = float(total_assets_growth) if total_assets_growth is not None else None
        col15_value = total_assets_numeric / 100 if total_assets_numeric is not None else ""
    except (ValueError, TypeError):
        col15_value = total_assets_growth if total_assets_growth is not None else ""
    print(f"[LOG] Columna 15 preparada: {col15_value}")
    
    print(f"[LOG] Preparando columna 16 (PP GROWTH) con valor: {net_pp_growth}")
    try:
        net_pp_numeric = float(net_pp_growth) if net_pp_growth is not None else None
        col16_value = net_pp_numeric / 100 if net_pp_numeric is not None else ""
    except (ValueError, TypeError):
        col16_value = net_pp_growth if net_pp_growth is not None else ""
    print(f"[LOG] Columna 16 preparada: {col16_value}")
    
    print(f"[LOG] Preparando columna 17 (EBITDA GROWTH) con valor: {ebitda_growth}")
    try:
        ebitda_numeric = float(ebitda_growth) if ebitda_growth is not None else None
        col17_value = ebitda_numeric / 100 if ebitda_numeric is not None else ""
    except (ValueError, TypeError):
        col17_value = ebitda_growth if ebitda_growth is not None else ""
    print(f"[LOG] Columna 17 preparada: {col17_value}")
    
    # Preparar datos en formato de lista para batch update
    values = [
        formula_link,  # Columna 1: Market (link)
        stock_name_clean,  # Columna 2: Name
        col3_value,  # Columna 3: DCF NEW (x5_value)
        col4_value,  # Columna 4: DCF VALUE (l14_value)
        target_price_2025 if target_price_2025 is not None else "",  # Columna 5: TARGET REG
        col6_value,  # Columna 6: UPSIDE REG
        ticker_empresa,  # Columna 7: TICKER
        shared_link,  # Columna 8: LINK
        currency,  # Columna 9: CURRENCY
        pais,  # Columna 10: COUNTRY
        exchange_rate,  # Columna 11: CHANGE
        tipo_empresa,  # Columna 12: BUSSINES
        market_cap,  # Columna 13: MARKET CAP
        col14_value,  # Columna 14: NET DEBT GROWTH
        col15_value,  # Columna 15: ASSETS GROWTH
        col16_value,  # Columna 16: PP GROWTH
        col17_value,  # Columna 17: EBITDA GROWTH
        "02/11/2025"  # Columna 18: DATE
    ]
    
    print(f"[LOG] Datos preparados para escritura: {values}")
    print(f"[LOG] Escribiendo datos en fila {row_index} usando update (batch)...")
    
    # Usar update para escribir todos los datos de una vez (más eficiente y confiable)
    range_name = f'A{row_index}:R{row_index}'
    print(f"[LOG] Rango de actualización: {range_name}")
    
    print(f"[LOG] Ejecutando update en Google Sheets...")
    sheet.update(range_name, [values], value_input_option='USER_ENTERED')
    print(f"[LOG] update completado correctamente")
    
    # Actualizar celda con fórmula de hipervínculo por separado (las fórmulas necesitan formato especial)
    print(f"[LOG] Actualizando celda A{row_index} con fórmula de hipervínculo...")
    sheet.update_cell(row_index, 1, formula_link)
    print(f"[LOG] Fórmula de hipervínculo actualizada correctamente")
    
    print(f"[LOG] Verificando datos escritos...")
    # Verificar que los datos se escribieron correctamente
    written_values = sheet.row_values(row_index)
    print(f"[LOG] Datos escritos en fila {row_index}: {written_values}")
    
    print(f"[LOG] ✓ PROCESO COMPLETADO: Datos actualizados en Google Sheets")
    print(f"[LOG] Stock Code: {stock_code}")
    print(f"[LOG] Stock Name: {stock_name_clean}")
    print(f"[LOG] DCF NEW (x5): {col3_value}")
    print(f"[LOG] DCF VALUE (l14): {col4_value}")
    print(f"[LOG] Target Price 2025: {target_price_2025}")
    print(f"[LOG] Upside 2025 (%): {col6_value}")
    print(f"[LOG] Ticker: {ticker_empresa}")
    print(f"[LOG] Link: {shared_link}")
    print(f"[LOG] Currency: {currency}")
    print(f"[LOG] Country: {pais}")
    print(f"[LOG] Exchange Rate: {exchange_rate}")
    print(f"[LOG] Business Type: {tipo_empresa}")
    print(f"[LOG] Market Cap: {market_cap}")
    print(f"[LOG] Net Debt Growth: {col14_value}")
    print(f"[LOG] Assets Growth: {col15_value}")
    print(f"[LOG] PP Growth: {col16_value}")
    print(f"[LOG] EBITDA Growth: {col17_value}")
    print(f"[LOG] ===== PROCESO COMPLETADO EXITOSAMENTE PARA {stock_code} =====")
    


def process_all_stock_codes(file_path, forecast_years, x2):
    stock_codes = pd.read_excel(file_path, usecols=[0], header=None).iloc[:, 0].tolist()
    drive = setup_drive() 

    with ThreadPoolExecutor(max_workers=1) as executor:  # Reducir a 1 worker para evitar demasiadas peticiones simultáneas
        futures = []
        for code in stock_codes:
            future = executor.submit(process_stock_code, code, forecast_years, x2, drive)
            futures.append(future)
            time.sleep(30)  # Aumentar el retraso a 30 segundos para evitar el error "Too Many Requests"

        # Esperar a que se completen todas las tareas
        for future in futures:
            try:
                future.result()
            except Exception as e:
                print(f"Error en el procesamiento del código de stock: {e}")

    gc.collect()




file_path = r'C:\Users\relim\Desktop\prueba\StockCodes.xlsx'
forecast_years = 11
x2 = 1

process_all_stock_codes(file_path, forecast_years, x2)